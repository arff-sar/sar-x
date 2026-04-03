import io
import base64
import hmac
import json
import mimetypes
import re
import secrets
import zipfile
from pathlib import Path
from urllib.parse import unquote, urlparse, urlunsplit
from datetime import datetime, timedelta

import pandas as pd
import sqlalchemy as sa
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from flask import Blueprint, abort, current_app, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required
from reportlab.rl_config import TTFSearchPath
from xhtml2pdf import pisa

from extensions import (
    audit_log,
    column_exists,
    create_approval_request,
    create_notification,
    create_notification_once,
    compact_log_detail,
    db,
    guvenli_metin,
    is_allowed_file,
    is_allowed_mime,
    limiter,
    log_kaydet,
    safe_display_filename,
    secure_upload_filename,
    shorten_external_reference,
)
from models import (
    ApprovalRequest,
    AssignmentHistoryEntry,
    AssignmentItem,
    AssignmentRecipient,
    AssignmentRecord,
    AssetSparePartLink,
    AssetOperationalState,
    AssetMeterReading,
    BakimKaydi,
    CalibrationRecord,
    CalibrationSchedule,
    ConsumableItem,
    ConsumableStockMovement,
    DemoSeedRecord,
    EquipmentTemplate,
    Havalimani,
    InventoryAsset,
    InventoryBulkImportJob,
    InventoryBulkImportRowResult,
    InventoryCategory,
    IslemLog,
    Kutu,
    Kullanici,
    MaintenanceFormTemplate,
    MaintenanceHistory,
    MaintenanceInstruction,
    MaintenancePlan,
    Malzeme,
    MaintenanceTriggerRule,
    PPERecord,
    PPERecordEvent,
    PPEAssignmentRecord,
    PPEAssignmentItem,
    SparePart,
    SparePartStock,
    TatbikatBelgesi,
    TR_TZ,
    Notification,
    WorkOrder,
    get_tr_now,
)
from extensions import table_exists
from qr_logic import generate_qr_data
from decorators import (
    CANONICAL_ROLE_SYSTEM,
    CANONICAL_ROLE_TEAM_LEAD,
    CANONICAL_ROLE_TEAM_MEMBER,
    get_effective_role,
    has_permission,
    permission_required,
)
from google_drive_service import GoogleDriveError, get_drill_drive_service
from reporting import build_dashboard_kpis
from storage import get_storage_adapter
from demo_data import apply_platform_demo_scope, platform_demo_is_active
from services.inventory_bulk_import_service import (
    normalize_lookup,
    normalize_person_name,
    parse_flexible_bool,
    parse_flexible_date,
    to_int,
)
from services.inventory_code_service import generate_inventory_code
from services.inventory_excel_service import (
    ExcelTemplateError,
    build_inventory_template_workbook,
    parse_inventory_workbook,
)
from services.inventory_template_service import form_label_map
from services.qr_service import assign_asset_qr
from services.text_normalization_service import turkish_contains, turkish_equals, turkish_upper


inventory_bp = Blueprint("inventory", __name__)
OFFLINE_MAINTENANCE_REQUEST_ID_RE = re.compile(r"^[A-Za-z0-9._:-]{8,64}$")
GOOGLE_DRIVE_OAUTH_STATE_SESSION_KEY = "google_drive_oauth_state"

ASSIGNMENT_STATUS_LABELS = {
    "active": "Aktif",
    "returned": "İade Edildi",
    "partial": "Kısmi İade",
    "cancelled": "İptal",
}
ASSIGNMENT_STATUS_ALIASES = {
    "partially_returned": "partial",
    "partial_returned": "partial",
    "kismi_iade": "partial",
}
PPE_ASSIGNMENT_STATUS_LABELS = {
    "active": "Aktif",
    "returned": "İade Edildi",
    "cancelled": "İptal",
}
PPE_STATUS_LABELS = {
    "aktif": "Aktif",
    "eksik": "Eksik",
    "hasarli": "Hasarlı",
    "kayip": "Kayıp",
    "kullanim_disi": "Kullanım Dışı",
    "degisim_talebi": "Değişim Talebi",
}
PPE_PHYSICAL_CONDITION_LABELS = {
    "yeni": "Yeni",
    "iyi": "İyi",
    "hasarli": "Hasarlı",
    "bakim_gerektiriyor": "Bakım Gerektiriyor",
}
PPE_CATEGORY_OPTIONS = {
    "Baş ve Yüz Koruması": [
        "Baret",
        "Aydınlatmalı Baret",
        "Aydınlatmasız Baret",
        "Koruyucu Gözlük",
        "Vizör",
    ],
    "Solunum Koruması": [
        "Toz Maskesi",
        "Yarım Yüz Gaz Maskesi",
        "Tam Yüz Gaz Maskesi",
        "Filtre",
    ],
    "Vücut Koruması": [
        "Reflektif Yelek",
        "Operasyon Tulumu",
        "Yağmurluk",
        "Isı Yalıtımlı Mont",
    ],
    "El Koruması": [
        "Mekanik Risk Eldiveni",
        "Kimyasal Eldiven",
        "Elektrik Eldiveni",
    ],
    "Ayak Koruması": [
        "Çelik Burunlu İş Botu",
        "Çelik Tabanlı İş Botu",
        "Çizme",
    ],
    "Yüksekte Çalışma ve Özel Donanım": [
        "Emniyet Kemeri / Harness",
        "Karabina",
        "Baret İçi Kulaklık",
        "Kafa Lambası",
    ],
}
PPE_APPAREL_SIZES = ["XS", "S", "M", "L", "XL", "XXL", "3XL"]
PPE_SHOE_SIZES = [str(size) for size in range(36, 49)]
PPE_APPAREL_CATEGORIES = {"Vücut Koruması", "El Koruması"}
PPE_SHOE_CATEGORIES = {"Ayak Koruması"}
PPE_EXPIRY_WARNING_DAYS = 30
SIGNED_ASSIGNMENT_ALLOWED_EXTENSIONS = {"pdf"}
PPE_ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "webp"}
DRILL_ALLOWED_EXTENSIONS = {"rar", "zip", "7z"}
DRILL_ALLOWED_MIME_TYPES = (
    "application/zip",
    "application/x-zip-compressed",
    "application/x-7z-compressed",
    "application/x-rar-compressed",
    "application/vnd.rar",
    "application/octet-stream",
)
ASSET_STATUS_ACTIVE = "aktif"
ASSET_STATUS_PASSIVE = "pasif"
ASSET_STATUS_VALUES = {ASSET_STATUS_ACTIVE, ASSET_STATUS_PASSIVE}
MAINTENANCE_MONTH_VALUES = list(range(1, 13))
CALIBRATION_PDF_MAX_BYTES = 15 * 1024 * 1024
INVENTORY_CATEGORY_OPTIONS = [
    "Elektronik",
    "Mekanik",
    "Hidrolik",
    "Kurtarma",
    "Koruyucu Donanım",
    "Haberleşme",
    "Aydınlatma",
    "Diğer",
]
_MISSING = object()


def havalimani_filtreli_sorgu(model_sinifi):
    if _can_view_all_operational_scope():
        query = model_sinifi.query.filter_by(is_deleted=False)
    else:
        query = model_sinifi.query.filter_by(havalimani_id=current_user.havalimani_id, is_deleted=False)
    if hasattr(model_sinifi, "id"):
        query = apply_platform_demo_scope(query, model_sinifi.__name__, model_sinifi.id)
    return query


def _can_view_all_operational_scope():
    actor_role = get_effective_role(current_user)
    if actor_role == CANONICAL_ROLE_SYSTEM:
        return True
    if actor_role in {CANONICAL_ROLE_TEAM_LEAD, CANONICAL_ROLE_TEAM_MEMBER}:
        return False
    return has_permission("logs.view") or has_permission("settings.manage")


def _visible_operational_airports():
    if _can_view_all_operational_scope():
        query = Havalimani.query.filter_by(is_deleted=False)
    else:
        query = Havalimani.query.filter_by(
            is_deleted=False,
            id=current_user.havalimani_id,
        )
    query = apply_platform_demo_scope(query, "Havalimani", Havalimani.id)
    return query.order_by(Havalimani.kodu.asc()).all()


def _visible_personnel_query(airport_id=None):
    query = Kullanici.query.filter_by(is_deleted=False)
    if _can_view_all_operational_scope():
        if airport_id == "global":
            scoped = query.filter(Kullanici.havalimani_id.is_(None))
            return apply_platform_demo_scope(scoped, "Kullanici", Kullanici.id)
        if airport_id:
            scoped = query.filter(Kullanici.havalimani_id == airport_id)
            return apply_platform_demo_scope(scoped, "Kullanici", Kullanici.id)
        return apply_platform_demo_scope(query, "Kullanici", Kullanici.id)
    if current_user.havalimani_id is None:
        scoped = query.filter(Kullanici.havalimani_id.is_(None))
        return apply_platform_demo_scope(scoped, "Kullanici", Kullanici.id)
    scoped = query.filter(Kullanici.havalimani_id == current_user.havalimani_id)
    return apply_platform_demo_scope(scoped, "Kullanici", Kullanici.id)


def _visible_material_query(airport_id=None):
    query = Malzeme.query.filter_by(is_deleted=False)
    if _can_view_all_operational_scope():
        if airport_id:
            scoped = query.filter(Malzeme.havalimani_id == airport_id)
            return apply_platform_demo_scope(scoped, "Malzeme", Malzeme.id)
        return apply_platform_demo_scope(query, "Malzeme", Malzeme.id)
    scoped = query.filter(Malzeme.havalimani_id == current_user.havalimani_id)
    return apply_platform_demo_scope(scoped, "Malzeme", Malzeme.id)


def _can_issue_assignments(actor=None):
    actor = actor or current_user
    return bool(
        getattr(actor, "is_authenticated", False)
        and get_effective_role(actor) in {CANONICAL_ROLE_SYSTEM, CANONICAL_ROLE_TEAM_LEAD}
    )


def _can_delete_assignments():
    if not getattr(current_user, "is_authenticated", False):
        return False
    actor_role = get_effective_role(current_user)
    if actor_role == CANONICAL_ROLE_SYSTEM:
        return has_permission("assignment.view")
    if actor_role == CANONICAL_ROLE_TEAM_LEAD:
        return has_permission("assignment.manage")
    return False


def _assignment_scope():
    query = AssignmentRecord.query.filter_by(is_deleted=False)
    if _can_view_all_operational_scope():
        return apply_platform_demo_scope(query, "AssignmentRecord", AssignmentRecord.id)
    if has_permission("assignment.manage") or has_permission("assignment.create"):
        scoped = query.filter(AssignmentRecord.airport_id == current_user.havalimani_id)
        return apply_platform_demo_scope(scoped, "AssignmentRecord", AssignmentRecord.id)
    scoped = query.join(AssignmentRecipient).filter(AssignmentRecipient.user_id == current_user.id).distinct()
    return apply_platform_demo_scope(scoped, "AssignmentRecord", AssignmentRecord.id)


def _can_issue_ppe_assignments(actor=None):
    actor = actor or current_user
    return bool(
        getattr(actor, "is_authenticated", False)
        and has_permission("ppe.manage", user=actor)
        and get_effective_role(actor) in {CANONICAL_ROLE_SYSTEM, CANONICAL_ROLE_TEAM_LEAD}
    )


def _ppe_assignment_scope():
    query = PPEAssignmentRecord.query.filter_by(is_deleted=False)
    if _can_view_all_operational_scope():
        return apply_platform_demo_scope(query, "PPEAssignmentRecord", PPEAssignmentRecord.id)
    if _can_issue_ppe_assignments(current_user):
        scoped = query.filter(PPEAssignmentRecord.airport_id == current_user.havalimani_id)
        return apply_platform_demo_scope(scoped, "PPEAssignmentRecord", PPEAssignmentRecord.id)
    scoped = query.filter(PPEAssignmentRecord.recipient_user_id == current_user.id)
    return apply_platform_demo_scope(scoped, "PPEAssignmentRecord", PPEAssignmentRecord.id)


def _ppe_linkable_assignments(airport_id=None):
    query = _ppe_assignment_scope().filter(PPEAssignmentRecord.status == "active")
    if airport_id:
        query = query.filter(PPEAssignmentRecord.airport_id == airport_id)
    return (
        query.options(
            joinedload(PPEAssignmentRecord.recipient_user),
            joinedload(PPEAssignmentRecord.airport),
        )
        .order_by(PPEAssignmentRecord.assignment_date.desc(), PPEAssignmentRecord.created_at.desc())
        .all()
    )


def _ppe_scope():
    query = PPERecord.query.filter_by(is_deleted=False)
    if _can_view_all_operational_scope():
        return apply_platform_demo_scope(query, "PPERecord", PPERecord.id)
    if has_permission("ppe.manage"):
        scoped = query.filter(PPERecord.airport_id == current_user.havalimani_id)
        return apply_platform_demo_scope(scoped, "PPERecord", PPERecord.id)
    scoped = query.filter(PPERecord.user_id == current_user.id)
    return apply_platform_demo_scope(scoped, "PPERecord", PPERecord.id)


def _ensure_kkd_schema_ready():
    missing_parts = []
    if not table_exists("ppe_record"):
        missing_parts.append("ppe_record tablosu")
    elif not column_exists("ppe_record", "ppe_assignment_id"):
        missing_parts.append("ppe_record.ppe_assignment_id kolonu")

    if not table_exists("ppe_assignment_record"):
        missing_parts.append("ppe_assignment_record tablosu")
    elif any(
        not column_exists("ppe_assignment_record", column_name)
        for column_name in ("returned_at", "returned_by_id", "returned_note")
    ):
        missing_parts.append("ppe_assignment_record iade kolonları")
    if not table_exists("ppe_assignment_item"):
        missing_parts.append("ppe_assignment_item tablosu")

    if not missing_parts:
        return

    detail = ", ".join(missing_parts)
    raise RuntimeError(
        f"KKD şeması güncel değil ({detail}). Lütfen `flask db upgrade` çalıştırın."
    )


def _drill_scope():
    query = TatbikatBelgesi.query.filter_by(is_deleted=False)
    if get_effective_role(current_user) == CANONICAL_ROLE_SYSTEM:
        return apply_platform_demo_scope(query, "TatbikatBelgesi", TatbikatBelgesi.id)
    if current_user.havalimani_id is None:
        scoped = query.filter(TatbikatBelgesi.havalimani_id.is_(None))
        return apply_platform_demo_scope(scoped, "TatbikatBelgesi", TatbikatBelgesi.id)
    scoped = query.filter(TatbikatBelgesi.havalimani_id == current_user.havalimani_id)
    return apply_platform_demo_scope(scoped, "TatbikatBelgesi", TatbikatBelgesi.id)


def _can_view_drills_for_airport(airport_id):
    if get_effective_role(current_user) == CANONICAL_ROLE_SYSTEM:
        return True
    return bool(current_user.havalimani_id and current_user.havalimani_id == airport_id)


def _can_manage_drills_for_airport(airport_id):
    if get_effective_role(current_user) == CANONICAL_ROLE_SYSTEM:
        return True
    return bool(
        get_effective_role(current_user) == CANONICAL_ROLE_TEAM_LEAD
        and current_user.havalimani_id
        and current_user.havalimani_id == airport_id
    )


def _visible_drill_airports():
    if get_effective_role(current_user) == CANONICAL_ROLE_SYSTEM:
        query = Havalimani.query.filter_by(is_deleted=False)
        query = apply_platform_demo_scope(query, "Havalimani", Havalimani.id)
        return query.order_by(Havalimani.kodu.asc()).all()
    if current_user.havalimani_id is None:
        return []
    query = Havalimani.query.filter_by(
        is_deleted=False,
        id=current_user.havalimani_id,
    )
    query = apply_platform_demo_scope(query, "Havalimani", Havalimani.id)
    return query.order_by(Havalimani.kodu.asc()).all()


def _drill_file_size(upload):
    stream = getattr(upload, "stream", None)
    if stream is None:
        return int(getattr(upload, "content_length", 0) or 0)
    try:
        position = stream.tell()
    except Exception:
        position = 0
    try:
        stream.seek(0, io.SEEK_END)
        size = stream.tell()
    except Exception:
        size = int(getattr(upload, "content_length", 0) or 0)
    finally:
        try:
            stream.seek(position)
        except Exception:
            pass
    return int(size or 0)


def _detect_drill_archive_signature(upload):
    stream = getattr(upload, "stream", None)
    if stream is None:
        return None

    try:
        position = stream.tell()
    except Exception:
        position = None

    try:
        if position is not None:
            stream.seek(0)
        header = stream.read(8) or b""
    except Exception:
        header = b""
    finally:
        try:
            if position is not None:
                stream.seek(position)
            else:
                stream.seek(0)
        except Exception:
            pass

    zip_signatures = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")
    if any(header.startswith(signature) for signature in zip_signatures):
        return "zip"
    if header.startswith(b"Rar!\x1a\x07\x00") or header.startswith(b"Rar!\x1a\x07\x01\x00"):
        return "rar"
    if header.startswith(b"7z\xbc\xaf\x27\x1c"):
        return "7z"
    return None


def _validate_drill_upload(upload):
    if not upload or not upload.filename:
        return None, None, "Yüklenecek dosya seçilmedi."
    safe_name = secure_upload_filename(upload.filename)
    if not safe_name:
        return None, None, "Dosya adı güvenli hale getirilemedi."
    if not is_allowed_file(safe_name, DRILL_ALLOWED_EXTENSIONS):
        return None, None, "Sadece RAR, ZIP veya 7Z arşiv dosyası yükleyebilirsiniz."
    extension = safe_name.rsplit(".", 1)[1].lower()

    file_size = _drill_file_size(upload)
    if file_size <= 0:
        return None, None, "Boş dosya yüklenemez."
    max_bytes = int(current_app.config.get("DRILL_MAX_FILE_SIZE") or current_app.config.get("MAX_CONTENT_LENGTH") or 0)
    if max_bytes and file_size > max_bytes:
        max_mb = max(1, int(max_bytes / (1024 * 1024)))
        return None, None, f"Tatbikat dosyası en fazla {max_mb} MB olabilir."

    mime_type = getattr(upload, "mimetype", None) or mimetypes.guess_type(safe_name)[0] or "application/octet-stream"
    if not any(mime_type.startswith(prefix) for prefix in DRILL_ALLOWED_MIME_TYPES):
        return None, None, "Arşiv dosyası türü doğrulanamadı. RAR, ZIP veya 7Z yükleyin."
    detected_archive_type = _detect_drill_archive_signature(upload)
    if detected_archive_type != extension:
        return None, None, "Arşiv dosyası türü doğrulanamadı. RAR, ZIP veya 7Z yükleyin."
    return safe_name, mime_type, None


def _build_drill_storage_filename(drill_date, safe_name):
    extension = Path(safe_name).suffix.lower()
    candidate = secure_upload_filename(f"{drill_date.strftime('%d.%m.%Y')}_tatbikat{extension}")
    if not candidate or not is_allowed_file(candidate, DRILL_ALLOWED_EXTENSIONS):
        raise ValueError("Tatbikat dosya adı güvenli biçimde oluşturulamadı.")
    return candidate


def _format_drill_file_size(size):
    amount = float(size or 0)
    for unit in ("B", "KB", "MB", "GB"):
        if amount < 1024 or unit == "GB":
            if unit == "B":
                return f"{int(amount)} {unit}"
            return f"{amount:.1f} {unit}"
        amount /= 1024
    return "0 B"


def _get_drill_document_or_403(document_id):
    document = TatbikatBelgesi.query.filter_by(id=document_id, is_deleted=False).first_or_404()
    if not _can_view_drills_for_airport(document.havalimani_id):
        abort(403)
    return document


def _google_drive_oauth_state_matches(request_state):
    expected_state = str(session.pop(GOOGLE_DRIVE_OAUTH_STATE_SESSION_KEY, "") or "").strip()
    provided_state = str(request_state or "").strip()
    if not expected_state or not provided_state:
        return False
    return hmac.compare_digest(expected_state, provided_state)


def _redirect_after_google_oauth():
    if current_user.is_authenticated:
        if getattr(current_user, "is_sahip", False) or has_permission("settings.manage"):
            return redirect(url_for("admin.site_yonetimi", tab="genel"))
        if has_permission("drill.view"):
            return redirect(url_for("inventory.tatbikat_listesi"))
    return redirect(url_for("auth.login"))


def _next_assignment_no():
    now = get_tr_now()
    return f"ZMT-{now.strftime('%Y%m%d%H%M%S')}-{now.microsecond % 1000:03d}"


def _next_ppe_assignment_no():
    now = get_tr_now()
    return f"KKD-{now.strftime('%Y%m%d%H%M%S')}-{now.microsecond % 1000:03d}"


def _storage_safe_segment(raw_value, default):
    safe_value = secure_upload_filename(guvenli_metin(raw_value or "").strip().replace(" ", "_"))
    return safe_value or default


def _signed_document_folder_for_person(airport_label, person_label):
    airport_folder = _storage_safe_segment(airport_label, "global")
    person_folder = _storage_safe_segment(person_label, "personel")
    return f"{airport_folder}/zimmet/{person_folder}"


def _signed_document_filename_for_person(person_label):
    person_slug = _storage_safe_segment(person_label, "personel").lower()
    timestamp = get_tr_now().strftime("%Y%m%d%H%M%S")
    return secure_upload_filename(f"kkd_{person_slug}_zimmet_{timestamp}.pdf")


def _assignment_primary_person_name(assignment):
    names = [
        recipient.user.tam_ad
        for recipient in getattr(assignment, "recipients", [])
        if getattr(recipient, "user", None) and not getattr(recipient.user, "is_deleted", False)
    ]
    if names:
        return names[0]
    delivered_by = getattr(getattr(assignment, "delivered_by", None), "tam_ad", "") or getattr(assignment, "delivered_by_name", "")
    return delivered_by or "personel"


def _assignment_signed_document_folder(assignment):
    airport_label = getattr(getattr(assignment, "airport", None), "kodu", "") or getattr(getattr(assignment, "airport", None), "ad", "") or "global"
    person_label = _assignment_primary_person_name(assignment)
    return _signed_document_folder_for_person(airport_label, person_label)


def _assignment_signed_document_filename(assignment):
    return _signed_document_filename_for_person(_assignment_primary_person_name(assignment))


def _parse_int_list(values):
    rows = []
    for value in values or []:
        try:
            rows.append(int(value))
        except (TypeError, ValueError):
            continue
    return rows


def _append_assignment_history(assignment, event_type, note):
    db.session.add(
        AssignmentHistoryEntry(
            assignment_id=assignment.id,
            event_type=event_type,
            event_note=note,
            created_by_id=current_user.id if current_user.is_authenticated else None,
        )
    )


def _register_assignment_for_active_demo_scope(assignment):
    if not assignment or not getattr(assignment, "id", None):
        return
    if not platform_demo_is_active() or not table_exists("demo_seed_record"):
        return
    existing = DemoSeedRecord.query.filter_by(
        seed_tag="demo_seed",
        model_name="AssignmentRecord",
        record_id=assignment.id,
    ).first()
    if existing:
        return
    db.session.add(
        DemoSeedRecord(
            seed_tag="demo_seed",
            model_name="AssignmentRecord",
            record_id=assignment.id,
            record_label=getattr(assignment, "assignment_no", None),
        )
    )


def _register_ppe_record_for_active_demo_scope(record):
    if not record or not getattr(record, "id", None):
        return
    if not platform_demo_is_active() or not table_exists("demo_seed_record"):
        return
    existing = DemoSeedRecord.query.filter_by(
        seed_tag="demo_seed",
        model_name="PPERecord",
        record_id=record.id,
    ).first()
    if existing:
        return
    db.session.add(
        DemoSeedRecord(
            seed_tag="demo_seed",
            model_name="PPERecord",
            record_id=record.id,
            record_label=getattr(record, "item_name", None),
        )
    )


def _register_ppe_assignment_for_active_demo_scope(assignment):
    if not assignment or not getattr(assignment, "id", None):
        return
    if not platform_demo_is_active() or not table_exists("demo_seed_record"):
        return
    existing = DemoSeedRecord.query.filter_by(
        seed_tag="demo_seed",
        model_name="PPEAssignmentRecord",
        record_id=assignment.id,
    ).first()
    if existing:
        return
    db.session.add(
        DemoSeedRecord(
            seed_tag="demo_seed",
            model_name="PPEAssignmentRecord",
            record_id=assignment.id,
            record_label=getattr(assignment, "assignment_no", None),
        )
    )


def _recalculate_assignment_status(assignment):
    if assignment.status == "cancelled":
        return assignment.status
    items = [item for item in assignment.items if not item.is_deleted]
    if not items:
        assignment.status = "cancelled"
        return assignment.status
    if all((item.remaining_quantity or 0) <= 0 for item in items):
        assignment.status = "returned"
    elif any(float(item.returned_quantity or 0) > 0 for item in items):
        assignment.status = "partial"
    else:
        assignment.status = "active"
    return assignment.status


def _assignment_status_label(value):
    normalized = _normalize_assignment_status(value)
    return ASSIGNMENT_STATUS_LABELS.get(normalized, value)


def _normalize_assignment_status(value):
    normalized = str(value or "").strip().lower()
    if not normalized:
        return ""
    return ASSIGNMENT_STATUS_ALIASES.get(normalized, normalized)


def _assignment_status_query_values(value):
    normalized = _normalize_assignment_status(value)
    if not normalized:
        return ()
    if normalized == "partial":
        return ("partial", "partially_returned", "partial_returned", "kismi_iade")
    return (normalized,)


def _work_order_status_label(value):
    labels = {
        "acik": "Açık",
        "atandi": "Atandı",
        "islemde": "İşlemde",
        "beklemede_parca": "Parça Bekleniyor",
        "beklemede_onay": "Onay Bekleniyor",
        "tamamlandi": "Tamamlandı",
        "iptal_edildi": "İptal Edildi",
    }
    return labels.get(value, value)


def _format_assignment_quantity(value):
    try:
        numeric_value = float(value or 0)
    except (TypeError, ValueError):
        return "0"
    if abs(numeric_value) < 1e-9:
        return "0"
    rounded_value = round(numeric_value)
    if abs(numeric_value - rounded_value) < 1e-9:
        return str(int(rounded_value))
    return f"{numeric_value:.2f}".rstrip("0").rstrip(".")


def _ppe_assignment_status_label(value):
    return PPE_ASSIGNMENT_STATUS_LABELS.get(value, value)


def _ppe_assignment_display_name(value):
    return turkish_upper(guvenli_metin(value or "").strip())


def _to_float_safe(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _ppe_available_quantity_map(airport_id=None):
    ppe_query = _ppe_scope().filter(PPERecord.is_active.is_(True))
    if airport_id:
        ppe_query = ppe_query.filter(PPERecord.airport_id == airport_id)
    ppe_rows = ppe_query.all()
    if not ppe_rows:
        return {}, {}

    ppe_rows_by_id = {row.id: row for row in ppe_rows}
    assignment_rows = (
        db.session.query(
            PPEAssignmentItem.ppe_record_id,
            sa.func.coalesce(sa.func.sum(PPEAssignmentItem.quantity), 0.0),
        )
        .join(PPEAssignmentRecord, PPEAssignmentRecord.id == PPEAssignmentItem.assignment_id)
        .filter(
            PPEAssignmentItem.is_deleted.is_(False),
            PPEAssignmentRecord.is_deleted.is_(False),
            PPEAssignmentRecord.status == "active",
            PPEAssignmentItem.ppe_record_id.in_(list(ppe_rows_by_id.keys())),
        )
        .group_by(PPEAssignmentItem.ppe_record_id)
        .all()
    )
    assigned_map = {row[0]: _to_float_safe(row[1]) for row in assignment_rows}
    available_map = {}
    for record_id, record in ppe_rows_by_id.items():
        available_qty = max(_to_float_safe(record.quantity, 0) - assigned_map.get(record_id, 0.0), 0.0)
        available_map[record_id] = available_qty
    return ppe_rows_by_id, available_map


def _ppe_assignment_signed_document_folder(assignment):
    airport_label = (
        getattr(getattr(assignment, "airport", None), "kodu", "")
        or getattr(getattr(assignment, "airport", None), "ad", "")
        or "global"
    )
    recipient_name = getattr(getattr(assignment, "recipient_user", None), "tam_ad", "") or "personel"
    return f"{_storage_safe_segment(airport_label, 'global')}/KKD/{_storage_safe_segment(recipient_name, 'personel')}"


def _ppe_assignment_signed_document_filename(assignment):
    person_slug = _storage_safe_segment(
        getattr(getattr(assignment, "recipient_user", None), "tam_ad", "") or "personel",
        "personel",
    ).lower()
    timestamp = get_tr_now().strftime("%Y%m%d")
    return secure_upload_filename(f"kkd_{person_slug}_{timestamp}.pdf")


def _upload_ppe_signed_document_to_drive(assignment, upload, safe_name):
    if not table_exists("havalimani"):
        return None
    airport = getattr(assignment, "airport", None)
    if not airport:
        return None
    service = get_drill_drive_service()
    airport_folder_id = service.ensure_airport_folder(airport)
    kkd_folder_id = service._find_folder("KKD", airport_folder_id) or service._create_folder("KKD", airport_folder_id)
    person_slug = _storage_safe_segment(
        getattr(getattr(assignment, "recipient_user", None), "tam_ad", "") or "personel",
        "personel",
    ).lower()
    date_str = get_tr_now().strftime("%Y%m%d")
    drive_filename = secure_upload_filename(f"kkd_{person_slug}_{date_str}.pdf") or safe_name
    upload_result = service.upload_file_to_folder(
        folder_id=kkd_folder_id,
        upload=upload,
        filename=drive_filename,
        mime_type="application/pdf",
    )
    return {
        "drive_file_id": upload_result.get("drive_file_id"),
        "drive_folder_id": kkd_folder_id,
    }


def _ppe_status_label(value):
    return PPE_STATUS_LABELS.get(value, value)


def _ppe_condition_label(value):
    return PPE_PHYSICAL_CONDITION_LABELS.get(value, value)


def _ppe_category_options():
    return PPE_CATEGORY_OPTIONS


def _ppe_subtype_options(category):
    return list(_ppe_category_options().get(category, []))


def _ppe_requires_apparel_size(category, subcategory):
    return category in PPE_APPAREL_CATEGORIES or subcategory in {
        "Reflektif Yelek",
        "Operasyon Tulumu",
        "Yağmurluk",
        "Isı Yalıtımlı Mont",
        "Mekanik Risk Eldiveni",
        "Kimyasal Eldiven",
        "Elektrik Eldiveni",
    }


def _ppe_requires_shoe_size(category, subcategory):
    return category in PPE_SHOE_CATEGORIES or subcategory in {
        "Çelik Burunlu İş Botu",
        "Çelik Tabanlı İş Botu",
        "Çizme",
    }


def _ppe_combined_size(apparel_size, shoe_size, fallback=None):
    if shoe_size:
        return shoe_size
    if apparel_size:
        return apparel_size
    return guvenli_metin(fallback or "").strip()


def _ppe_brand_model_display(record):
    return " ".join(part for part in [record.brand, record.model_name] if part).strip() or (record.brand_model or "-")


def _ppe_size_display(record):
    return record.size_display or "-"


def _ppe_alert_state(record):
    today = get_tr_now().date()
    expiry_date = getattr(record, "expiry_date", None)
    if expiry_date:
        if expiry_date < today:
            return ("danger", "Süresi doldu")
        if expiry_date <= (today + timedelta(days=PPE_EXPIRY_WARNING_DAYS)):
            return ("warning", "Süresi yaklaşıyor")
    if getattr(record, "physical_condition", "") == "bakim_gerektiriyor":
        return ("warning", "Bakım gerekiyor")
    if getattr(record, "physical_condition", "") == "hasarli":
        return ("danger", "Hasarlı")
    return ("success", "Uygun")


def _ppe_import_feedback_session_key():
    return "ppe_import_feedback"


def _validate_upload(upload, allowed_extensions, allowed_mime_prefixes):
    if not upload or not upload.filename:
        return None, "Yüklenecek dosya seçilmedi."
    safe_name = secure_upload_filename(upload.filename)
    if not safe_name:
        return None, "Dosya adı güvenli hale getirilemedi."
    if not is_allowed_file(safe_name, allowed_extensions):
        return None, "Dosya uzantısı desteklenmiyor."
    if not is_allowed_mime(safe_name, allowed_mime_prefixes=allowed_mime_prefixes, upload=upload):
        return None, "Dosya türü desteklenmiyor."
    return safe_name, None


def _is_valid_xlsx_workbook(upload):
    stream = getattr(upload, "stream", None)
    if stream is None:
        return False

    try:
        position = stream.tell()
    except Exception:
        position = None

    try:
        if position is not None:
            stream.seek(0)
        with zipfile.ZipFile(stream) as workbook_archive:
            names = set(workbook_archive.namelist())
    except Exception:
        return False
    finally:
        try:
            if position is not None:
                stream.seek(position)
            else:
                stream.seek(0)
        except Exception:
            pass

    required_entries = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
    return required_entries.issubset(names)


def _pdf_link_callback(uri, _rel):
    if not uri:
        return uri
    if uri.startswith("file://"):
        return unquote(urlparse(uri).path)
    if uri.startswith("/"):
        return uri
    return uri


def _assignment_pdf_font_uris():
    search_roots = []
    for candidate in list(TTFSearchPath) + [
        "/usr/share/fonts",
        "/usr/local/share/fonts",
        "/usr/share/fonts/truetype",
        "/usr/share/fonts/truetype/dejavu",
        "/usr/share/fonts/truetype/liberation2",
        "/Library/Fonts",
        "/System/Library/Fonts",
        "/System/Library/Fonts/Supplemental",
        str(Path.home() / "Library/Fonts"),
    ]:
        path = Path(candidate)
        if path.exists() and path not in search_roots:
            search_roots.append(path)

    font_candidates = [
        ("Vera.ttf", "VeraBd.ttf"),
        ("DejaVuSans.ttf", "DejaVuSans-Bold.ttf"),
        ("Arial Unicode.ttf", "Arial Unicode.ttf"),
        ("Arial.ttf", "Arial Bold.ttf"),
    ]

    def _find_font(filename):
        for root in search_roots:
            direct = root / filename
            if direct.exists():
                return direct
            nested = list(root.rglob(filename))
            if nested:
                return nested[0]
        return None

    for regular_name, bold_name in font_candidates:
        regular = _find_font(regular_name)
        if not regular:
            continue
        bold = _find_font(bold_name) or regular
        return {
            "regular": regular.resolve().as_uri(),
            "bold": bold.resolve().as_uri(),
        }

    return {"regular": "", "bold": ""}


def _assignment_pdf_logo_uri():
    candidate_paths = [
        Path(current_app.root_path) / "static" / "img" / "arfflogo.png",
        Path(current_app.root_path) / "static" / "img" / "logo_guncell.png",
        Path(current_app.root_path) / "static" / "img" / "logo_guncel.png",
        Path(current_app.root_path) / "static" / "img" / "icon-512.png",
        Path(current_app.root_path) / "static" / "img" / "icon-192.png",
        Path(current_app.root_path) / "static" / "favicon.png",
    ]
    seen = set()
    for candidate in candidate_paths:
        try:
            resolved = candidate.resolve()
        except Exception:
            continue
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists() and resolved.is_file():
            # xhtml2pdf bazı ortamlarda file:// URI yerine doğrudan mutlak yol ile
            # daha stabil render ettiği için logo kaynağını path olarak döndürüyoruz.
            return str(resolved)
    return ""


def _asset_scope():
    query = InventoryAsset.query.filter_by(is_deleted=False)
    if _can_view_all_operational_scope():
        scoped = query
    else:
        scoped = query.filter_by(havalimani_id=current_user.havalimani_id)
    return apply_platform_demo_scope(scoped, "InventoryAsset", InventoryAsset.id)


def _asset_qr_url(asset):
    return url_for("inventory.quick_asset_view", asset_id=asset.id, _external=True)


def _asset_qr_payload(asset):
    # Persisted qr_code alanında legacy/plain değerler kalabildiği için
    # QR payload her zaman gerçek detay URL'sinden üretilir.
    return _asset_qr_url(asset)


def _asset_qr_context(asset):
    return {
        "qr_payload": _asset_qr_payload(asset),
        "asset_code": generate_inventory_code(asset),
        "airport_name": asset.qr_label_airport_name,
    }


def _box_scope(include_deleted=False):
    query = Kutu.query
    if not include_deleted:
        query = query.filter_by(is_deleted=False)
    if _can_view_all_box_scope():
        scoped = query
    else:
        scoped = query.filter_by(havalimani_id=current_user.havalimani_id)
    return apply_platform_demo_scope(scoped, "Kutu", Kutu.id)


def _can_view_all_box_scope():
    return get_effective_role(current_user) == CANONICAL_ROLE_SYSTEM


def _can_manage_box_airport(airport_id):
    actor_role = get_effective_role(current_user)
    if actor_role == CANONICAL_ROLE_SYSTEM:
        return True
    return bool(
        actor_role == CANONICAL_ROLE_TEAM_LEAD
        and current_user.havalimani_id
        and current_user.havalimani_id == airport_id
    )


def _validate_box_write_access(airport_id):
    if not _can_manage_box_airport(airport_id):
        abort(403)


def _extract_box_sequence(code, airport_code):
    if not code:
        return None
    normalized = str(code).strip().upper()
    prefix = f"{airport_code}-BOX-"
    if not normalized.startswith(prefix):
        return None
    serial_part = normalized[len(prefix):]
    if not serial_part.isdigit():
        return None
    return int(serial_part)


def _next_box_code_for_airport(airport):
    airport_code = (airport.kodu or "").strip().upper()
    if not airport_code:
        raise ValueError("Havalimanı kodu bulunamadı.")
    existing_codes = [
        row[0]
        for row in db.session.query(Kutu.kodu)
        .filter(
            Kutu.havalimani_id == airport.id,
        )
        .all()
    ]
    sequences = [seq for seq in (_extract_box_sequence(code, airport_code) for code in existing_codes) if seq is not None]
    next_serial = (max(sequences) + 1) if sequences else 1
    return f"{airport_code}-BOX-{next_serial:02d}"


def _create_box_with_generated_code(airport_id, marka=None):
    airport = db.session.get(Havalimani, airport_id)
    if not airport or airport.is_deleted:
        raise ValueError("Geçerli bir havalimanı bulunamadı.")

    normalized_brand = guvenli_metin(marka or "").strip() or None

    for _ in range(8):
        generated_code = _next_box_code_for_airport(airport)
        kutu = Kutu(
            kodu=generated_code,
            marka=normalized_brand,
            konum=generated_code,
            havalimani_id=airport.id,
        )
        db.session.add(kutu)
        try:
            db.session.flush()
            return kutu
        except IntegrityError:
            db.session.rollback()
    raise ValueError("Kutu kodu üretilemedi. Lütfen tekrar deneyin.")


def _box_qr_context(box):
    return {
        "qr_payload": _box_qr_payload(box),
        "box_code": box.qr_code_label,
        "airport_name": box.qr_label_airport_name,
    }


def _box_qr_url(box):
    return url_for("inventory.kutu_detay", kodu=box.kodu, _external=True)


def _box_qr_payload(box):
    # Asset QR ile aynı standardı korumak için kutu QR payload'u da doğrudan detay URL'si olur.
    return _box_qr_url(box)


def _sync_asset_location(material):
    if not material or not material.linked_asset:
        return
    material.linked_asset.depot_location = material.kutu.kodu if material.kutu else material.linked_asset.depot_location
    material.linked_asset.havalimani_id = material.havalimani_id
    material.linked_asset.unit_count = material.stok_miktari or material.linked_asset.unit_count


def _ensure_operational_state(asset):
    if asset.operational_state:
        return asset.operational_state
    state = AssetOperationalState(asset_id=asset.id, lifecycle_status="active")
    db.session.add(state)
    db.session.flush()
    return state


def _consumable_scope():
    query = ConsumableItem.query.filter_by(is_deleted=False, is_active=True)
    query = apply_platform_demo_scope(query, "ConsumableItem", ConsumableItem.id)
    return query.order_by(ConsumableItem.title.asc())


def _consumable_balance(consumable_id, airport_id):
    movements = ConsumableStockMovement.query.filter_by(
        consumable_id=consumable_id,
        airport_id=airport_id,
        is_deleted=False,
    ).all()
    balance = 0.0
    for movement in movements:
        sign = 1 if movement.movement_type in {"in", "adjust", "transfer"} else -1
        balance += sign * float(movement.quantity or 0)
    return round(balance, 2)


def _lifecycle_to_status(lifecycle_status):
    mapping = {
        "planned": "pasif",
        "received": "pasif",
        "active": "aktif",
        "in_maintenance": "pasif",
        "calibration_due": "pasif",
        "out_of_service": "pasif",
        "decommissioned": "pasif",
        "disposed": "pasif",
        "transferred": "aktif",
    }
    return mapping.get(lifecycle_status, "aktif")


def _create_operational_alerts():
    today = get_tr_now().date()
    asset_rows = _asset_scope().all()
    if table_exists("consumable_item") and table_exists("consumable_stock_movement"):
        for consumable in _consumable_scope().all():
            balance = _consumable_balance(consumable.id, current_user.havalimani_id) if current_user.havalimani_id else 0
            if current_user.havalimani_id and balance <= float(consumable.critical_level or 0):
                create_notification_once(
                    current_user.id,
                    "consumable_critical_stock",
                    "Kritik sarf stoğu",
                    f"{consumable.title} kritik stok seviyesine indi.",
                    link_url=url_for("inventory.consumables"),
                    severity="danger",
                )
            elif current_user.havalimani_id and balance <= float(consumable.min_stock_level or 0):
                create_notification_once(
                    current_user.id,
                    "consumable_low_stock",
                    "Düşük sarf stoğu",
                    f"{consumable.title} minimum stok seviyesine yaklaştı.",
                    link_url=url_for("inventory.consumables"),
                    severity="warning",
                )

    for asset in asset_rows:
        if asset.next_calibration_date and asset.next_calibration_date < today:
            create_notification_once(
                current_user.id,
                "calibration_overdue",
                "Geciken kalibrasyon",
                f"{asset.asset_code} için kalibrasyon gecikti.",
                link_url=url_for("inventory.calibration_records"),
                severity="warning",
            )
        elif asset.next_calibration_date and asset.next_calibration_date <= (today + timedelta(days=15)):
            create_notification_once(
                current_user.id,
                "calibration_upcoming",
                "Yaklaşan kalibrasyon",
                f"{asset.asset_code} için kalibrasyon tarihi yaklaşıyor.",
                link_url=url_for("inventory.calibration_records"),
                severity="info",
            )
        if asset.warranty_end_date and today <= asset.warranty_end_date <= (today + timedelta(days=30)):
            create_notification_once(
                current_user.id,
                "warranty_expiring",
                "Garanti bitişi yaklaşıyor",
                f"{asset.asset_code} için garanti bitiş tarihi yaklaşıyor.",
                link_url=url_for("inventory.asset_lifecycle"),
                severity="warning",
            )
        if asset.is_critical and asset.lifecycle_status == "out_of_service":
            asset_label = asset.asset_code or "Kritik ekipman"
            create_notification_once(
                current_user.id,
                "critical_out_of_service",
                "Kritik ekipman hizmet dışı",
                f"{asset_label} hizmet dışı durumda.",
                link_url=url_for("inventory.asset_lifecycle"),
                severity="danger",
            )
def _parse_date(raw_value):
    if not raw_value:
        return None
    normalized = str(raw_value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except (TypeError, ValueError):
            continue
    return None


def _parse_positive_int(raw_value, default=1):
    try:
        value = int(str(raw_value).strip())
    except (TypeError, ValueError, AttributeError):
        return default
    return max(value, 1)


def _parse_non_negative_int(raw_value, default=0):
    try:
        value = int(str(raw_value).strip())
    except (TypeError, ValueError, AttributeError):
        return default
    return max(value, 0)


def _parse_month_period(raw_value, default=6):
    value = _parse_positive_int(raw_value, default=default)
    return min(max(value, 1), 12)


def _to_bool(raw_value):
    try:
        return parse_flexible_bool(raw_value)
    except ValueError:
        return False


def _field_present(form_data, key):
    if form_data is None:
        return False
    try:
        return key in form_data
    except Exception:
        return form_data.get(key) is not None


def _first_value(form_data, *keys, default=None):
    for key in keys:
        if _field_present(form_data, key):
            return form_data.get(key)
    return default


def _canonical_asset_payload(form_data, *, mode):
    payload = {}

    payload["airport_id"] = _first_value(form_data, "airport_id", "havalimani_id")
    payload["asset_name"] = _first_value(form_data, "asset_name", "ad")
    payload["serial_no"] = _first_value(form_data, "serial_no", "seri_no")
    payload["unit_count"] = _first_value(form_data, "unit_count", "stok")
    payload["status"] = _first_value(form_data, "status", "durum")
    payload["is_demirbas"] = _first_value(form_data, "is_demirbas")
    payload["demirbas_no"] = _first_value(form_data, "demirbas_no", "asset_tag")
    payload["last_maintenance_date"] = _first_value(
        form_data,
        "last_maintenance_date",
        "son_bakim_tarihi",
        "bakim",
    )
    payload["maintenance_period_months"] = _first_value(form_data, "maintenance_period_months", "bakim_periyodu_ay")
    payload["calibration_required"] = _first_value(form_data, "calibration_required")
    payload["calibration_period_months"] = _first_value(form_data, "calibration_period_months", "calibration_periyodu_ay")
    payload["last_calibration_date"] = _first_value(form_data, "last_calibration_date")
    payload["next_calibration_date"] = _first_value(form_data, "next_calibration_date")
    payload["acquired_date"] = _first_value(form_data, "acquired_date", "edinim_tarihi")
    payload["warranty_end_date"] = _first_value(form_data, "warranty_end_date", "garanti_bitis_tarihi")
    payload["box_id"] = _first_value(form_data, "box_id", "kutu_id")
    payload["box_code"] = _first_value(form_data, "kutu_kodu")
    payload["manual_url"] = _first_value(form_data, "manual_url")
    payload["technical_specs"] = _first_value(form_data, "technical_specs", "teknik")
    payload["notes"] = _first_value(form_data, "notes", "notlar", "note")
    payload["parent_asset_id"] = _first_value(form_data, "parent_asset_id")

    payload["catalog_mode"] = "existing_template"
    if _to_bool(_first_value(form_data, "central_catalog")):
        payload["catalog_mode"] = "create_template_owner"
    elif _field_present(form_data, "use_template_mode") and not _to_bool(_first_value(form_data, "use_template_mode")):
        payload["catalog_mode"] = "create_template_owner"

    payload["template_id"] = _first_value(form_data, "template_id")
    payload["maintenance_form_template_id"] = _first_value(form_data, "maintenance_form_template_id", "bakim_formu_id")
    payload["category"] = _first_value(form_data, "category", "kategori")
    payload["brand"] = _first_value(form_data, "brand", "marka")
    payload["model_code"] = _first_value(form_data, "model_code", "model")
    payload["template_name_seed"] = _first_value(form_data, "template_name_seed", "asset_name", "ad")

    if mode in {"asset_duzenle", "quick_detail"}:
        payload["airport_id"] = None
        payload["asset_name"] = None
        payload["acquired_date"] = None
        payload["warranty_end_date"] = None
        payload["box_id"] = None
        payload["box_code"] = None
        payload["technical_specs"] = None
        payload["template_id"] = None
        payload["catalog_mode"] = None
        payload["maintenance_form_template_id"] = None
        payload["category"] = None
        payload["brand"] = None
        payload["model_code"] = None
        payload["template_name_seed"] = None
    return payload


def _months_to_days(period_months):
    months = _parse_month_period(period_months, default=6)
    return months * 30


def _is_valid_url(raw_value):
    value = str(raw_value or "").strip()
    if not value:
        return True
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def _safe_internal_redirect_target(raw_target, fallback_target):
    fallback = str(fallback_target or "/").strip() or "/"
    target = str(raw_target or "").strip()
    if not target or target.startswith("//"):
        return fallback

    parsed = urlparse(target)
    if parsed.scheme or parsed.netloc:
        request_origin = urlparse(request.host_url)
        if (parsed.scheme or "").lower() != (request_origin.scheme or "").lower():
            return fallback
        if (parsed.hostname or "").lower() != (request_origin.hostname or "").lower():
            return fallback
        try:
            request_port = request_origin.port
            parsed_port = parsed.port
        except ValueError:
            return fallback
        if (request_port or None) != (parsed_port or None):
            return fallback
        safe_path = parsed.path if str(parsed.path or "").startswith("/") else f"/{parsed.path or ''}"
        return urlunsplit(("", "", safe_path or "/", parsed.query, parsed.fragment))

    if not target.startswith("/"):
        return fallback
    return target


def _safe_asset_display_name(asset):
    base_name = (
        (asset.equipment_template.name if asset and asset.equipment_template else "")
        or (asset.legacy_material.ad if asset and asset.legacy_material else "")
        or "asset"
    )
    return secure_upload_filename(base_name).strip().lower() or "asset"


def _can_manage_asset_registry():
    return bool(getattr(current_user, "is_sahip", False) or getattr(current_user, "is_airport_manager", False))


def _resolve_allowed_categories():
    managed_rows = (
        InventoryCategory.query.filter_by(is_deleted=False, is_active=True)
        .order_by(InventoryCategory.name.asc())
        .all()
    )
    seen = [guvenli_metin(item.name or "").strip() for item in managed_rows if guvenli_metin(item.name or "").strip()]

    rows = (
        db.session.query(EquipmentTemplate.category)
        .filter(
            EquipmentTemplate.is_deleted.is_(False),
            EquipmentTemplate.category.isnot(None),
            EquipmentTemplate.category != "",
        )
        .distinct()
        .order_by(EquipmentTemplate.category.asc())
        .all()
    )
    for row in rows:
        value = guvenli_metin(row[0] or "").strip()
        if value and value not in seen:
            seen.append(value)
    for item in INVENTORY_CATEGORY_OPTIONS:
        if item not in seen:
            seen.append(item)
    return seen


def _normalize_status(status_value):
    value = str(status_value or "").strip().lower()
    if value in {"aktif", "active"}:
        return ASSET_STATUS_ACTIVE
    if value in {"pasif", "passive", "bakimda", "arizali", "hurda"}:
        return ASSET_STATUS_PASSIVE
    if value in {"bakımda", "arızalı"}:
        return ASSET_STATUS_PASSIVE
    return ASSET_STATUS_ACTIVE


def _display_status(status_value):
    mapping = {
        ASSET_STATUS_ACTIVE: "Aktif",
        ASSET_STATUS_PASSIVE: "Pasif",
    }
    return mapping.get(status_value, "Aktif")


def _status_label_tr(status_value):
    normalized = _normalize_status(status_value)
    return _display_status(normalized)


def _maintenance_label_tr(next_maintenance_date, *, today=None):
    if not next_maintenance_date:
        return "Planlanmadı"
    today = today or get_tr_now().date()
    if next_maintenance_date < today:
        return "Gecikmiş"
    if next_maintenance_date <= (today + timedelta(days=15)):
        return "Yaklaşan"
    return "Planlı"


def _sync_maintenance_plan_for_asset(asset, template=None):
    if not asset:
        return
    template = template or asset.equipment_template
    active_plan = (
        MaintenancePlan.query.filter_by(asset_id=asset.id, is_deleted=False)
        .order_by(MaintenancePlan.created_at.desc())
        .first()
    )

    if _normalize_status(asset.status) != ASSET_STATUS_ACTIVE:
        plans = MaintenancePlan.query.filter_by(asset_id=asset.id, is_deleted=False).all()
        for plan in plans:
            plan.is_active = False
        asset.next_maintenance_date = None
        if asset.legacy_material:
            asset.legacy_material.gelecek_bakim_tarihi = None
        return

    period_months = _parse_month_period(
        asset.maintenance_period_months
        or (template.maintenance_period_months if template else None)
        or max(int((asset.maintenance_period_days or (template.maintenance_period_days if template else 180) or 180) / 30), 1),
        default=6,
    )
    period_days = _months_to_days(period_months)
    asset.maintenance_period_months = period_months
    asset.maintenance_period_days = period_days
    reference_date = asset.last_maintenance_date or get_tr_now().date()

    if active_plan is None:
        active_plan = MaintenancePlan(
            name=f"{template.name if template else 'Ekipman'} Periyodik Bakım Planı",
            equipment_template_id=template.id if template else None,
            asset_id=asset.id,
            owner_airport_id=asset.havalimani_id,
            period_days=period_days,
            start_date=get_tr_now().date(),
            last_maintenance_date=asset.last_maintenance_date,
            is_active=True,
        )
        db.session.add(active_plan)
    else:
        active_plan.is_active = True
        active_plan.period_days = period_days
        active_plan.last_maintenance_date = asset.last_maintenance_date
        if template and not active_plan.equipment_template_id:
            active_plan.equipment_template_id = template.id

    active_plan.recalculate_next_due_date(reference_date)
    asset.next_maintenance_date = active_plan.next_due_date
    if asset.legacy_material:
        asset.legacy_material.gelecek_bakim_tarihi = active_plan.next_due_date


def _sync_calibration_schedule_for_asset(asset):
    if not table_exists("calibration_schedule"):
        return None
    schedule = CalibrationSchedule.query.filter_by(asset_id=asset.id, is_deleted=False).first()
    if not asset.calibration_required:
        if schedule:
            schedule.is_active = False
        return schedule
    if schedule is None:
        schedule = CalibrationSchedule(
            asset_id=asset.id,
            period_days=asset.calibration_period_days or 180,
            warning_days=15,
            provider="",
            is_active=True,
            note="",
        )
        db.session.add(schedule)
        db.session.flush()
    else:
        schedule.period_days = asset.calibration_period_days or schedule.period_days or 180
        schedule.is_active = True
    return schedule


def _validate_calibration_certificate_upload(upload):
    if not upload or not upload.filename:
        return None, "Kalibrasyon sertifikası seçilmedi."
    safe_name = secure_upload_filename(upload.filename)
    if not safe_name.lower().endswith(".pdf"):
        return None, "Kalibrasyon sertifikası sadece PDF olabilir."
    if not is_allowed_mime(safe_name, allowed_mime_prefixes=("application/pdf",), upload=upload):
        return None, "Yüklenen dosya geçerli bir PDF değil."
    stream = getattr(upload, "stream", None)
    if stream is not None:
        current_pos = stream.tell()
        stream.seek(0, io.SEEK_END)
        file_size = int(stream.tell() or 0)
        stream.seek(current_pos)
    else:
        file_size = int(getattr(upload, "content_length", 0) or 0)
    if file_size > CALIBRATION_PDF_MAX_BYTES:
        return None, "Kalibrasyon sertifikası en fazla 15 MB olabilir."
    return safe_name, None


def _upload_calibration_certificate(asset, calibration_date, upload):
    safe_name, upload_error = _validate_calibration_certificate_upload(upload)
    if upload_error:
        raise ValueError(upload_error)

    service = get_drill_drive_service()
    root_name = "Kalibrasyonlar"
    parent_folder_id = service._config("GOOGLE_DRIVE_PARENT_FOLDER_ID", "root")
    root_id = service._find_folder(root_name, parent_folder_id) or service._create_folder(root_name, parent_folder_id)
    airport_name = f"{asset.airport.kodu} - {asset.airport.ad}" if asset.airport else "Birim"
    airport_folder_id = service._find_folder(airport_name, root_id) or service._create_folder(airport_name, root_id)
    asset_folder_name = f"{asset.asset_code or 'ASSET'}-{_safe_asset_display_name(asset)}"
    asset_folder_id = service._find_folder(asset_folder_name, airport_folder_id) or service._create_folder(asset_folder_name, airport_folder_id)

    date_str = calibration_date.strftime("%Y-%m-%d")
    base_filename = f"{_safe_asset_display_name(asset)}_{date_str}.pdf"
    candidate = base_filename
    suffix = 1
    while service._request_json(
        "GET",
        f"{service.DRIVE_API_BASE}/files",
        params={
            "q": (
                f"trashed = false and "
                f"name = '{service._escape_query(candidate)}' and "
                f"'{asset_folder_id}' in parents"
            ),
            "fields": "files(id,name)",
            "pageSize": 1,
        },
    ).get("files"):
        suffix += 1
        candidate = f"{_safe_asset_display_name(asset)}_{date_str}_{suffix}.pdf"

    upload_result = service.upload_file_to_folder(
        folder_id=asset_folder_id,
        upload=upload,
        filename=candidate,
        mime_type="application/pdf",
    )
    return {
        "certificate_file": upload_result["filename"],
        "drive_file_id": upload_result["drive_file_id"],
        "drive_folder_id": asset_folder_id,
        "mime_type": "application/pdf",
        "size_bytes": upload_result.get("file_size"),
    }


def _resolve_box_from_payload(canonical_payload, airport_id, *, allow_auto_create=False):
    raw_box_id = canonical_payload.get("box_id")
    box_id = None
    try:
        if raw_box_id not in (None, ""):
            box_id = int(raw_box_id)
    except (TypeError, ValueError):
        box_id = None
    if box_id:
        box = Kutu.query.filter_by(id=box_id, is_deleted=False).first()
        if not box:
            raise ValueError("Seçilen kutu bulunamadı.")
        if airport_id and box.havalimani_id != airport_id:
            raise ValueError("Seçilen kutu, seçilen havalimanına ait değil.")
        return box

    box_code = guvenli_metin(canonical_payload.get("box_code") or "").strip().upper()
    if not box_code:
        if allow_auto_create and airport_id:
            return _ensure_fallback_box_for_airport(airport_id)
        raise ValueError("Kutu seçimi zorunludur.")
    box = Kutu.query.filter_by(kodu=box_code, havalimani_id=airport_id, is_deleted=False).first()
    if box:
        return box
    if not allow_auto_create:
        raise ValueError("Seçilen kutu bulunamadı.")
    return _ensure_box(box_code, airport_id)


def _normalized_asset_contract_values(canonical_payload, *, mode, current_asset=None):
    values = {}

    if mode == "create":
        try:
            values["airport_id"] = int(canonical_payload.get("airport_id")) if canonical_payload.get("airport_id") not in (None, "") else None
        except (TypeError, ValueError):
            values["airport_id"] = None
        values["asset_name"] = guvenli_metin(canonical_payload.get("asset_name") or "").strip()
        values["acquired_date"] = _parse_date(canonical_payload.get("acquired_date"))
        values["warranty_end_date"] = _parse_date(canonical_payload.get("warranty_end_date"))
        values["technical_specs"] = guvenli_metin(canonical_payload.get("technical_specs") or "").strip()
        values["template_id"] = int(canonical_payload.get("template_id")) if canonical_payload.get("template_id") not in (None, "") else None
        values["catalog_mode"] = (canonical_payload.get("catalog_mode") or "existing_template").strip()
        values["maintenance_form_template_id"] = (
            int(canonical_payload.get("maintenance_form_template_id"))
            if canonical_payload.get("maintenance_form_template_id") not in (None, "")
            else None
        )
        values["category"] = guvenli_metin(canonical_payload.get("category") or "").strip()
        values["brand"] = guvenli_metin(canonical_payload.get("brand") or "").strip()
        values["model_code"] = guvenli_metin(canonical_payload.get("model_code") or "").strip()
        values["template_name_seed"] = guvenli_metin(canonical_payload.get("template_name_seed") or "").strip()

    values["serial_no"] = guvenli_metin(canonical_payload.get("serial_no") or (current_asset.serial_no if current_asset else "")).strip()
    values["unit_count"] = _parse_positive_int(canonical_payload.get("unit_count"), default=(current_asset.unit_count if current_asset else 1) or 1)
    values["status"] = _normalize_status(canonical_payload.get("status") or (current_asset.status if current_asset else ASSET_STATUS_ACTIVE))
    values["is_demirbas"] = _to_bool(canonical_payload.get("is_demirbas"))
    values["demirbas_no"] = guvenli_metin(canonical_payload.get("demirbas_no") or "").strip() if values["is_demirbas"] else ""
    values["last_maintenance_date"] = _parse_date(canonical_payload.get("last_maintenance_date"))
    values["maintenance_period_months"] = _parse_month_period(
        canonical_payload.get("maintenance_period_months"),
        default=(current_asset.maintenance_period_months if current_asset else 6) or 6,
    )
    values["maintenance_period_days"] = _months_to_days(values["maintenance_period_months"])
    values["calibration_required"] = _to_bool(canonical_payload.get("calibration_required"))
    values["calibration_period_months"] = _parse_month_period(canonical_payload.get("calibration_period_months"), default=6)
    values["calibration_period_days"] = _months_to_days(values["calibration_period_months"]) if values["calibration_required"] else None
    values["last_calibration_date"] = _parse_date(canonical_payload.get("last_calibration_date")) if values["calibration_required"] else None
    values["next_calibration_date"] = _parse_date(canonical_payload.get("next_calibration_date")) if values["calibration_required"] else None
    values["manual_url"] = str(
        guvenli_metin(canonical_payload.get("manual_url") or (current_asset.manual_url if current_asset else "")) or ""
    ).strip()
    values["notes"] = guvenli_metin(canonical_payload.get("notes") or "").strip()

    raw_parent_id = canonical_payload.get("parent_asset_id")
    try:
        values["parent_asset_id"] = int(raw_parent_id) if raw_parent_id not in (None, "") else None
    except (TypeError, ValueError):
        values["parent_asset_id"] = None
    return values


def _validate_asset_contract_values(values, *, mode, current_asset=None):
    if not _is_valid_url(values.get("manual_url")):
        raise ValueError("Kullanım kılavuzu linki geçerli bir URL olmalıdır.")
    if values.get("status") not in ASSET_STATUS_VALUES:
        raise ValueError("Durum sadece aktif veya pasif olabilir.")
    if int(values.get("unit_count") or 1) < 1:
        raise ValueError("Stok birim sayısı en az 1 olmalıdır.")
    if mode in {"asset_duzenle", "quick_detail"} and current_asset and values.get("parent_asset_id") == current_asset.id:
        raise ValueError("Bir ekipman kendisine bağlanamaz.")


def _ensure_box(kutu_kodu, havalimani_id, marka=None):
    kutu = Kutu.query.filter_by(kodu=kutu_kodu, havalimani_id=havalimani_id, is_deleted=False).first()
    if kutu:
        if marka:
            kutu.marka = guvenli_metin(marka).strip()
        return kutu
    kutu = Kutu(
        kodu=kutu_kodu,
        marka=guvenli_metin(marka or "").strip() or None,
        havalimani_id=havalimani_id,
        konum=kutu_kodu,
    )
    db.session.add(kutu)
    db.session.flush()
    return kutu


def _fallback_box_code_for_airport(airport):
    airport_code = guvenli_metin(getattr(airport, "kodu", "") or "").strip().upper()
    if not airport_code:
        raise ValueError("Kutu otomatik oluşturulamadı: havalimanı kodu bulunamadı.")
    return f"{airport_code}-ATANMADI"


def _ensure_fallback_box_for_airport(airport_id):
    airport = db.session.get(Havalimani, airport_id)
    if not airport or airport.is_deleted:
        raise ValueError("Kutu otomatik oluşturulamadı: geçerli bir havalimanı bulunamadı.")
    return _ensure_box(_fallback_box_code_for_airport(airport), airport.id)


def _ensure_template_from_form(form_data, selected_template_id, *, allow_create=False):
    if selected_template_id:
        template = db.session.get(EquipmentTemplate, selected_template_id)
        if template and not template.is_deleted and template.is_active:
            return template
        return None

    if not allow_create:
        return None

    template_name = guvenli_metin(_first_value(form_data, "template_name_seed", "asset_name", "ad") or "").strip()
    if not template_name:
        return None
    category = guvenli_metin(_first_value(form_data, "category", "kategori") or "").strip()
    category = category or "Diğer"
    period_months = _parse_month_period(_first_value(form_data, "maintenance_period_months", "bakim_periyodu_ay"), default=6)

    template = EquipmentTemplate(
        name=template_name,
        category=category,
        brand=guvenli_metin(_first_value(form_data, "brand", "marka") or "").strip(),
        model_code=guvenli_metin(_first_value(form_data, "model_code", "model") or "").strip(),
        description=guvenli_metin(form_data.get("aciklama") or "").strip(),
        technical_specs=guvenli_metin(_first_value(form_data, "technical_specs", "teknik") or "").strip(),
        maintenance_period_months=period_months,
        maintenance_period_days=_months_to_days(period_months),
        criticality_level=(form_data.get("kritik_seviye") or "normal").strip(),
        default_maintenance_form_id=(
            form_data.get("maintenance_form_template_id", type=int)
            or form_data.get("bakim_formu_id", type=int)
            or None
        ),
        is_active=True,
    )
    db.session.add(template)
    db.session.flush()
    return template


def _ensure_default_equipment_template():
    fallback_name = "Genel Ekipman"
    template = (
        EquipmentTemplate.query.filter_by(
            is_deleted=False,
            is_active=True,
            name=fallback_name,
        )
        .order_by(EquipmentTemplate.id.asc())
        .first()
    )
    if template:
        return template

    template = EquipmentTemplate(
        name=fallback_name,
        category="Diğer",
        brand="",
        model_code="",
        description="Formda zorunlu alanlar boş bırakıldığında kullanılan varsayılan şablon.",
        technical_specs="",
        maintenance_period_months=6,
        maintenance_period_days=_months_to_days(6),
        criticality_level="normal",
        default_maintenance_form_id=None,
        is_active=True,
    )
    db.session.add(template)
    db.session.flush()
    return template


def _create_asset_and_legacy_material(template, kutu, havalimani_id, form_data):
    def _raw(key, default=""):
        value = form_data.get(key, default)
        return default if value is None else value

    def _as_int(key, default=None):
        value = _raw(key, default)
        try:
            return int(value) if value not in (None, "") else default
        except (TypeError, ValueError):
            return default

    canonical_payload = _canonical_asset_payload(form_data, mode="create")
    canonical_payload["airport_id"] = havalimani_id
    values = _normalized_asset_contract_values(canonical_payload, mode="create")
    _validate_asset_contract_values(values, mode="create")
    serial_no = values["serial_no"] or None
    status_display = _display_status(values["status"])
    status_internal = values["status"]
    is_demirbas = values["is_demirbas"]
    demirbas_no = values["demirbas_no"]
    stock_count = values["unit_count"]
    maintenance_period_months = values["maintenance_period_months"]
    maintenance_period_days = values["maintenance_period_days"]
    calibration_required = values["calibration_required"]
    calibration_period_days = values["calibration_period_days"]
    manual_url = values["manual_url"]

    legacy_material = Malzeme(
        ad=values["asset_name"] or guvenli_metin(form_data.get("ad") or template.name),
        seri_no=serial_no,
        teknik_ozellikler=values["technical_specs"] or guvenli_metin(form_data.get("teknik") or template.technical_specs),
        stok_miktari=stock_count,
        durum=status_display,
        son_bakim_tarihi=values["last_maintenance_date"],
        gelecek_bakim_tarihi=_parse_date(form_data.get("gelecek_bakim")),
        kalibrasyon_tarihi=values["last_calibration_date"] if calibration_required else None,
        kutu_id=kutu.id,
        havalimani_id=havalimani_id,
    )
    db.session.add(legacy_material)
    db.session.flush()

    parent_asset_id = _as_int("parent_asset_id")
    parent_asset = None
    if parent_asset_id:
        parent_asset = InventoryAsset.query.filter_by(
            id=parent_asset_id,
            havalimani_id=havalimani_id,
            is_deleted=False,
        ).first()

    asset = InventoryAsset(
        equipment_template_id=template.id,
        havalimani_id=havalimani_id,
        legacy_material_id=legacy_material.id,
        parent_asset_id=parent_asset.id if parent_asset else None,
        asset_type="spare_part" if parent_asset else "equipment",
        serial_no=serial_no,
        qr_code="",
        asset_tag=demirbas_no,
        is_demirbas=is_demirbas,
        unit_count=stock_count,
        depot_location=guvenli_metin(form_data.get("depo_konumu") or kutu.kodu).strip() or kutu.kodu,
        status=status_internal,
        maintenance_state="normal",
        last_maintenance_date=values["last_maintenance_date"],
        next_maintenance_date=None,
        calibration_required=calibration_required,
        calibration_period_days=calibration_period_days,
        last_calibration_date=values["last_calibration_date"] if calibration_required else None,
        next_calibration_date=values["next_calibration_date"] if calibration_required else None,
        acquired_date=values["acquired_date"],
        warranty_end_date=values["warranty_end_date"],
        manual_url=manual_url,
        notes=values["notes"],
        maintenance_period_days=maintenance_period_days,
        maintenance_period_months=maintenance_period_months,
    )
    db.session.add(asset)
    db.session.flush()
    assign_asset_qr(asset, force=True)
    _sync_maintenance_plan_for_asset(asset, template=template)

    return asset, legacy_material


def _is_system_owner():
    return get_effective_role(current_user) == CANONICAL_ROLE_SYSTEM


def _import_allowed_airports():
    return _visible_operational_airports() if _is_system_owner() else ([current_user.havalimani] if current_user.havalimani else [])


def _build_excel_lookup_context():
    templates = EquipmentTemplate.query.filter_by(is_deleted=False, is_active=True).all()
    forms = MaintenanceFormTemplate.query.filter_by(is_deleted=False, is_active=True).all()
    airports = [item for item in _import_allowed_airports() if item]
    airport_ids = [item.id for item in airports]

    box_query = Kutu.query.filter(Kutu.is_deleted.is_(False))
    if airport_ids:
        box_query = box_query.filter(Kutu.havalimani_id.in_(airport_ids))
    else:
        box_query = box_query.filter(sa.text("1=0"))
    boxes = box_query.all()

    def _key(value):
        return normalize_lookup(value)

    airport_map = {}
    for airport in airports:
        airport_map[_key(airport.kodu)] = airport
        airport_map[_key(airport.ad)] = airport
        airport_map[_key(f"{airport.kodu}-{airport.ad}")] = airport
        airport_map[_key(f"{airport.kodu} - {airport.ad}")] = airport

    template_map = {_key(item.name): item for item in templates}
    form_map = {_key(item.name): item for item in forms}
    box_map = {_key(item.kodu): item for item in boxes}

    category_values = _resolve_allowed_categories()
    category_map = {_key(item): item for item in category_values}

    return {
        "templates": templates,
        "template_map": template_map,
        "forms": forms,
        "form_map": form_map,
        "airports": airports,
        "airport_map": airport_map,
        "boxes": boxes,
        "box_map": box_map,
        "categories": category_values,
        "category_map": category_map,
    }


def _resolve_row_airport(raw_airport_value, lookup_ctx):
    if not _is_system_owner():
        if not current_user.havalimani_id:
            raise ValueError("Kullanıcıya havalimanı atanmadığı için import yapılamaz.")
        text = guvenli_metin(raw_airport_value or "").strip()
        if text:
            requested = lookup_ctx["airport_map"].get(normalize_lookup(text))
            if not requested or requested.id != current_user.havalimani_id:
                raise ValueError("Satır havalimanı için yetkiniz yok.")
        return current_user.havalimani_id
    text = guvenli_metin(raw_airport_value or "").strip()
    if not text:
        raise ValueError("havalimani alanı zorunludur.")
    airport = lookup_ctx["airport_map"].get(normalize_lookup(text))
    if not airport:
        raise ValueError(f"Yetkili havalimanı bulunamadı: {text}")
    return airport.id


def _resolve_row_template(row_values, lookup_ctx, *, category_value, maintenance_form_id, allow_create):
    template_name = guvenli_metin(row_values.get("merkezi_sablon") or "").strip()
    create_from_template = _to_bool(row_values.get("merkezi_sablondan_olustur"))
    template = lookup_ctx["template_map"].get(normalize_lookup(template_name)) if template_name else None
    if create_from_template and not template:
        raise ValueError("Merkezi şablondan oluşturma seçili ama şablon bulunamadı.")
    if template:
        return template
    if not allow_create:
        raise ValueError("Bu rol için merkezi şablon seçimi zorunludur.")

    template_seed = guvenli_metin(row_values.get("malzeme_adi") or "").strip()
    if not template_seed:
        raise ValueError("malzeme_adi zorunludur.")
    period_months = _parse_month_period(row_values.get("bakim_periyodu"), default=6)
    template = EquipmentTemplate(
        name=template_seed,
        category=category_value or "Diğer",
        brand=guvenli_metin(row_values.get("marka") or "").strip(),
        model_code=guvenli_metin(row_values.get("model") or "").strip(),
        technical_specs=guvenli_metin(row_values.get("teknik_ozellikler") or "").strip(),
        maintenance_period_months=period_months,
        maintenance_period_days=_months_to_days(period_months),
        default_maintenance_form_id=maintenance_form_id,
        criticality_level="normal",
        is_active=True,
    )
    db.session.add(template)
    db.session.flush()
    return template


def _build_form_payload_from_excel_row(row_values, lookup_ctx):
    airport_id = _resolve_row_airport(row_values.get("havalimani"), lookup_ctx)

    category_raw = guvenli_metin(row_values.get("kategori") or "").strip()
    category_value = lookup_ctx["category_map"].get(normalize_lookup(category_raw))
    if not category_value:
        raise ValueError(f"Kategori listede yok: {category_raw or '-'}")

    form_name = guvenli_metin(row_values.get("bakim_formu") or "").strip()
    maintenance_form = lookup_ctx["form_map"].get(normalize_lookup(form_name)) if form_name else None
    maintenance_form_id = maintenance_form.id if maintenance_form else None

    template = _resolve_row_template(
        row_values,
        lookup_ctx,
        category_value=category_value,
        maintenance_form_id=maintenance_form_id,
        allow_create=_is_system_owner(),
    )

    box_text = guvenli_metin(row_values.get("kutu_kodu") or "").strip()
    box = lookup_ctx["box_map"].get(normalize_lookup(box_text))
    if not box:
        box = _ensure_box(box_text, airport_id)
        lookup_ctx["box_map"][normalize_lookup(box.kodu)] = box
    if box.havalimani_id != airport_id:
        raise ValueError("Seçilen kutu satırdaki havalimanına ait değil.")

    person_name = normalize_person_name(row_values.get("ad_soyad"))
    note_lines = [guvenli_metin(row_values.get("aciklama_notlar") or "").strip()]
    if person_name:
        note_lines.append(f"SORUMLU: {person_name}")

    payload = {
        "havalimani_id": airport_id,
        "template_id": template.id,
        "kategori": category_value,
        "ad": guvenli_metin(row_values.get("malzeme_adi") or "").strip() or template.name,
        "marka": guvenli_metin(row_values.get("marka") or "").strip(),
        "model": guvenli_metin(row_values.get("model") or "").strip(),
        "is_demirbas": _to_bool(row_values.get("demirbas_mi")),
        "demirbas_no": guvenli_metin(row_values.get("demirbas_no") or "").strip(),
        "seri_no": guvenli_metin(row_values.get("seri_no") or "").strip(),
        "stok": to_int(row_values.get("stok_birim_sayisi"), default=1, min_value=1),
        "durum": guvenli_metin(row_values.get("kullanim_durumu") or "aktif").strip() or "aktif",
        "calibration_required": parse_flexible_bool(row_values.get("kalibrasyon_gerekli_mi")),
        "calibration_periyodu_ay": to_int(row_values.get("kalibrasyon_periyodu_ay"), default=6, min_value=1),
        "last_calibration_date": parse_flexible_date(row_values.get("son_kalibrasyon_tarihi")),
        "next_calibration_date": parse_flexible_date(row_values.get("sonraki_kalibrasyon_tarihi")),
        "bakim_formu_id": maintenance_form_id,
        "bakim_periyodu_ay": to_int(row_values.get("bakim_periyodu"), default=6, min_value=1),
        "edinim_tarihi": parse_flexible_date(row_values.get("edinim_tarihi")),
        "garanti_bitis_tarihi": parse_flexible_date(row_values.get("garanti_bitis_tarihi")),
        "kutu_id": box.id,
        "manual_url": guvenli_metin(row_values.get("kullanim_kilavuzu_linki") or "").strip(),
        "teknik": guvenli_metin(row_values.get("teknik_ozellikler") or "").strip(),
        "notlar": "\n".join([line for line in note_lines if line]),
    }
    yedek_link = guvenli_metin(row_values.get("yedek_parca_baglantisi") or "").strip()
    if yedek_link:
        parent_asset = _asset_scope().filter(
            sa.or_(
                InventoryAsset.serial_no == yedek_link,
                InventoryAsset.qr_code == yedek_link,
            )
        ).first()
        if parent_asset:
            payload["parent_asset_id"] = parent_asset.id
    return payload, template, box, airport_id


@inventory_bp.route("/dashboard")
@login_required
@permission_required("dashboard.view")
def dashboard():
    if _can_view_all_operational_scope():
        h_ad = "Genel Müdürlük / Tüm Birimler"
    else:
        h_ad = current_user.havalimani.ad if current_user.havalimani else "Atanmamış Birim"

    bugun = datetime.now(TR_TZ).date()
    on_bes_gun_sonra = bugun + timedelta(days=15)
    trend_days = request.args.get("trend_days", type=int) or 30
    if trend_days not in {7, 30, 90}:
        trend_days = 30

    bakim_sorgu = havalimani_filtreli_sorgu(Malzeme).filter(
        Malzeme.gelecek_bakim_tarihi <= on_bes_gun_sonra,
        Malzeme.durum != "Pasif",
    )
    ariza_sorgu = havalimani_filtreli_sorgu(Malzeme).filter_by(durum="Pasif")
    bakim_uyari_sayi = bakim_sorgu.count()
    arizali_malzeme_sayi = ariza_sorgu.count()

    asset_query = _asset_scope()
    due_today_count = asset_query.filter(
        InventoryAsset.next_maintenance_date.isnot(None),
        InventoryAsset.next_maintenance_date >= bugun,
        InventoryAsset.next_maintenance_date <= on_bes_gun_sonra,
    ).count()
    overdue_count = asset_query.filter(
        InventoryAsset.next_maintenance_date.isnot(None),
        InventoryAsset.next_maintenance_date < bugun,
        InventoryAsset.status != "pasif",
    ).count()

    open_work_order_query = WorkOrder.query.filter_by(is_deleted=False).join(InventoryAsset).filter(
        InventoryAsset.is_deleted.is_(False),
        WorkOrder.status.in_(["acik", "atandi", "islemde"]),
    )
    if not _can_view_all_operational_scope():
        open_work_order_query = open_work_order_query.filter(
            InventoryAsset.havalimani_id == current_user.havalimani_id
        )
    open_work_order_count = open_work_order_query.count()

    low_stock_query = SparePartStock.query.filter_by(is_deleted=False, is_active=True)
    if not _can_view_all_operational_scope():
        low_stock_query = low_stock_query.filter(SparePartStock.airport_id == current_user.havalimani_id)
    low_stock_items = low_stock_query.options(joinedload(SparePartStock.spare_part)).all()
    low_stock_count = sum(
        1
        for stock in low_stock_items
        if stock.available_quantity
        <= float(stock.reorder_point if stock.reorder_point is not None else (stock.spare_part.min_stock_level if stock.spare_part else 0))
    )

    meter_rule_query = MaintenanceTriggerRule.query.filter_by(is_deleted=False, is_active=True).join(
        InventoryAsset, MaintenanceTriggerRule.asset_id == InventoryAsset.id, isouter=True
    )
    if not _can_view_all_operational_scope():
        meter_rule_query = meter_rule_query.filter(
            (MaintenanceTriggerRule.asset_id.is_(None))
            | (InventoryAsset.havalimani_id == current_user.havalimani_id)
        )
    meter_rules = meter_rule_query.filter(
        MaintenanceTriggerRule.asset_id.isnot(None),
        MaintenanceTriggerRule.meter_definition_id.isnot(None),
    ).all()
    meter_warning_count = 0
    if meter_rules:
        rule_pairs = {
            (rule.asset_id, rule.meter_definition_id): rule
            for rule in meter_rules
            if rule.asset_id and rule.meter_definition_id
        }
        if rule_pairs:
            pair_asset_ids = {item[0] for item in rule_pairs.keys()}
            pair_meter_ids = {item[1] for item in rule_pairs.keys()}
            latest_reading_subquery = (
                db.session.query(
                    AssetMeterReading.asset_id.label("asset_id"),
                    AssetMeterReading.meter_definition_id.label("meter_definition_id"),
                    sa.func.max(AssetMeterReading.reading_at).label("max_reading_at"),
                )
                .filter(
                    AssetMeterReading.is_deleted.is_(False),
                    AssetMeterReading.asset_id.in_(pair_asset_ids),
                    AssetMeterReading.meter_definition_id.in_(pair_meter_ids),
                )
                .group_by(AssetMeterReading.asset_id, AssetMeterReading.meter_definition_id)
                .subquery()
            )
            latest_rows = (
                db.session.query(
                    AssetMeterReading.asset_id,
                    AssetMeterReading.meter_definition_id,
                    AssetMeterReading.reading_value,
                )
                .join(
                    latest_reading_subquery,
                    sa.and_(
                        AssetMeterReading.asset_id == latest_reading_subquery.c.asset_id,
                        AssetMeterReading.meter_definition_id == latest_reading_subquery.c.meter_definition_id,
                        AssetMeterReading.reading_at == latest_reading_subquery.c.max_reading_at,
                    ),
                )
                .all()
            )
            for asset_id, meter_definition_id, reading_value in latest_rows:
                rule = rule_pairs.get((asset_id, meter_definition_id))
                if not rule:
                    continue
                warning_threshold = float(rule.threshold_value or 0) - float(rule.warning_lead_value or 0)
                if float(reading_value or 0) >= max(warning_threshold, 0):
                    meter_warning_count += 1

    auto_work_order_query = WorkOrder.query.filter_by(is_deleted=False, source_type="meter_trigger").join(InventoryAsset)
    auto_work_order_query = auto_work_order_query.filter(
        InventoryAsset.is_deleted.is_(False),
        WorkOrder.status.in_(["acik", "atandi", "islemde", "beklemede_parca", "beklemede_onay"]),
    )
    if not _can_view_all_operational_scope():
        auto_work_order_query = auto_work_order_query.filter(InventoryAsset.havalimani_id == current_user.havalimani_id)
    auto_work_order_count = auto_work_order_query.count()

    child_fault_query = _asset_scope().filter(
        InventoryAsset.parent_asset_id.isnot(None),
        InventoryAsset.status == "pasif",
    )
    child_fault_count = child_fault_query.count()

    calibration_overdue_count = _asset_scope().filter(
        InventoryAsset.next_calibration_date.isnot(None),
        InventoryAsset.next_calibration_date < bugun,
        InventoryAsset.status != "pasif",
    ).count()
    calibration_upcoming_count = _asset_scope().filter(
        InventoryAsset.next_calibration_date.isnot(None),
        InventoryAsset.next_calibration_date >= bugun,
        InventoryAsset.next_calibration_date <= on_bes_gun_sonra,
    ).count()
    warranty_expiring_count = _asset_scope().filter(
        InventoryAsset.warranty_end_date.isnot(None),
        InventoryAsset.warranty_end_date >= bugun,
        InventoryAsset.warranty_end_date <= (bugun + timedelta(days=30)),
    ).count()
    low_consumable_count = 0
    critical_consumable_count = 0
    if table_exists("consumable_item") and table_exists("consumable_stock_movement") and current_user.havalimani_id:
        consumables = _consumable_scope().all()
        consumable_ids = [consumable.id for consumable in consumables]
        balance_map = {}
        if consumable_ids:
            signed_quantity = sa.case(
                (
                    ConsumableStockMovement.movement_type.in_(["in", "adjust", "transfer"]),
                    sa.func.coalesce(ConsumableStockMovement.quantity, 0.0),
                ),
                else_=-sa.func.coalesce(ConsumableStockMovement.quantity, 0.0),
            )
            balance_rows = (
                db.session.query(
                    ConsumableStockMovement.consumable_id,
                    sa.func.coalesce(sa.func.sum(signed_quantity), 0.0),
                )
                .filter(
                    ConsumableStockMovement.is_deleted.is_(False),
                    ConsumableStockMovement.airport_id == current_user.havalimani_id,
                    ConsumableStockMovement.consumable_id.in_(consumable_ids),
                )
                .group_by(ConsumableStockMovement.consumable_id)
                .all()
            )
            balance_map = {consumable_id: float(balance or 0) for consumable_id, balance in balance_rows}
        for consumable in consumables:
            balance = balance_map.get(consumable.id, 0.0)
            if balance <= float(consumable.critical_level or 0):
                critical_consumable_count += 1
            if balance <= float(consumable.min_stock_level or 0):
                low_consumable_count += 1
    total_asset_count = asset_query.count()

    urgent_actions = []
    for material in ariza_sorgu.order_by(Malzeme.updated_at.desc()).limit(3).all():
        urgent_actions.append(
            {
                "name": material.ad,
                "meta": f"Kutu: {material.kutu.kodu if material.kutu else '-'} · Durum: {material.durum}",
                "tag": "ARIZA",
                "tag_class": "tag-red",
                "url": url_for("inventory.kutu_detay", kodu=material.kutu.kodu) if material.kutu else url_for("inventory.envanter"),
            }
        )
    for material in bakim_sorgu.order_by(Malzeme.gelecek_bakim_tarihi.asc()).limit(4).all():
        if len(urgent_actions) >= 5:
            break
        urgent_actions.append(
            {
                "name": material.ad,
                "meta": f"Kutu: {material.kutu.kodu if material.kutu else '-'} · Tarih: {material.gelecek_bakim_tarihi.strftime('%d.%m.%Y') if material.gelecek_bakim_tarihi else '-'}",
                "tag": "BAKIM",
                "tag_class": "tag-amber",
                "url": url_for("inventory.kutu_detay", kodu=material.kutu.kodu) if material.kutu else url_for("inventory.envanter"),
            }
        )

    pending_approval_count = 0
    unread_notification_count = 0
    critical_role_change_count = 0
    qr_regen_pending_count = 0
    critical_audit_24h_count = 0
    if table_exists("approval_request"):
        pending_approval_count = ApprovalRequest.query.filter_by(status="pending").count()
        critical_role_change_count = ApprovalRequest.query.filter_by(status="pending", approval_type="role_change").count()
        qr_regen_pending_count = ApprovalRequest.query.filter_by(status="pending", approval_type="qr_regenerate").count()
    if table_exists("notification"):
        unread_notification_count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    if table_exists("islem_log") and column_exists("islem_log", "event_key") and column_exists("islem_log", "outcome"):
        last_day = get_tr_now().replace(tzinfo=None) - timedelta(hours=24)
        critical_audit_24h_count = IslemLog.query.filter(
            IslemLog.zaman >= last_day,
            IslemLog.outcome.in_(["failed", "success"]),
            IslemLog.event_key.in_([
                "role.assignment.change",
                "role.assignment.pending",
                "permission.matrix.update",
                "inventory.qr.regenerate",
                "inventory.asset_code.change_attempt",
            ]),
        ).count()

    return render_template(
        "dashboard.html",
        havalimani_ad=h_ad,
        bakim_uyari_sayi=bakim_uyari_sayi,
        arizali_malzeme_sayi=arizali_malzeme_sayi,
        toplam_ekipman_sayi=total_asset_count,
        bugun=bugun,
        bugun_bakim_yaklasan_sayi=due_today_count,
        geciken_bakim_sayi=overdue_count,
        acik_is_emri_sayi=open_work_order_count,
        dusuk_stok_parca_sayi=low_stock_count,
        sayac_yaklasan_bakim_sayi=meter_warning_count,
        otomatik_is_emri_sayi=auto_work_order_count,
        child_asset_ariza_sayi=child_fault_count,
        kalibrasyon_gecikme_sayi=calibration_overdue_count,
        yaklasan_kalibrasyon_sayi=calibration_upcoming_count,
        garanti_yaklasan_sayi=warranty_expiring_count,
        dusuk_sarf_sayi=low_consumable_count,
        kritik_sarf_sayi=critical_consumable_count,
        approval_bekleyen_sayi=pending_approval_count,
        okunmamis_bildirim_sayi=unread_notification_count,
        kritik_rol_degisim_sayi=critical_role_change_count,
        qr_yenileme_talep_sayi=qr_regen_pending_count,
        kritik_audit_24h_sayi=critical_audit_24h_count,
        acil_aksiyonlar=urgent_actions,
        dashboard_trends=build_dashboard_kpis(current_user, {"trend_days": trend_days, "demo_scope": "all"})["kpis"],
        trend_days=trend_days,
    )


@inventory_bp.route("/dashboard/alerts/sync", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("dashboard.view")
def dashboard_alert_sync():
    _create_operational_alerts()
    return ("", 204)


@inventory_bp.route("/envanter")
@login_required
@permission_required("inventory.view")
def envanter():
    selected_airport = request.args.get("havalimani_id", type=int)
    selected_category = request.args.get("kategori", "").strip()

    query = havalimani_filtreli_sorgu(Malzeme).options(
        joinedload(Malzeme.havalimani),
        joinedload(Malzeme.kutu),
        joinedload(Malzeme.linked_asset).joinedload(InventoryAsset.equipment_template),
        joinedload(Malzeme.linked_asset).joinedload(InventoryAsset.operational_state),
    )
    if _can_view_all_operational_scope() and selected_airport:
        query = query.filter(Malzeme.havalimani_id == selected_airport)
    query = query.outerjoin(Malzeme.linked_asset).outerjoin(InventoryAsset.operational_state).filter(
        sa.or_(
            InventoryAsset.id.is_(None),
            sa.and_(
                AssetOperationalState.id.isnot(None),
                AssetOperationalState.lifecycle_status.isnot(None),
                AssetOperationalState.lifecycle_status.notin_(["disposed", "decommissioned"]),
            ),
            sa.and_(
                sa.or_(
                    AssetOperationalState.id.is_(None),
                    AssetOperationalState.lifecycle_status.is_(None),
                ),
                sa.or_(
                    InventoryAsset.status.is_(None),
                    InventoryAsset.status.notin_(["hurda", "pasif"]),
                ),
            ),
        )
    )
    if selected_category:
        query = query.join(InventoryAsset.equipment_template).filter(EquipmentTemplate.category == selected_category)
    malzemeler = query.order_by(Malzeme.created_at.desc()).all()

    if _can_view_all_operational_scope():
        h_ad = "Genel Envanter (Tüm Birimler)"
        havalimanlari = _visible_operational_airports()
    else:
        h_ad = current_user.havalimani.ad if current_user.havalimani else "Atanmamış Birim"
        havalimanlari = [current_user.havalimani] if current_user.havalimani else []

    categories = (
        db.session.query(EquipmentTemplate.category)
        .filter(
            EquipmentTemplate.is_deleted.is_(False),
            EquipmentTemplate.category.isnot(None),
            EquipmentTemplate.category != "",
        )
        .distinct()
        .all()
    )

    return render_template(
        "envanter.html",
        malzemeler=malzemeler,
        havalimani_ad=h_ad,
        havalimanlari=havalimanlari,
        categories=[row[0] for row in categories],
        selected_airport=selected_airport,
        selected_category=selected_category,
        status_label_tr=_status_label_tr,
        maintenance_label_tr=_maintenance_label_tr,
        today=get_tr_now().date(),
    )


def _malzeme_create_page_context():
    templates = EquipmentTemplate.query.filter_by(is_deleted=False, is_active=True).order_by(
        EquipmentTemplate.name.asc()
    ).all()
    form_templates = MaintenanceFormTemplate.query.filter_by(is_deleted=False, is_active=True).order_by(
        MaintenanceFormTemplate.name.asc()
    ).all()
    category_options = _resolve_allowed_categories()
    airport_options = _visible_operational_airports() if current_user.is_sahip else ([current_user.havalimani] if current_user.havalimani else [])
    airport_ids = [item.id for item in airport_options if item]
    box_query = Kutu.query.filter(Kutu.is_deleted.is_(False))
    if airport_ids:
        box_query = box_query.filter(Kutu.havalimani_id.in_(airport_ids))
    else:
        box_query = box_query.filter(sa.text("1=0"))
    box_options = box_query.order_by(Kutu.kodu.asc()).all()
    parent_candidates = _asset_scope().order_by(InventoryAsset.created_at.desc()).limit(200).all()
    default_parent_id = request.args.get("parent_asset_id", type=int)
    preselected_asset_type = (request.args.get("asset_type") or "equipment").strip()
    preselected_airport_id = current_user.havalimani_id if not current_user.is_sahip else (airport_options[0].id if airport_options else None)

    return {
        "templates": templates,
        "form_templates": form_templates,
        "havalimanlari": airport_options,
        "categories": category_options,
        "boxes": box_options,
        "parent_candidates": parent_candidates,
        "default_parent_id": default_parent_id,
        "preselected_asset_type": preselected_asset_type,
        "preselected_airport_id": preselected_airport_id,
        "can_manage_template_catalog": bool(current_user.is_sahip),
        "can_manage_category_catalog": _can_manage_asset_registry(),
        "maintenance_month_values": MAINTENANCE_MONTH_VALUES,
        "field_labels": form_label_map(),
    }


@inventory_bp.route("/malzeme-ekle", methods=["GET", "POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"), methods=["POST"])
@permission_required("inventory.create")
def malzeme_ekle():
    if not _can_manage_asset_registry():
        abort(403)

    page_ctx = _malzeme_create_page_context()

    if request.method == "POST":
        canonical_payload = _canonical_asset_payload(request.form, mode="create")
        if current_user.is_sahip:
            havalimani_id = request.form.get("airport_id", type=int) or request.form.get("havalimani_id", type=int)
            if not havalimani_id:
                havalimani_id = current_user.havalimani_id or 1
        else:
            havalimani_id = current_user.havalimani_id
        canonical_payload["airport_id"] = havalimani_id

        try:
            kutu = _resolve_box_from_payload(canonical_payload, havalimani_id, allow_auto_create=True)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("inventory.malzeme_ekle"))

        category_value = guvenli_metin(canonical_payload.get("category") or "").strip() or "Diğer"
        canonical_payload["category"] = category_value

        template_id = request.form.get("template_id", type=int)
        central_catalog_flag = (canonical_payload.get("catalog_mode") or "") == "create_template_owner"
        can_manage_template_catalog = bool(current_user.is_sahip)
        template = _ensure_template_from_form(request.form, template_id, allow_create=can_manage_template_catalog and central_catalog_flag)

        if central_catalog_flag and not can_manage_template_catalog:
            flash("Merkezi kataloga yeni şablon ekleme yetkisi sadece sahip rolünde.", "danger")
            return redirect(url_for("inventory.malzeme_ekle"))

        if not template:
            template = _ensure_template_from_form(request.form, None, allow_create=True)
        if template is None:
            template = _ensure_default_equipment_template()

        try:
            asset, legacy_material = _create_asset_and_legacy_material(
                template=template,
                kutu=kutu,
                havalimani_id=havalimani_id,
                form_data=request.form,
            )
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(url_for("inventory.malzeme_ekle"))
        except IntegrityError:
            db.session.rollback()
            flash("Kayıt kaydedilemedi. Seri no veya ilişkili verilerde çakışma olabilir.", "danger")
            return redirect(url_for("inventory.malzeme_ekle"))

        if asset.calibration_required:
            calibration_date = asset.last_calibration_date or get_tr_now().date()
            period_days = asset.calibration_period_days or 180
            next_calibration_date = asset.next_calibration_date or (calibration_date + timedelta(days=period_days))
            schedule = _sync_calibration_schedule_for_asset(asset)

            certificate_upload = request.files.get("calibration_certificate")
            certificate_payload = None
            if certificate_upload and certificate_upload.filename:
                try:
                    certificate_payload = _upload_calibration_certificate(asset, calibration_date, certificate_upload)
                except (ValueError, GoogleDriveError) as exc:
                    db.session.rollback()
                    flash(str(exc), "danger")
                    return redirect(url_for("inventory.malzeme_ekle"))

            calibration_record = CalibrationRecord(
                asset_id=asset.id,
                calibration_schedule_id=schedule.id,
                calibration_date=calibration_date,
                next_calibration_date=next_calibration_date,
                calibrated_by_id=current_user.id,
                provider=guvenli_metin(request.form.get("calibration_provider") or ""),
                certificate_no=guvenli_metin(request.form.get("certificate_no") or ""),
                certificate_file=certificate_payload["certificate_file"] if certificate_payload else None,
                certificate_drive_file_id=certificate_payload["drive_file_id"] if certificate_payload else None,
                certificate_drive_folder_id=certificate_payload["drive_folder_id"] if certificate_payload else None,
                certificate_mime_type=certificate_payload["mime_type"] if certificate_payload else None,
                certificate_size_bytes=certificate_payload["size_bytes"] if certificate_payload else None,
                result_status="passed",
                note=guvenli_metin(request.form.get("calibration_note") or ""),
            )
            db.session.add(calibration_record)
            asset.last_calibration_date = calibration_date
            asset.next_calibration_date = next_calibration_date
        else:
            _sync_calibration_schedule_for_asset(asset)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Kayıt kaydedilemedi. Seri no veya ilişkili verilerde çakışma olabilir.", "danger")
            return redirect(url_for("inventory.malzeme_ekle"))
        log_kaydet(
            "Envanter",
            f"Yeni ekipman eklendi: {legacy_material.ad} ({legacy_material.havalimani.kodu}) / Şablon: {template.name}",
        )
        audit_log("inventory.asset.create", outcome="success", asset_id=asset.id, airport_id=havalimani_id)
        if asset.parent_asset_id:
            audit_log(
                "inventory.asset.link_child",
                outcome="success",
                asset_id=asset.id,
                parent_asset_id=asset.parent_asset_id,
            )
        flash("Malzeme başarıyla eklendi ve bakım varlığı oluşturuldu.", "success")
        return redirect(url_for("inventory.envanter"))

    return render_template("malzeme_ekle.html", **page_ctx)


@inventory_bp.route("/malzeme-ekle/excel-sablon")
@login_required
@permission_required("inventory.create")
def malzeme_excel_sablon_indir():
    if not _can_manage_asset_registry():
        abort(403)
    lookup_ctx = _build_excel_lookup_context()
    workbook = build_inventory_template_workbook(
        lists_context={
            "templates": [item.name for item in lookup_ctx["templates"]],
            "airports": [f"{item.kodu} - {item.ad}" for item in lookup_ctx["airports"]],
            "categories": lookup_ctx["categories"],
            "statuses": ["aktif", "pasif"],
            "maintenance_forms": [item.name for item in lookup_ctx["forms"]],
            "month_values": MAINTENANCE_MONTH_VALUES,
            "boxes": [item.kodu for item in lookup_ctx["boxes"]],
        }
    )
    return send_file(
        workbook,
        as_attachment=True,
        download_name=f"malzeme_toplu_import_sablonu_{get_tr_now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@inventory_bp.route("/malzeme-ekle/excel-yukle", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("inventory.create")
def malzeme_excel_yukle():
    if not _can_manage_asset_registry():
        abort(403)

    upload = request.files.get("excel_file")
    if not upload or not upload.filename:
        flash("Excel dosyası seçilmedi.", "danger")
        return redirect(url_for("inventory.malzeme_ekle"))
    safe_name = secure_upload_filename(upload.filename or "")
    if not safe_name.lower().endswith(".xlsx"):
        flash("Sadece .xlsx formatı desteklenir.", "danger")
        return redirect(url_for("inventory.malzeme_ekle"))
    if not _is_valid_xlsx_workbook(upload):
        flash("Excel dosyası okunamadı.", "danger")
        return redirect(url_for("inventory.malzeme_ekle"))

    try:
        rows = parse_inventory_workbook(upload)
    except ExcelTemplateError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("inventory.malzeme_ekle"))

    lookup_ctx = _build_excel_lookup_context()
    job = InventoryBulkImportJob(
        requested_by_user_id=current_user.id,
        havalimani_id=current_user.havalimani_id,
        source_filename=safe_name,
        status="processing",
        total_rows=len(rows),
    )
    db.session.add(job)
    db.session.flush()

    success_count = 0
    fail_count = 0
    row_feedback = []

    for row in rows:
        row_no = row["row_no"]
        values = row["values"]
        savepoint = db.session.begin_nested()
        try:
            payload, template, kutu, airport_id = _build_form_payload_from_excel_row(values, lookup_ctx)
            asset, _legacy = _create_asset_and_legacy_material(
                template=template,
                kutu=kutu,
                havalimani_id=airport_id,
                form_data=payload,
            )
            _sync_calibration_schedule_for_asset(asset)
            assign_asset_qr(asset)
            db.session.add(
                InventoryBulkImportRowResult(
                    job_id=job.id,
                    row_no=row_no,
                    status="success",
                    message="Satır başarıyla işlendi.",
                    serial_no=asset.serial_no,
                    asset_id=asset.id,
                )
            )
            savepoint.commit()
            success_count += 1
        except Exception as exc:
            savepoint.rollback()
            fail_count += 1
            err_text = guvenli_metin(str(exc) or "Bilinmeyen hata").strip() or "Bilinmeyen hata"
            db.session.add(
                InventoryBulkImportRowResult(
                    job_id=job.id,
                    row_no=row_no,
                    status="failed",
                    message=err_text,
                    serial_no=guvenli_metin(values.get("seri_no") or "").strip() or None,
                )
            )
            row_feedback.append({"row_no": row_no, "error": err_text})

    job.status = "completed"
    job.success_rows = success_count
    job.failed_rows = fail_count
    job.summary_note = f"Toplam {len(rows)} satır işlendi. Başarılı: {success_count}, Hatalı: {fail_count}"

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("Toplu import sırasında kritik hata oluştu.", "danger")
        return redirect(url_for("inventory.malzeme_ekle"))

    flash(f"Toplu import tamamlandı. Başarılı: {success_count} / Hatalı: {fail_count}", "success" if fail_count == 0 else "warning")
    page_ctx = _malzeme_create_page_context()
    page_ctx["import_summary"] = {
        "job_id": job.id,
        "total": len(rows),
        "success": success_count,
        "failed": fail_count,
        "errors": row_feedback[:50],
    }
    return render_template("malzeme_ekle.html", **page_ctx)


@inventory_bp.route("/merkezi-katalog")
@login_required
@permission_required("inventory.create")
def merkezi_katalog():
    from routes.maintenance import _build_instruction_catalog_options

    templates = EquipmentTemplate.query.filter_by(is_deleted=False, is_active=True).order_by(
        EquipmentTemplate.name.asc()
    ).all()
    form_templates = MaintenanceFormTemplate.query.filter_by(is_deleted=False, is_active=True).order_by(
        MaintenanceFormTemplate.name.asc()
    ).all()
    airports = _visible_operational_airports()
    selectable_catalog, selectable_categories = _build_instruction_catalog_options()
    return render_template(
        "ekipman_sablonlari.html",
        templates=templates,
        form_templates=form_templates,
        airports=airports,
        selectable_catalog=selectable_catalog,
        selectable_categories=selectable_categories,
    )


def _find_active_inventory_category(name):
    normalized_name = normalize_lookup(name)
    if not normalized_name:
        return None
    for item in InventoryCategory.query.filter(InventoryCategory.is_deleted.is_(False)).all():
        if normalize_lookup(item.name) == normalized_name:
            return item
    return None


def _ensure_inventory_category_exists(name, *, description="", created_by_user_id=None):
    cleaned_name = guvenli_metin(name or "").strip()
    if not cleaned_name:
        return None

    existing = _find_active_inventory_category(cleaned_name)
    if existing:
        return existing

    created = InventoryCategory(
        name=cleaned_name,
        description=guvenli_metin(description or "").strip(),
        created_by_user_id=created_by_user_id,
        is_active=True,
    )
    db.session.add(created)
    db.session.flush()
    return created


@inventory_bp.route("/envanter/kategori-ekle", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("inventory.create")
def envanter_kategori_ekle():
    if not _can_manage_asset_registry():
        abort(403)
    name = guvenli_metin(request.form.get("name") or "").strip()
    if not name:
        flash("Kategori adı zorunludur.", "danger")
        return redirect(url_for("inventory.malzeme_ekle"))

    exists = _find_active_inventory_category(name)
    if exists:
        flash("Bu kategori zaten mevcut.", "warning")
        return redirect(url_for("inventory.malzeme_ekle"))

    _ensure_inventory_category_exists(
        name,
        description=request.form.get("description"),
        created_by_user_id=current_user.id,
    )
    db.session.commit()
    flash("Kategori eklendi.", "success")
    return redirect(url_for("inventory.malzeme_ekle"))


@inventory_bp.route("/envanter/merkezi-sablon-ekle", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("inventory.create")
def merkezi_sablon_ekle():
    if not _is_system_owner():
        abort(403)

    name = guvenli_metin(request.form.get("name") or "").strip()
    if not name:
        flash("Şablon adı zorunludur.", "danger")
        return redirect(url_for("inventory.malzeme_ekle"))

    selected_category = guvenli_metin(request.form.get("category") or "").strip()
    new_category_name = guvenli_metin(request.form.get("new_category_name") or "").strip()
    if selected_category == "__new__":
        category = new_category_name
        if not category:
            flash("Yeni kategori adı zorunludur.", "danger")
            return redirect(url_for("inventory.malzeme_ekle"))
    else:
        category = selected_category or "Diğer"

    if selected_category == "__new__":
        _ensure_inventory_category_exists(
            category,
            description="Merkezi şablon oluşturma ekranından eklendi.",
            created_by_user_id=current_user.id,
        )
    period_months = _parse_month_period(request.form.get("maintenance_period_months"), default=6)
    maintenance_form_id = request.form.get("default_maintenance_form_id", type=int) or None

    template = EquipmentTemplate(
        name=name,
        category=category,
        brand=guvenli_metin(request.form.get("brand") or "").strip(),
        model_code=guvenli_metin(request.form.get("model_code") or "").strip(),
        description=guvenli_metin(request.form.get("description") or "").strip(),
        technical_specs=guvenli_metin(request.form.get("technical_specs") or "").strip(),
        maintenance_period_months=period_months,
        maintenance_period_days=_months_to_days(period_months),
        default_maintenance_form_id=maintenance_form_id,
        criticality_level="normal",
        is_active=True,
    )
    db.session.add(template)
    db.session.commit()
    flash("Merkezi şablon eklendi.", "success")
    return redirect(url_for("inventory.malzeme_ekle"))


@inventory_bp.route("/merkezi-sablondan-envantere-ekle/<int:template_id>", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("inventory.create")
def merkezi_sablondan_envantere_ekle(template_id):
    if not _can_manage_asset_registry():
        abort(403)
    template = EquipmentTemplate.query.filter_by(id=template_id, is_deleted=False, is_active=True).first_or_404()

    if current_user.is_sahip:
        havalimani_id = request.form.get("havalimani_id", type=int) or current_user.havalimani_id or 1
    else:
        havalimani_id = current_user.havalimani_id

    canonical_payload = _canonical_asset_payload(request.form, mode="create")
    canonical_payload["airport_id"] = havalimani_id
    try:
        kutu = _resolve_box_from_payload(canonical_payload, havalimani_id, allow_auto_create=True)
    except ValueError as exc:
        flash(str(exc) or "Merkezi şablondan ekleme için kutu seçimi zorunludur.", "danger")
        return redirect(url_for("maintenance.ekipman_sablonlari"))

    asset, legacy_material = _create_asset_and_legacy_material(
        template=template,
        kutu=kutu,
        havalimani_id=havalimani_id,
        form_data={
            **request.form,
            "bakim_periyodu_ay": str(_parse_month_period(template.maintenance_period_months or max(int((template.maintenance_period_days or 180) / 30), 1))),
            "kategori": template.category or "",
        },
    )
    legacy_material.ad = template.name

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash("Merkezi katalogdan ekipman eklenemedi. Seri no veya ilişkili verilerde çakışma olabilir.", "danger")
        return redirect(url_for("maintenance.ekipman_sablonlari"))
    log_kaydet(
        "Merkezi Şablon",
        f"Merkezi şablondan birime ekipman eklendi: {template.name} -> {legacy_material.havalimani.kodu}",
    )
    if asset.parent_asset_id:
        audit_log(
            "inventory.asset.link_child",
            outcome="success",
            asset_id=asset.id,
            parent_asset_id=asset.parent_asset_id,
        )
    flash("Merkezi katalogdan ekipman envantere eklendi.", "success")
    return redirect(url_for("inventory.envanter"))


@inventory_bp.route("/asset-duzenle/<int:asset_id>", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("inventory.edit")
def asset_duzenle(asset_id):
    if not _can_manage_asset_registry():
        abort(403)
    asset = _asset_scope().filter(InventoryAsset.id == asset_id).first_or_404()
    requested_asset_code = guvenli_metin(request.form.get("asset_code") or "").strip()
    if requested_asset_code and requested_asset_code != asset.asset_code:
        log_kaydet(
            "Güvenlik",
            f"Asset code değişikliği denemesi engellendi: {asset.asset_code} -> {requested_asset_code}",
            event_key="inventory.asset_code.change_attempt",
            target_model="InventoryAsset",
            target_id=asset.id,
            outcome="failed",
        )
        audit_log("inventory.asset_code.change_attempt", outcome="failed", asset_id=asset.id)
        flash("Asset code doğrudan değiştirilemez.", "danger")
        return redirect(url_for("inventory.envanter"))
    canonical_payload = _canonical_asset_payload(request.form, mode="asset_duzenle")
    values = _normalized_asset_contract_values(canonical_payload, mode="asset_duzenle", current_asset=asset)
    try:
        _validate_asset_contract_values(values, mode="asset_duzenle", current_asset=asset)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("inventory.envanter"))

    asset.serial_no = values["serial_no"] or asset.serial_no
    asset.unit_count = values["unit_count"]
    asset.depot_location = guvenli_metin(request.form.get("depo_konumu") or asset.depot_location)
    asset.status = values["status"]
    asset.last_maintenance_date = values["last_maintenance_date"] or asset.last_maintenance_date
    if values["notes"]:
        asset.notes = values["notes"]
    asset.manual_url = values["manual_url"]
    asset.maintenance_period_months = values["maintenance_period_months"]
    asset.maintenance_period_days = values["maintenance_period_days"]
    asset.calibration_required = values["calibration_required"]
    asset.calibration_period_days = values["calibration_period_days"]
    asset.last_calibration_date = values["last_calibration_date"] if asset.calibration_required else None
    asset.next_calibration_date = values["next_calibration_date"] if asset.calibration_required else None
    asset.is_demirbas = values["is_demirbas"]
    asset.asset_tag = values["demirbas_no"] if asset.is_demirbas else ""

    new_parent_id = values["parent_asset_id"]
    if new_parent_id:
        parent_candidate = _asset_scope().filter(InventoryAsset.id == new_parent_id).first()
        if not parent_candidate:
            flash("Üst ekipman seçimi geçersiz.", "danger")
            return redirect(url_for("inventory.envanter"))
        asset.parent_asset_id = parent_candidate.id
    elif request.form.get("parent_asset_id") == "":
        asset.parent_asset_id = None

    _sync_maintenance_plan_for_asset(asset)
    _sync_calibration_schedule_for_asset(asset)
    if asset.legacy_material:
        asset.legacy_material.seri_no = asset.serial_no
        asset.legacy_material.stok_miktari = asset.unit_count
        asset.legacy_material.durum = _display_status(asset.status)
        asset.legacy_material.son_bakim_tarihi = asset.last_maintenance_date
        asset.legacy_material.gelecek_bakim_tarihi = asset.next_maintenance_date
        asset.legacy_material.kalibrasyon_tarihi = asset.last_calibration_date if asset.calibration_required else None
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash("Envanter kaydı güncellenemedi. Seri no veya ilişkili verilerde çakışma olabilir.", "danger")
        return redirect(url_for("inventory.envanter"))
    log_kaydet("Envanter", f"Yerel asset güncellendi: ID {asset.id}")
    if asset.parent_asset_id:
        audit_log(
            "inventory.asset.link_child",
            outcome="success",
            asset_id=asset.id,
            parent_asset_id=asset.parent_asset_id,
        )
    flash("Envanter kaydı güncellendi.", "success")
    return redirect(url_for("inventory.envanter"))


@inventory_bp.route("/bakim-kaydet/<int:id>", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("maintenance.edit")
def bakim_kaydet(id):
    malzeme = Malzeme.query.filter_by(id=id, is_deleted=False).first_or_404()
    if not has_permission("settings.manage") and malzeme.havalimani_id != current_user.havalimani_id:
        flash("Farklı bir birimin malzemesine bakım girişi yapamazsınız.", "danger")
        abort(403)

    offline_sync_mode = (request.headers.get("X-SARX-Offline-Sync") or "").strip() == "1"
    raw_request_id = str(request.headers.get("X-SARX-Offline-Request-Id") or "").strip()[:64]
    offline_request_id = raw_request_id if OFFLINE_MAINTENANCE_REQUEST_ID_RE.match(raw_request_id) else ""

    if offline_sync_mode and offline_request_id and table_exists("islem_log"):
        duplicate_log = (
            IslemLog.query.filter_by(
                event_key="inventory.maintenance.offline_sync",
                request_id=offline_request_id,
                outcome="success",
            )
            .order_by(IslemLog.id.desc())
            .first()
        )
        if duplicate_log:
            return jsonify({"status": "success", "duplicate": True}), 200

    guvenli_not = guvenli_metin(request.form.get("not"))
    yeni_kayit = BakimKaydi(
        malzeme_id=id,
        yapan_personel_id=current_user.id,
        islem_notu=guvenli_not,
        maliyet=float(request.form.get("maliyet", 0)),
    )

    malzeme.son_bakim_tarihi = get_tr_now().date()
    yeni_gelecek = request.form.get("gelecek_bakim")
    if yeni_gelecek:
        malzeme.gelecek_bakim_tarihi = datetime.strptime(yeni_gelecek, "%Y-%m-%d").date()

    db.session.add(yeni_kayit)

    if malzeme.linked_asset:
        asset = malzeme.linked_asset
        asset.last_maintenance_date = malzeme.son_bakim_tarihi
        if malzeme.gelecek_bakim_tarihi:
            asset.next_maintenance_date = malzeme.gelecek_bakim_tarihi

        history = MaintenanceHistory(
            asset_id=asset.id,
            performed_by_id=current_user.id,
            maintenance_type="bakim",
            result="Bakım kaydı tamamlandı",
            notes=guvenli_not,
            next_maintenance_date=asset.next_maintenance_date,
        )
        db.session.add(history)

    db.session.commit()

    if offline_sync_mode:
        log_kaydet(
            "Bakım",
            f"{malzeme.ad} için çevrimdışı bakım kaydı işlendi ({malzeme.havalimani.kodu})",
            event_key="inventory.maintenance.offline_sync",
            outcome="success",
            request_id=offline_request_id or None,
            target_model="BakimKaydi",
            target_id=yeni_kayit.id,
        )
        audit_log(
            "inventory.maintenance.offline_sync",
            outcome="success",
            maintenance_id=yeni_kayit.id,
            material_id=malzeme.id,
            request_id=offline_request_id or None,
            user_id=current_user.id,
        )
        return jsonify({"status": "success", "maintenance_id": yeni_kayit.id}), 200

    log_kaydet("Bakım", f"{malzeme.ad} için bakım kaydı girildi ({malzeme.havalimani.kodu})")
    flash("Bakım kaydı başarıyla işlendi.", "success")
    return redirect(url_for("inventory.envanter"))


@inventory_bp.route("/envanter/excel")
@login_required
@permission_required("inventory.export")
def envanter_excel():
    malzemeler = havalimani_filtreli_sorgu(Malzeme).options(
        joinedload(Malzeme.havalimani),
        joinedload(Malzeme.kutu),
    ).all()
    if len(malzemeler) > int(current_app.config.get("MAX_EXPORT_ROWS", 10000)):
        abort(413)
    data = [
        {
            "Birim": m.havalimani.kodu,
            "Kutu": m.kutu.kodu,
            "Malzeme Adı": m.ad,
            "Seri No": m.seri_no,
            "Durum": m.durum,
            "Son Bakım": m.son_bakim_tarihi.strftime("%d.%m.%Y") if m.son_bakim_tarihi else "-",
            "Gelecek Bakım": m.gelecek_bakim_tarihi.strftime("%d.%m.%Y") if m.gelecek_bakim_tarihi else "-",
        }
        for m in malzemeler
    ]

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    output.seek(0)

    log_kaydet("Rapor", f"Envanter Excel raporu oluşturuldu ({current_user.rol})")
    return send_file(
        output,
        download_name=f"SAR_Envanter_{datetime.now(TR_TZ).strftime('%Y%m%d')}.xlsx",
        as_attachment=True,
    )


@inventory_bp.route("/malzeme-sil/<int:id>", methods=["GET"], endpoint="malzeme_sil_legacy")
@login_required
@permission_required("inventory.delete")
def malzeme_sil_legacy(id):
    flash("Bu işlem yalnızca form gönderimi ile yapılabilir.", "warning")
    return redirect(_safe_internal_redirect_target(request.referrer, url_for("inventory.dashboard")))


@inventory_bp.route("/malzeme-sil/<int:id>", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"), methods=["POST"])
@permission_required("inventory.delete")
def malzeme_sil(id):
    malzeme = havalimani_filtreli_sorgu(Malzeme).filter(Malzeme.id == id).first()
    if malzeme and not malzeme.is_deleted:
        malzeme_adi = malzeme.ad
        malzeme.soft_delete()

        if malzeme.linked_asset and not malzeme.linked_asset.is_deleted:
            for child in malzeme.linked_asset.child_assets:
                if not child.is_deleted:
                    child.parent_asset_id = None
            malzeme.linked_asset.soft_delete()

        db.session.commit()

        log_kaydet("Envanter", f"Malzeme arşive gönderildi: {malzeme_adi}")
        flash(f"'{malzeme_adi}' başarıyla silindi ve arşive taşındı.", "info")
    else:
        flash("Hata: Malzeme bulunamadı veya zaten silinmiş.", "danger")

    return redirect(_safe_internal_redirect_target(request.referrer, url_for("inventory.dashboard")))


@inventory_bp.route("/envanter/pdf")
@login_required
@permission_required("inventory.export")
def envanter_pdf():
    malzemeler = havalimani_filtreli_sorgu(Malzeme).options(
        joinedload(Malzeme.havalimani),
        joinedload(Malzeme.kutu),
    ).all()
    if len(malzemeler) > int(current_app.config.get("MAX_EXPORT_ROWS", 10000)):
        abort(413)
    html = render_template("pdf_sablonu.html", malzemeler=malzemeler, tarih=datetime.now(TR_TZ))
    output = io.BytesIO()
    pisa.CreatePDF(html, dest=output)
    output.seek(0)

    log_kaydet("Rapor", f"Envanter PDF raporu oluşturuldu ({current_user.rol})")
    return send_file(
        output,
        download_name=f"SAR_Rapor_{datetime.now(TR_TZ).strftime('%Y%m%d')}.pdf",
        as_attachment=True,
    )


@inventory_bp.route("/zimmetler", methods=["GET", "POST"])
@login_required
@permission_required("assignment.view")
def zimmetler():
    selected_airport = request.args.get("airport_id", type=int)
    selected_status = _normalize_assignment_status(request.args.get("status"))
    selected_recipient = request.args.get("recipient_id", type=int)

    visible_airports = _visible_operational_airports()
    if not _can_view_all_operational_scope():
        selected_airport = current_user.havalimani_id

    can_create_assignment = has_permission("assignment.create") and _can_issue_assignments(current_user)
    can_delete_assignment = _can_delete_assignments()

    if request.method == "POST":
        if not can_create_assignment:
            abort(403)

        raw_airport_id = request.form.get("airport_id", type=int)
        airport_id = raw_airport_id or current_user.havalimani_id
        selected_item_ids = list(dict.fromkeys(_parse_int_list(request.form.getlist("item_ids"))))
        if _can_view_all_operational_scope():
            airport_allowed = airport_id is None or any(airport.id == airport_id for airport in visible_airports)
        else:
            airport_allowed = airport_id == current_user.havalimani_id
        if not airport_allowed:
            flash("Seçilen havalimanı için zimmet oluşturma yetkiniz yok.", "danger")
            return redirect(url_for("inventory.zimmetler", airport_id=selected_airport or None))

        delivered_by_name = guvenli_metin(request.form.get("delivered_by_name") or "").strip()
        if not delivered_by_name:
            delivered_by_name = guvenli_metin(getattr(current_user, "tam_ad", "") or "").strip()
        recipient_scope_airport = airport_id if not _can_view_all_operational_scope() else None
        visible_users = {user.id: user for user in _visible_personnel_query(recipient_scope_airport).all()}
        visible_user_ids = set(visible_users.keys())

        recipient_ids = list(dict.fromkeys(_parse_int_list(request.form.getlist("recipient_ids"))))
        recipient_ids = [user_id for user_id in recipient_ids if user_id in visible_user_ids]
        if not recipient_ids:
            flash("En az bir teslim alan personel seçin.", "danger")
            return redirect(url_for("inventory.zimmetler", airport_id=airport_id or None))

        if _can_view_all_operational_scope():
            recipient_airport_ids = {
                visible_users[user_id].havalimani_id
                for user_id in recipient_ids
                if user_id in visible_users
            }
            recipient_airport_ids_non_null = {airport_id for airport_id in recipient_airport_ids if airport_id is not None}
            if len(recipient_airport_ids_non_null) == 1:
                # Kullanıcı seçim sırasından bağımsız olarak tek havalimanı kapsamını otomatik hizala.
                airport_id = next(iter(recipient_airport_ids_non_null))
            if len(recipient_airport_ids_non_null) > 1:
                flash("Seçilen personeller tek bir havalimanı kapsamında olmalıdır.", "danger")
                return redirect(url_for("inventory.zimmetler", airport_id=airport_id or None))
            if airport_id is None and not selected_item_ids:
                flash("Zimmet oluşturmak için bir havalimanı seçin.", "danger")
                return redirect(url_for("inventory.zimmetler"))
            if airport_id is not None and not any(airport.id == airport_id for airport in visible_airports):
                flash("Seçilen havalimanı için zimmet oluşturma yetkiniz yok.", "danger")
                return redirect(url_for("inventory.zimmetler", airport_id=selected_airport or None))

        if _can_view_all_operational_scope() and selected_item_ids:
            selected_material_rows = (
                _visible_material_query(None)
                .filter(Malzeme.id.in_(selected_item_ids))
                .all()
            )
            selected_material_airport_ids = {
                row.havalimani_id
                for row in selected_material_rows
                if row.havalimani_id is not None
            }
            if len(selected_material_airport_ids) > 1:
                flash("Seçilen malzemeler tek bir havalimanı kapsamında olmalıdır.", "danger")
                return redirect(url_for("inventory.zimmetler", airport_id=airport_id or None))

            recipient_airport_id = next(iter(recipient_airport_ids)) if len(recipient_airport_ids) == 1 else None
            material_airport_id = next(iter(selected_material_airport_ids)) if len(selected_material_airport_ids) == 1 else None
            if recipient_airport_id and material_airport_id and recipient_airport_id != material_airport_id:
                flash("Seçilen personel ve malzemeler aynı havalimanı kapsamında olmalıdır.", "danger")
                return redirect(url_for("inventory.zimmetler", airport_id=airport_id or None))

            resolved_airport_id = recipient_airport_id or material_airport_id or airport_id
            if resolved_airport_id is not None:
                airport_id = resolved_airport_id
            if airport_id is None:
                flash("Zimmet oluşturmak için bir havalimanı seçin.", "danger")
                return redirect(url_for("inventory.zimmetler"))
            if not any(airport.id == airport_id for airport in visible_airports):
                flash("Seçilen havalimanı için zimmet oluşturma yetkiniz yok.", "danger")
                return redirect(url_for("inventory.zimmetler", airport_id=selected_airport or None))

        visible_materials = {
            item.id: item
            for item in _visible_material_query(airport_id).order_by(Malzeme.ad.asc()).all()
        }
        selected_items = [visible_materials[item_id] for item_id in selected_item_ids if item_id in visible_materials]
        if not selected_items:
            flash("En az bir malzeme seçin.", "danger")
            return redirect(url_for("inventory.zimmetler", airport_id=airport_id or None))

        try:
            assignment = AssignmentRecord(
                assignment_no=_next_assignment_no(),
                assignment_date=_parse_date(request.form.get("assignment_date")) or get_tr_now().date(),
                delivered_by_id=current_user.id,
                delivered_by_name=delivered_by_name,
                airport_id=airport_id,
                note=guvenli_metin(request.form.get("note") or ""),
                status="active",
                created_by_id=current_user.id,
            )
            db.session.add(assignment)
            db.session.flush()

            for user_id in recipient_ids:
                db.session.add(AssignmentRecipient(assignment_id=assignment.id, user_id=user_id))

            created_item_count = 0
            for material in selected_items:
                quantity = request.form.get(f"item_qty_{material.id}", type=float) or float(material.stok_miktari or 1)
                quantity = max(quantity, 0)
                if quantity <= 0:
                    continue
                db.session.add(
                    AssignmentItem(
                        assignment_id=assignment.id,
                        material_id=material.id,
                        asset_id=material.linked_asset.id if material.linked_asset else None,
                        item_name=material.ad,
                        quantity=quantity,
                        unit=guvenli_metin(request.form.get(f"item_unit_{material.id}") or "adet") or "adet",
                        note=guvenli_metin(request.form.get(f"item_note_{material.id}") or ""),
                    )
                )
                created_item_count += 1

            if created_item_count == 0:
                db.session.rollback()
                flash("Seçilen malzemeler için geçerli miktar bulunamadı.", "danger")
                return redirect(url_for("inventory.zimmetler", airport_id=airport_id or None))

            _append_assignment_history(assignment, "created", "Zimmet kaydı oluşturuldu.")
            _register_assignment_for_active_demo_scope(assignment)
            db.session.commit()
        except Exception:
            db.session.rollback()
            current_app.logger.exception(
                "Zimmet olusturma basarisiz | user_id=%s airport_id=%s",
                getattr(current_user, "id", None),
                airport_id,
            )
            flash("Zimmet kaydı oluşturulamadı. Lütfen tekrar deneyin.", "danger")
            audit_log("assignment.create", outcome="failed", airport_id=airport_id, user_id=getattr(current_user, "id", None))
            return redirect(url_for("inventory.zimmetler", airport_id=airport_id or None))
        log_kaydet(
            "Zimmet",
            f"Zimmet oluşturuldu: {assignment.assignment_no}",
            event_key="assignment.create",
            target_model="AssignmentRecord",
            target_id=assignment.id,
        )
        audit_log("assignment.create", outcome="success", assignment_id=assignment.id, airport_id=airport_id)
        flash("Zimmet kaydı oluşturuldu.", "success")
        flash("Zimmet formu PDF olarak otomatik indiriliyor.", "info")
        return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id, auto_pdf=1))

    query = _assignment_scope()
    if selected_status:
        query_values = _assignment_status_query_values(selected_status)
        if query_values:
            query = query.filter(AssignmentRecord.status.in_(query_values))
    if selected_recipient:
        query = query.join(AssignmentRecipient).filter(AssignmentRecipient.user_id == selected_recipient)
    if selected_airport:
        query = query.filter(AssignmentRecord.airport_id == selected_airport)
    assignments = query.order_by(AssignmentRecord.assignment_date.desc(), AssignmentRecord.created_at.desc()).all()

    recipient_query = _visible_personnel_query(selected_airport)
    if not can_create_assignment and not has_permission("assignment.manage"):
        recipient_query = recipient_query.filter(Kullanici.id == current_user.id)
    recipient_options = recipient_query.order_by(Kullanici.tam_ad.asc()).all()
    material_options = _visible_material_query(selected_airport).order_by(Malzeme.ad.asc()).limit(250).all()
    recipient_lookup = {user.id: user for user in recipient_options}
    selected_recipient_user = recipient_lookup.get(selected_recipient)
    if not selected_recipient_user and not can_create_assignment and getattr(current_user, "id", None) in recipient_lookup:
        selected_recipient_user = recipient_lookup[current_user.id]

    recipient_active_assignments = []
    if selected_recipient_user:
        recipient_assignment_query = _assignment_scope()
        if can_create_assignment or selected_recipient_user.id != getattr(current_user, "id", None):
            recipient_assignment_query = recipient_assignment_query.join(AssignmentRecipient).filter(
                AssignmentRecipient.user_id == selected_recipient_user.id,
            )
        recipient_active_assignments = (
            recipient_assignment_query
            .filter(
                AssignmentRecord.status.in_(
                    _assignment_status_query_values("active") + _assignment_status_query_values("partial")
                )
            )
            .order_by(AssignmentRecord.assignment_date.desc(), AssignmentRecord.created_at.desc())
            .distinct()
            .all()
        )

    return render_template(
        "zimmetler.html",
        assignments=assignments,
        airports=visible_airports,
        recipient_options=recipient_options,
        material_options=material_options,
        assignment_status_labels=ASSIGNMENT_STATUS_LABELS,
        selected_airport=selected_airport,
        selected_status=selected_status,
        selected_recipient=selected_recipient,
        selected_recipient_user=selected_recipient_user,
        recipient_active_assignments=recipient_active_assignments,
        can_create_assignment=can_create_assignment,
        can_delete_assignment=can_delete_assignment,
        assignment_status_label=_assignment_status_label,
        normalize_assignment_status=_normalize_assignment_status,
        format_assignment_qty=_format_assignment_quantity,
    )


@inventory_bp.route("/zimmetler/<int:assignment_id>")
@login_required
@permission_required("assignment.view")
def zimmet_detay(assignment_id):
    assignment = _assignment_scope().filter(AssignmentRecord.id == assignment_id).first()
    if assignment is None:
        flash("Zimmet kaydı bulunamadı veya erişim izniniz yok.", "warning")
        return redirect(url_for("inventory.zimmetler"))
    return render_template(
        "zimmet_detay.html",
        assignment=assignment,
        assignment_status_label=_assignment_status_label,
        can_manage_assignment=has_permission("assignment.manage"),
        can_delete_assignment=_can_delete_assignments(),
        can_upload_assignment_document=has_permission("assignment.document.upload"),
        can_download_assignment_pdf=has_permission("assignment.pdf"),
        format_assignment_qty=_format_assignment_quantity,
    )


@inventory_bp.route("/zimmet/<int:assignment_id>")
@login_required
@permission_required("assignment.view")
def zimmet_detay_legacy(assignment_id):
    return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment_id), code=302)


@inventory_bp.route("/zimmetler/<int:assignment_id>/pdf")
@login_required
@permission_required("assignment.pdf")
def zimmet_pdf(assignment_id):
    assignment = _assignment_scope().filter(AssignmentRecord.id == assignment_id).first()
    if assignment is None:
        flash("PDF oluşturulacak zimmet kaydı bulunamadı.", "warning")
        return redirect(url_for("inventory.zimmetler"))
    font_uris = _assignment_pdf_font_uris()
    html = render_template(
        "zimmet_pdf.html",
        assignment=assignment,
        assignment_status_label=_assignment_status_label,
        generated_at=get_tr_now(),
        pdf_font_regular=font_uris["regular"],
        pdf_font_bold=font_uris["bold"],
        pdf_logo_uri=_assignment_pdf_logo_uri(),
        format_assignment_qty=_format_assignment_quantity,
    )
    output = io.BytesIO()
    pdf_result = pisa.CreatePDF(
        html,
        dest=output,
        encoding="utf-8",
        link_callback=_pdf_link_callback,
    )
    if pdf_result.err:
        current_app.logger.error("Zimmet PDF olusturulamadi | assignment_id=%s", assignment.id)
        abort(500)
    output.seek(0)
    audit_log("assignment.pdf", outcome="success", assignment_id=assignment.id)
    return send_file(
        output,
        download_name=f"{assignment.assignment_no}.pdf",
        as_attachment=True,
        mimetype="application/pdf",
    )


@inventory_bp.route("/zimmet/<int:assignment_id>/pdf")
@login_required
@permission_required("assignment.pdf")
def zimmet_pdf_legacy(assignment_id):
    return redirect(url_for("inventory.zimmet_pdf", assignment_id=assignment_id), code=302)


@inventory_bp.route("/zimmetler/<int:assignment_id>/signed-document", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("assignment.document.upload")
def zimmet_imzali_belge_yukle(assignment_id):
    assignment = _assignment_scope().filter(AssignmentRecord.id == assignment_id).first_or_404()
    upload = request.files.get("signed_document")
    safe_name, error = _validate_upload(
        upload,
        SIGNED_ASSIGNMENT_ALLOWED_EXTENSIONS,
        ("application/pdf",),
    )
    if error:
        flash(error, "danger")
        return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))

    filename = _assignment_signed_document_filename(assignment)
    display_name = safe_display_filename(
        getattr(upload, "filename", "") or filename,
        fallback=filename,
        default_extension=".pdf",
        max_length=180,
    )
    folder = _assignment_signed_document_folder(assignment)
    stored = get_storage_adapter().save_upload(upload, folder=folder, filename=filename)
    assignment.signed_document_key = stored.storage_key
    assignment.signed_document_url = stored.public_url
    assignment.signed_document_name = display_name
    _append_assignment_history(assignment, "signed_upload", "İmzalı belge yüklendi.")
    db.session.commit()

    log_kaydet(
        "Zimmet",
        f"İmzalı belge yüklendi: {assignment.assignment_no}",
        event_key="assignment.document.upload",
        target_model="AssignmentRecord",
        target_id=assignment.id,
    )
    flash("İmzalı zimmet belgesi yüklendi.", "success")
    return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))


@inventory_bp.route("/zimmetler/<int:assignment_id>/signed-document/download")
@login_required
@permission_required("assignment.view")
def zimmet_imzali_belge_indir(assignment_id):
    assignment = _assignment_scope().filter(AssignmentRecord.id == assignment_id).first_or_404()
    storage_adapter = get_storage_adapter()
    storage_key = (assignment.signed_document_key or "").strip()
    if not storage_key and assignment.signed_document_url:
        storage_key = storage_adapter.storage_key_from_public_url(assignment.signed_document_url) or ""
    if not storage_key:
        flash("Bu zimmet için yüklü imzalı belge bulunmuyor.", "warning")
        return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))
    try:
        signed_document = storage_adapter.read_bytes(storage_key)
    except FileNotFoundError:
        current_app.logger.warning(
            "Zimmet imzali belge bulunamadi | assignment_id=%s storage_key=%s",
            assignment.id,
            storage_key,
        )
        flash("Bu zimmet için yüklü imzalı belgeye erişilemiyor.", "warning")
        return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))
    except Exception:
        current_app.logger.exception(
            "Zimmet imzali belge indirilirken hata olustu | assignment_id=%s storage_key=%s",
            assignment.id,
            storage_key,
        )
        flash("Bu zimmet için yüklü imzalı belgeye erişilemiyor.", "warning")
        return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))
    download_name = safe_display_filename(
        assignment.signed_document_name,
        fallback=f"{assignment.assignment_no}.pdf",
        default_extension=".pdf",
        max_length=180,
    )
    return send_file(
        io.BytesIO(signed_document),
        download_name=download_name,
        as_attachment=True,
        mimetype="application/pdf",
    )


@inventory_bp.route("/zimmetler/<int:assignment_id>/iade", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("assignment.manage")
def zimmet_iade(assignment_id):
    assignment = _assignment_scope().filter(AssignmentRecord.id == assignment_id).first_or_404()
    processed = 0
    for item in assignment.items:
        returned_amount = request.form.get(f"return_qty_{item.id}", type=float) or 0
        if returned_amount <= 0:
            continue
        item.returned_quantity = min(float(item.quantity or 0), float(item.returned_quantity or 0) + returned_amount)
        item.returned_at = get_tr_now()
        item.returned_by_id = current_user.id
        note = guvenli_metin(request.form.get(f"return_note_{item.id}") or "")
        if note:
            item.return_note = note
        processed += 1

    if processed == 0:
        flash("İade işlemi için en az bir kalemde miktar girin.", "danger")
        return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))

    _recalculate_assignment_status(assignment)
    _append_assignment_history(
        assignment,
        "return",
        f"{processed} kalem için iade işlemi kaydedildi. Durum: {_assignment_status_label(assignment.status)}",
    )
    db.session.commit()
    flash("İade işlemi kaydedildi.", "success")
    return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))


@inventory_bp.route("/zimmetler/<int:assignment_id>/sil", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("assignment.view")
def zimmet_sil(assignment_id):
    if not _can_delete_assignments():
        abort(403)
    assignment = _assignment_scope().filter(AssignmentRecord.id == assignment_id).first_or_404()
    assignment_no = assignment.assignment_no
    assignment.soft_delete()
    db.session.commit()
    log_kaydet(
        "Zimmet",
        f"Zimmet kaydı arşive taşındı: {assignment_no}",
        event_key="assignment.delete",
        target_model="AssignmentRecord",
        target_id=assignment.id,
    )
    audit_log("assignment.delete", outcome="success", assignment_id=assignment.id)
    flash("Zimmet kaydı silindi ve arşive taşındı.", "warning")
    return redirect(url_for("inventory.zimmetler"))


@inventory_bp.route("/zimmetler/<int:assignment_id>/durum", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("assignment.manage")
def zimmet_durum_guncelle(assignment_id):
    assignment = _assignment_scope().filter(AssignmentRecord.id == assignment_id).first_or_404()
    new_status = _normalize_assignment_status(request.form.get("status"))
    if new_status not in ASSIGNMENT_STATUS_LABELS:
        flash("Geçersiz zimmet durumu.", "danger")
        return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))
    assignment.status = new_status
    _append_assignment_history(assignment, "status", f"Durum güncellendi: {_assignment_status_label(new_status)}")
    db.session.commit()
    flash("Zimmet durumu güncellendi.", "success")
    return redirect(url_for("inventory.zimmet_detay", assignment_id=assignment.id))


def _ppe_airport_display(airport):
    if not airport:
        return ""
    if getattr(airport, "kodu", None) and getattr(airport, "ad", None):
        return f"{airport.kodu} - {airport.ad}"
    return getattr(airport, "kodu", None) or getattr(airport, "ad", None) or ""


def _ppe_visible_users(selected_airport=None):
    query = _visible_personnel_query(selected_airport)
    if not has_permission("ppe.manage"):
        query = query.filter(Kullanici.id == current_user.id)
    return query.order_by(Kullanici.tam_ad.asc()).all()


def _ppe_field_requirements(category, subcategory):
    return {
        "apparel": _ppe_requires_apparel_size(category, subcategory),
        "shoe": _ppe_requires_shoe_size(category, subcategory),
    }


def _ppe_form_get(source, key, *, default=None, cast=None):
    value = default
    if hasattr(source, "get"):
        try:
            value = source.get(key, default=default, type=cast) if cast else source.get(key, default)
        except TypeError:
            value = source.get(key, default)
    if cast and value not in (None, "") and not isinstance(value, cast):
        try:
            value = cast(value)
        except (TypeError, ValueError):
            return default
    return value


def _normalize_ppe_form_payload(form, *, visible_users, allow_unassigned=False, default_airport_id=None):
    user_id = _ppe_form_get(form, "user_id", cast=int)
    user = visible_users.get(user_id) if user_id else None
    if not user and not allow_unassigned:
        raise ValueError("KKD kaydı için geçerli bir personel seçin.")

    category = guvenli_metin(_ppe_form_get(form, "category") or "").strip()
    subcategory = guvenli_metin(_ppe_form_get(form, "subcategory") or "").strip()
    item_name = guvenli_metin(_ppe_form_get(form, "item_name") or "").strip()
    brand = guvenli_metin(_ppe_form_get(form, "brand") or "").strip()
    model_name = guvenli_metin(_ppe_form_get(form, "model_name") or "").strip()
    serial_no = guvenli_metin(_ppe_form_get(form, "serial_no") or "").strip()
    apparel_size = guvenli_metin(_ppe_form_get(form, "apparel_size") or "").strip()
    shoe_size = guvenli_metin(_ppe_form_get(form, "shoe_size") or "").strip()
    manufacturer_url = guvenli_metin(_ppe_form_get(form, "manufacturer_url") or "").strip()
    physical_condition = str(_ppe_form_get(form, "physical_condition") or "iyi").strip()
    is_active = str(_ppe_form_get(form, "is_active") or "1").strip() in {"1", "true", "True", "evet", "Evet", "on"}
    requirements = _ppe_field_requirements(category, subcategory)

    if category not in _ppe_category_options():
        raise ValueError("Geçerli bir KKD kategorisi seçin.")
    if subcategory not in _ppe_subtype_options(category):
        raise ValueError("Seçilen kategori için geçerli bir alt tür seçin.")
    if not item_name:
        raise ValueError("KKD donanım adı zorunludur.")
    if physical_condition not in PPE_PHYSICAL_CONDITION_LABELS:
        raise ValueError("Geçerli bir fiziksel durum seçin.")
    if manufacturer_url and not _is_valid_url(manufacturer_url):
        raise ValueError("Üretici sayfası için geçerli bir bağlantı girin.")
    if requirements["apparel"]:
        if apparel_size not in PPE_APPAREL_SIZES:
            raise ValueError("Bu KKD türü için geçerli bir beden seçin.")
        shoe_size = ""
    elif requirements["shoe"]:
        if shoe_size not in PPE_SHOE_SIZES:
            raise ValueError("Bu KKD türü için geçerli bir ayakkabı numarası seçin.")
        apparel_size = ""
    else:
        apparel_size = ""
        shoe_size = ""

    delivered_at = _parse_date(_ppe_form_get(form, "delivered_at")) or get_tr_now().date()
    production_date = _parse_date(_ppe_form_get(form, "production_date"))
    expiry_date = _parse_date(_ppe_form_get(form, "expiry_date"))
    if production_date and expiry_date and expiry_date < production_date:
        raise ValueError("Son kullanma tarihi üretim tarihinden önce olamaz.")

    ppe_assignment_id = _ppe_form_get(form, "ppe_assignment_id", cast=int)
    linked_ppe_assignment = None
    if ppe_assignment_id:
        linked_ppe_assignment = (
            _ppe_assignment_scope()
            .filter(
                PPEAssignmentRecord.id == ppe_assignment_id,
                PPEAssignmentRecord.status == "active",
            )
            .first()
        )
        if linked_ppe_assignment is None:
            ppe_assignment_id = None

    resolved_airport_id = (
        (user.havalimani_id if user else None)
        or (linked_ppe_assignment.airport_id if linked_ppe_assignment else None)
        or default_airport_id
        or current_user.havalimani_id
    )
    if not resolved_airport_id:
        raise ValueError("KKD kaydı için geçerli bir havalimanı seçin.")
    if linked_ppe_assignment and linked_ppe_assignment.airport_id and linked_ppe_assignment.airport_id != resolved_airport_id:
        raise ValueError("Bağlı aktif KKD zimmeti seçilen havalimanı ile uyumlu olmalıdır.")

    return {
        "user": user,
        "user_id": user.id if user else None,
        "airport_id": resolved_airport_id,
        "assignment_id": None,
        "ppe_assignment_id": ppe_assignment_id,
        "category": category,
        "subcategory": subcategory,
        "item_name": item_name,
        "brand": brand,
        "model_name": model_name,
        "serial_no": serial_no,
        "brand_model": " ".join(part for part in [brand, model_name] if part).strip(),
        "apparel_size": apparel_size,
        "shoe_size": shoe_size,
        "size_info": _ppe_combined_size(apparel_size, shoe_size),
        "delivered_at": delivered_at,
        "production_date": production_date,
        "expiry_date": expiry_date,
        "quantity": max(_ppe_form_get(form, "quantity", cast=int, default=1) or 1, 1),
        "status": "aktif" if is_active else "kullanim_disi",
        "physical_condition": physical_condition,
        "is_active": is_active,
        "manufacturer_url": manufacturer_url,
    }


def _ppe_group_records(records, visible_users, selected_user):
    user_lookup = {user.id: user for user in visible_users}
    grouped = {}
    for record in records:
        grouped.setdefault(record.user_id, []).append(record)
    unassigned_records = grouped.get(None, [])
    items = []
    for user_id, user in user_lookup.items():
        user_records = grouped.get(user_id, [])
        if selected_user and user_id != selected_user and not user_records:
            continue
        items.append(
            {
                "user": user,
                "records": user_records,
                "open": bool(selected_user == user_id),
            }
        )
    if unassigned_records and not selected_user:
        pool_airport = None
        for row in unassigned_records:
            if getattr(row, "airport", None):
                pool_airport = row.airport
                break
        pool_user = type(
            "PPEPoolGroup",
            (),
            {
                "tam_ad": "Havuza Eklenen KKD Kayıtları",
                "havalimani": pool_airport,
                "rol": "kkd_havuz",
            },
        )()
        items.insert(
            0,
            {
                "user": pool_user,
                "records": unassigned_records,
                "open": False,
            },
        )
    if selected_user and selected_user not in grouped and selected_user in user_lookup:
        return items
    if not items and not selected_user and current_user.is_authenticated and getattr(current_user, "id", None) in grouped:
        items.append({"user": current_user, "records": grouped.get(current_user.id, []), "open": False})
    return items


def _ppe_template_headers():
    return [
        "havalimani",
        "personel",
        "kategori",
        "alt_tur",
        "donanim_adi",
        "marka",
        "model",
        "seri_no",
        "beden",
        "ayakkabi_numarasi",
        "teslim_tarihi",
        "uretim_tarihi",
        "son_kullanim_tarihi",
        "miktar",
        "fiziksel_durum",
        "aktif_mi",
        "uretici_linki",
    ]


def _build_ppe_template_workbook(*, airports, users):
    workbook = Workbook()
    ws_data = workbook.active
    ws_data.title = "VERI_GIRISI"
    ws_lists = workbook.create_sheet("LISTELER")
    ws_help = workbook.create_sheet("YARDIM")
    headers = _ppe_template_headers()
    ws_data.append(headers)
    ws_data.freeze_panes = "A2"

    list_columns = [
        ("A", "havalimani", [_ppe_airport_display(item) for item in airports]),
        ("B", "personel", [item.tam_ad for item in users]),
        ("C", "kategori", list(_ppe_category_options().keys())),
        ("D", "alt_tur", sorted({sub for values in _ppe_category_options().values() for sub in values})),
        ("E", "beden", PPE_APPAREL_SIZES),
        ("F", "ayakkabi_numarasi", PPE_SHOE_SIZES),
        ("G", "fiziksel_durum", list(PPE_PHYSICAL_CONDITION_LABELS.values())),
        ("H", "evet_hayir", ["Evet", "Hayır"]),
    ]
    for column, title, values in list_columns:
        ws_lists[f"{column}1"] = title
        for idx, value in enumerate(values, start=2):
            ws_lists[f"{column}{idx}"] = value

    header_to_col = {header: idx + 1 for idx, header in enumerate(headers)}

    def _add_dropdown(header, formula):
        column_letter = ws_data.cell(row=1, column=header_to_col[header]).column_letter
        validator = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws_data.add_data_validation(validator)
        validator.add(f"{column_letter}2:{column_letter}5000")

    _add_dropdown("havalimani", "'LISTELER'!$A$2:$A$5000")
    _add_dropdown("personel", "'LISTELER'!$B$2:$B$5000")
    _add_dropdown("kategori", "'LISTELER'!$C$2:$C$5000")
    _add_dropdown("alt_tur", "'LISTELER'!$D$2:$D$5000")
    _add_dropdown("beden", "'LISTELER'!$E$2:$E$5000")
    _add_dropdown("ayakkabi_numarasi", "'LISTELER'!$F$2:$F$5000")
    _add_dropdown("fiziksel_durum", "'LISTELER'!$G$2:$G$5000")
    _add_dropdown("aktif_mi", "'LISTELER'!$H$2:$H$3")

    ws_help["A1"] = "KKD EXCEL ŞABLONU"
    ws_help["A2"] = "Kategori ve alt tür değerleri sabit katalogla eşleşmelidir."
    ws_help["A3"] = "Tarih formatı: YYYY-MM-DD veya DD.MM.YYYY"
    ws_help["A4"] = "Aktif mi alanı için: Evet/Hayır"
    ws_help["A5"] = "Üretici bağlantısı varsa http/https ile başlamalıdır."
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _parse_ppe_workbook(file_obj):
    try:
        workbook = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # pragma: no cover
        raise ExcelTemplateError("KKD Excel dosyası okunamadı.") from exc

    for name in ("VERI_GIRISI", "LISTELER", "YARDIM"):
        if name not in workbook.sheetnames:
            raise ExcelTemplateError(f"Eksik sheet: {name}")
    worksheet = workbook["VERI_GIRISI"]
    rows = list(worksheet.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ExcelTemplateError("VERI_GIRISI sheet'i boş.")
    headers = _ppe_template_headers()
    actual_headers = [str(cell or "").strip() for cell in rows[0]]
    if actual_headers != headers:
        raise ExcelTemplateError("KKD Excel başlık yapısı beklenen şablonla eşleşmiyor.")

    parsed_rows = []
    for row_no, row_values in enumerate(rows[1:], start=2):
        payload = {headers[idx]: row_values[idx] if idx < len(row_values) else None for idx in range(len(headers))}
        if not any(value not in (None, "") for value in payload.values()):
            continue
        parsed_rows.append({"row_no": row_no, "values": payload})
    return parsed_rows


def _resolve_ppe_import_airport(raw_value, visible_airports):
    value = guvenli_metin(raw_value or "").strip()
    if not value:
        return None
    normalized = normalize_lookup(value)
    for airport in visible_airports:
        labels = {
            normalize_lookup(_ppe_airport_display(airport)),
            normalize_lookup(airport.kodu or ""),
            normalize_lookup(airport.ad or ""),
        }
        if normalized in labels:
            return airport
    raise ValueError("Havalimanı eşleşmedi.")


def _resolve_ppe_import_user(raw_value, users, airport_id=None):
    value = guvenli_metin(raw_value or "").strip()
    if not value:
        raise ValueError("Personel alanı zorunludur.")
    normalized = normalize_lookup(value)
    matches = [
        user for user in users
        if normalize_lookup(user.tam_ad or "") == normalized and (airport_id is None or user.havalimani_id == airport_id)
    ]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise ValueError("Personel adı birden fazla kayıtla eşleşti.")
    raise ValueError("Personel bulunamadı.")


@inventory_bp.route("/kkd", methods=["GET", "POST"], endpoint="kkd_listesi")
@login_required
@permission_required("ppe.view")
def kkd_listesi():
    try:
        _ensure_kkd_schema_ready()
    except RuntimeError as exc:
        current_app.logger.error("KKD ekranı şema kontrolü başarısız: %s", exc)
        flash(str(exc), "danger")
        return redirect(url_for("inventory.envanter"))

    selected_status = (request.args.get("status") or "").strip()
    selected_user = request.args.get("user_id", type=int)
    selected_airport = request.args.get("airport_id", type=int)

    visible_airports = _visible_operational_airports()
    can_manage_ppe = has_permission("ppe.manage")
    can_create_ppe_assignment = _can_issue_ppe_assignments(current_user)
    if not _can_view_all_operational_scope() and can_manage_ppe:
        selected_airport = current_user.havalimani_id
    visible_users_list = _ppe_visible_users(selected_airport)
    visible_users = {item.id: item for item in visible_users_list}
    if selected_user and selected_user not in visible_users and not can_manage_ppe:
        selected_user = current_user.id

    if request.method == "POST":
        if not can_manage_ppe:
            abort(403)

        try:
            payload = _normalize_ppe_form_payload(
                request.form,
                visible_users=visible_users,
                allow_unassigned=True,
                default_airport_id=selected_airport,
            )
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None, user_id=selected_user or None))

        upload = request.files.get("photo_document")
        photo_key = None
        photo_url = None
        if upload and upload.filename:
            safe_name, error = _validate_upload(
                upload,
                PPE_ALLOWED_EXTENSIONS,
                ("application/pdf", "image/"),
            )
            if error:
                flash(error, "danger")
                return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None, user_id=payload["user_id"]))
            user_token = payload["user_id"] or "pool"
            filename = f"ppe_{user_token}_{int(get_tr_now().timestamp())}_{safe_name}"
            stored = get_storage_adapter().save_upload(upload, folder="ppe", filename=filename)
            photo_key = stored.storage_key
            photo_url = stored.public_url

        record = PPERecord(
            user_id=payload["user_id"],
            airport_id=payload["airport_id"],
            assignment_id=payload["assignment_id"],
            ppe_assignment_id=payload["ppe_assignment_id"],
            category=payload["category"],
            subcategory=payload["subcategory"],
            item_name=payload["item_name"],
            brand=payload["brand"],
            model_name=payload["model_name"],
            serial_no=payload["serial_no"],
            brand_model=payload["brand_model"],
            apparel_size=payload["apparel_size"],
            shoe_size=payload["shoe_size"],
            size_info=payload["size_info"],
            delivered_at=payload["delivered_at"],
            production_date=payload["production_date"],
            expiry_date=payload["expiry_date"],
            quantity=payload["quantity"],
            status=payload["status"],
            physical_condition=payload["physical_condition"],
            is_active=payload["is_active"],
            manufacturer_url=payload["manufacturer_url"],
            photo_storage_key=photo_key,
            photo_url=photo_url,
            created_by_id=current_user.id,
        )

        db.session.add(record)
        db.session.flush()
        _register_ppe_record_for_active_demo_scope(record)
        db.session.add(
            PPERecordEvent(
                ppe_record_id=record.id,
                event_type="create",
                status_after=record.status,
                event_note="KKD havuz kaydı oluşturuldu.",
                created_by_id=current_user.id,
            )
        )
        db.session.commit()
        flash("KKD kaydı oluşturuldu.", "success")
        return redirect(
            url_for(
                "inventory.kkd_listesi",
                user_id=payload["user_id"] or None,
                airport_id=payload["airport_id"],
            )
        )

    query = _ppe_scope().options(
        joinedload(PPERecord.user),
        joinedload(PPERecord.airport),
        joinedload(PPERecord.assignment),
        joinedload(PPERecord.ppe_assignment),
        joinedload(PPERecord.events).joinedload(PPERecordEvent.created_by),
    )
    if selected_status:
        query = query.filter(PPERecord.status == selected_status)
    if selected_user:
        query = query.filter(PPERecord.user_id == selected_user)
    if selected_airport:
        query = query.filter(PPERecord.airport_id == selected_airport)
    records = query.order_by(PPERecord.delivered_at.desc(), PPERecord.created_at.desc()).all()
    grouped_records = _ppe_group_records(records, visible_users_list, selected_user)
    ppe_rows_by_id, available_qty_map = _ppe_available_quantity_map(selected_airport)
    available_ppe_records = []
    for row in sorted(
        ppe_rows_by_id.values(),
        key=lambda item: (str(item.category or ""), str(item.item_name or ""), int(item.id or 0)),
    ):
        available_quantity = available_qty_map.get(row.id, 0)
        if available_quantity <= 0:
            continue
        if selected_user and row.user_id and row.user_id != selected_user:
            continue
        available_ppe_records.append(
            {
                "record": row,
                "available_quantity": _format_assignment_quantity(available_quantity),
            }
        )
    import_feedback = session.pop(_ppe_import_feedback_session_key(), None)
    ppe_assignments = (
        _ppe_assignment_scope()
        .options(
            joinedload(PPEAssignmentRecord.recipient_user),
            joinedload(PPEAssignmentRecord.airport),
            joinedload(PPEAssignmentRecord.items),
        )
        .order_by(PPEAssignmentRecord.assignment_date.desc(), PPEAssignmentRecord.created_at.desc())
        .limit(20)
        .all()
    )
    active_ppe_assignments = _ppe_linkable_assignments(selected_airport)

    return render_template(
        "kkd.html",
        records=records,
        grouped_records=grouped_records,
        visible_users=visible_users_list,
        visible_airports=visible_airports,
        active_ppe_assignments=active_ppe_assignments,
        ppe_status_labels=PPE_STATUS_LABELS,
        ppe_condition_labels=PPE_PHYSICAL_CONDITION_LABELS,
        ppe_categories=_ppe_category_options(),
        ppe_catalog_json=json.dumps(_ppe_category_options(), ensure_ascii=False),
        ppe_apparel_sizes=PPE_APPAREL_SIZES,
        ppe_shoe_sizes=PPE_SHOE_SIZES,
        ppe_alert_state=_ppe_alert_state,
        ppe_brand_model_display=_ppe_brand_model_display,
        ppe_size_display=_ppe_size_display,
        selected_status=selected_status,
        selected_user=selected_user,
        selected_airport=selected_airport,
        can_manage_ppe=can_manage_ppe,
        can_create_ppe_assignment=can_create_ppe_assignment,
        can_request_ppe=has_permission("ppe.request"),
        import_feedback=import_feedback,
        available_ppe_records=available_ppe_records,
        assignment_recipients=visible_users_list,
        ppe_assignments=ppe_assignments,
        ppe_assignment_status_label=_ppe_assignment_status_label,
        ppe_assignment_display_name=_ppe_assignment_display_name,
    )


@inventory_bp.route("/kkd/tahsis", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("ppe.manage")
def kkd_tahsis_olustur():
    try:
        _ensure_kkd_schema_ready()
    except RuntimeError as exc:
        current_app.logger.error("KKD tahsis oluşturma şema kontrolü başarısız: %s", exc)
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    if not _can_issue_ppe_assignments(current_user):
        abort(403)

    selected_airport = request.form.get("airport_id", type=int) or current_user.havalimani_id
    visible_users = {user.id: user for user in _ppe_visible_users(selected_airport)}
    recipient_user_id = request.form.get("recipient_user_id", type=int)
    recipient_user = visible_users.get(recipient_user_id)
    if not recipient_user:
        flash("Teslim alan personel seçimi zorunludur.", "danger")
        return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None))

    if selected_airport and recipient_user.havalimani_id and recipient_user.havalimani_id != selected_airport:
        flash("Teslim alan personel seçilen havalimanı ile uyumlu olmalıdır.", "danger")
        return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None))

    delivered_by_name = guvenli_metin(request.form.get("delivered_by_name") or "").strip()
    if not delivered_by_name:
        delivered_by_name = guvenli_metin(getattr(current_user, "tam_ad", "") or "").strip()
    if not delivered_by_name:
        flash("Teslim eden bilgisi zorunludur.", "danger")
        return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None))

    selected_item_ids = list(dict.fromkeys(_parse_int_list(request.form.getlist("ppe_record_ids"))))
    if not selected_item_ids:
        flash("Tahsis için en az bir aktif KKD kaydı seçin.", "danger")
        return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None))

    ppe_rows_by_id, available_qty_map = _ppe_available_quantity_map(selected_airport)
    selected_records = []
    for record_id in selected_item_ids:
        record = ppe_rows_by_id.get(record_id)
        if not record:
            continue
        if recipient_user.havalimani_id and record.airport_id and recipient_user.havalimani_id != record.airport_id:
            flash("KKD tahsisinde seçilen kalemler, teslim alan personel ile aynı havalimanında olmalıdır.", "danger")
            return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None))
        available_quantity = _to_float_safe(available_qty_map.get(record_id))
        if available_quantity <= 0:
            continue
        requested_quantity = request.form.get(f"ppe_qty_{record_id}", type=float)
        if requested_quantity is None:
            requested_quantity = 1.0
        requested_quantity = max(min(_to_float_safe(requested_quantity, 1.0), available_quantity), 0.0)
        if requested_quantity <= 0:
            continue
        selected_records.append(
            {
                "record": record,
                "quantity": requested_quantity,
                "unit": guvenli_metin(request.form.get(f"ppe_unit_{record_id}") or "adet").strip() or "adet",
                "note": guvenli_metin(request.form.get(f"ppe_note_{record_id}") or ""),
            }
        )

    if not selected_records:
        flash("Seçilen KKD kalemleri için geçerli miktar bulunamadı.", "danger")
        return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None))

    assignment_date = _parse_date(request.form.get("assignment_date")) or get_tr_now().date()
    airport_id = recipient_user.havalimani_id or selected_airport
    note = guvenli_metin(request.form.get("note") or "")

    try:
        assignment = PPEAssignmentRecord(
            assignment_no=_next_ppe_assignment_no(),
            assignment_date=assignment_date,
            delivered_by_id=current_user.id,
            delivered_by_name=delivered_by_name,
            recipient_user_id=recipient_user.id,
            airport_id=airport_id,
            note=note,
            status="active",
            created_by_id=current_user.id,
        )
        db.session.add(assignment)
        db.session.flush()
        _register_ppe_assignment_for_active_demo_scope(assignment)

        for item in selected_records:
            record = item["record"]
            db.session.add(
                PPEAssignmentItem(
                    assignment_id=assignment.id,
                    ppe_record_id=record.id,
                    item_name=record.item_name,
                    category=record.category,
                    subcategory=record.subcategory,
                    brand=record.brand,
                    model_name=record.model_name,
                    serial_no=record.serial_no,
                    size_info=record.size_info,
                    quantity=item["quantity"],
                    unit=item["unit"],
                    note=item["note"],
                )
            )
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception(
            "KKD tahsisi olusturma basarisiz | user_id=%s recipient_id=%s",
            getattr(current_user, "id", None),
            recipient_user_id,
        )
        flash("KKD tahsisi oluşturulamadı. Lütfen tekrar deneyin.", "danger")
        return redirect(url_for("inventory.kkd_listesi", airport_id=selected_airport or None))

    log_kaydet(
        "KKD",
        f"KKD tahsisi oluşturuldu: {assignment.assignment_no}",
        event_key="ppe.assignment.create",
        target_model="PPEAssignmentRecord",
        target_id=assignment.id,
    )
    audit_log(
        "ppe.assignment.create",
        outcome="success",
        assignment_id=assignment.id,
        recipient_user_id=assignment.recipient_user_id,
    )
    flash("KKD tahsisi oluşturuldu.", "success")
    flash("KKD teslim formu PDF olarak otomatik indiriliyor.", "info")
    return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id, auto_pdf=1))


@inventory_bp.route("/kkd/tahsisler/<int:assignment_id>")
@login_required
@permission_required("ppe.view")
def kkd_tahsis_detay(assignment_id):
    try:
        _ensure_kkd_schema_ready()
    except RuntimeError as exc:
        current_app.logger.error("KKD tahsis detay şema kontrolü başarısız: %s", exc)
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    assignment = _ppe_assignment_scope().filter(PPEAssignmentRecord.id == assignment_id).first()
    if assignment is None:
        flash("KKD tahsis kaydı bulunamadı veya erişim izniniz yok.", "warning")
        return redirect(url_for("inventory.kkd_listesi"))
    return render_template(
        "kkd_tahsis_detay.html",
        assignment=assignment,
        ppe_assignment_status_label=_ppe_assignment_status_label,
        ppe_assignment_display_name=_ppe_assignment_display_name,
        format_assignment_qty=_format_assignment_quantity,
        can_manage_ppe_assignment=_can_issue_ppe_assignments(current_user),
    )


@inventory_bp.route("/kkd/tahsisler/<int:assignment_id>/iade", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("ppe.manage")
def kkd_tahsis_iade(assignment_id):
    try:
        _ensure_kkd_schema_ready()
    except RuntimeError as exc:
        current_app.logger.error("KKD tahsis iade şema kontrolü başarısız: %s", exc)
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    if not _can_issue_ppe_assignments(current_user):
        abort(403)

    assignment = _ppe_assignment_scope().filter(PPEAssignmentRecord.id == assignment_id).first_or_404()
    if assignment.status != "active":
        flash("Sadece aktif KKD tahsisleri iade edilebilir.", "warning")
        return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id))

    assignment.status = "returned"
    assignment.returned_at = get_tr_now()
    assignment.returned_by_id = current_user.id
    assignment.returned_note = guvenli_metin(request.form.get("return_note") or "") or None
    db.session.commit()

    log_kaydet(
        "KKD",
        f"KKD tahsisi iade edildi: {assignment.assignment_no}",
        event_key="ppe.assignment.return",
        target_model="PPEAssignmentRecord",
        target_id=assignment.id,
    )
    audit_log("ppe.assignment.return", outcome="success", assignment_id=assignment.id)
    flash("KKD tahsisi iade edildi.", "success")
    return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id))


@inventory_bp.route("/kkd/tahsisler/<int:assignment_id>/sil", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("ppe.manage")
def kkd_tahsis_sil(assignment_id):
    try:
        _ensure_kkd_schema_ready()
    except RuntimeError as exc:
        current_app.logger.error("KKD tahsis silme şema kontrolü başarısız: %s", exc)
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    if not _can_issue_ppe_assignments(current_user):
        abort(403)

    assignment = _ppe_assignment_scope().filter(PPEAssignmentRecord.id == assignment_id).first_or_404()
    if assignment.status == "active":
        flash("Aktif KKD tahsisi silinemez. Önce iade işlemi yapın.", "danger")
        return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id))

    assignment_no = assignment.assignment_no
    assignment.soft_delete()
    db.session.commit()

    log_kaydet(
        "KKD",
        f"KKD tahsis kaydı arşive taşındı: {assignment_no}",
        event_key="ppe.assignment.delete",
        target_model="PPEAssignmentRecord",
        target_id=assignment.id,
    )
    audit_log("ppe.assignment.delete", outcome="success", assignment_id=assignment.id)
    flash("KKD tahsis kaydı silindi ve arşive taşındı.", "warning")
    return redirect(url_for("inventory.kkd_listesi"))


@inventory_bp.route("/kkd/tahsisler/<int:assignment_id>/pdf")
@login_required
@permission_required("ppe.view")
def kkd_tahsis_pdf(assignment_id):
    try:
        _ensure_kkd_schema_ready()
    except RuntimeError as exc:
        current_app.logger.error("KKD tahsis pdf şema kontrolü başarısız: %s", exc)
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    assignment = _ppe_assignment_scope().filter(PPEAssignmentRecord.id == assignment_id).first()
    if assignment is None:
        flash("PDF oluşturulacak KKD tahsis kaydı bulunamadı.", "warning")
        return redirect(url_for("inventory.kkd_listesi"))

    font_uris = _assignment_pdf_font_uris()
    if not font_uris.get("regular"):
        current_app.logger.error("KKD tahsis PDF fontu bulunamadi | assignment_id=%s", assignment.id)
        abort(500)

    html = render_template(
        "kkd_tahsis_pdf.html",
        assignment=assignment,
        generated_at=get_tr_now(),
        ppe_assignment_status_label=_ppe_assignment_status_label,
        ppe_assignment_display_name=_ppe_assignment_display_name,
        format_assignment_qty=_format_assignment_quantity,
        pdf_font_regular=font_uris["regular"],
        pdf_font_bold=font_uris["bold"],
        pdf_logo_uri=_assignment_pdf_logo_uri(),
    )
    output = io.BytesIO()
    pdf_result = pisa.CreatePDF(
        html,
        dest=output,
        encoding="utf-8",
        link_callback=_pdf_link_callback,
    )
    if pdf_result.err:
        current_app.logger.error("KKD tahsis PDF olusturulamadi | assignment_id=%s", assignment.id)
        abort(500)
    output.seek(0)
    audit_log("ppe.assignment.pdf", outcome="success", assignment_id=assignment.id)
    return send_file(
        output,
        download_name=f"{assignment.assignment_no}.pdf",
        as_attachment=True,
        mimetype="application/pdf",
    )


@inventory_bp.route("/kkd/tahsisler/<int:assignment_id>/signed-document", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("ppe.manage")
def kkd_tahsis_imzali_belge_yukle(assignment_id):
    try:
        _ensure_kkd_schema_ready()
    except RuntimeError as exc:
        current_app.logger.error("KKD imzalı belge yükleme şema kontrolü başarısız: %s", exc)
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    if not _can_issue_ppe_assignments(current_user):
        abort(403)
    assignment = _ppe_assignment_scope().filter(PPEAssignmentRecord.id == assignment_id).first_or_404()
    upload = request.files.get("signed_document")
    safe_name, error = _validate_upload(
        upload,
        SIGNED_ASSIGNMENT_ALLOWED_EXTENSIONS,
        ("application/pdf",),
    )
    if error:
        flash(error, "danger")
        return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id))

    filename = _ppe_assignment_signed_document_filename(assignment)
    display_name = safe_display_filename(
        getattr(upload, "filename", "") or filename,
        fallback=filename,
        default_extension=".pdf",
        max_length=180,
    )
    folder = _ppe_assignment_signed_document_folder(assignment)
    stored = get_storage_adapter().save_upload(upload, folder=folder, filename=filename)
    assignment.signed_document_key = stored.storage_key
    assignment.signed_document_url = stored.public_url
    assignment.signed_document_name = display_name

    drive_payload = None
    try:
        stream = getattr(upload, "stream", None)
        if stream is not None:
            stream.seek(0)
        drive_payload = _upload_ppe_signed_document_to_drive(assignment, upload, safe_name)
    except GoogleDriveError as exc:
        current_app.logger.warning(
            "KKD imzali belge Drive yuklemesi basarisiz | assignment_id=%s reason=%s",
            assignment.id,
            exc,
        )
        flash("Belge yerel depoya kaydedildi, Drive yüklemesi tamamlanamadı.", "warning")
    except Exception:
        current_app.logger.exception(
            "KKD imzali belge Drive yuklemesinde beklenmeyen hata | assignment_id=%s",
            assignment.id,
        )
        flash("Belge yerel depoya kaydedildi, Drive yüklemesinde hata oluştu.", "warning")

    if drive_payload:
        assignment.signed_document_drive_file_id = drive_payload.get("drive_file_id")
        assignment.signed_document_drive_folder_id = drive_payload.get("drive_folder_id")
    db.session.commit()

    log_kaydet(
        "KKD",
        f"KKD imzalı teslim belgesi yüklendi: {assignment.assignment_no}",
        event_key="ppe.assignment.document.upload",
        target_model="PPEAssignmentRecord",
        target_id=assignment.id,
    )
    flash("İmzalı KKD teslim belgesi yüklendi.", "success")
    return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id))


@inventory_bp.route("/kkd/tahsisler/<int:assignment_id>/signed-document/download")
@login_required
@permission_required("ppe.view")
def kkd_tahsis_imzali_belge_indir(assignment_id):
    try:
        _ensure_kkd_schema_ready()
    except RuntimeError as exc:
        current_app.logger.error("KKD imzalı belge indirme şema kontrolü başarısız: %s", exc)
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    assignment = _ppe_assignment_scope().filter(PPEAssignmentRecord.id == assignment_id).first_or_404()
    storage_adapter = get_storage_adapter()
    storage_key = (assignment.signed_document_key or "").strip()
    if not storage_key and assignment.signed_document_url:
        storage_key = storage_adapter.storage_key_from_public_url(assignment.signed_document_url) or ""
    if not storage_key:
        flash("Bu KKD tahsisi için yüklü imzalı belge bulunmuyor.", "warning")
        return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id))
    try:
        signed_document = storage_adapter.read_bytes(storage_key)
    except FileNotFoundError:
        current_app.logger.warning(
            "KKD tahsis imzali belge bulunamadi | assignment_id=%s storage_key=%s",
            assignment.id,
            storage_key,
        )
        flash("Bu KKD tahsisi için yüklü imzalı belgeye erişilemiyor.", "warning")
        return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id))
    except Exception:
        current_app.logger.exception(
            "KKD tahsis imzali belge indirilirken hata | assignment_id=%s storage_key=%s",
            assignment.id,
            storage_key,
        )
        flash("Bu KKD tahsisi için yüklü imzalı belgeye erişilemiyor.", "warning")
        return redirect(url_for("inventory.kkd_tahsis_detay", assignment_id=assignment.id))

    download_name = safe_display_filename(
        assignment.signed_document_name,
        fallback=f"{assignment.assignment_no}.pdf",
        default_extension=".pdf",
        max_length=180,
    )
    return send_file(
        io.BytesIO(signed_document),
        download_name=download_name,
        as_attachment=True,
        mimetype="application/pdf",
    )


@inventory_bp.route("/kkd/<int:record_id>/bildirim", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("ppe.request")
def kkd_bildirim(record_id):
    record = _ppe_scope().filter(PPERecord.id == record_id).first_or_404()
    if not has_permission("ppe.manage") and record.user_id != current_user.id:
        abort(403)

    new_status = (request.form.get("status") or "").strip()
    if has_permission("ppe.manage"):
        allowed_statuses = set(PPE_STATUS_LABELS.keys())
    else:
        allowed_statuses = {"eksik", "hasarli", "kayip", "kullanim_disi", "degisim_talebi"}
    if new_status not in allowed_statuses:
        flash("Geçersiz KKD durum bildirimi.", "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    note = guvenli_metin(request.form.get("event_note") or "")
    record.status = new_status
    if note:
        existing = (record.description or "").strip()
        record.description = f"{existing}\n{note}".strip() if existing else note

    db.session.add(
        PPERecordEvent(
            ppe_record_id=record.id,
            event_type="status_update" if has_permission("ppe.manage") else "user_report",
            status_after=new_status,
            event_note=note or "Durum güncellendi.",
            created_by_id=current_user.id,
        )
    )
    db.session.commit()
    flash("KKD bildirimi kaydedildi.", "success")
    return redirect(url_for("inventory.kkd_listesi", user_id=record.user_id if has_permission("ppe.manage") else None))


@inventory_bp.route("/kkd/<int:record_id>/signed-document", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("ppe.manage")
def kkd_signed_document_upload(record_id):
    record = _ppe_scope().filter(PPERecord.id == record_id).first_or_404()
    upload = request.files.get("signed_document")
    safe_name, error = _validate_upload(
        upload,
        SIGNED_ASSIGNMENT_ALLOWED_EXTENSIONS,
        ("application/pdf",),
    )
    if error:
        flash(error, "danger")
        return redirect(url_for("inventory.kkd_listesi", user_id=record.user_id, airport_id=record.airport_id))

    airport_label = getattr(getattr(record, "airport", None), "kodu", "") or getattr(getattr(record, "airport", None), "ad", "") or "global"
    person_label = getattr(getattr(record, "user", None), "tam_ad", "") or "personel"
    filename = _signed_document_filename_for_person(person_label)
    display_name = safe_display_filename(
        getattr(upload, "filename", "") or filename,
        fallback=filename,
        default_extension=".pdf",
        max_length=180,
    )
    folder = _signed_document_folder_for_person(airport_label, person_label)
    stored = get_storage_adapter().save_upload(upload, folder=folder, filename=filename)
    record.signed_document_key = stored.storage_key
    record.signed_document_url = stored.public_url
    record.signed_document_name = display_name
    db.session.add(
        PPERecordEvent(
            ppe_record_id=record.id,
            event_type="signed_upload",
            status_after=record.status,
            event_note="İmzalı KKD zimmet PDF'i yüklendi.",
            created_by_id=current_user.id,
        )
    )
    db.session.commit()
    flash("İmzalı KKD zimmet PDF'i yüklendi.", "success")
    return redirect(url_for("inventory.kkd_listesi", user_id=record.user_id, airport_id=record.airport_id))


@inventory_bp.route("/kkd/<int:record_id>/signed-document/download")
@login_required
@permission_required("ppe.view")
def kkd_signed_document_download(record_id):
    record = _ppe_scope().filter(PPERecord.id == record_id).first_or_404()
    storage_adapter = get_storage_adapter()
    storage_key = (record.signed_document_key or "").strip()
    if not storage_key and record.signed_document_url:
        storage_key = storage_adapter.storage_key_from_public_url(record.signed_document_url) or ""
    if not storage_key:
        flash("Bu KKD kaydı için yüklü imzalı PDF bulunmuyor.", "warning")
        return redirect(url_for("inventory.kkd_listesi", user_id=record.user_id if has_permission("ppe.manage") else None))
    try:
        signed_document = storage_adapter.read_bytes(storage_key)
    except FileNotFoundError:
        current_app.logger.warning(
            "KKD imzali belge bulunamadi | record_id=%s storage_key=%s",
            record.id,
            storage_key,
        )
        flash("Bu KKD kaydı için yüklü imzalı PDF'e erişilemiyor.", "warning")
        return redirect(url_for("inventory.kkd_listesi", user_id=record.user_id if has_permission("ppe.manage") else None))
    except Exception:
        current_app.logger.exception(
            "KKD imzali belge indirilirken hata olustu | record_id=%s storage_key=%s",
            record.id,
            storage_key,
        )
        flash("Bu KKD kaydı için yüklü imzalı PDF'e erişilemiyor.", "warning")
        return redirect(url_for("inventory.kkd_listesi", user_id=record.user_id if has_permission("ppe.manage") else None))
    download_name = safe_display_filename(
        record.signed_document_name,
        fallback="kkd_imzali_belge.pdf",
        default_extension=".pdf",
        max_length=180,
    )
    return send_file(
        io.BytesIO(signed_document),
        download_name=download_name,
        as_attachment=True,
        mimetype="application/pdf",
    )


@inventory_bp.route("/kkd/excel-sablon")
@login_required
@permission_required("ppe.manage")
def kkd_excel_sablon_indir():
    visible_airports = _visible_operational_airports()
    visible_users = _ppe_visible_users(request.args.get("airport_id", type=int))
    workbook = _build_ppe_template_workbook(airports=visible_airports, users=visible_users)
    return send_file(
        workbook,
        as_attachment=True,
        download_name=f"kkd_toplu_import_sablonu_{get_tr_now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@inventory_bp.route("/kkd/excel-yukle", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("ppe.manage")
def kkd_excel_yukle():
    upload = request.files.get("excel_file")
    if not upload or not upload.filename:
        flash("Excel dosyası seçilmedi.", "danger")
        return redirect(url_for("inventory.kkd_listesi"))
    safe_name = secure_upload_filename(upload.filename or "")
    if not safe_name.lower().endswith(".xlsx"):
        flash("Sadece .xlsx formatı desteklenir.", "danger")
        return redirect(url_for("inventory.kkd_listesi"))
    if not _is_valid_xlsx_workbook(upload):
        flash("KKD Excel dosyası okunamadı.", "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    try:
        rows = _parse_ppe_workbook(upload)
    except ExcelTemplateError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kkd_listesi"))

    visible_airports = _visible_operational_airports()
    visible_users = _ppe_visible_users()
    success_count = 0
    failure_count = 0
    feedback = []

    for row in rows:
        values = row["values"]
        savepoint = db.session.begin_nested()
        try:
            airport = _resolve_ppe_import_airport(values.get("havalimani"), visible_airports)
            user = _resolve_ppe_import_user(values.get("personel"), visible_users, getattr(airport, "id", None))
            physical_label = guvenli_metin(values.get("fiziksel_durum") or "").strip()
            physical_condition = next(
                (key for key, label in PPE_PHYSICAL_CONDITION_LABELS.items() if normalize_lookup(label) == normalize_lookup(physical_label)),
                None,
            )
            if not physical_condition:
                raise ValueError("Fiziksel durum değeri geçersiz.")
            form_payload = _normalize_ppe_form_payload(
                {
                    "user_id": user.id,
                    "category": values.get("kategori"),
                    "subcategory": values.get("alt_tur"),
                    "item_name": values.get("donanim_adi"),
                    "brand": values.get("marka"),
                    "model_name": values.get("model"),
                    "serial_no": values.get("seri_no"),
                    "apparel_size": values.get("beden"),
                    "shoe_size": values.get("ayakkabi_numarasi"),
                    "delivered_at": values.get("teslim_tarihi"),
                    "production_date": values.get("uretim_tarihi"),
                    "expiry_date": values.get("son_kullanim_tarihi"),
                    "quantity": values.get("miktar"),
                    "physical_condition": physical_condition,
                    "is_active": parse_flexible_bool(values.get("aktif_mi")),
                    "manufacturer_url": values.get("uretici_linki"),
                },
                visible_users={user.id: user},
            )
            record = PPERecord(
                user_id=form_payload["user_id"],
                airport_id=form_payload["airport_id"],
                category=form_payload["category"],
                subcategory=form_payload["subcategory"],
                item_name=form_payload["item_name"],
                brand=form_payload["brand"],
                model_name=form_payload["model_name"],
                serial_no=form_payload["serial_no"],
                brand_model=form_payload["brand_model"],
                apparel_size=form_payload["apparel_size"],
                shoe_size=form_payload["shoe_size"],
                size_info=form_payload["size_info"],
                delivered_at=form_payload["delivered_at"],
                production_date=form_payload["production_date"],
                expiry_date=form_payload["expiry_date"],
                quantity=form_payload["quantity"],
                status=form_payload["status"],
                physical_condition=form_payload["physical_condition"],
                is_active=form_payload["is_active"],
                manufacturer_url=form_payload["manufacturer_url"],
                created_by_id=current_user.id,
            )
            db.session.add(record)
            db.session.flush()
            db.session.add(
                PPERecordEvent(
                    ppe_record_id=record.id,
                    event_type="import",
                    status_after=record.status,
                    event_note=f"{safe_name} içe aktarma satırı alındı.",
                    created_by_id=current_user.id,
                )
            )
            savepoint.commit()
            success_count += 1
        except Exception as exc:  # noqa: BLE001
            savepoint.rollback()
            failure_count += 1
            feedback.append({"row_no": row["row_no"], "message": str(exc)})
            continue
    if success_count:
        db.session.commit()
        flash(f"{success_count} KKD satırı içe aktarıldı.", "success")
    else:
        db.session.rollback()
    if failure_count:
        flash(f"{failure_count} satır doğrulama hatası nedeniyle alınmadı.", "warning")
    session[_ppe_import_feedback_session_key()] = feedback[:50]
    return redirect(url_for("inventory.kkd_listesi"))


@inventory_bp.route("/kkd/export/<string:fmt>")
@login_required
@permission_required("ppe.manage")
def kkd_export(fmt):
    selected_status = (request.args.get("status") or "").strip()
    selected_airport = request.args.get("airport_id", type=int)
    selected_user = request.args.get("user_id", type=int)

    query = _ppe_scope()
    if selected_status:
        query = query.filter(PPERecord.status == selected_status)
    if selected_airport:
        query = query.filter(PPERecord.airport_id == selected_airport)
    if selected_user:
        query = query.filter(PPERecord.user_id == selected_user)

    records = query.options(
        joinedload(PPERecord.user),
        joinedload(PPERecord.airport),
    ).order_by(PPERecord.delivered_at.desc()).all()

    airport_label = "Tüm görünür havalimanları"
    if selected_airport:
        selected_airport_obj = next((item for item in _visible_operational_airports() if item.id == selected_airport), None)
        if selected_airport_obj:
            airport_label = f"{selected_airport_obj.kodu} - {selected_airport_obj.ad}"

    user_label = "Tüm personel / havuz"
    if selected_user:
        selected_user_obj = _visible_personnel_query(selected_airport).filter(Kullanici.id == selected_user).first()
        if selected_user_obj:
            user_label = selected_user_obj.tam_ad

    rows = [
        {
            "Personel": record.user.tam_ad if record.user else "Havuz Kaydı",
            "Havalimanı": record.airport.ad if record.airport else "-",
            "Kategori": record.category or "-",
            "Alt Tür": record.subcategory or "-",
            "KKD": record.item_name,
            "Marka": record.brand or "-",
            "Model": record.model_name or "-",
            "Seri No": record.serial_no or "-",
            "Beden / Numara": _ppe_size_display(record),
            "Teslim Tarihi": record.delivered_at.strftime("%d.%m.%Y") if record.delivered_at else "-",
            "Üretim Tarihi": record.production_date.strftime("%d.%m.%Y") if record.production_date else "-",
            "Son Kullanma": record.expiry_date.strftime("%d.%m.%Y") if record.expiry_date else "-",
            "Miktar": record.quantity,
            "Durum": _ppe_status_label(record.status),
            "Fiziksel Durum": _ppe_condition_label(record.physical_condition),
            "Aktif/Pasif": "Aktif" if record.is_active else "Pasif",
            "Üretici Linki": record.manufacturer_url or "-",
        }
        for record in records
    ]

    if fmt == "xlsx":
        payload = io.BytesIO()
        with pd.ExcelWriter(payload, engine="openpyxl") as writer:
            pd.DataFrame(rows).to_excel(writer, index=False)
        payload.seek(0)
        return send_file(
            payload,
            as_attachment=True,
            download_name=f"kkd_raporu_{get_tr_now().strftime('%Y%m%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    if fmt == "pdf":
        font_uris = _assignment_pdf_font_uris()
        if not font_uris.get("regular"):
            current_app.logger.error("KKD rapor PDF fontu bulunamadi")
            abort(500)
        html = render_template(
            "kkd_report_pdf.html",
            records=records,
            ppe_status_label=_ppe_status_label,
            ppe_condition_label=_ppe_condition_label,
            generated_at=get_tr_now(),
            selected_status=selected_status,
            selected_status_label=_ppe_status_label(selected_status) if selected_status else "Tüm durumlar",
            selected_airport_label=airport_label,
            selected_user_label=user_label,
            total_records=len(records),
            pdf_font_regular=font_uris["regular"],
            pdf_font_bold=font_uris["bold"],
            pdf_logo_uri=_assignment_pdf_logo_uri(),
        )
        payload = io.BytesIO()
        pdf_result = pisa.CreatePDF(
            html,
            dest=payload,
            encoding="utf-8",
            link_callback=_pdf_link_callback,
        )
        if pdf_result.err:
            current_app.logger.error("KKD rapor PDF olusturulamadi")
            abort(500)
        payload.seek(0)
        return send_file(
            payload,
            as_attachment=True,
            download_name=f"kkd_raporu_{get_tr_now().strftime('%Y%m%d')}.pdf",
            mimetype="application/pdf",
        )
    flash("Desteklenmeyen export formatı.", "danger")
    return redirect(url_for("inventory.kkd_listesi"))


@inventory_bp.route("/tatbikatlar")
@login_required
@permission_required("drill.view")
def tatbikat_listesi():
    selected_airport = request.args.get("airport_id", type=int)
    airports = _visible_drill_airports()
    visible_airport_ids = {airport.id for airport in airports}
    if selected_airport and selected_airport not in visible_airport_ids:
        abort(403)
    if not current_user.is_sahip:
        selected_airport = current_user.havalimani_id

    query = _drill_scope().order_by(TatbikatBelgesi.tatbikat_tarihi.desc(), TatbikatBelgesi.created_at.desc())
    if selected_airport:
        query = query.filter(TatbikatBelgesi.havalimani_id == selected_airport)

    documents = query.all()
    drill_airport_select_enabled = bool(current_user.is_sahip)
    can_manage_drills = has_permission("drill.manage") and _can_manage_drills_for_airport(selected_airport or current_user.havalimani_id)
    return render_template(
        "tatbikatlar.html",
        documents=documents,
        airports=airports,
        selected_airport=selected_airport,
        can_manage_drills=can_manage_drills,
        drill_airport_select_enabled=drill_airport_select_enabled,
        format_file_size=_format_drill_file_size,
    )


@inventory_bp.route("/tatbikatlar/yukle", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("drill.manage")
def tatbikat_yukle():
    airport_id = request.form.get("airport_id", type=int)
    if not current_user.is_sahip:
        airport_id = current_user.havalimani_id
    if airport_id is None:
        flash("Tatbikat belgesi için havalimanı seçin.", "danger")
        return redirect(url_for("inventory.tatbikat_listesi"))
    if not _can_manage_drills_for_airport(airport_id):
        abort(403)

    airport = apply_platform_demo_scope(
        Havalimani.query.filter_by(id=airport_id, is_deleted=False),
        "Havalimani",
        Havalimani.id,
    ).first_or_404()
    title = guvenli_metin(request.form.get("title") or "")
    description = guvenli_metin(request.form.get("description") or "")
    upload = request.files.get("document")
    safe_name, mime_type, error = _validate_drill_upload(upload)
    if error:
        flash(error, "danger")
        return redirect(url_for("inventory.tatbikat_listesi", airport_id=airport_id))
    drill_date_raw = request.form.get("drill_date")
    drill_date = _parse_date(drill_date_raw)
    if not drill_date:
        flash("Tatbikat tarihi zorunludur. Geçerli bir tarih seçin.", "danger")
        return redirect(url_for("inventory.tatbikat_listesi", airport_id=airport_id))
    try:
        storage_filename = _build_drill_storage_filename(drill_date, safe_name)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("inventory.tatbikat_listesi", airport_id=airport_id))
    if not title:
        title = Path(storage_filename).stem

    drive_service = get_drill_drive_service()
    try:
        drive_result = drive_service.upload_file(
            airport=airport,
            upload=upload,
            filename=storage_filename,
            mime_type=mime_type,
        )
        record = TatbikatBelgesi(
            havalimani_id=airport.id,
            yukleyen_kullanici_id=current_user.id,
            baslik=title,
            tatbikat_tarihi=drill_date,
            aciklama=description,
            dosya_adi=drive_result["filename"],
            drive_file_id=drive_result["drive_file_id"],
            drive_folder_id=drive_result["drive_folder_id"],
            mime_type=drive_result["mime_type"],
            dosya_boyutu=drive_result["file_size"],
        )
        db.session.add(record)
        db.session.commit()
    except GoogleDriveError as exc:
        db.session.rollback()
        current_app.logger.warning("Tatbikat belgesi Drive'a yuklenemedi: %s", compact_log_detail(exc, limit=160))
        flash("Tatbikat belgesi Google Drive'a yüklenemedi.", "danger")
        return redirect(url_for("inventory.tatbikat_listesi", airport_id=airport_id))
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Tatbikat belgesi kaydedilemedi.")
        try:
            if "drive_result" in locals():
                drive_service.delete_file(drive_result["drive_file_id"])
        except Exception:
            current_app.logger.warning("Tatbikat belgesi icin Drive temizleme islemi basarisiz.")
        flash("Tatbikat belgesi kaydedilemedi.", "danger")
        return redirect(url_for("inventory.tatbikat_listesi", airport_id=airport_id))

    log_kaydet(
        "Tatbikat",
        f"Tatbikat belgesi yüklendi: {compact_log_detail(record.baslik, limit=80)}",
        event_key="drill.upload",
        target_model="TatbikatBelgesi",
        target_id=record.id,
    )
    flash("Tatbikat belgesi Google Drive'a yüklendi.", "success")
    return redirect(url_for("inventory.tatbikat_detay", document_id=record.id))


@inventory_bp.route("/tatbikatlar/<int:document_id>")
@login_required
@permission_required("drill.view")
def tatbikat_detay(document_id):
    document = _get_drill_document_or_403(document_id)
    can_manage_document = has_permission("drill.manage") and _can_manage_drills_for_airport(document.havalimani_id)
    return render_template(
        "tatbikat_detay.html",
        document=document,
        can_manage_document=can_manage_document,
        format_file_size=_format_drill_file_size,
    )


@inventory_bp.route("/tatbikatlar/<int:document_id>/indir")
@login_required
@permission_required("drill.view")
def tatbikat_indir(document_id):
    document = _get_drill_document_or_403(document_id)
    try:
        payload = get_drill_drive_service().download_file(document.drive_file_id)
    except GoogleDriveError:
        current_app.logger.warning(
            "Tatbikat belgesi indirilemedi: %s",
            shorten_external_reference(document.drive_file_id),
        )
        flash("Tatbikat belgesine şu an erişilemiyor.", "danger")
        return redirect(url_for("inventory.tatbikat_detay", document_id=document.id))

    return send_file(
        io.BytesIO(payload["content"]),
        mimetype=payload["mime_type"],
        as_attachment=True,
        download_name=document.dosya_adi,
    )


@inventory_bp.route("/tatbikatlar/<int:document_id>/goruntule")
@login_required
@permission_required("drill.view")
def tatbikat_goruntule(document_id):
    document = _get_drill_document_or_403(document_id)
    try:
        payload = get_drill_drive_service().download_file(document.drive_file_id)
    except GoogleDriveError:
        current_app.logger.warning(
            "Tatbikat belgesi goruntulenemedi: %s",
            shorten_external_reference(document.drive_file_id),
        )
        flash("Tatbikat belgesine şu an erişilemiyor.", "danger")
        return redirect(url_for("inventory.tatbikat_detay", document_id=document.id))

    inline_mime = payload["mime_type"].startswith("image/") or payload["mime_type"] == "application/pdf"
    return send_file(
        io.BytesIO(payload["content"]),
        mimetype=payload["mime_type"],
        as_attachment=not inline_mime,
        download_name=document.dosya_adi,
    )


@inventory_bp.route("/tatbikatlar/<int:document_id>/sil", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("drill.manage")
def tatbikat_sil(document_id):
    document = _get_drill_document_or_403(document_id)
    if not _can_manage_drills_for_airport(document.havalimani_id):
        abort(403)

    try:
        get_drill_drive_service().delete_file(document.drive_file_id)
    except GoogleDriveError:
        current_app.logger.warning(
            "Tatbikat belgesi Drive'dan silinemedi: %s",
            shorten_external_reference(document.drive_file_id),
        )
        flash("Tatbikat belgesi Google Drive'dan silinemedi.", "danger")
        return redirect(url_for("inventory.tatbikat_detay", document_id=document.id))

    document.soft_delete()
    db.session.commit()
    log_kaydet(
        "Tatbikat",
        f"Tatbikat belgesi silindi: {compact_log_detail(document.baslik, limit=80)}",
        event_key="drill.delete",
        target_model="TatbikatBelgesi",
        target_id=document.id,
    )
    flash("Tatbikat belgesi kaldırıldı.", "success")
    return redirect(url_for("inventory.tatbikat_listesi", airport_id=document.havalimani_id))


@inventory_bp.route("/google-drive/oauth/baslat")
@login_required
def google_drive_oauth_start():
    oauth_state = secrets.token_urlsafe(32)
    session[GOOGLE_DRIVE_OAUTH_STATE_SESSION_KEY] = oauth_state
    session.modified = True
    try:
        auth_url = get_drill_drive_service().build_authorization_url(oauth_state)
    except GoogleDriveError as exc:
        current_app.logger.warning(
            "Google Drive OAuth başlatılamadı: %s",
            compact_log_detail(exc, limit=160),
        )
        flash("Google Drive yetkilendirme bağlantısı oluşturulamadı.", "danger")
        return _redirect_after_google_oauth()
    return redirect(auth_url)


@inventory_bp.route("/google-drive/oauth/callback")
@login_required
def google_drive_oauth_callback():
    require_state = bool(current_app.config.get("GOOGLE_DRIVE_OAUTH_REQUIRE_STATE", True))
    request_state = request.args.get("state")
    if require_state and not _google_drive_oauth_state_matches(request_state):
        current_app.logger.warning("Google Drive OAuth callback state doğrulaması başarısız.")
        flash("Google Drive oturum doğrulaması başarısız oldu. Lütfen işlemi tekrar başlatın.", "danger")
        return _redirect_after_google_oauth()

    error_code = str(request.args.get("error") or "").strip()
    if error_code:
        current_app.logger.warning("Google Drive OAuth reddedildi: %s", error_code)
        flash("Google Drive yetkilendirmesi tamamlanamadı.", "warning")
        return _redirect_after_google_oauth()

    code = str(request.args.get("code") or "").strip()
    if not code:
        current_app.logger.warning("Google Drive OAuth callback kod olmadan geldi.")
        flash("Google Drive dönüşünde yetkilendirme kodu alınamadı.", "danger")
        return _redirect_after_google_oauth()

    try:
        token_payload = get_drill_drive_service().exchange_authorization_code(code)
    except GoogleDriveError as exc:
        current_app.logger.warning("Google Drive OAuth callback başarısız: %s", compact_log_detail(exc, limit=160))
        flash("Google Drive yetkilendirmesi alınamadı.", "danger")
        return _redirect_after_google_oauth()

    if token_payload.get("refresh_token"):
        current_app.logger.info("Google Drive OAuth callback başarıyla tamamlandı.")
        flash("Google Drive yetkilendirmesi başarıyla alındı.", "success")
    else:
        current_app.logger.warning("Google Drive OAuth callback refresh token dönmedi.")
        flash("Google Drive yetkilendirmesi tamamlandı ancak kalıcı refresh token dönmedi.", "warning")
    return _redirect_after_google_oauth()


@inventory_bp.route("/asset/<int:asset_id>/hiyerarsi")
@login_required
@permission_required("inventory.view")
def asset_hierarchy_detail(asset_id):
    asset = _asset_scope().filter(InventoryAsset.id == asset_id).first_or_404()
    siblings = []
    if asset.parent_asset_id:
        siblings = _asset_scope().filter(
            InventoryAsset.parent_asset_id == asset.parent_asset_id,
            InventoryAsset.id != asset.id,
        ).all()
    return render_template(
        "asset_hierarchy_detail.html",
        asset=asset,
        parent_asset=asset.parent_asset,
        child_assets=asset.child_assets,
        sibling_assets=siblings,
    )


@inventory_bp.route("/asset/<int:asset_id>/quick", methods=["GET", "POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"), methods=["POST"])
@permission_required("inventory.view")
def quick_asset_view(asset_id):
    return _asset_detail_view(asset_id, detail_mode=False)


@inventory_bp.route("/asset/<int:asset_id>/detay", methods=["GET", "POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"), methods=["POST"])
@permission_required("inventory.view")
def asset_detail(asset_id):
    return _asset_detail_view(asset_id, detail_mode=True)


def _asset_maintenance_summary(asset):
    today = get_tr_now().date()
    next_maintenance = asset.next_maintenance_date
    if not next_maintenance:
        return "Planlanmadı"
    if next_maintenance < today:
        return "Gecikmiş"
    if next_maintenance <= today + timedelta(days=15):
        return "Yaklaşan"
    return "Planlı"


def _asset_detail_endpoint(detail_mode):
    return "inventory.asset_detail" if detail_mode else "inventory.quick_asset_view"


def _parse_optional_float(raw_value):
    if raw_value in (None, ""):
        return None
    try:
        return float(raw_value)
    except (TypeError, ValueError):
        return None


def _upsert_spare_part_stock(
    asset,
    spare_part,
    *,
    quantity_on_hand=None,
    quantity_reserved=None,
    reorder_point=None,
    shelf_location=None,
    create_if_missing=False,
):
    if not asset.havalimani_id:
        return None

    stock = SparePartStock.query.filter_by(
        spare_part_id=spare_part.id,
        airport_id=asset.havalimani_id,
        is_deleted=False,
    ).first()
    should_create = create_if_missing or any(
        value is not None for value in (quantity_on_hand, quantity_reserved, reorder_point)
    ) or bool(str(shelf_location or "").strip())
    if stock is None and should_create:
        stock = SparePartStock(
            spare_part_id=spare_part.id,
            airport_id=asset.havalimani_id,
            quantity_on_hand=0,
            quantity_reserved=0,
            reorder_point=spare_part.min_stock_level or 0,
            is_active=True,
        )
        db.session.add(stock)

    if stock is None:
        return None

    if quantity_on_hand is not None:
        stock.quantity_on_hand = max(quantity_on_hand, 0)
    if quantity_reserved is not None:
        stock.quantity_reserved = max(quantity_reserved, 0)
    if reorder_point is not None:
        stock.reorder_point = max(reorder_point, 0)
    if shelf_location is not None:
        stock.shelf_location = guvenli_metin(shelf_location).strip()
    stock.is_active = True
    return stock


def _asset_spare_parts_payload(asset, include_available=True):
    if not table_exists("asset_spare_part_link"):
        return [], [], {}

    linked_parts = (
        AssetSparePartLink.query.join(SparePart, SparePart.id == AssetSparePartLink.spare_part_id)
        .filter(
            AssetSparePartLink.asset_id == asset.id,
            AssetSparePartLink.is_deleted.is_(False),
            AssetSparePartLink.is_active.is_(True),
            SparePart.is_deleted.is_(False),
        )
        .order_by(SparePart.title.asc(), SparePart.part_code.asc())
        .all()
    )
    linked_part_ids = {link.spare_part_id for link in linked_parts}

    available_parts = []
    if include_available:
        available_query = SparePart.query.filter(
            SparePart.is_deleted.is_(False),
            SparePart.is_active.is_(True),
        )
        if linked_part_ids:
            available_query = available_query.filter(~SparePart.id.in_(linked_part_ids))
        available_parts = available_query.order_by(SparePart.title.asc(), SparePart.part_code.asc()).limit(250).all()

    spare_stock_map = {}
    if asset.havalimani_id:
        stock_rows = SparePartStock.query.filter_by(
            airport_id=asset.havalimani_id,
            is_deleted=False,
        ).all()
        spare_stock_map = {row.spare_part_id: row for row in stock_rows}

    return linked_parts, available_parts, spare_stock_map


def _handle_asset_spare_part_action(asset, action):
    if action == "spare_link_existing":
        part_id = request.form.get("spare_part_id", type=int)
        part = SparePart.query.filter_by(id=part_id, is_deleted=False).first()
        if not part:
            raise ValueError("Seçilen yedek parça bulunamadı.")

        quantity_required = _parse_optional_float(request.form.get("quantity_required"))
        quantity_required = max(quantity_required if quantity_required is not None else 1.0, 0.1)
        link_note = guvenli_metin(request.form.get("link_note") or "").strip()

        link = AssetSparePartLink.query.filter_by(asset_id=asset.id, spare_part_id=part.id).first()
        if link is None:
            link = AssetSparePartLink(
                asset_id=asset.id,
                spare_part_id=part.id,
                quantity_required=quantity_required,
                note=link_note,
                is_active=True,
            )
            db.session.add(link)
            action_label = "bağlandı"
        else:
            if link.is_deleted:
                link.is_deleted = False
                link.deleted_at = None
            link.is_active = True
            link.quantity_required = quantity_required
            link.note = link_note
            action_label = "güncellendi"

        _upsert_spare_part_stock(
            asset,
            part,
            quantity_on_hand=_parse_optional_float(request.form.get("quantity_on_hand")),
            quantity_reserved=_parse_optional_float(request.form.get("quantity_reserved")),
            reorder_point=_parse_optional_float(request.form.get("reorder_point")),
            shelf_location=request.form.get("shelf_location"),
            create_if_missing=False,
        )
        return {
            "message": f"Yedek parça bağlantısı {action_label}: {part.part_code}",
            "log": f"Asset {asset.id} yedek parça bağlantısı {action_label}: {part.part_code}",
        }

    if action == "spare_create_linked":
        part_code = guvenli_metin(request.form.get("part_code") or "").strip().upper()
        title = guvenli_metin(request.form.get("title") or "").strip()
        if not part_code or not title:
            raise ValueError("Parça kodu ve parça adı zorunludur.")
        if SparePart.query.filter_by(part_code=part_code).first():
            raise ValueError("Bu parça kodu zaten kayıtlı.")

        part = SparePart(
            part_code=part_code,
            title=title,
            category=guvenli_metin(request.form.get("category") or "").strip(),
            compatible_asset_type=guvenli_metin(request.form.get("compatible_asset_type") or "").strip(),
            manufacturer=guvenli_metin(request.form.get("manufacturer") or "").strip(),
            model_code=guvenli_metin(request.form.get("model_code") or "").strip(),
            description=guvenli_metin(request.form.get("description") or "").strip(),
            unit=guvenli_metin(request.form.get("unit") or "").strip() or "adet",
            min_stock_level=max(_parse_optional_float(request.form.get("min_stock_level")) or 0, 0),
            critical_level=max(_parse_optional_float(request.form.get("critical_level")) or 0, 0),
            is_active=True,
        )
        db.session.add(part)
        db.session.flush()

        link = AssetSparePartLink(
            asset_id=asset.id,
            spare_part_id=part.id,
            quantity_required=max(_parse_optional_float(request.form.get("quantity_required")) or 1, 0.1),
            note=guvenli_metin(request.form.get("link_note") or "").strip(),
            is_active=True,
        )
        db.session.add(link)

        _upsert_spare_part_stock(
            asset,
            part,
            quantity_on_hand=_parse_optional_float(request.form.get("quantity_on_hand")) or 0,
            quantity_reserved=_parse_optional_float(request.form.get("quantity_reserved")) or 0,
            reorder_point=_parse_optional_float(request.form.get("reorder_point")),
            shelf_location=request.form.get("shelf_location"),
            create_if_missing=True,
        )
        return {
            "message": f"Yeni yedek parça oluşturuldu ve envantere bağlandı: {part.part_code}",
            "log": f"Asset {asset.id} için yeni yedek parça oluşturuldu: {part.part_code}",
        }

    if action == "spare_link_update":
        link_id = request.form.get("link_id", type=int)
        link = AssetSparePartLink.query.filter_by(
            id=link_id,
            asset_id=asset.id,
            is_deleted=False,
        ).first()
        if not link or not link.spare_part or link.spare_part.is_deleted:
            raise ValueError("Güncellenecek yedek parça bağlantısı bulunamadı.")

        link.quantity_required = max(_parse_optional_float(request.form.get("quantity_required")) or link.quantity_required or 1, 0.1)
        link.note = guvenli_metin(request.form.get("link_note") or "").strip()
        link.is_active = request.form.get("is_active") == "on" or request.form.get("is_active") is None

        _upsert_spare_part_stock(
            asset,
            link.spare_part,
            quantity_on_hand=_parse_optional_float(request.form.get("quantity_on_hand")),
            quantity_reserved=_parse_optional_float(request.form.get("quantity_reserved")),
            reorder_point=_parse_optional_float(request.form.get("reorder_point")),
            shelf_location=request.form.get("shelf_location"),
            create_if_missing=False,
        )
        return {
            "message": f"Bağlı yedek parça güncellendi: {link.spare_part.part_code}",
            "log": f"Asset {asset.id} yedek parça bağlantısı güncellendi: {link.spare_part.part_code}",
        }

    if action == "spare_link_archive":
        link_id = request.form.get("link_id", type=int)
        link = AssetSparePartLink.query.filter_by(
            id=link_id,
            asset_id=asset.id,
            is_deleted=False,
        ).first()
        if not link or not link.spare_part:
            raise ValueError("Arşivlenecek bağlantı bulunamadı.")

        link.is_active = False
        link.soft_delete()
        return {
            "message": f"Yedek parça bağlantısı arşive alındı: {link.spare_part.part_code}",
            "log": f"Asset {asset.id} yedek parça bağlantısı arşive alındı: {link.spare_part.part_code}",
        }

    raise ValueError("Geçersiz yedek parça işlemi gönderildi.")


def _asset_detail_view(asset_id, detail_mode=False):
    asset = _asset_scope().filter(InventoryAsset.id == asset_id).first_or_404()
    target_endpoint = _asset_detail_endpoint(detail_mode)

    if request.method == "POST":
        action = (request.form.get("action") or "asset_quick_update").strip()
        if action.startswith("spare_"):
            if not has_permission("parts.edit"):
                abort(403)
            if not table_exists("asset_spare_part_link"):
                flash("Yedek parça bağlantı şeması hazır değil. Lütfen migration adımını tamamlayın.", "danger")
                return redirect(url_for(target_endpoint, asset_id=asset.id))

            try:
                result = _handle_asset_spare_part_action(asset, action)
                db.session.commit()
                log_kaydet("Yedek Parça", result["log"], event_key=f"asset.spare.{action}", target_model="InventoryAsset", target_id=asset.id)
                flash(result["message"], "success")
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("Yedek parça bağlantısı kaydedilemedi. Gönderilen verileri kontrol edin.", "danger")
            except Exception:
                db.session.rollback()
                current_app.logger.exception("Asset yedek parça işlemi başarısız oldu. asset_id=%s action=%s", asset.id, action)
                flash("Yedek parça işlemi sırasında beklenmeyen bir hata oluştu.", "danger")
            return redirect(url_for(target_endpoint, asset_id=asset.id))

        if not has_permission("inventory.edit"):
            abort(403)
        if not _can_manage_asset_registry():
            abort(403)

        canonical_payload = _canonical_asset_payload(request.form, mode="quick_detail")
        values = _normalized_asset_contract_values(canonical_payload, mode="quick_detail", current_asset=asset)
        try:
            _validate_asset_contract_values(values, mode="quick_detail", current_asset=asset)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for(target_endpoint, asset_id=asset.id))

        asset.status = values["status"]
        if values["last_maintenance_date"]:
            asset.last_maintenance_date = values["last_maintenance_date"]
        asset.manual_url = values["manual_url"]
        asset.maintenance_period_months = values["maintenance_period_months"]
        asset.maintenance_period_days = values["maintenance_period_days"]
        asset.is_demirbas = values["is_demirbas"]
        asset.asset_tag = values["demirbas_no"] if asset.is_demirbas else ""
        asset.calibration_required = values["calibration_required"]
        asset.calibration_period_days = values["calibration_period_days"]
        asset.last_calibration_date = values["last_calibration_date"] if asset.calibration_required else None
        asset.next_calibration_date = values["next_calibration_date"] if asset.calibration_required else None
        if values["notes"]:
            existing = (asset.notes or "").strip()
            asset.notes = f"{existing}\n{values['notes']}".strip() if existing else values["notes"]

        _sync_maintenance_plan_for_asset(asset)
        _sync_calibration_schedule_for_asset(asset)
        if asset.legacy_material:
            asset.legacy_material.durum = _display_status(asset.status)
            asset.legacy_material.son_bakim_tarihi = asset.last_maintenance_date
            asset.legacy_material.gelecek_bakim_tarihi = asset.next_maintenance_date
            asset.legacy_material.kalibrasyon_tarihi = asset.last_calibration_date if asset.calibration_required else None

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Hızlı güncelleme kaydedilemedi. Seri no veya ilişkili verilerde çakışma olabilir.", "danger")
            return redirect(url_for(target_endpoint, asset_id=asset.id))
        log_kaydet("Saha Hızlı Güncelleme", f"Asset hızlı güncellendi: ID {asset.id} / durum={asset.status}")
        flash("Hızlı ekipman güncellemesi kaydedildi.", "success")
        return redirect(url_for(target_endpoint, asset_id=asset.id))

    related_work_orders = (
        WorkOrder.query.filter_by(asset_id=asset.id, is_deleted=False)
        .order_by(WorkOrder.opened_at.desc())
        .limit(10)
        .all()
    )
    assignment_history = (
        AssignmentItem.query.join(AssignmentRecord)
        .filter(
            AssignmentItem.asset_id == asset.id,
            AssignmentRecord.is_deleted.is_(False),
        )
        .order_by(AssignmentRecord.assignment_date.desc(), AssignmentRecord.created_at.desc())
        .limit(8)
        .all()
    )
    open_work_order = next((order for order in related_work_orders if order.status in {"acik", "atandi", "islemde", "beklemede_parca", "beklemede_onay"}), None)
    linked_box = asset.legacy_material.kutu if asset.legacy_material else None
    can_view_parts = has_permission("parts.view")
    can_manage_parts = has_permission("parts.edit")
    spare_part_links, available_spare_parts, spare_stock_map = _asset_spare_parts_payload(
        asset,
        include_available=can_manage_parts,
    ) if can_view_parts else ([], [], {})
    return render_template(
        "quick_asset_view.html",
        asset=asset,
        related_work_orders=related_work_orders,
        assignment_history=assignment_history,
        maintenance_instruction=asset.equipment_template.maintenance_instruction if asset.equipment_template else None,
        assignment_status_label=_assignment_status_label,
        work_order_status_label=_work_order_status_label,
        format_assignment_qty=_format_assignment_quantity,
        qr_context=_asset_qr_context(asset),
        open_work_order=open_work_order,
        detail_mode=detail_mode,
        linked_box=linked_box,
        maintenance_summary=_asset_maintenance_summary(asset),
        spare_part_links=spare_part_links,
        available_spare_parts=available_spare_parts,
        spare_stock_map=spare_stock_map,
        spare_link_table_ready=table_exists("asset_spare_part_link"),
        can_view_parts=can_view_parts,
        can_manage_parts=can_manage_parts,
        status_label_tr=_status_label_tr,
    )


@inventory_bp.route("/kutu/<string:kodu>")
@login_required
@permission_required("inventory.view")
def kutu_detay(kodu):
    kutu = _box_scope().filter(Kutu.kodu == kodu).first()
    if not kutu:
        flash("Biriminizde böyle bir kutu bulunamadı veya yetkiniz yok.", "danger")
        return redirect(url_for("inventory.dashboard"))

    available_materials = havalimani_filtreli_sorgu(Malzeme).filter(
        Malzeme.kutu_id != kutu.id,
    ).order_by(Malzeme.ad.asc()).limit(250).all()
    current_sequence = _extract_box_sequence(
        kutu.kodu,
        guvenli_metin(kutu.havalimani.kodu if kutu.havalimani else "").strip().upper(),
    )
    return render_template(
        "kutu_detay.html",
        kutu=kutu,
        materials=kutu.active_materials,
        available_materials=available_materials,
        qr_context=_box_qr_context(kutu),
        can_manage_box=_can_manage_box_airport(kutu.havalimani_id),
        current_box_sequence=current_sequence,
    )


@inventory_bp.route("/kutular")
@login_required
@permission_required("inventory.view")
def kutular():
    selected_airport = request.args.get("havalimani_id", type=int)
    selected_brand = guvenli_metin(request.args.get("marka") or "").strip()
    if current_user.is_sahip:
        havalimanlari = _visible_operational_airports()
    elif current_user.havalimani:
        havalimanlari = [current_user.havalimani]
    else:
        havalimanlari = []
    visible_airport_ids = {airport.id for airport in havalimanlari}
    if selected_airport and selected_airport not in visible_airport_ids:
        abort(403)

    query = _box_scope()
    if selected_airport:
        query = query.filter(Kutu.havalimani_id == selected_airport)
    kutular_listesi = query.order_by(Kutu.kodu.asc()).all()
    if selected_brand:
        kutular_listesi = [
            box for box in kutular_listesi
            if turkish_contains(box.marka, selected_brand)
        ]
    manageable_box_ids = {box.id for box in kutular_listesi if _can_manage_box_airport(box.havalimani_id)}
    can_create_box = bool(has_permission("inventory.create") and (current_user.is_sahip or current_user.is_airport_manager))
    return render_template(
        "kutular.html",
        kutular=kutular_listesi,
        havalimanlari=havalimanlari,
        selected_airport=selected_airport,
        selected_brand=selected_brand,
        manageable_box_ids=manageable_box_ids,
        can_create_box=can_create_box,
    )


@inventory_bp.route("/kutular/yeni", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"), methods=["POST"])
@permission_required("inventory.create")
def kutu_olustur():
    if current_user.is_sahip:
        airport_id = request.form.get("havalimani_id", type=int) or current_user.havalimani_id
    else:
        airport_id = current_user.havalimani_id
    if not airport_id:
        flash("Kutu oluşturmak için geçerli bir havalimanı seçilmelidir.", "danger")
        return redirect(url_for("inventory.kutular"))
    _validate_box_write_access(airport_id)

    marka = request.form.get("marka")
    try:
        kutu = _create_box_with_generated_code(
            airport_id=airport_id,
            marka=marka,
        )
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("inventory.kutular"))

    db.session.commit()
    log_kaydet("Kutu", f"Yeni kutu oluşturuldu: {kutu.kodu}", event_key="box.create", target_model="Kutu", target_id=kutu.id)
    audit_log("box.create", outcome="success", box_id=kutu.id, airport_id=kutu.havalimani_id, box_code=kutu.kodu)
    flash(f"Yeni kutu oluşturuldu: {kutu.kodu}", "success")
    return redirect(url_for("inventory.kutu_detay", kodu=kutu.kodu))


@inventory_bp.route("/kutu/<string:kodu>/guncelle", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"), methods=["POST"])
@permission_required("inventory.edit")
def kutu_guncelle(kodu):
    kutu = _box_scope().filter(Kutu.kodu == kodu).first_or_404()
    _validate_box_write_access(kutu.havalimani_id)

    old_code = kutu.kodu
    kutu.marka = guvenli_metin(request.form.get("marka") or "").strip() or None
    requested_sequence_raw = guvenli_metin(request.form.get("box_sequence") or "").strip()
    if requested_sequence_raw:
        try:
            requested_sequence = int(requested_sequence_raw)
        except (TypeError, ValueError):
            flash("Kutu sıra numarası sadece sayı olabilir.", "danger")
            return redirect(url_for("inventory.kutu_detay", kodu=kodu))

        if requested_sequence < 1:
            flash("Kutu sıra numarası 1 veya daha büyük olmalıdır.", "danger")
            return redirect(url_for("inventory.kutu_detay", kodu=kodu))

        airport_code = guvenli_metin(kutu.havalimani.kodu if kutu.havalimani else "").strip().upper()
        if not airport_code:
            flash("Havalimanı kodu bulunamadığı için kutu kodu güncellenemedi.", "danger")
            return redirect(url_for("inventory.kutu_detay", kodu=kodu))

        next_code = f"{airport_code}-BOX-{requested_sequence:02d}"
        collision = Kutu.query.filter(
            Kutu.kodu == next_code,
            Kutu.id != kutu.id,
        ).first()
        if collision:
            flash(f"{next_code} kodu başka bir kutu tarafından kullanılıyor.", "danger")
            return redirect(url_for("inventory.kutu_detay", kodu=kodu))

        kutu.kodu = next_code
        if not kutu.konum or kutu.konum == old_code:
            kutu.konum = next_code
        for material in kutu.active_materials:
            if material.linked_asset:
                material.linked_asset.depot_location = next_code

    db.session.commit()
    log_kaydet("Kutu", f"Kutu bilgisi güncellendi: {old_code} -> {kutu.kodu}", event_key="box.update", target_model="Kutu", target_id=kutu.id)
    audit_log("box.update", outcome="success", box_id=kutu.id, airport_id=kutu.havalimani_id)
    flash("Kutu bilgileri güncellendi.", "success")
    return redirect(url_for("inventory.kutu_detay", kodu=kutu.kodu))


@inventory_bp.route("/kutu/<string:kodu>/arsivle", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"), methods=["POST"])
@permission_required("inventory.delete")
def kutu_arsivle(kodu):
    kutu = _box_scope().filter(Kutu.kodu == kodu).first_or_404()
    _validate_box_write_access(kutu.havalimani_id)
    if kutu.active_materials:
        flash("İçinde aktif malzeme bulunan kutu arşivlenemez.", "danger")
        return redirect(url_for("inventory.kutu_detay", kodu=kodu))
    kutu.soft_delete()
    db.session.commit()
    log_kaydet("Kutu", f"Kutu arşivlendi: {kutu.kodu}", event_key="box.archive", target_model="Kutu", target_id=kutu.id)
    audit_log("box.archive", outcome="success", box_id=kutu.id, airport_id=kutu.havalimani_id)
    flash("Kutu arşive alındı.", "success")
    return redirect(url_for("inventory.kutular"))


@inventory_bp.route("/kutu/<string:kodu>/sil", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"), methods=["POST"])
@permission_required("inventory.delete")
def kutu_sil(kodu):
    kutu = _box_scope(include_deleted=True).filter(Kutu.kodu == kodu).first_or_404()
    _validate_box_write_access(kutu.havalimani_id)
    if kutu.active_materials:
        flash("İçinde aktif malzeme bulunan kutu silinemez. Önce içeriği taşıyın.", "danger")
        return redirect(url_for("inventory.kutu_detay", kodu=kodu))
    db.session.delete(kutu)
    db.session.commit()
    log_kaydet("Kutu", f"Kutu kalıcı silindi: {kodu}", event_key="box.delete", target_model="Kutu")
    audit_log("box.delete", outcome="success", box_code=kodu)
    flash("Kutu kalıcı olarak silindi.", "warning")
    return redirect(url_for("inventory.kutular"))


@inventory_bp.route("/sarf-malzemeler", methods=["GET", "POST"])
@login_required
@permission_required("inventory.view")
def consumables():
    if request.method == "POST":
        if not has_permission("inventory.edit"):
            abort(403)
        item_id = request.form.get("item_id", type=int)
        if item_id:
            item = db.session.get(ConsumableItem, item_id)
        else:
            item = ConsumableItem(
                code=guvenli_metin(request.form.get("code") or "").upper(),
                title=guvenli_metin(request.form.get("title") or ""),
                category=guvenli_metin(request.form.get("category") or ""),
                unit=guvenli_metin(request.form.get("unit") or "adet"),
                min_stock_level=float(request.form.get("min_stock_level") or 0),
                critical_level=float(request.form.get("critical_level") or 0),
                description=guvenli_metin(request.form.get("description") or ""),
                is_active=True,
            )
            db.session.add(item)
            db.session.flush()
        quantity = float(request.form.get("quantity") or 0)
        movement_type = (request.form.get("movement_type") or "in").strip()
        movement = ConsumableStockMovement(
            consumable_id=item.id,
            airport_id=current_user.havalimani_id or request.form.get("airport_id", type=int) or 1,
            kutu_id=request.form.get("kutu_id", type=int),
            movement_type=movement_type,
            quantity=abs(quantity),
            reference_note=guvenli_metin(request.form.get("reference_note") or ""),
            performed_by_id=current_user.id,
        )
        db.session.add(movement)
        db.session.commit()
        balance = _consumable_balance(item.id, movement.airport_id)
        if balance <= float(item.critical_level or 0):
            create_notification_once(current_user.id, "consumable_critical_stock", "Kritik sarf stoğu", f"{item.title} kritik stok seviyesine indi.", link_url=url_for("inventory.consumables"), severity="danger")
        elif balance <= float(item.min_stock_level or 0):
            create_notification_once(current_user.id, "consumable_low_stock", "Düşük sarf stoğu", f"{item.title} minimum stok seviyesine yaklaştı.", link_url=url_for("inventory.consumables"), severity="warning")
        log_kaydet("Sarf", f"Sarf hareketi işlendi: {item.title} / {movement_type} / {quantity}", event_key="consumable.movement", target_model="ConsumableItem", target_id=item.id)
        audit_log("consumable.movement", outcome="success", consumable_id=item.id, movement_type=movement_type, quantity=quantity)
        flash("Sarf hareketi işlendi.", "success")
        return redirect(url_for("inventory.consumables"))

    consumables_list = _consumable_scope().all() if table_exists("consumable_item") else []
    balances = {item.id: _consumable_balance(item.id, current_user.havalimani_id) if current_user.havalimani_id else 0 for item in consumables_list}
    kutular_listesi = _box_scope().order_by(Kutu.kodu.asc()).all()
    return render_template("consumables.html", consumables=consumables_list, balances=balances, kutular=kutular_listesi)


@inventory_bp.route("/kalibrasyon", methods=["GET", "POST"])
@login_required
@permission_required("maintenance.view")
def calibration_records():
    if request.method == "POST":
        if not has_permission("maintenance.edit"):
            abort(403)
        asset = _asset_scope().filter(InventoryAsset.id == request.form.get("asset_id", type=int)).first_or_404()
        schedule = CalibrationSchedule.query.filter_by(asset_id=asset.id, is_deleted=False, is_active=True).first()
        if not schedule:
            schedule = CalibrationSchedule(
                asset_id=asset.id,
                period_days=request.form.get("period_days", type=int) or 180,
                warning_days=request.form.get("warning_days", type=int) or 15,
                provider=guvenli_metin(request.form.get("provider") or ""),
                is_active=True,
                note=guvenli_metin(request.form.get("note") or ""),
            )
            db.session.add(schedule)
            db.session.flush()
        calibration_date = _parse_date(request.form.get("calibration_date")) or get_tr_now().date()
        next_calibration_date = _parse_date(request.form.get("next_calibration_date")) or (calibration_date + timedelta(days=schedule.period_days or 180))
        record = CalibrationRecord(
            asset_id=asset.id,
            calibration_schedule_id=schedule.id,
            calibration_date=calibration_date,
            next_calibration_date=next_calibration_date,
            calibrated_by_id=current_user.id,
            provider=guvenli_metin(request.form.get("provider") or schedule.provider),
            certificate_no=guvenli_metin(request.form.get("certificate_no") or ""),
            result_status=guvenli_metin(request.form.get("result_status") or "passed"),
            note=guvenli_metin(request.form.get("note") or ""),
        )
        asset.last_calibration_date = calibration_date
        asset.next_calibration_date = next_calibration_date
        db.session.add(record)
        db.session.commit()
        log_kaydet("Kalibrasyon", f"Kalibrasyon kaydı işlendi: {asset.asset_code}", event_key="calibration.record", target_model="InventoryAsset", target_id=asset.id)
        flash("Kalibrasyon kaydı kaydedildi.", "success")
        return redirect(url_for("inventory.calibration_records"))

    assets = _asset_scope().order_by(InventoryAsset.created_at.desc()).all()
    schedules = CalibrationSchedule.query.filter_by(is_deleted=False, is_active=True).all() if table_exists("calibration_schedule") else []
    records = CalibrationRecord.query.filter_by(is_deleted=False).order_by(CalibrationRecord.calibration_date.desc()).all() if table_exists("calibration_record") else []
    return render_template("calibration_records.html", assets=assets, schedules=schedules, records=records, today=get_tr_now().date())


@inventory_bp.route("/asset-lifecycle", methods=["GET", "POST"])
@login_required
@permission_required("inventory.view")
def asset_lifecycle():
    if request.method == "POST":
        if not has_permission("inventory.edit"):
            abort(403)
        asset = _asset_scope().filter(InventoryAsset.id == request.form.get("asset_id", type=int)).first_or_404()
        lifecycle_status = (request.form.get("lifecycle_status") or "active").strip()
        target_airport_id = request.form.get("target_airport_id", type=int)
        note = guvenli_metin(request.form.get("lifecycle_note") or "")
        approval_needed = lifecycle_status in {"disposed", "decommissioned", "transferred"} and not has_permission("workorder.approve")
        if approval_needed:
            payload = json.dumps({"asset_id": asset.id, "lifecycle_status": lifecycle_status, "target_airport_id": target_airport_id, "note": note}, ensure_ascii=False)
            approval = create_approval_request("asset_lifecycle", "InventoryAsset", asset.id, current_user.id, payload, commit=False)
            if approval:
                create_notification_once(current_user.id, "lifecycle_pending", "Lifecycle değişimi onay bekliyor", f"{asset.asset_code} için lifecycle değişimi onaya gönderildi.", link_url=url_for("admin.approvals"), severity="warning", commit=False)
                db.session.commit()
                flash("Lifecycle değişimi onaya gönderildi.", "warning")
                return redirect(url_for("inventory.asset_lifecycle"))

        state = _ensure_operational_state(asset)
        state.lifecycle_status = lifecycle_status
        state.lifecycle_note = note
        state.last_service_date = _parse_date(request.form.get("last_service_date")) or state.last_service_date
        state.warranty_start = _parse_date(request.form.get("warranty_start")) or state.warranty_start
        state.service_provider = guvenli_metin(request.form.get("service_provider") or state.service_provider or "")
        state.service_note = guvenli_metin(request.form.get("service_note") or state.service_note or "")
        asset.status = _lifecycle_to_status(lifecycle_status)
        if request.form.get("warranty_end"):
            asset.warranty_end_date = _parse_date(request.form.get("warranty_end"))
        if lifecycle_status == "transferred" and target_airport_id:
            asset.havalimani_id = target_airport_id
            if asset.legacy_material:
                asset.legacy_material.havalimani_id = target_airport_id
        db.session.commit()
        log_kaydet("Lifecycle", f"Lifecycle güncellendi: {asset.asset_code} -> {lifecycle_status}", event_key="asset.lifecycle.change", target_model="InventoryAsset", target_id=asset.id)
        audit_log("asset.lifecycle.change", outcome="success", asset_id=asset.id, lifecycle_status=lifecycle_status)
        flash("Lifecycle bilgisi güncellendi.", "success")
        return redirect(url_for("inventory.asset_lifecycle"))

    assets = [asset for asset in _asset_scope().order_by(InventoryAsset.created_at.desc()).all()]
    airports = _visible_operational_airports()
    return render_template("asset_lifecycle.html", assets=assets, airports=airports, lifecycle_statuses=["planned", "received", "active", "in_maintenance", "calibration_due", "out_of_service", "decommissioned", "disposed", "transferred"])


@inventory_bp.route("/kutu-bul", methods=["POST"])
@login_required
@permission_required("inventory.view")
def kutu_bul():
    kodu = request.form.get("kutu_kodu", "").strip().upper()
    if kodu:
        kutu = _box_scope().filter(Kutu.kodu == kodu).first()
        if kutu:
            return redirect(url_for("inventory.kutu_detay", kodu=kutu.kodu))
        flash(f"'{kodu}' koduna ait bir ünite bulunamadı veya erişim yetkiniz yok.", "danger")
    return redirect(url_for("inventory.dashboard"))


@inventory_bp.route("/kutu/<string:kodu>/malzeme-ekle", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("inventory.edit")
def kutu_malzeme_ekle(kodu):
    kutu = _box_scope().filter(Kutu.kodu == kodu).first_or_404()
    material_id = request.form.get("material_id", type=int)
    material = havalimani_filtreli_sorgu(Malzeme).filter(Malzeme.id == material_id).first()
    if not material:
        flash("Seçilen malzeme bulunamadı.", "danger")
        return redirect(url_for("inventory.kutu_detay", kodu=kutu.kodu))

    material.kutu_id = kutu.id
    material.havalimani_id = kutu.havalimani_id
    _sync_asset_location(material)
    db.session.commit()
    log_kaydet(
        "Kutu",
        f"Malzeme kutuya eklendi: {material.ad} -> {kutu.kodu}",
        event_key="box.content.add",
        target_model="Kutu",
        target_id=kutu.id,
    )
    flash("Malzeme kutuya eklendi.", "success")
    return redirect(url_for("inventory.kutu_detay", kodu=kutu.kodu))


@inventory_bp.route("/kutu/<string:kodu>/icerik-guncelle/<int:malzeme_id>", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("inventory.edit")
def kutu_icerik_guncelle(kodu, malzeme_id):
    kutu = _box_scope().filter(Kutu.kodu == kodu).first_or_404()
    material = Malzeme.query.filter_by(id=malzeme_id, kutu_id=kutu.id, is_deleted=False).first_or_404()
    new_quantity = _parse_positive_int(request.form.get("stok_miktari"), material.stok_miktari or 1)
    new_box_id = request.form.get("target_kutu_id", type=int)
    material.stok_miktari = new_quantity
    if new_box_id and new_box_id != kutu.id:
        target_box = _box_scope().filter(Kutu.id == new_box_id).first()
        if target_box:
            material.stok_miktari = new_quantity
            material.kutu_id = target_box.id
            material.havalimani_id = target_box.havalimani_id
            _sync_asset_location(material)
            Malzeme.query.filter_by(id=material.id).update(
                {
                    "stok_miktari": new_quantity,
                    "kutu_id": target_box.id,
                    "havalimani_id": target_box.havalimani_id,
                },
                synchronize_session=False,
            )
            db.session.commit()
            log_kaydet(
                "Kutu",
                f"Malzeme kutular arasında taşındı: {material.ad} -> {target_box.kodu}",
                event_key="box.content.move",
                target_model="Kutu",
                target_id=target_box.id,
            )
            flash("Malzeme hedef kutuya taşındı.", "success")
            return redirect(url_for("inventory.kutu_detay", kodu=target_box.kodu))

    _sync_asset_location(material)
    Malzeme.query.filter_by(id=material.id).update(
        {"stok_miktari": new_quantity},
        synchronize_session=False,
    )
    db.session.commit()
    log_kaydet(
        "Kutu",
        f"Kutu içeriği güncellendi: {material.ad} / miktar={material.stok_miktari}",
        event_key="box.content.update",
        target_model="Kutu",
        target_id=kutu.id,
    )
    flash("Kutu içeriği güncellendi.", "success")
    return redirect(url_for("inventory.kutu_detay", kodu=kutu.kodu))


@inventory_bp.route("/kutu/<string:kodu>/malzeme-cikar/<int:malzeme_id>", methods=["POST"])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required("inventory.edit")
def kutu_malzeme_cikar(kodu, malzeme_id):
    kutu = _box_scope().filter(Kutu.kodu == kodu).first_or_404()
    material = Malzeme.query.filter_by(id=malzeme_id, kutu_id=kutu.id, is_deleted=False).first_or_404()
    fallback_box = _ensure_box(f"{kutu.havalimani.kodu}-ATANMADI", kutu.havalimani_id)
    material.kutu_id = fallback_box.id
    if material.linked_asset:
        material.linked_asset.depot_location = fallback_box.kodu
    db.session.commit()
    log_kaydet(
        "Kutu",
        f"Malzeme kutudan çıkarıldı: {material.ad} / {kutu.kodu}",
        event_key="box.content.remove",
        target_model="Kutu",
        target_id=kutu.id,
    )
    flash("Malzeme kutudan çıkarıldı.", "warning")
    return redirect(url_for("inventory.kutu_detay", kodu=kutu.kodu))


@inventory_bp.route("/qr-uret/asset/<int:asset_id>")
@login_required
@permission_required("qr.generate")
def qr_uret(asset_id):
    asset = _asset_scope().filter(InventoryAsset.id == asset_id).first_or_404()
    log_kaydet("QR", f"QR etiketi goruntulendi: {asset.asset_code}")
    audit_log("inventory.qr.render", outcome="success", asset_id=asset.id, asset_code=asset.asset_code)
    return render_template("qr_yazdir.html", asset=asset, qr_context=_asset_qr_context(asset))


@inventory_bp.route("/qr-uret/toplu/envanter", methods=["POST"])
@login_required
@permission_required("qr.generate")
def qr_uret_toplu_envanter():
    selected_ids = []
    for raw_id in request.form.getlist("asset_ids"):
        try:
            parsed_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if parsed_id > 0 and parsed_id not in selected_ids:
            selected_ids.append(parsed_id)

    if not selected_ids:
        flash("Toplu QR yazdırmak için en az bir ekipman seçin.", "warning")
        return redirect(url_for("inventory.envanter"))
    if len(selected_ids) > 18:
        flash("Toplu yazdırmada bir sayfada en fazla 18 ekipman seçilebilir.", "danger")
        return redirect(url_for("inventory.envanter"))

    rows = (
        _asset_scope()
        .options(
            joinedload(InventoryAsset.equipment_template),
            joinedload(InventoryAsset.legacy_material),
        )
        .filter(InventoryAsset.id.in_(selected_ids))
        .all()
    )
    row_by_id = {row.id: row for row in rows}
    selected_assets = [row_by_id[row_id] for row_id in selected_ids if row_id in row_by_id]
    if not selected_assets:
        flash("Seçilen ekipmanlar için erişilebilir QR kaydı bulunamadı.", "danger")
        return redirect(url_for("inventory.envanter"))

    if len(selected_assets) != len(selected_ids):
        flash("Bazı ekipmanlar kapsam dışında olduğu için yazdırma listesine alınmadı.", "warning")

    log_kaydet("QR", f"Toplu envanter QR etiketi görüntülendi (adet: {len(selected_assets)})", event_key="inventory.qr.bulk.render")
    audit_log("inventory.qr.bulk.render", outcome="success", item_count=len(selected_assets))
    return render_template(
        "qr_toplu_envanter_yazdir.html",
        assets=selected_assets[:18],
    )


@inventory_bp.route("/qr-yenile/asset/<int:asset_id>", methods=["POST"])
@login_required
@permission_required("qr.generate")
def qr_yenile(asset_id):
    asset = _asset_scope().filter(InventoryAsset.id == asset_id).first_or_404()
    payload = json.dumps(
        {"asset_id": asset.id, "requested_by_id": current_user.id},
        ensure_ascii=False,
    )
    approval = create_approval_request(
        approval_type="qr_regenerate",
        target_model="InventoryAsset",
        target_id=asset.id,
        requested_by_id=current_user.id,
        request_payload=payload,
        commit=False,
    )
    if approval:
        log_kaydet(
            "QR",
            f"QR yeniden üretim talebi açıldı: {asset.asset_code}",
            event_key="inventory.qr.regenerate.pending",
            target_model="InventoryAsset",
            target_id=asset.id,
            commit=False,
        )
        create_notification(
            current_user.id,
            "approval_pending",
            "QR yeniden üretimi onay bekliyor",
            f"{asset.asset_code} için QR yeniden üretim talebi oluşturuldu.",
            link_url=url_for("admin.approvals"),
            severity="warning",
            commit=False,
        )
        db.session.commit()
        flash("QR yeniden üretimi onaya gönderildi.", "warning")
    else:
        flash("QR yeniden üretim talebi oluşturulamadı.", "danger")
    return redirect(url_for("inventory.quick_asset_view", asset_id=asset.id))


@inventory_bp.route("/qr-uret/kutu/<int:box_id>")
@login_required
@permission_required("qr.generate")
def kutu_qr_uret(box_id):
    kutu = _box_scope().filter(Kutu.id == box_id).first_or_404()
    log_kaydet("QR", f"Kutu QR etiketi görüntülendi: {kutu.qr_code_label}", event_key="box.qr.render", target_model="Kutu", target_id=kutu.id)
    audit_log("box.qr.render", outcome="success", box_id=kutu.id, box_code=kutu.qr_code_label)
    return render_template("kutu_qr_yazdir.html", kutu=kutu, qr_context=_box_qr_context(kutu))


@inventory_bp.route("/qr-uret/toplu/kutular", methods=["POST"])
@login_required
@permission_required("qr.generate")
def qr_uret_toplu_kutular():
    selected_ids = []
    for raw_id in request.form.getlist("box_ids"):
        try:
            parsed_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if parsed_id > 0 and parsed_id not in selected_ids:
            selected_ids.append(parsed_id)

    if not selected_ids:
        flash("Toplu QR yazdırmak için en az bir kutu seçin.", "warning")
        return redirect(url_for("inventory.kutular"))
    if len(selected_ids) > 10:
        flash("Toplu yazdırmada bir sayfada en fazla 10 kutu seçilebilir.", "danger")
        return redirect(url_for("inventory.kutular"))

    rows = _box_scope().filter(Kutu.id.in_(selected_ids)).all()
    row_by_id = {row.id: row for row in rows}
    selected_boxes = [row_by_id[row_id] for row_id in selected_ids if row_id in row_by_id]
    if not selected_boxes:
        flash("Seçilen kutular için erişilebilir QR kaydı bulunamadı.", "danger")
        return redirect(url_for("inventory.kutular"))

    if len(selected_boxes) != len(selected_ids):
        flash("Bazı kutular kapsam dışında olduğu için yazdırma listesine alınmadı.", "warning")

    log_kaydet("QR", f"Toplu kutu QR etiketi görüntülendi (adet: {len(selected_boxes)})", event_key="box.qr.bulk.render")
    audit_log("box.qr.bulk.render", outcome="success", item_count=len(selected_boxes))
    return render_template(
        "qr_toplu_kutu_yazdir.html",
        boxes=selected_boxes[:10],
    )


@inventory_bp.route("/api/qr-img/kutu/<int:box_id>")
@login_required
@permission_required("qr.generate")
def kutu_qr_img(box_id):
    kutu = _box_scope().filter(Kutu.id == box_id).first_or_404()
    img_io = generate_qr_data(_box_qr_payload(kutu))
    return send_file(img_io, mimetype="image/png") if img_io else ("QR Hatası", 500)


@inventory_bp.route("/kutu/<string:kodu>/etiket")
@login_required
@permission_required("qr.generate")
def kutu_etiket(kodu):
    kutu = _box_scope().filter(Kutu.kodu == kodu).first_or_404()
    log_kaydet("QR", f"Kutu etiketi görüntülendi: {kutu.qr_code_label}", event_key="box.label.view", target_model="Kutu", target_id=kutu.id)
    return render_template("kutu_etiket.html", kutu=kutu, materials=kutu.active_materials, qr_context=_box_qr_context(kutu))


@inventory_bp.route("/kutu/<string:kodu>/etiket/pdf")
@login_required
@permission_required("qr.generate")
def kutu_etiket_pdf(kodu):
    kutu = _box_scope().filter(Kutu.kodu == kodu).first_or_404()
    qr_img = generate_qr_data(_box_qr_payload(kutu))
    qr_data_uri = None
    if qr_img:
        qr_data_uri = "data:image/png;base64," + base64.b64encode(qr_img.getvalue()).decode("ascii")
    html = render_template(
        "kutu_export_pdf.html",
        kutu=kutu,
        materials=kutu.active_materials,
        qr_context=_box_qr_context(kutu),
        qr_data_uri=qr_data_uri,
    )
    output = io.BytesIO()
    pisa.CreatePDF(html, dest=output)
    output.seek(0)
    log_kaydet("Rapor", f"Kutu etiketi PDF oluşturuldu: {kutu.qr_code_label}", event_key="box.label.export", target_model="Kutu", target_id=kutu.id)
    return send_file(
        output,
        download_name=f"{kutu.qr_code_label}_etiket.pdf",
        as_attachment=True,
    )


@inventory_bp.route("/api/qr-img/asset/<int:asset_id>")
@login_required
@permission_required("qr.generate")
def qr_img(asset_id):
    asset = _asset_scope().filter(InventoryAsset.id == asset_id).first_or_404()
    img_io = generate_qr_data(_asset_qr_payload(asset))
    return send_file(img_io, mimetype="image/png") if img_io else ("QR Hatası", 500)


@inventory_bp.route("/api/qr-img/<string:kodu>")
@login_required
@permission_required("qr.generate")
def qr_img_legacy(kodu):
    normalized_code = (kodu or "").strip().upper()
    if not normalized_code:
        abort(404)

    kutu = _box_scope().filter(Kutu.kodu == normalized_code).first()
    if kutu:
        img_io = generate_qr_data(_box_qr_payload(kutu))
        return send_file(img_io, mimetype="image/png") if img_io else ("QR Hatası", 500)

    asset = _asset_scope().filter(InventoryAsset.qr_code == normalized_code).first()
    if not asset and normalized_code.startswith("ARFF-SAR-"):
        serial = normalized_code.replace("ARFF-SAR-", "", 1)
        if serial.isdigit():
            asset = _asset_scope().filter(InventoryAsset.id == int(serial)).first()

    if not asset:
        abort(404)

    img_io = generate_qr_data(_asset_qr_payload(asset))
    return send_file(img_io, mimetype="image/png") if img_io else ("QR Hatası", 500)
