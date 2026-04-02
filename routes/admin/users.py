import io
import json
import re

import pandas as pd
from flask import current_app, render_template, request, redirect, send_file, session, url_for, flash, abort
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import func

from extensions import (
    audit_log,
    create_approval_request,
    create_notification,
    db,
    is_allowed_file,
    limiter,
    log_kaydet,
    guvenli_metin,
    secure_upload_filename,
)
from models import Havalimani, Kullanici, TR_UPPER_MAP
from . import admin_bp
from decorators import (
    CANONICAL_ROLE_ADMIN,
    CANONICAL_ROLE_SYSTEM,
    CANONICAL_ROLE_TEAM_LEAD,
    CANONICAL_ROLE_TEAM_MEMBER,
    actor_can_manage_target,
    actor_can_view_target_user,
    can_assign_role,
    expand_role_keys,
    get_effective_role,
    get_effective_permissions,
    has_permission,
    get_permission_catalog,
    get_role_labels,
    get_role_options,
    get_user_permission_overrides,
    is_impersonation_mode,
    permission_required,
    update_user_permission_overrides,
)
from services.text_normalization_service import normalize_lookup_key, turkish_contains_all


GLOBAL_ROLES = {CANONICAL_ROLE_SYSTEM, CANONICAL_ROLE_ADMIN}
AIRPORT_ROLES = {CANONICAL_ROLE_TEAM_LEAD, CANONICAL_ROLE_TEAM_MEMBER}
STATUS_OPTIONS = [
    {"key": "active", "label": "Aktif kayıtlar"},
    {"key": "archived", "label": "Arşivdekiler"},
    {"key": "all", "label": "Tüm durumlar"},
]
BULK_IMPORT_COLUMNS = [
    "ad",
    "soyad",
    "e-posta",
    "telefon",
    "rol",
    "havalimani",
    "aktif/pasif",
    "not",
    "gecici_sifre",
]
BULK_IMPORT_OPTIONAL_COLUMNS = [
    "kan_grubu_harf",
    "kan_grubu_rh",
    "boy_cm",
    "kilo_kg",
    "ayak_numarasi",
    "ust_beden",
    "alt_beden",
]
BULK_IMPORT_TEMPLATE_COLUMNS = BULK_IMPORT_COLUMNS + BULK_IMPORT_OPTIONAL_COLUMNS
EMAIL_PATTERN = re.compile(
    r"^[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@"
    r"[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?"
    r"(?:\.[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?)+$",
    re.IGNORECASE,
)
USER_PASSWORD_PATTERN = re.compile(r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9]).{8,}$")
BLOOD_TYPE_LETTER_OPTIONS = ("A", "B", "AB", "0")
RH_FACTOR_OPTIONS = ("+", "-")
BODY_SIZE_OPTIONS = ("XS", "S", "M", "L", "XL", "XXL", "3XL")
SHOE_SIZE_OPTIONS = tuple(
    str(size).rstrip("0").rstrip(".")
    for size in (34 + (step * 0.5) for step in range(33))
)


def _visible_users_query(actor):
    query = Kullanici.query.outerjoin(Havalimani)
    actor_role = get_effective_role(actor)
    if actor_role in {CANONICAL_ROLE_SYSTEM, CANONICAL_ROLE_ADMIN}:
        return query
    if actor.havalimani_id is None:
        return query.filter(Kullanici.havalimani_id.is_(None))
    return query.filter(Kullanici.havalimani_id == actor.havalimani_id)


def _normalize_search_term(raw_value):
    return " ".join((guvenli_metin(raw_value or "")).split())


def _user_matches_search(user, raw_search):
    haystack = " ".join(
        str(field or "")
        for field in [
            getattr(user, "tam_ad", ""),
            getattr(user, "kullanici_adi", ""),
            getattr(getattr(user, "havalimani", None), "ad", ""),
            getattr(getattr(user, "havalimani", None), "kodu", ""),
        ]
    )
    return turkish_contains_all(haystack, raw_search)


def _normalize_user_email(raw_value):
    return guvenli_metin(raw_value or "").strip().lower()


def _normalize_full_name(raw_value):
    cleaned = " ".join(guvenli_metin(raw_value or "").split())
    if not cleaned:
        return ""
    return cleaned.translate(TR_UPPER_MAP).upper()


def _validate_user_email(raw_value):
    normalized = _normalize_user_email(raw_value)
    if not normalized or not EMAIL_PATTERN.match(normalized):
        return None, "Geçerli bir e-posta adresi girin."
    return normalized, None


def _validate_user_password(raw_value):
    password = str(raw_value or "")
    if USER_PASSWORD_PATTERN.match(password):
        return None
    return (
        "Şifre en az 8 karakter uzunluğunda olmalı; "
        "1 büyük harf, 1 küçük harf, 1 rakam ve 1 özel karakter içermelidir."
    )


def _find_user_by_email(raw_value):
    normalized = _normalize_user_email(raw_value)
    if not normalized:
        return None
    return Kullanici.query.filter(func.lower(func.trim(Kullanici.kullanici_adi)) == normalized).first()


def _normalize_phone_number(raw_value):
    cleaned = guvenli_metin(raw_value or "").strip()
    if not cleaned:
        return None, None

    digits = re.sub(r"\D", "", cleaned)
    if digits.startswith("90"):
        digits = digits[2:]
    elif digits.startswith("0"):
        digits = digits[1:]

    if len(digits) != 10 or not digits.startswith("5"):
        return None, "Telefon numarasını +90 5xx xxx xx xx formatında girin."
    return f"+90{digits}", None


def _visible_airports(actor):
    if get_effective_role(actor) in {CANONICAL_ROLE_SYSTEM, CANONICAL_ROLE_ADMIN}:
        return Havalimani.query.filter_by(is_deleted=False).all()
    return Havalimani.query.filter_by(is_deleted=False, id=actor.havalimani_id).all()


def _create_role_options_for_actor(actor, role_options):
    if get_effective_role(actor) == CANONICAL_ROLE_TEAM_LEAD:
        return [role for role in role_options if role["key"] == CANONICAL_ROLE_TEAM_MEMBER]
    return role_options


def _assignable_role_options_for_actor(actor, role_options):
    return [role for role in role_options if can_assign_role(actor, role["key"])]


def _can_manage_user_phone_on_create(actor):
    return bool(getattr(actor, "is_sahip", False) and get_effective_role(actor) == CANONICAL_ROLE_SYSTEM)


def _bulk_import_session_key():
    return "bulk_user_import_preview"


def _default_import_password():
    return "Gecici@123"


def _safe_import_phone(raw_value):
    if not current_user.is_sahip:
        return None, None
    return _normalize_phone_number(raw_value)


def _normalize_optional_positive_int(raw_value, field_label, *, min_value=None, max_value=None):
    cleaned = guvenli_metin(raw_value or "").strip()
    if not cleaned:
        return None, None
    try:
        value = int(cleaned)
    except (TypeError, ValueError):
        return None, f"{field_label} sayısal olmalıdır."
    if min_value is not None and value < min_value:
        return None, f"{field_label} en az {min_value} olmalıdır."
    if max_value is not None and value > max_value:
        return None, f"{field_label} en fazla {max_value} olabilir."
    return value, None


def _normalize_shoe_size(raw_value):
    cleaned = guvenli_metin(raw_value or "").strip().replace(",", ".")
    if not cleaned:
        return None, None
    if cleaned not in SHOE_SIZE_OPTIONS:
        return None, "Ayak numarası listeden seçilmelidir."
    try:
        return float(cleaned), None
    except (TypeError, ValueError):
        return None, "Ayak numarası doğrulanamadı."


def _collect_user_profile_fields(form):
    blood_letter = guvenli_metin(form.get("kan_grubu_harf") or "").strip()
    rh_factor = guvenli_metin(form.get("kan_grubu_rh") or "").strip()
    legacy_body_size = guvenli_metin(form.get("beden") or "").strip().upper()
    upper_body_size = guvenli_metin(form.get("ust_beden") or "").strip().upper()
    lower_body_size = guvenli_metin(form.get("alt_beden") or "").strip().upper()

    if rh_factor == "Rh+":
        rh_factor = "+"
    elif rh_factor == "Rh-":
        rh_factor = "-"

    if blood_letter and blood_letter not in BLOOD_TYPE_LETTER_OPTIONS:
        return None, "Kan grubu harfi listeden seçilmelidir."
    if rh_factor and rh_factor not in RH_FACTOR_OPTIONS:
        return None, "Rh alanı listeden seçilmelidir."
    if bool(blood_letter) != bool(rh_factor):
        return None, "Kan grubu için harf ve Rh alanlarını birlikte seçin."
    if legacy_body_size and legacy_body_size not in BODY_SIZE_OPTIONS:
        return None, "Beden alanı listeden seçilmelidir."
    if upper_body_size and upper_body_size not in BODY_SIZE_OPTIONS:
        return None, "Üst beden alanı listeden seçilmelidir."
    if lower_body_size and lower_body_size not in BODY_SIZE_OPTIONS:
        return None, "Alt beden alanı listeden seçilmelidir."

    if not upper_body_size and not lower_body_size and legacy_body_size:
        upper_body_size = legacy_body_size

    boy_cm, error = _normalize_optional_positive_int(form.get("boy_cm"), "Boy", min_value=90, max_value=260)
    if error:
        return None, error
    kilo_kg, error = _normalize_optional_positive_int(form.get("kilo_kg"), "Kilo", min_value=30, max_value=250)
    if error:
        return None, error
    ayak_numarasi, error = _normalize_shoe_size(form.get("ayak_numarasi"))
    if error:
        return None, error

    return {
        "kan_grubu_harf": blood_letter or None,
        "kan_grubu_rh": rh_factor or None,
        "boy_cm": boy_cm,
        "kilo_kg": kilo_kg,
        "ayak_numarasi": ayak_numarasi,
        "ust_beden": upper_body_size or None,
        "alt_beden": lower_body_size or None,
        "beden": upper_body_size or lower_body_size or legacy_body_size or None,
    }, None


def _apply_user_profile_fields(user, profile_fields):
    for key, value in (profile_fields or {}).items():
        setattr(user, key, value)


def _resolve_airport_for_import(actor, airport_text, role_key, visible_airports):
    cleaned = guvenli_metin(airport_text or "").strip()
    if role_key in GLOBAL_ROLES:
        return None, None
    if not cleaned:
        return None, "Saha rolleri için havalimanı bilgisi zorunludur."

    lookup = {}
    for airport in visible_airports:
        lookup[normalize_lookup_key(airport.id)] = airport
        lookup[normalize_lookup_key(airport.kodu)] = airport
        lookup[normalize_lookup_key(airport.ad)] = airport
    airport = lookup.get(normalize_lookup_key(cleaned))
    if not airport:
        return None, "Havalimanı değeri görünür kapsamınızda bulunamadı."
    if not actor.is_sahip and airport.id != actor.havalimani_id:
        return None, "Bu havalimanı için kullanıcı içe aktarma yetkiniz yok."
    return airport.id, None


def _build_user_import_preview(actor, frame, role_options, visible_airports):
    lower_columns = {str(column).strip().lower(): column for column in frame.columns}
    missing_columns = [column for column in BULK_IMPORT_COLUMNS if column not in lower_columns]
    if missing_columns:
        return {
            "errors": [{"row": "Şablon", "reason": f"Eksik kolonlar: {', '.join(missing_columns)}"}],
            "valid_rows": [],
            "summary": {"total": 0, "valid": 0, "invalid": 0},
        }

    role_keys = {item["key"] for item in role_options}
    role_lookup = {}
    for item in role_options:
        for role_key in expand_role_keys(item["key"]):
            role_lookup[role_key] = item["key"]
    existing_emails = {
        item.kullanici_adi.strip().lower()
        for item in Kullanici.query.all()
        if item.kullanici_adi
    }
    seen_emails = set()
    errors = []
    valid_rows = []

    for index, row in frame.fillna("").iterrows():
        row_number = index + 2
        first_name = guvenli_metin(row[lower_columns["ad"]]).strip()
        last_name = guvenli_metin(row[lower_columns["soyad"]]).strip()
        email = guvenli_metin(row[lower_columns["e-posta"]]).strip().lower()
        role_key = role_lookup.get(guvenli_metin(row[lower_columns["rol"]]).strip(), guvenli_metin(row[lower_columns["rol"]]).strip())

        if not first_name or not last_name or not email or not role_key:
            errors.append({"row": row_number, "reason": "Ad, soyad, e-posta ve rol alanları zorunludur."})
            continue
        if role_key not in role_keys:
            errors.append({"row": row_number, "reason": f"Rol geçersiz: {role_key}"})
            continue
        if not can_assign_role(actor, role_key):
            errors.append({"row": row_number, "reason": f"Bu rolü içe aktarma yetkiniz yok: {role_key}"})
            continue
        if email in existing_emails or email in seen_emails:
            errors.append({"row": row_number, "reason": f"E-posta zaten kullanımda: {email}"})
            continue

        airport_id, airport_error = _resolve_airport_for_import(
            actor,
            row[lower_columns["havalimani"]],
            role_key,
            visible_airports,
        )
        if airport_error:
            errors.append({"row": row_number, "reason": airport_error})
            continue

        phone_value, phone_error = _safe_import_phone(row[lower_columns["telefon"]])
        if phone_error:
            errors.append({"row": row_number, "reason": phone_error})
            continue

        profile_payload = {}
        for column in BULK_IMPORT_OPTIONAL_COLUMNS:
            column_key = lower_columns.get(column)
            if not column_key:
                continue
            raw_value = row[column_key]
            if raw_value in (None, ""):
                profile_payload[column] = ""
            elif isinstance(raw_value, float) and raw_value.is_integer():
                profile_payload[column] = str(int(raw_value))
            else:
                profile_payload[column] = str(raw_value).strip()
        profile_fields, profile_error = _collect_user_profile_fields(profile_payload)
        if profile_error:
            errors.append({"row": row_number, "reason": profile_error})
            continue

        active_text = guvenli_metin(row[lower_columns["aktif/pasif"]]).strip().lower()
        is_deleted = active_text in {"pasif", "arsiv", "archived", "0", "hayir", "false"}
        temp_password = guvenli_metin(row[lower_columns["gecici_sifre"]]).strip() or _default_import_password()

        valid_rows.append(
            {
                "row": row_number,
                "tam_ad": f"{first_name} {last_name}".strip(),
                "kullanici_adi": email,
                "rol": role_key,
                "havalimani_id": airport_id,
                "telefon_numarasi": phone_value,
                "is_deleted": is_deleted,
                "gecici_sifre": temp_password,
                "not": guvenli_metin(row[lower_columns["not"]]).strip(),
                "profile_fields": profile_fields,
            }
        )
        seen_emails.add(email)

    return {
        "errors": errors,
        "valid_rows": valid_rows,
        "summary": {
            "total": int(len(frame.index)),
            "valid": len(valid_rows),
            "invalid": len(errors),
        },
    }


@admin_bp.route('/kullanicilar')
@login_required
@permission_required('users.manage')
def kullanicilar():
    """Kullanıcıları görünürlük ve filtre kurallarına göre listeler."""
    role_options = get_role_options()
    permission_catalog = get_permission_catalog()
    permission_lookup = {
        permission["key"]: permission["label"]
        for permissions in permission_catalog.values()
        for permission in permissions
    }
    havalimanlari = _visible_airports(current_user)
    bulk_import_preview = session.get(_bulk_import_session_key())
    valid_airport_keys = {str(item.id) for item in havalimanlari}
    role_filter_lookup = {}
    for role in role_options:
        for role_key in expand_role_keys(role["key"]):
            role_filter_lookup[role_key] = role["key"]
    valid_role_keys = {role["key"] for role in role_options}
    assignable_role_options = _assignable_role_options_for_actor(current_user, role_options)
    search_term = _normalize_search_term(request.args.get("q"))
    selected_role_key = role_filter_lookup.get((request.args.get("role") or "").strip(), "")
    selected_airport_key = (request.args.get("airport_id") or "").strip()
    selected_status_key = (request.args.get("status") or "active").strip() or "active"
    if selected_role_key not in valid_role_keys:
        selected_role_key = ""
    if selected_airport_key not in valid_airport_keys | {"global", ""}:
        selected_airport_key = ""
    valid_status_keys = {item["key"] for item in STATUS_OPTIONS}
    if selected_status_key not in valid_status_keys:
        selected_status_key = "active"
    selected_bulk_airport_key = (request.args.get("bulk_airport_id") or "").strip()
    if selected_bulk_airport_key not in valid_airport_keys:
        selected_bulk_airport_key = ""
    selected_role_label = get_role_labels().get(selected_role_key, "") if selected_role_key else ""
    if selected_airport_key == "global":
        selected_airport_label = "Global erişim"
    else:
        selected_airport_label = next(
            (item.ad for item in havalimanlari if str(item.id) == selected_airport_key),
            "",
        )
    selected_status_label = next(
        (item["label"] for item in STATUS_OPTIONS if item["key"] == selected_status_key),
        STATUS_OPTIONS[0]["label"],
    )
    selected_bulk_airport_label = next(
        (item.ad for item in havalimanlari if str(item.id) == selected_bulk_airport_key),
        "",
    )
    has_active_filters = bool(
        search_term
        or selected_role_key
        or selected_airport_key
        or selected_status_key != "active"
    )

    scope_query = _visible_users_query(current_user)
    filtered_query = scope_query
    if selected_status_key == "active":
        filtered_query = filtered_query.filter(Kullanici.is_deleted.is_(False))
    elif selected_status_key == "archived":
        filtered_query = filtered_query.filter(Kullanici.is_deleted.is_(True))
    if selected_airport_key == "global":
        filtered_query = filtered_query.filter(Kullanici.havalimani_id.is_(None))
    elif selected_airport_key:
        filtered_query = filtered_query.filter(Kullanici.havalimani_id == int(selected_airport_key))
    if selected_role_key:
        filtered_query = filtered_query.filter(Kullanici.rol.in_(sorted(expand_role_keys(selected_role_key))))

    liste = filtered_query.order_by(
        func.lower(Kullanici.tam_ad).asc(),
        func.lower(Kullanici.kullanici_adi).asc(),
    ).all()
    if search_term:
        liste = [item for item in liste if _user_matches_search(item, search_term)]
    filter_result_count = len(liste)
    selected_user_id = request.args.get("user_id", type=int)
    selected_user = None
    if selected_user_id:
        selected_user = next((item for item in liste if item.id == selected_user_id), None)
        if selected_user is None:
            scoped_user = scope_query.filter(Kullanici.id == selected_user_id).first()
            if scoped_user and actor_can_view_target_user(current_user, scoped_user):
                flash("Seçilen kayıt geçerli filtre içinde görüntülenemedi.", "warning")
            else:
                flash("Yetkiniz olmayan kayıt görüntülenemedi.", "danger")

    selected_override_allow = set()
    selected_override_deny = set()
    selected_effective_permissions = set()
    selected_user_effective_role = ""
    bulk_airport_staff = []
    if selected_user:
        overrides = get_user_permission_overrides(selected_user)
        selected_override_allow = overrides["allow"]
        selected_override_deny = overrides["deny"]
        selected_effective_permissions = get_effective_permissions(selected_user)
        selected_user_effective_role = get_effective_role(selected_user)
    elif selected_bulk_airport_key:
        bulk_airport_staff = (
            _visible_users_query(current_user)
            .filter(
                Kullanici.is_deleted.is_(False),
                Kullanici.havalimani_id == int(selected_bulk_airport_key),
            )
            .order_by(
                func.lower(Kullanici.tam_ad).asc(),
                func.lower(Kullanici.kullanici_adi).asc(),
            )
            .all()
        )

    create_role_options = _create_role_options_for_actor(current_user, role_options)
    create_default_role = create_role_options[0]["key"] if create_role_options else ""
    create_lock_to_airport = get_effective_role(current_user) == CANONICAL_ROLE_TEAM_LEAD
    create_default_airport_id = str(current_user.havalimani_id or "") if create_lock_to_airport else ""
    can_edit_role_scope_permissions = bool(current_user.is_sahip)

    return render_template(
        'admin/kullanicilar.html',
        kullanicilar=liste,
        havalimanlari=havalimanlari,
        role_options=role_options,
        blood_type_letter_options=BLOOD_TYPE_LETTER_OPTIONS,
        rh_factor_options=RH_FACTOR_OPTIONS,
        body_size_options=BODY_SIZE_OPTIONS,
        shoe_size_options=SHOE_SIZE_OPTIONS,
        status_options=STATUS_OPTIONS,
        permission_catalog=permission_catalog,
        permission_lookup=permission_lookup,
        selected_user=selected_user,
        selected_user_id=selected_user_id,
        search_term=search_term,
        has_active_filters=has_active_filters,
        filter_result_count=filter_result_count,
        selected_role_key=selected_role_key,
        selected_role_label=selected_role_label,
        selected_airport_key=selected_airport_key,
        selected_airport_label=selected_airport_label,
        selected_status_key=selected_status_key,
        selected_status_label=selected_status_label,
        selected_override_allow=selected_override_allow,
        selected_override_deny=selected_override_deny,
        selected_effective_permissions=selected_effective_permissions,
        selected_user_effective_role=selected_user_effective_role,
        create_role_options=create_role_options,
        create_default_role=create_default_role,
        create_lock_to_airport=create_lock_to_airport,
        create_default_airport_id=create_default_airport_id,
        can_edit_role_scope_permissions=can_edit_role_scope_permissions,
        assignable_role_options=assignable_role_options,
        selected_bulk_airport_key=selected_bulk_airport_key,
        selected_bulk_airport_label=selected_bulk_airport_label,
        bulk_airport_staff=bulk_airport_staff,
        create_can_manage_phone=_can_manage_user_phone_on_create(current_user),
        create_is_impersonation=is_impersonation_mode(current_user),
        bulk_import_preview=bulk_import_preview,
        can_download_user_template=has_permission("users.template.download"),
        can_import_users=has_permission("users.import"),
    )


@admin_bp.route('/kullanicilar/toplu-yetki-guncelle', methods=['POST'])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required('users.manage')
def kullanici_toplu_yetki_guncelle():
    if not current_user.is_sahip:
        abort(403)

    selected_bulk_airport_key = (request.form.get("bulk_airport_id") or "").strip()
    visible_airports = _visible_airports(current_user)
    airport_lookup = {str(item.id): item for item in visible_airports}
    selected_airport = airport_lookup.get(selected_bulk_airport_key)
    if not selected_airport:
        flash("Toplu düzenleme için geçerli bir havalimanı seçin.", "danger")
        return redirect(url_for('admin.kullanicilar'))

    selected_user_ids = []
    for raw_value in request.form.getlist("bulk_user_ids"):
        try:
            parsed_id = int(raw_value)
        except (TypeError, ValueError):
            continue
        if parsed_id > 0:
            selected_user_ids.append(parsed_id)
    selected_user_ids = sorted(set(selected_user_ids))
    if not selected_user_ids:
        flash("Toplu düzenleme için en az bir personel seçin.", "warning")
        return redirect(url_for('admin.kullanicilar', bulk_airport_id=selected_airport.id))

    role_lookup = {}
    for role in get_role_options():
        for role_key in expand_role_keys(role["key"]):
            role_lookup[role_key] = role["key"]
    assignable_role_options = _assignable_role_options_for_actor(current_user, get_role_options())
    assignable_role_keys = {item["key"] for item in assignable_role_options}

    selected_roles = []
    for submitted in request.form.getlist("bulk_roles"):
        canonical_role = role_lookup.get((submitted or "").strip(), (submitted or "").strip())
        if canonical_role in assignable_role_keys and canonical_role not in selected_roles:
            selected_roles.append(canonical_role)

    if not selected_roles:
        flash("Toplu düzenleme için en az bir rol seçin.", "warning")
        return redirect(url_for('admin.kullanicilar', bulk_airport_id=selected_airport.id))

    target_role = selected_roles[0]
    target_scope_id = None if target_role in GLOBAL_ROLES else selected_airport.id
    users = (
        _visible_users_query(current_user)
        .filter(
            Kullanici.is_deleted.is_(False),
            Kullanici.id.in_(selected_user_ids),
            Kullanici.havalimani_id == selected_airport.id,
        )
        .all()
    )

    updated_count = 0
    skipped_count = max(0, len(selected_user_ids) - len(users))
    for user in users:
        if not actor_can_manage_target(current_user, user):
            skipped_count += 1
            continue
        if user.rol == target_role and user.havalimani_id == target_scope_id:
            continue
        user.rol = target_role
        user.havalimani_id = target_scope_id
        updated_count += 1

    if updated_count:
        db.session.commit()
        log_kaydet(
            'Güvenlik',
            f'{updated_count} kullanıcı için toplu rol/kapsam güncellendi ({selected_airport.kodu} -> {target_role}).',
            event_key='role.assignment.bulk',
            target_model='Havalimani',
            target_id=selected_airport.id,
        )

    role_label = get_role_labels().get(target_role, target_role)
    if len(selected_roles) > 1:
        flash(f"Birden fazla rol seçildiği için ilk rol uygulandı: {role_label}.", "info")
    if updated_count:
        flash(f"{updated_count} kullanıcı için toplu rol/kapsam güncellendi.", "success")
    else:
        flash("Seçili kullanıcılar zaten aynı rol/kapsamda olduğu için değişiklik yapılmadı.", "info")
    if skipped_count:
        flash(f"{skipped_count} kullanıcı kapsam/yetki kontrolü nedeniyle atlandı.", "warning")
    return redirect(url_for('admin.kullanicilar', bulk_airport_id=selected_airport.id))


@admin_bp.route('/kullanicilar/template.xlsx')
@login_required
@permission_required('users.template.download')
def kullanici_import_sablonu():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kullanicilar"
    sheet.append(BULK_IMPORT_TEMPLATE_COLUMNS)
    sheet.append([
        "Ayse",
        "Yilmaz",
        "ayse.yilmaz@example.com",
        "+90 555 111 22 33",
        "ekip_uyesi",
        "ERZ",
        "aktif",
        "ARFF vardiya personeli",
        "Gecici@123",
        "A",
        "+",
        "168",
        "62",
        "39",
        "M",
        "M",
    ])

    header_fill = PatternFill(fill_type="solid", fgColor="0F2D4A")
    required_fill = PatternFill(fill_type="solid", fgColor="1E4D78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_alignment = Alignment(vertical="center")

    for idx, header in enumerate(BULK_IMPORT_TEMPLATE_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        if header in BULK_IMPORT_COLUMNS:
            cell.fill = required_fill

    width_map = {
        "A": 16,
        "B": 16,
        "C": 30,
        "D": 20,
        "E": 20,
        "F": 26,
        "G": 14,
        "H": 30,
        "I": 20,
        "J": 15,
        "K": 12,
        "L": 10,
        "M": 10,
        "N": 16,
        "O": 12,
        "P": 12,
        "Q": 12,
    }
    for col, width in width_map.items():
        sheet.column_dimensions[col].width = width
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(BULK_IMPORT_TEMPLATE_COLUMNS))}1"
    sheet.freeze_panes = "A2"

    lists = workbook.create_sheet("Listeler")
    lists.append(["rol", "aktif_pasit", "kan_grubu_harf", "kan_grubu_rh", "ayak_numarasi", "ust_beden", "alt_beden"])
    role_values = sorted({item["key"] for item in get_role_options()})
    max_len = max(
        len(role_values),
        2,
        len(BLOOD_TYPE_LETTER_OPTIONS),
        len(RH_FACTOR_OPTIONS),
        len(SHOE_SIZE_OPTIONS),
        len(BODY_SIZE_OPTIONS),
    )
    for idx in range(max_len):
        lists.append(
            [
                role_values[idx] if idx < len(role_values) else "",
                ["aktif", "pasif"][idx] if idx < 2 else "",
                BLOOD_TYPE_LETTER_OPTIONS[idx] if idx < len(BLOOD_TYPE_LETTER_OPTIONS) else "",
                RH_FACTOR_OPTIONS[idx] if idx < len(RH_FACTOR_OPTIONS) else "",
                SHOE_SIZE_OPTIONS[idx] if idx < len(SHOE_SIZE_OPTIONS) else "",
                BODY_SIZE_OPTIONS[idx] if idx < len(BODY_SIZE_OPTIONS) else "",
                BODY_SIZE_OPTIONS[idx] if idx < len(BODY_SIZE_OPTIONS) else "",
            ]
        )

    def _add_dropdown(column_letter, formula):
        validation = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=True)
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}5000")

    _add_dropdown("E", "'Listeler'!$A$2:$A$5000")
    _add_dropdown("G", "'Listeler'!$B$2:$B$3")
    _add_dropdown("J", "'Listeler'!$C$2:$C$5")
    _add_dropdown("K", "'Listeler'!$D$2:$D$3")
    _add_dropdown("N", "'Listeler'!$E$2:$E$5000")
    _add_dropdown("O", "'Listeler'!$F$2:$F$5000")
    _add_dropdown("P", "'Listeler'!$G$2:$G$5000")
    _add_dropdown("Q", "'Listeler'!$G$2:$G$5000")
    lists.sheet_state = "hidden"

    notes = workbook.create_sheet("Aciklamalar")
    notes.append(["Alan", "Açıklama"])
    notes.append(["rol", "Sistemde tanımlı rol anahtarı kullanılmalıdır. Örn: ekip_uyesi, ekip_sorumlusu, admin"])
    notes.append(["havalimani", "Kod, ad veya görünür havalimanı ID değeri kullanılabilir. Global roller için boş bırakılabilir."])
    notes.append(["aktif/pasif", "aktif veya pasif değerlerinden biri kullanılmalıdır."])
    notes.append(["telefon", "Telefon yalnızca site sahibi içe aktarıyorsa kaydedilir."])
    notes.append(["kan_grubu_harf", "A, B, AB veya 0 değerlerinden biri seçilebilir."])
    notes.append(["kan_grubu_rh", "+ veya - seçilmelidir; harf doluysa bu alan da doldurulmalıdır."])
    notes.append(["boy_cm / kilo_kg", "Boy 90-260 cm, kilo 30-250 kg aralığında olmalıdır."])
    notes.append(["ayak_numarasi", "Listeden seçilmelidir."])
    notes.append(["ust_beden / alt_beden", "XS, S, M, L, XL, XXL, 3XL seçenekleri kullanılmalıdır."])
    notes.append(["Not", "Lacivert başlıklar zorunlu sütunları gösterir."])
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 96

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="kullanici_import_sablonu.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route('/kullanicilar/import/preview', methods=['POST'])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "10 per minute"))
@permission_required('users.import')
def kullanici_import_preview():
    upload = request.files.get("import_file")
    if not upload or not upload.filename:
        flash("İçe aktarma için bir Excel dosyası seçin.", "danger")
        return redirect(url_for('admin.kullanicilar'))

    safe_name = secure_upload_filename(upload.filename)
    if not is_allowed_file(safe_name, {"xlsx", "xls"}):
        flash("Yalnızca Excel dosyaları yüklenebilir.", "danger")
        return redirect(url_for('admin.kullanicilar'))

    try:
        frame = pd.read_excel(upload)
    except Exception:
        flash("Excel dosyası okunamadı. Şablon dosyasını kullanıp tekrar deneyin.", "danger")
        return redirect(url_for('admin.kullanicilar'))

    preview = _build_user_import_preview(current_user, frame, get_role_options(), _visible_airports(current_user))
    session[_bulk_import_session_key()] = preview
    session.modified = True

    if preview["summary"]["valid"]:
        flash(f"{preview['summary']['valid']} satır önizleme için hazırlandı.", "success")
    if preview["summary"]["invalid"]:
        flash(f"{preview['summary']['invalid']} satır doğrulama hatası içeriyor.", "warning")
    return redirect(url_for('admin.kullanicilar'))


@admin_bp.route('/kullanicilar/import/commit', methods=['POST'])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "10 per minute"))
@permission_required('users.import')
def kullanici_import_commit():
    preview = session.get(_bulk_import_session_key()) or {}
    valid_rows = preview.get("valid_rows") or []
    if not valid_rows:
        flash("Onaylanacak geçerli kullanıcı satırı bulunmuyor.", "danger")
        return redirect(url_for('admin.kullanicilar'))

    created_count = 0
    skipped_count = 0
    for row in valid_rows:
        existing = _find_user_by_email(row["kullanici_adi"])
        if existing:
            skipped_count += 1
            continue
        user = Kullanici(
            tam_ad=row["tam_ad"],
            kullanici_adi=row["kullanici_adi"],
            rol=row["rol"],
            havalimani_id=row["havalimani_id"],
            telefon_numarasi=row["telefon_numarasi"] if current_user.is_sahip else None,
            is_deleted=bool(row["is_deleted"]),
        )
        user.sifre_set(row["gecici_sifre"] or _default_import_password())
        _apply_user_profile_fields(user, row.get("profile_fields"))
        db.session.add(user)
        created_count += 1

    db.session.commit()
    session.pop(_bulk_import_session_key(), None)
    session.modified = True

    flash(f"{created_count} kullanıcı içe aktarıldı.", "success")
    if skipped_count:
        flash(f"{skipped_count} satır mevcut e-posta çakışması nedeniyle atlandı.", "warning")
    return redirect(url_for('admin.kullanicilar'))


@admin_bp.route('/kullanicilar/import/clear', methods=['POST'])
@login_required
@permission_required('users.import')
def kullanici_import_clear():
    session.pop(_bulk_import_session_key(), None)
    session.modified = True
    flash("İçe aktarma önizlemesi temizlendi.", "info")
    return redirect(url_for('admin.kullanicilar'))


@admin_bp.route('/kullanici-ekle', methods=['POST'])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required('users.manage')
def kullanici_ekle():
    """Yeni kullanıcı ekler."""
    tam_ad = _normalize_full_name(request.form.get('tam_ad'))
    k_adi, email_error = _validate_user_email(request.form.get('k_adi'))
    submitted_role = (request.form.get('rol') or "").strip()
    role_lookup = {}
    for role in get_role_options():
        for role_key in expand_role_keys(role["key"]):
            role_lookup[role_key] = role["key"]
    rol = role_lookup.get(submitted_role, submitted_role)
    h_id = request.form.get('h_id', type=int)
    sifre = request.form.get('sifre')
    telefon_numarasi = None
    profile_fields, profile_error = _collect_user_profile_fields(request.form)
    actor_role = get_effective_role(current_user)

    if not tam_ad:
        flash("Ad soyad alanını doldurun.", "danger")
        return redirect(url_for('admin.kullanicilar'))
    if email_error:
        flash(email_error, "danger")
        return redirect(url_for('admin.kullanicilar'))
    password_error = _validate_user_password(sifre)
    if password_error:
        flash(password_error, "danger")
        return redirect(url_for('admin.kullanicilar'))
    if profile_error:
        flash(profile_error, "danger")
        return redirect(url_for('admin.kullanicilar'))

    if _can_manage_user_phone_on_create(current_user):
        telefon_numarasi, phone_error = _normalize_phone_number(request.form.get('telefon_numarasi'))
        if request.form.get('telefon_numarasi') and phone_error:
            flash(phone_error, "danger")
            return redirect(url_for('admin.kullanicilar'))
    
    if actor_role == CANONICAL_ROLE_TEAM_LEAD:
        rol = CANONICAL_ROLE_TEAM_MEMBER
        h_id = current_user.havalimani_id
    elif not can_assign_role(current_user, rol):
        abort(403)

    if rol in GLOBAL_ROLES:
        h_id = None
    elif rol in AIRPORT_ROLES and not h_id:
        flash("Saha personeli için birim seçimi zorunludur!", "danger")
        return redirect(url_for('admin.kullanicilar'))

    # Kullanıcı adı kontrolü
    mevcut = _find_user_by_email(k_adi)
    if mevcut:
        flash("Bu e-posta/kullanıcı adı zaten kullanımda!", "warning")
        return redirect(url_for('admin.kullanicilar'))

    yeni = Kullanici(
        tam_ad=tam_ad, 
        kullanici_adi=k_adi, 
        rol=rol, 
        havalimani_id=h_id,
        telefon_numarasi=telefon_numarasi,
    )
    _apply_user_profile_fields(yeni, profile_fields)
    yeni.sifre_set(sifre)
    db.session.add(yeni)
    db.session.commit()
    rol_etiketi = get_role_labels().get(rol, rol)
    
    log_kaydet('Güvenlik', f'Yeni kullanıcı ({rol_etiketi}) eklendi: {k_adi}', event_key='user.create', target_model='Kullanici', target_id=yeni.id)
    flash(f"{tam_ad} personeli sisteme eklendi.", "success")
    if telefon_numarasi:
        flash("Telefon numarası kaydedildi.", "success")
    return redirect(url_for('admin.kullanicilar', user_id=yeni.id))


@admin_bp.route('/kullanici-guncelle/<int:id>', methods=['POST'])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required('users.manage')
def kullanici_guncelle(id):
    user = db.session.get(Kullanici, id)
    if not user or user.is_deleted or not actor_can_manage_target(current_user, user):
        abort(403)

    yeni_rol = request.form.get('rol') or user.rol
    role_labels = get_role_labels()
    yeni_tam_ad = _normalize_full_name(request.form.get('tam_ad') or user.tam_ad)
    yeni_email, email_error = _validate_user_email(request.form.get('k_adi') or user.kullanici_adi)

    h_id = request.form.get('h_id', type=int)
    role_lookup = {}
    for role in get_role_options():
        for role_key in expand_role_keys(role["key"]):
            role_lookup[role_key] = role["key"]
    yeni_rol = role_lookup.get(yeni_rol, yeni_rol)

    if yeni_rol in GLOBAL_ROLES:
        h_id = None
    elif not h_id:
        h_id = user.havalimani_id or current_user.havalimani_id

    eski_rol = user.rol
    allow_permissions = request.form.getlist('allow_permissions')
    deny_permissions = request.form.getlist('deny_permissions')
    yeni_telefon_numarasi = user.telefon_numarasi
    profile_fields, profile_error = _collect_user_profile_fields(request.form)
    if current_user.is_sahip:
        yeni_telefon_numarasi, phone_error = _normalize_phone_number(request.form.get('telefon_numarasi'))
        if request.form.get('telefon_numarasi') and phone_error:
            flash(phone_error, "danger")
            return redirect(url_for('admin.kullanicilar', user_id=user.id))
    phone_changed = yeni_telefon_numarasi != user.telefon_numarasi

    if not yeni_tam_ad:
        flash("Ad soyad alanını doldurun.", "danger")
        return redirect(url_for('admin.kullanicilar', user_id=user.id))
    if email_error:
        flash(email_error, "danger")
        return redirect(url_for('admin.kullanicilar', user_id=user.id))
    if profile_error:
        flash(profile_error, "danger")
        return redirect(url_for('admin.kullanicilar', user_id=user.id))

    current_overrides = get_user_permission_overrides(user)
    role_or_scope_changed = (
        yeni_rol != user.rol
        or h_id != user.havalimani_id
        or set(allow_permissions) != set(current_overrides["allow"])
        or set(deny_permissions) != set(current_overrides["deny"])
    )
    if role_or_scope_changed and not current_user.is_sahip:
        abort(403)

    approval_required = False
    if not approval_required and not can_assign_role(current_user, yeni_rol):
        abort(403)
    if approval_required:
        payload = json.dumps(
            {
                "user_id": user.id,
                "tam_ad": yeni_tam_ad,
                "k_adi": yeni_email,
                "rol": yeni_rol,
                "h_id": h_id,
                "allow_permissions": allow_permissions,
                "deny_permissions": deny_permissions,
            },
            ensure_ascii=False,
        )
        approval = create_approval_request(
            approval_type="role_change",
            target_model="Kullanici",
            target_id=user.id,
            requested_by_id=current_user.id,
            request_payload=payload,
            commit=False,
        )
        if approval:
            create_notification(
                user.id,
                "approval_pending",
                "Rol değişikliği onaya gönderildi",
                f"Hesabınız için talep edilen {role_labels.get(yeni_rol, yeni_rol)} rolü yönetici onayı bekliyor.",
                link_url=url_for('admin.approvals'),
                severity="warning",
                commit=False,
            )
            log_kaydet(
                'Güvenlik',
                f'Kullanıcı rol değişikliği approval bekliyor: {user.kullanici_adi} ({eski_rol} -> {yeni_rol})',
                event_key='role.assignment.pending',
                target_model='Kullanici',
                target_id=user.id,
                commit=False,
            )
            db.session.commit()
            audit_log('role.assignment.pending', outcome='success', target_user_id=user.id, requested_role=yeni_rol)
            flash("Rol değişikliği onaya gönderildi.", "warning")
            return redirect(url_for('admin.kullanicilar', user_id=user.id))

    existing = _find_user_by_email(yeni_email)
    if existing and existing.id != user.id:
        flash("Bu e-posta/kullanıcı adı zaten kullanımda!", "warning")
        return redirect(url_for('admin.kullanicilar', user_id=user.id))

    user.tam_ad = yeni_tam_ad
    user.kullanici_adi = yeni_email
    user.rol = yeni_rol
    user.havalimani_id = h_id
    if current_user.is_sahip:
        user.telefon_numarasi = yeni_telefon_numarasi
    _apply_user_profile_fields(user, profile_fields)
    update_user_permission_overrides(user.id, allow_permissions, deny_permissions)
    db.session.commit()

    log_kaydet(
        'Güvenlik',
        f'Kullanıcı rol/yetki güncellendi: {user.kullanici_adi} ({eski_rol} -> {yeni_rol})',
        event_key='role.assignment.change',
        target_model='Kullanici',
        target_id=user.id,
    )
    audit_log('role.assignment.change', outcome='success', target_user_id=user.id, old_role=eski_rol, new_role=yeni_rol)
    create_notification(
        user.id,
        "role_change",
        "Rol ve yetkiler güncellendi",
        f"Hesabınız için yeni rol ataması yapıldı: {role_labels.get(yeni_rol, yeni_rol)}",
        link_url=url_for('admin.kullanicilar'),
        severity="info",
    )
    flash("Kullanıcı yetkileri güncellendi.", "success")
    if phone_changed and current_user.is_sahip:
        flash("Telefon numarası kaydedildi.", "success")
    return redirect(url_for('admin.kullanicilar', user_id=user.id))

@admin_bp.route('/kullanici-sil/<int:id>', methods=['GET'], endpoint='kullanici_sil_legacy')
@login_required
@permission_required('users.manage')
def kullanici_sil_legacy(id):
    flash("Bu işlem yalnızca form gönderimi ile yapılabilir.", "warning")
    return redirect(url_for('admin.kullanicilar'))


@admin_bp.route('/kullanici-sil/<int:id>', methods=['POST'])
@login_required
@limiter.limit(lambda: current_app.config.get("CRITICAL_POST_RATE_LIMIT", "20 per minute"))
@permission_required('users.manage')
def kullanici_sil(id):
    """Kullanıcıyı soft-delete ile arşivler."""
    user = db.session.get(Kullanici, id)
    
    if not user or user.is_deleted:
        flash("Kullanıcı bulunamadı!", "danger")
    elif not actor_can_manage_target(current_user, user):
        abort(403)
    elif _normalize_user_email(user.kullanici_adi) == 'mehmetcinocevi@gmail.com':
        flash("Ana yönetici hesabı silinemez!", "danger")
    else:
        k_adi = user.kullanici_adi
        
        # ✅ SOFT DELETE: db.session.delete yerine kendi metodumuzu çağırıyoruz
        user.soft_delete()
        db.session.commit()
        
        log_kaydet('Güvenlik', f'Kullanıcı silindi (Arşivlendi): {k_adi}')
        flash(f"{k_adi} kullanıcısı sistemden kaldırıldı.", "info")
        
    return redirect(url_for('admin.kullanicilar'))
