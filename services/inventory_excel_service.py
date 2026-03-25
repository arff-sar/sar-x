import io

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from services.inventory_template_service import (
    FIELD_SCHEMA,
    SHEET_DATA,
    SHEET_HELP,
    SHEET_LISTS,
    excel_headers,
    required_excel_headers,
)


class ExcelTemplateError(ValueError):
    pass


def build_inventory_template_workbook(*, lists_context):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = SHEET_DATA
    ws_lists = wb.create_sheet(SHEET_LISTS)
    ws_help = wb.create_sheet(SHEET_HELP)

    headers = excel_headers()
    ws_data.append(headers)
    ws_data.freeze_panes = "A2"

    list_columns = [
        ("A", "merkezi_sablon", lists_context.get("templates", [])),
        ("B", "havalimani", lists_context.get("airports", [])),
        ("C", "kategori", lists_context.get("categories", [])),
        ("D", "kullanim_durumu", lists_context.get("statuses", ["aktif", "pasif"])),
        ("E", "bakim_formu", lists_context.get("maintenance_forms", [])),
        ("F", "bakim_periyodu", [str(item) for item in lists_context.get("month_values", list(range(1, 13)))]),
        ("G", "kutu_kodu", lists_context.get("boxes", [])),
        ("H", "evet_hayir", ["Evet", "Hayır"]),
    ]
    for col, title, values in list_columns:
        ws_lists[f"{col}1"] = title
        for idx, value in enumerate(values, start=2):
            ws_lists[f"{col}{idx}"] = value

    header_to_col = {header: idx + 1 for idx, header in enumerate(headers)}

    def _add_dropdown(header, formula):
        col_idx = header_to_col[header]
        col_letter = ws_data.cell(row=1, column=col_idx).column_letter
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws_data.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}5000")

    _add_dropdown("merkezi_sablon", f"'{SHEET_LISTS}'!$A$2:$A$5000")
    _add_dropdown("havalimani", f"'{SHEET_LISTS}'!$B$2:$B$5000")
    _add_dropdown("kategori", f"'{SHEET_LISTS}'!$C$2:$C$5000")
    _add_dropdown("kullanim_durumu", f"'{SHEET_LISTS}'!$D$2:$D$5000")
    _add_dropdown("bakim_formu", f"'{SHEET_LISTS}'!$E$2:$E$5000")
    _add_dropdown("bakim_periyodu", f"'{SHEET_LISTS}'!$F$2:$F$5000")
    _add_dropdown("kutu_kodu", f"'{SHEET_LISTS}'!$G$2:$G$5000")
    _add_dropdown("demirbas_mi", f"'{SHEET_LISTS}'!$H$2:$H$3")
    _add_dropdown("kalibrasyon_gerekli_mi", f"'{SHEET_LISTS}'!$H$2:$H$3")
    _add_dropdown("merkezi_sablondan_olustur", f"'{SHEET_LISTS}'!$H$2:$H$3")

    ws_help["A1"] = "KULLANIM TALIMATI"
    ws_help["A2"] = "Zorunlu alanlar: " + ", ".join(required_excel_headers())
    ws_help["A3"] = "Tarih formatı: YYYY-MM-DD (alternatif: DD.MM.YYYY)"
    ws_help["A4"] = "Evet/Hayır alanları: Evet/Hayır, 1/0, true/false, x/boş"
    ws_help["A5"] = "Merkezi şablondan oluşturmak için: merkezi_sablondan_olustur=Evet"
    ws_help["A6"] = "Başlıklar değiştirilemez. Beklenen kolonlar: " + ", ".join(excel_headers())

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_inventory_workbook(file_obj):
    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # pragma: no cover
        raise ExcelTemplateError("Excel dosyası okunamadı.") from exc

    missing = [name for name in (SHEET_DATA, SHEET_LISTS, SHEET_HELP) if name not in wb.sheetnames]
    if missing:
        raise ExcelTemplateError(f"Eksik sheet: {', '.join(missing)}")

    ws = wb[SHEET_DATA]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ExcelTemplateError("VERI_GIRISI sheet'i boş.")

    expected = excel_headers()
    actual = [str(cell or "").strip() for cell in rows[0]]
    if actual != expected:
        raise ExcelTemplateError("Başlık yapısı beklenen şablonla eşleşmiyor.")

    parsed = []
    for row_no, row_values in enumerate(rows[1:], start=2):
        payload = {expected[idx]: row_values[idx] if idx < len(row_values) else None for idx in range(len(expected))}
        if not any(value not in (None, "") for value in payload.values()):
            continue
        parsed.append({"row_no": row_no, "values": payload})
    return parsed

