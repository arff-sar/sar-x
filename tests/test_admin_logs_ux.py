from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.dialects import postgresql

from decorators import update_user_permission_overrides
from extensions import db
from models import IslemLog
from routes.admin.logs import _format_timestamp_label
from tests.factories import HavalimaniFactory, KullaniciFactory


def _login(client, user_id):
    with client.session_transaction() as session:
        session.clear()
        session["_user_id"] = str(user_id)
        session["_fresh"] = True


def test_event_type_filter_is_selectable_and_really_filters_results(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-logs@sarx.com", havalimani=airport)
        db.session.add_all([airport, owner])
        db.session.flush()
        db.session.add_all(
            [
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Giriş",
                    event_key="auth.login",
                    detay="Kullanıcı sisteme giriş yaptı.",
                    outcome="success",
                ),
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Rapor",
                    event_key="reports.export.completed",
                    detay="PDF dışa aktarma tamamlandı.",
                    target_model="InventoryAsset",
                    target_id=15,
                    outcome="success",
                ),
            ]
        )
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/islem-loglari?event_type=Rapor")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'name="event_type"' in html
    assert "Rapor ve Dışa Aktarma" in html
    assert "PDF dışa aktarma tamamlandı." in html
    assert "Kullanıcı sisteme giriş yaptı." not in html
    assert "1 kayıt" in html
    assert "Beklenmedik Bir Hata" not in html
    assert "Sistem bağlantı hatası oluştu." not in html


def test_logs_screen_translates_outcomes_and_hides_technical_labels(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-logs-tr@sarx.com")
        db.session.add(owner)
        db.session.flush()
        db.session.add_all(
            [
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Güvenlik",
                    event_key="auth.login",
                    detay="Başarılı giriş denemesi kaydedildi.",
                    outcome="success",
                ),
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Yetki",
                    event_key="permission.matrix.update",
                    detay="Yetki matrisi güncellendi.",
                    target_model="Role",
                    target_id=3,
                    outcome="failed",
                ),
                IslemLog(
                    kullanici_id=None,
                    islem_tipi="Sistem",
                    detay="Eski yapıdan taşınan kayıt.",
                    outcome="legacy",
                ),
            ]
        )
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/islem-loglari")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "İşlem Kayıtları" in html
    assert "Başarılı" in html
    assert "Başarısız" in html
    assert "Eski kayıt" in html
    assert "Event Key" not in html
    assert "Hedef Model" not in html
    assert "Success" not in html
    assert "Failed" not in html
    assert "LEGACY" not in html


def test_target_model_filter_and_empty_state_render_cleanly(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-logs-empty@sarx.com")
        db.session.add(owner)
        db.session.flush()
        db.session.add_all(
            [
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Yetki",
                    event_key="role.assignment.change",
                    detay="Rol ataması güncellendi.",
                    target_model="Kullanici",
                    target_id=7,
                    outcome="success",
                ),
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Rapor",
                    event_key="reports.export.failed",
                    detay="Dışa aktarma başarısız oldu.",
                    target_model="InventoryAsset",
                    target_id=11,
                    outcome="failed",
                ),
            ]
        )
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    filtered = client.get("/islem-loglari?target_model=Kullanici")
    filtered_html = filtered.data.decode("utf-8")

    assert filtered.status_code == 200
    assert "İlgili Kayıt Türü" in filtered_html
    assert "Kullanıcı" in filtered_html
    assert "Rol ataması güncellendi." in filtered_html
    assert "Dışa aktarma başarısız oldu." not in filtered_html

    empty = client.get("/islem-loglari?event_type=Yetki&outcome=failed")
    empty_html = empty.data.decode("utf-8")

    assert empty.status_code == 200
    assert "Kayıt bulunamadı" in empty_html
    assert "Seçili filtrelerle eşleşen işlem kaydı yok." in empty_html


def test_user_filter_renders_selected_user_without_crashing(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Samsun Havalimanı", kodu="SZF")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-logs-user@sarx.com", havalimani=airport)
        other = KullaniciFactory(rol="personel", is_deleted=False, kullanici_adi="staff-logs-user@sarx.com", havalimani=airport)
        db.session.add_all([airport, owner, other])
        db.session.flush()
        db.session.add_all(
            [
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Giriş",
                    detay="Sahip kullanıcısı sisteme giriş yaptı.",
                    outcome="success",
                ),
                IslemLog(
                    kullanici_id=other.id,
                    islem_tipi="Bakım",
                    detay="Personel bakım akışı başlattı.",
                    outcome="success",
                ),
            ]
        )
        db.session.commit()
        owner_id = owner.id
        other_id = other.id

    _login(client, owner_id)
    response = client.get(f"/islem-loglari?user_id={other_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Personel bakım akışı başlattı." in html
    assert "Sahip kullanıcısı sisteme giriş yaptı." not in html
    assert "Kullanıcı: " in html


def test_team_lead_only_sees_own_airport_logs_and_filter_options(client, app):
    with app.app_context():
        own_airport = HavalimaniFactory(ad="Dalaman Havalimanı", kodu="DLM")
        other_airport = HavalimaniFactory(ad="Trabzon Havalimanı", kodu="TZX")
        team_lead = KullaniciFactory(
            rol="ekip_sorumlusu",
            is_deleted=False,
            kullanici_adi="dlm-logs@sarx.com",
            havalimani=own_airport,
        )
        own_user = KullaniciFactory(
            rol="ekip_uyesi",
            is_deleted=False,
            kullanici_adi="dlm-staff@sarx.com",
            havalimani=own_airport,
        )
        other_user = KullaniciFactory(
            rol="ekip_uyesi",
            is_deleted=False,
            kullanici_adi="tzx-staff@sarx.com",
            havalimani=other_airport,
        )
        db.session.add_all([own_airport, other_airport, team_lead, own_user, other_user])
        db.session.flush()
        update_user_permission_overrides(team_lead.id, ["logs.view"], [])
        db.session.add_all(
            [
                IslemLog(
                    kullanici_id=own_user.id,
                    havalimani_id=None,
                    islem_tipi="Envanter",
                    detay="Kendi havalimanı envanter güncellemesi.",
                    outcome="success",
                ),
                IslemLog(
                    kullanici_id=other_user.id,
                    havalimani_id=None,
                    islem_tipi="Bakım",
                    detay="Diğer havalimanı bakım kaydı.",
                    outcome="success",
                ),
                IslemLog(
                    kullanici_id=None,
                    havalimani_id=own_airport.id,
                    islem_tipi="Rapor",
                    detay="Havalimanı bazlı sistem raporu.",
                    outcome="success",
                ),
            ]
        )
        db.session.commit()
        team_lead_id = team_lead.id

    _login(client, team_lead_id)
    response = client.get("/islem-loglari")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Kendi havalimanı envanter güncellemesi." in html
    assert "Havalimanı bazlı sistem raporu." in html
    assert "Diğer havalimanı bakım kaydı." not in html
    assert 'value="Bakım"' not in html
    assert "dlm-staff@sarx.com" not in html
    assert "tzx-staff@sarx.com" not in html


def test_distinct_filter_queries_stay_postgresql_safe(app):
    with app.app_context():
        event_type_sql = str(
            IslemLog.query.with_entities(IslemLog.islem_tipi)
            .filter(IslemLog.islem_tipi.isnot(None))
            .distinct()
            .statement.compile(dialect=postgresql.dialect())
        )
        target_model_sql = str(
            IslemLog.query.with_entities(IslemLog.target_model)
            .filter(IslemLog.target_model.isnot(None))
            .distinct()
            .statement.compile(dialect=postgresql.dialect())
        )

    assert "ORDER BY lower(" not in event_type_sql
    assert "ORDER BY lower(" not in target_model_sql


def test_logs_timestamp_formatter_supports_string_values():
    assert _format_timestamp_label("2026-03-30 10:45:12") == "30.03.2026 10:45:12"


def test_logs_timestamp_formatter_converts_utc_offset_to_tr():
    assert _format_timestamp_label("2026-03-31T09:15:00+00:00") == "31.03.2026 12:15:00"


def test_logs_resolve_actor_name_from_email_when_fk_is_missing(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Kars Havalimanı", kodu="KSY")
        owner = KullaniciFactory(
            rol="sahip",
            is_deleted=False,
            kullanici_adi="actor-email@sarx.com",
            tam_ad="Actor Email Kullanıcısı",
            havalimani=airport,
        )
        db.session.add_all([airport, owner])
        db.session.flush()
        db.session.add_all(
            [
                IslemLog(
                    kullanici_id=None,
                    user_email=owner.kullanici_adi,
                    islem_tipi="Giriş",
                    event_key="auth.login",
                    detay="Email fallback ile giriş kaydı.",
                    outcome="success",
                ),
                IslemLog(
                    kullanici_id=None,
                    islem_tipi="Sistem",
                    detay="Tamamen sistem kaydı.",
                    outcome="info",
                ),
            ]
        )
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/islem-loglari")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Actor Email Kullanıcısı" in html
    assert "Tamamen sistem kaydı." in html
    assert "Sistem" in html


def test_logs_pagination_keeps_filters_and_slices_results(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-logs-pages@sarx.com")
        db.session.add(owner)
        db.session.flush()
        db.session.add(
            IslemLog(
                kullanici_id=owner.id,
                islem_tipi="Giriş",
                detay="Filtre dışı kayıt",
                outcome="success",
            )
        )
        for index in range(25):
            db.session.add(
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Envanter",
                    detay=f"Sayfalanan kayıt {index}",
                    target_model="Kutu",
                    target_id=index + 1,
                    outcome="success",
                )
            )
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/islem-loglari?event_type=Envanter&target_model=Kutu&page=2")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Sayfalanan kayıt 24" not in html
    assert "Sayfalanan kayıt 4" in html
    assert "Filtre dışı kayıt" not in html
    assert "event_type=Envanter" in html
    assert "target_model=Kutu" in html
    assert "Sayfa 2 / 2" in html


def test_logs_excel_export_respects_filters(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-logs-export@sarx.com")
        db.session.add(owner)
        db.session.flush()
        db.session.add_all(
            [
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Rapor",
                    detay="Excel dışa aktarma kaydı",
                    outcome="success",
                ),
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Giriş",
                    detay="Filtre dışı giriş kaydı",
                    outcome="success",
                ),
            ]
        )
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/islem-loglari/excel?event_type=Rapor")

    assert response.status_code == 200
    assert "spreadsheetml.sheet" in response.headers["Content-Type"]

    workbook = load_workbook(filename=BytesIO(response.data))
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    assert headers == ["Tarih", "Kullanıcı", "İşlem", "İlgili Kayıt", "Sonuç", "Açıklama"]
    assert len(rows) == 1
    assert rows[0][2] == "Rapor ve Dışa Aktarma"
    assert rows[0][5] == "Excel dışa aktarma kaydı"
