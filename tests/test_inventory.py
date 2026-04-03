import pytest
import io
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook
from tests.factories import (
    EquipmentTemplateFactory,
    InventoryAssetFactory,
    MalzemeFactory,
    KullaniciFactory,
    HavalimaniFactory,
    KutuFactory,
)
from extensions import db, safe_display_filename
from datetime import date

from models import (
    AssignmentRecord,
    AssignmentRecipient,
    BakimKaydi,
    EquipmentTemplate,
    InventoryAsset,
    InventoryCategory,
    Malzeme,
    PPERecord,
)


def _login(client, user_id):
    with client.session_transaction() as sess:
        sess["_user_id"] = str(user_id)
        sess["_fresh"] = True


def test_safe_display_filename_preserves_turkish_names_for_visible_downloads():
    assert safe_display_filename("Şule Işık.pdf", fallback="fallback.pdf", default_extension=".pdf") == "Şule Işık.pdf"
    assert safe_display_filename("Çağrı Göğüş Formu.pdf", fallback="fallback.pdf", default_extension=".pdf") == "Çağrı Göğüş Formu.pdf"
    assert safe_display_filename("İzmir Çiğli Zimmet.pdf", fallback="fallback.pdf", default_extension=".pdf") == "İzmir Çiğli Zimmet.pdf"


def test_safe_display_filename_uses_secure_legacy_fallback_for_path_like_values():
    assert safe_display_filename("AYT/zimmet/Belge_Alan/kkd_belge.pdf", fallback="ZMT-001.pdf", default_extension=".pdf") == "kkd_belge.pdf"
    assert safe_display_filename("", fallback="ZMT-001.pdf", default_extension=".pdf") == "ZMT-001.pdf"

# 1. YETKİ KONTROLÜ: Giriş yapmayan kullanıcı envanteri göremez
def test_envanter_access_required_login(client):
    response = client.get('/envanter')
    assert response.status_code == 302

# 2. SOFT DELETE KONTROLÜ: Silinmiş (Arşivlenmiş) malzemeler listede çıkmaz
def test_envanter_list_active_only(client, app):
    user = KullaniciFactory(rol="sistem_sorumlusu")
    m1 = MalzemeFactory(ad="Aktif Ekipman", is_deleted=False)
    m2 = MalzemeFactory(ad="Arşivlenmiş Ekipman", is_deleted=True)
    
    db.session.add_all([user, m1, m2])
    db.session.commit()
    
    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id) 
        sess['_fresh'] = True
    
    response = client.get('/envanter')
    data_str = response.data.decode('utf-8')
    
    assert response.status_code == 200
    assert "Aktif Ekipman" in data_str
    assert "Arşivlenmiş Ekipman" not in data_str
    assert 'href="/zimmetler" class="row-btn" style="padding:16px; border-radius:16px; justify-content:flex-start;"' not in data_str
    assert 'href="/kkd" class="row-btn" style="padding:16px; border-radius:16px; justify-content:flex-start;"' not in data_str

# 3. BİRİM FİLTRELEME: Personel sadece kendi havalimanındaki malzemeyi görür
def test_havalimani_isolation(client, app):
    h1 = HavalimaniFactory(kodu="ESB", ad="Ankara")
    h2 = HavalimaniFactory(kodu="SAW", ad="İstanbul")
    
    m1 = MalzemeFactory(ad="Ankara Cihazı", havalimani=h1)
    m2 = MalzemeFactory(ad="İstanbul Cihazı", havalimani=h2)
    
    user = KullaniciFactory(rol="ekip_uyesi", havalimani=h1)
    
    db.session.add_all([h1, h2, m1, m2, user])
    db.session.commit() 
    
    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True
    
    response = client.get('/envanter')
    data_str = response.data.decode('utf-8')
    
    assert response.status_code == 200
    assert "Ankara Cihazı" in data_str
    assert "İstanbul Cihazı" not in data_str


def test_malzeme_sil_restricts_deletion_to_user_airport_scope(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    airport_a = HavalimaniFactory(kodu="AAA")
    airport_b = HavalimaniFactory(kodu="BBB")
    box_b = KutuFactory(havalimani=airport_b)
    scoped_manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport_a, is_deleted=False)
    foreign_material = MalzemeFactory(ad="Kapsam Dışı Malzeme", kutu=box_b, is_deleted=False)
    db.session.add_all([airport_a, airport_b, box_b, scoped_manager, foreign_material])
    db.session.commit()

    _login(client, scoped_manager.id)
    response = client.post(f"/malzeme-sil/{foreign_material.id}", follow_redirects=True)

    assert response.status_code == 200
    assert "Hata: Malzeme bulunamadı veya zaten silinmiş." in response.data.decode("utf-8")
    db.session.refresh(foreign_material)
    assert foreign_material.is_deleted is False


def test_malzeme_sil_rejects_external_referrer_redirect(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    owner = KullaniciFactory(rol="sistem_sorumlusu", is_deleted=False)
    material = MalzemeFactory(ad="Redirect Test Malzeme", is_deleted=False)
    db.session.add_all([owner, material])
    db.session.commit()

    _login(client, owner.id)
    response = client.post(
        f"/malzeme-sil/{material.id}",
        headers={"Referer": "https://attacker.example/phish"},
        follow_redirects=False,
    )

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/dashboard")

# 4. YAZMA YETKİSİ: Yetkili kullanıcı malzeme ekleyebilir (Kapsam Artırıcı)
def test_malzeme_ekle_success(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    user = KullaniciFactory(rol="sistem_sorumlusu")
    h = HavalimaniFactory(kodu="IST")
    db.session.add_all([user, h])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    response = client.post('/malzeme-ekle', data={
        'ad': 'Test Malzemesi',
        'seri_no': 'SN12345',
        'kutu_kodu': 'K-99',
        'stok': 10,
        'durum': 'Aktif',
        'havalimani_id': h.id
    }, follow_redirects=True)

    assert response.status_code == 200
    assert "Malzeme başarıyla eklendi" in response.data.decode('utf-8')
    # Veritabanına gerçekten yazılmış mı kontrol et
    assert Malzeme.query.filter_by(ad="Test Malzemesi").first() is not None


def test_malzeme_ekle_allows_non_owner_without_template_selection(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    airport = HavalimaniFactory(kodu="ADB")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False)
    box = KutuFactory(kodu="ADB-K-21", havalimani=airport)
    db.session.add_all([airport, manager, box])
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(manager.id)
        sess["_fresh"] = True

    response = client.post(
        "/malzeme-ekle",
        data={
            "kategori": "Elektronik",
            "ad": "Şablonsuz Ekipman",
            "marka": "Demo",
            "model": "X1",
            "seri_no": "ADB-NO-TEMPLATE-001",
            "kutu_id": box.id,
            "stok": 1,
            "durum": "aktif",
            "bakim_periyodu_ay": 6,
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    html = response.data.decode("utf-8")
    assert "Malzeme başarıyla eklendi" in html
    asset = InventoryAsset.query.filter_by(serial_no="ADB-NO-TEMPLATE-001").first()
    assert asset is not None
    assert asset.equipment_template is not None
    assert asset.equipment_template.name == "Şablonsuz Ekipman"


def test_malzeme_ekle_allows_blank_required_like_fields_via_fallbacks(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    airport = HavalimaniFactory(kodu="BJV")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False)
    db.session.add_all([airport, manager])
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(manager.id)
        sess["_fresh"] = True

    response = client.post(
        "/malzeme-ekle",
        data={
            "ad": "",
            "kategori": "",
            "kutu_id": "",
            "template_id": "",
            "stok": "",
            "seri_no": "",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Malzeme başarıyla eklendi" in response.data.decode("utf-8")
    created_material = (
        Malzeme.query.filter_by(havalimani_id=airport.id, is_deleted=False)
        .order_by(Malzeme.id.desc())
        .first()
    )
    assert created_material is not None
    assert created_material.ad == "Genel Ekipman"
    assert created_material.kutu is not None
    assert created_material.kutu.kodu == "BJV-ATANMADI"
    created_asset = InventoryAsset.query.filter_by(legacy_material_id=created_material.id).first()
    assert created_asset is not None
    assert created_asset.equipment_template is not None
    assert created_asset.equipment_template.name == "Genel Ekipman"

    second_response = client.post(
        "/malzeme-ekle",
        data={
            "ad": "",
            "kategori": "",
            "kutu_id": "",
            "template_id": "",
            "stok": "",
            "seri_no": "",
        },
        follow_redirects=True,
    )

    assert second_response.status_code == 200
    assert "Malzeme başarıyla eklendi" in second_response.data.decode("utf-8")
    created_materials = (
        Malzeme.query.filter_by(havalimani_id=airport.id, is_deleted=False)
        .order_by(Malzeme.id.asc())
        .all()
    )
    assert len(created_materials) == 2
    assert created_materials[0].seri_no is None
    assert created_materials[1].seri_no is None


def test_envanter_kategori_ekle_allows_team_lead_role(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    airport = HavalimaniFactory(kodu="GZT")
    team_lead = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False)
    db.session.add_all([airport, team_lead])
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(team_lead.id)
        sess["_fresh"] = True

    response = client.post(
        "/envanter/kategori-ekle",
        data={"name": "Araç Üstü Sistemler", "description": "Ekip sorumlusu ekledi"},
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Kategori eklendi." in response.data.decode("utf-8")
    created = InventoryCategory.query.filter_by(name="Araç Üstü Sistemler", is_deleted=False).first()
    assert created is not None


def test_merkezi_sablon_ekle_accepts_new_category_option(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    owner = KullaniciFactory(rol="sistem_sorumlusu", is_deleted=False)
    db.session.add(owner)
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner.id)
        sess["_fresh"] = True

    response = client.post(
        "/envanter/merkezi-sablon-ekle",
        data={
            "name": "Yeni Kategori Şablonu",
            "category": "__new__",
            "new_category_name": "Robotik Sistemler",
            "brand": "ARFF",
            "model_code": "RX-1",
            "maintenance_period_months": 6,
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Merkezi şablon eklendi." in response.data.decode("utf-8")
    created_template = EquipmentTemplate.query.filter_by(name="Yeni Kategori Şablonu", is_deleted=False).first()
    assert created_template is not None
    assert created_template.category == "Robotik Sistemler"
    created_category = InventoryCategory.query.filter_by(name="Robotik Sistemler", is_deleted=False).first()
    assert created_category is not None


def test_single_create_assigns_asset_code_and_qr(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="ESB")
    box = KutuFactory(kodu="ESB-KUTU-11", havalimani=airport)
    template = EquipmentTemplateFactory(name="Kod QR Test", category="Elektronik")
    db.session.add_all([user, airport, box, template])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    response = client.post(
        '/malzeme-ekle',
        data={
            'havalimani_id': airport.id,
            'template_id': template.id,
            'kategori': 'Elektronik',
            'ad': 'Kod QR Test',
            'seri_no': 'SINGLE-CODE-001',
            'kutu_id': box.id,
            'stok': 1,
            'durum': 'aktif',
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    asset = InventoryAsset.query.filter_by(serial_no="SINGLE-CODE-001").first()
    assert asset is not None
    assert (asset.asset_code or "").startswith("ARFF-SAR-")
    assert (asset.qr_code or "").startswith("http")
    assert f"/asset/{asset.id}/quick" in (asset.qr_code or "")


def test_asset_qr_image_endpoint_returns_png(client, app):
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="GZT")
    template = EquipmentTemplateFactory(name="QR Endpoint Template")
    asset = InventoryAssetFactory(equipment_template=template, airport=airport, status="aktif")
    db.session.add_all([user, airport, template, asset])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    response = client.get(f"/api/qr-img/asset/{asset.id}")
    assert response.status_code == 200
    assert response.mimetype == "image/png"


def test_kkd_page_renders_catalog_and_keeps_pool_outside_general_inventory(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Antalya Havalimanı", kodu="AYT")
        owner = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-owner@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="KKD Personeli")
        db.session.add_all([airport, owner, recipient])
        db.session.flush()
        record = PPERecord(
            user_id=recipient.id,
            airport_id=airport.id,
            category="Baş ve Yüz Koruması",
            subcategory="Baret",
            item_name="Operasyon Bareti",
            quantity=1,
            status="aktif",
            physical_condition="iyi",
            is_active=True,
            created_by_id=owner.id,
        )
        db.session.add(record)
        db.session.commit()
        owner_id = owner.id
        recipient_id = recipient.id

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner_id)
        sess["_fresh"] = True

    response = client.get(f"/kkd?user_id={recipient_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Personel Bazlı KKD Akışı" in html
    assert "Baş ve Yüz Koruması" in html
    assert "Baret" in html
    assert "Haberleşme Ekipmanı" not in html
    assert "Operasyon Bareti" in html

    inventory_html = client.get("/envanter").data.decode("utf-8")
    assert "Operasyon Bareti" not in inventory_html


def test_kkd_personnel_flow_starts_compact_until_selection(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Esenboğa Havalimanı", kodu="ESB")
        owner = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-compact@sarx.com")
        first_user = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Birinci Personel")
        second_user = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="İkinci Personel")
        db.session.add_all([airport, owner, first_user, second_user])
        db.session.flush()
        db.session.add_all(
            [
                PPERecord(
                    user_id=first_user.id,
                    airport_id=airport.id,
                    category="Baş ve Yüz Koruması",
                    subcategory="Baret",
                    item_name="Birinci Baret",
                    quantity=1,
                    status="aktif",
                    physical_condition="iyi",
                    is_active=True,
                    created_by_id=owner.id,
                ),
                PPERecord(
                    user_id=second_user.id,
                    airport_id=airport.id,
                    category="Ayak Koruması",
                    subcategory="Çizme",
                    item_name="İkinci Çizme",
                    quantity=1,
                    status="aktif",
                    physical_condition="iyi",
                    is_active=True,
                    created_by_id=owner.id,
                ),
            ]
        )
        db.session.commit()
        owner_id = owner.id

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner_id)
        sess["_fresh"] = True

    response = client.get("/kkd")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'id="ppeSelectionShell"' in html
    assert "max-height:min(58vh, 700px);" in html
    assert html.count('class="ppe-user-accordion" open') == 0


def test_kkd_personnel_selection_opens_only_selected_accordion(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Adana Havalimanı", kodu="ADA")
        owner = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-selected@sarx.com")
        selected_person = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Seçili KKD Personeli")
        other_person = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Kapalı KKD Personeli")
        db.session.add_all([airport, owner, selected_person, other_person])
        db.session.flush()
        selected_record = PPERecord(
            user_id=selected_person.id,
            airport_id=airport.id,
            category="Baş ve Yüz Koruması",
            subcategory="Vizör",
            item_name="Seçili Vizör",
            quantity=1,
            status="aktif",
            physical_condition="hasarli",
            is_active=True,
            created_by_id=owner.id,
            signed_document_url="https://example.com/ada.pdf",
        )
        other_record = PPERecord(
            user_id=other_person.id,
            airport_id=airport.id,
            category="Vücut Koruması",
            subcategory="Yağmurluk",
            item_name="Kapalı Yağmurluk",
            quantity=1,
            status="aktif",
            physical_condition="iyi",
            is_active=True,
            created_by_id=owner.id,
        )
        db.session.add_all([selected_record, other_record])
        db.session.commit()
        owner_id = owner.id
        selected_user_id = selected_person.id

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner_id)
        sess["_fresh"] = True

    response = client.get(f"/kkd?user_id={selected_user_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert html.count('class="ppe-user-accordion" open') == 0
    assert "Seçili Vizör" in html
    assert "Kapalı Yağmurluk" not in html
    assert "İmzalı PDF" in html
    assert "Hasarlı" in html


def test_inventory_excel_upload_rejects_fake_xlsx_payload(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    owner = KullaniciFactory(rol="sistem_sorumlusu", is_deleted=False, kullanici_adi="inventory-fake-xlsx@sarx.com")
    db.session.add(owner)
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner.id)
        sess["_fresh"] = True

    response = client.post(
        "/malzeme-ekle/excel-yukle",
        data={"excel_file": (io.BytesIO(b"not-a-real-xlsx"), "import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Excel dosyası okunamadı." in response.data.decode("utf-8")


def test_kkd_create_validates_manufacturer_url_and_size_rules(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    with app.app_context():
        airport = HavalimaniFactory(ad="İzmir Havalimanı", kodu="ADB")
        owner = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-create@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="URL Test Personeli")
        db.session.add_all([airport, owner, recipient])
        db.session.commit()
        owner_id = owner.id
        recipient_id = recipient.id

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner_id)
        sess["_fresh"] = True

    invalid_response = client.post(
        "/kkd",
        data={
            "user_id": recipient_id,
            "category": "Ayak Koruması",
            "subcategory": "Çelik Burunlu İş Botu",
            "item_name": "Bot",
            "manufacturer_url": "gecersiz-link",
            "physical_condition": "iyi",
            "is_active": "1",
        },
        follow_redirects=True,
    )
    invalid_html = invalid_response.data.decode("utf-8")
    assert invalid_response.status_code == 200
    assert "Üretici sayfası için geçerli bir bağlantı girin." in invalid_html

    size_response = client.post(
        "/kkd",
        data={
            "user_id": recipient_id,
            "category": "Ayak Koruması",
            "subcategory": "Çelik Burunlu İş Botu",
            "item_name": "Bot",
            "manufacturer_url": "https://example.com/bot",
            "physical_condition": "iyi",
            "is_active": "1",
        },
        follow_redirects=True,
    )
    assert "geçerli bir ayakkabı numarası" in size_response.data.decode("utf-8")

    valid_response = client.post(
        "/kkd",
        data={
            "user_id": recipient_id,
            "category": "Ayak Koruması",
            "subcategory": "Çelik Burunlu İş Botu",
            "item_name": "Bot",
            "shoe_size": "42",
            "manufacturer_url": "https://example.com/bot",
            "physical_condition": "iyi",
            "is_active": "1",
        },
        follow_redirects=True,
    )
    assert valid_response.status_code == 200
    assert "KKD kaydı oluşturuldu." in valid_response.data.decode("utf-8")
    with app.app_context():
        assert PPERecord.query.filter_by(item_name="Bot", shoe_size="42").first() is not None


def test_kkd_excel_template_and_import_flow(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    with app.app_context():
        airport = HavalimaniFactory(ad="Muğla Dalaman Havalimanı", kodu="DLM")
        owner = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-excel@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Excel Personeli")
        db.session.add_all([airport, owner, recipient])
        db.session.commit()
        owner_id = owner.id

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner_id)
        sess["_fresh"] = True

    template_response = client.get("/kkd/excel-sablon")
    assert template_response.status_code == 200
    workbook = load_workbook(io.BytesIO(template_response.data))
    assert workbook.sheetnames == ["VERI_GIRISI", "LISTELER", "YARDIM"]
    assert workbook["VERI_GIRISI"]["A1"].value == "havalimani"

    payload = io.BytesIO(template_response.data)
    workbook = load_workbook(payload)
    sheet = workbook["VERI_GIRISI"]
    sheet.append([
        "DLM - Muğla Dalaman Havalimanı",
        "Excel Personeli",
        "Vücut Koruması",
        "Reflektif Yelek",
        "Reflektif Yelek A",
        "MarkaX",
        "ModelY",
        "SERI-01",
        "L",
        "",
        "2026-03-30",
        "2026-01-01",
        "2026-12-31",
        1,
        "İyi",
        "Evet",
        "https://example.com/yelek",
    ])
    sheet.append([
        "DLM - Muğla Dalaman Havalimanı",
        "Excel Personeli",
        "Ayak Koruması",
        "Çizme",
        "Bozuk Satır",
        "",
        "",
        "",
        "",
        "",
        "2026-03-30",
        "",
        "",
        1,
        "İyi",
        "Evet",
        "not-a-url",
    ])
    out = io.BytesIO()
    workbook.save(out)
    out.seek(0)

    response = client.post(
        "/kkd/excel-yukle",
        data={"excel_file": (out, "kkd-import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    html = response.data.decode("utf-8")
    assert response.status_code == 200
    assert "1 KKD satırı içe aktarıldı." in html
    assert "1 satır doğrulama hatası nedeniyle alınmadı." in html
    assert "KKD Excel Hata Raporu" in html
    with app.app_context():
        assert PPERecord.query.filter_by(item_name="Reflektif Yelek A").first() is not None


def test_kkd_signed_document_upload_uses_pdf_only_folder_standard(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    with app.app_context():
        airport = HavalimaniFactory(ad="Antalya Havalimanı", kodu="AYT")
        owner = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-pdf@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Belge Alan")
        db.session.add_all([airport, owner, recipient])
        db.session.flush()
        record = PPERecord(
            user_id=recipient.id,
            airport_id=airport.id,
            category="Baş ve Yüz Koruması",
            subcategory="Baret",
            item_name="Belge Bareti",
            quantity=1,
            status="aktif",
            physical_condition="iyi",
            is_active=True,
            created_by_id=owner.id,
        )
        db.session.add(record)
        db.session.commit()
        owner_id = owner.id
        record_id = record.id

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner_id)
        sess["_fresh"] = True

    with patch("routes.inventory.get_storage_adapter") as mocked_storage:
        mocked_storage.return_value.save_upload.return_value = SimpleNamespace(
            storage_key="AYT/zimmet/Belge_Alan/kkd_belge_alan_zimmet_20260330010101.pdf",
            public_url="https://example.com/uploads/AYT/zimmet/Belge_Alan/kkd_belge_alan_zimmet_20260330010101.pdf",
        )
        response = client.post(
            f"/kkd/{record_id}/signed-document",
            data={"signed_document": (io.BytesIO(b"%PDF-1.4 ppe"), "Çağrı Göğüş Formu.pdf", "application/pdf")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )

    assert response.status_code == 200
    assert "İmzalı KKD zimmet PDF" in response.data.decode("utf-8")
    upload_kwargs = mocked_storage.return_value.save_upload.call_args.kwargs
    assert upload_kwargs["folder"] == "AYT/zimmet/Belge_Alan"
    assert upload_kwargs["filename"].startswith("kkd_belge_alan_zimmet_")
    assert upload_kwargs["filename"].endswith(".pdf")
    with app.app_context():
        stored = db.session.get(PPERecord, record_id)
        assert stored is not None
        assert stored.signed_document_name == "Çağrı Göğüş Formu.pdf"

    reject_response = client.post(
        f"/kkd/{record_id}/signed-document",
        data={"signed_document": (io.BytesIO(b"png"), "signed.png", "image/png")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert "Dosya uzantısı desteklenmiyor." in reject_response.data.decode("utf-8")


def test_kkd_excel_upload_rejects_fake_xlsx_payload(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    owner = KullaniciFactory(rol="sistem_sorumlusu", is_deleted=False, kullanici_adi="kkd-fake-xlsx@sarx.com")
    db.session.add(owner)
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner.id)
        sess["_fresh"] = True

    response = client.post(
        "/kkd/excel-yukle",
        data={"excel_file": (io.BytesIO(b"not-a-real-xlsx"), "kkd-import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "KKD Excel dosyası okunamadı." in response.data.decode("utf-8")


def test_kkd_signed_document_download_streams_pdf_for_manager_scope_user(client, app, tmp_path):
    pdf_bytes = b"%PDF-1.4 ppe-manager"
    app.config["STORAGE_BACKEND"] = "local"
    app.config["LOCAL_UPLOAD_ROOT"] = str(tmp_path)

    with app.app_context():
        airport = HavalimaniFactory(ad="Antalya Havalimanı", kodu="AYT")
        manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-manager@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="KKD Alan")
        db.session.add_all([airport, manager, recipient])
        db.session.flush()
        record = PPERecord(
            user_id=recipient.id,
            airport_id=airport.id,
            category="Baş ve Yüz Koruması",
            subcategory="Baret",
            item_name="İmzalı Baret",
            quantity=1,
            status="aktif",
            physical_condition="iyi",
            is_active=True,
            created_by_id=manager.id,
            signed_document_key="AYT/zimmet/KKD_Alan/kkd_manager.pdf",
            signed_document_url="/static/uploads/AYT/zimmet/KKD_Alan/kkd_manager.pdf",
            signed_document_name="kkd_manager.pdf",
        )
        db.session.add(record)
        db.session.commit()
        manager_id = manager.id
        record_id = record.id

    target_path = tmp_path / "AYT" / "zimmet" / "KKD_Alan" / "kkd_manager.pdf"
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(pdf_bytes)

    with client.session_transaction() as sess:
        sess["_user_id"] = str(manager_id)
        sess["_fresh"] = True

    response = client.get(f"/kkd/{record_id}/signed-document/download")

    assert response.status_code == 200
    assert response.data == pdf_bytes
    assert response.mimetype == "application/pdf"
    assert "attachment;" in response.headers.get("Content-Disposition", "")
    assert "kkd_manager.pdf" in response.headers.get("Content-Disposition", "")
    assert response.headers.get("Location") is None
    assert response.headers.get("Cache-Control") == "no-store, no-cache, must-revalidate, max-age=0, private"


def test_kkd_signed_document_download_uses_person_scope_and_local_public_url_fallback(client, app, tmp_path):
    pdf_bytes = b"%PDF-1.4 ppe-person"
    app.config["STORAGE_BACKEND"] = "local"
    app.config["LOCAL_UPLOAD_ROOT"] = str(tmp_path)

    with app.app_context():
        airport = HavalimaniFactory(ad="Adana Havalimanı", kodu="ADA")
        manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-owner@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="KKD Belge Sahibi")
        db.session.add_all([airport, manager, recipient])
        db.session.flush()
        record = PPERecord(
            user_id=recipient.id,
            airport_id=airport.id,
            category="Baş ve Yüz Koruması",
            subcategory="Vizör",
            item_name="İmzalı Vizör",
            quantity=1,
            status="aktif",
            physical_condition="iyi",
            is_active=True,
            created_by_id=manager.id,
            signed_document_key=None,
            signed_document_url="/static/uploads/ADA/zimmet/KKD_Belge_Sahibi/kkd_person.pdf",
            signed_document_name="kkd_person.pdf",
        )
        db.session.add(record)
        db.session.commit()
        recipient_id = recipient.id
        record_id = record.id

    target_path = tmp_path / "ADA" / "zimmet" / "KKD_Belge_Sahibi" / "kkd_person.pdf"
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(pdf_bytes)

    with client.session_transaction() as sess:
        sess["_user_id"] = str(recipient_id)
        sess["_fresh"] = True

    response = client.get(f"/kkd/{record_id}/signed-document/download")

    assert response.status_code == 200
    assert response.data == pdf_bytes
    assert response.mimetype == "application/pdf"
    assert "kkd_person.pdf" in response.headers.get("Content-Disposition", "")


def test_kkd_signed_document_download_denies_unrelated_user(client, app, tmp_path):
    app.config["STORAGE_BACKEND"] = "local"
    app.config["LOCAL_UPLOAD_ROOT"] = str(tmp_path)

    with app.app_context():
        airport = HavalimaniFactory(ad="İzmir Havalimanı", kodu="ADB")
        manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="kkd-manager-2@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Yetkili KKD")
        unrelated = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Yetkisiz KKD")
        db.session.add_all([airport, manager, recipient, unrelated])
        db.session.flush()
        record = PPERecord(
            user_id=recipient.id,
            airport_id=airport.id,
            category="Vücut Koruması",
            subcategory="Reflektif Yelek",
            item_name="İmzalı Yelek",
            quantity=1,
            status="aktif",
            physical_condition="iyi",
            is_active=True,
            created_by_id=manager.id,
            signed_document_key="ADB/zimmet/Yetkili_KKD/kkd_deny.pdf",
            signed_document_url="/static/uploads/ADB/zimmet/Yetkili_KKD/kkd_deny.pdf",
            signed_document_name="kkd_deny.pdf",
        )
        db.session.add(record)
        db.session.commit()
        unrelated_id = unrelated.id
        record_id = record.id

    target_path = tmp_path / "ADB" / "zimmet" / "Yetkili_KKD" / "kkd_deny.pdf"
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(b"%PDF-1.4 ppe-deny")

    with client.session_transaction() as sess:
        sess["_user_id"] = str(unrelated_id)
        sess["_fresh"] = True

    response = client.get(f"/kkd/{record_id}/signed-document/download")

    assert response.status_code == 404


def test_assignment_signed_document_download_streams_pdf_for_airport_scope_user(client, app, tmp_path):
    pdf_bytes = b"%PDF-1.4 assignment-owner"
    app.config["STORAGE_BACKEND"] = "local"
    app.config["LOCAL_UPLOAD_ROOT"] = str(tmp_path)

    with app.app_context():
        airport = HavalimaniFactory(ad="Antalya Havalimanı", kodu="AYT")
        owner = KullaniciFactory(rol="sistem_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="zimmet-owner@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Belge Alan")
        db.session.add_all([airport, owner, recipient])
        db.session.flush()
        assignment = AssignmentRecord(
            assignment_no="ZMT-OWNER-001",
            airport_id=airport.id,
            created_by_id=owner.id,
            signed_document_key="AYT/zimmet/Belge_Alan/zimmet_owner.pdf",
            signed_document_url="/static/uploads/AYT/zimmet/Belge_Alan/zimmet_owner.pdf",
            signed_document_name="zimmet_owner.pdf",
        )
        db.session.add(assignment)
        db.session.flush()
        db.session.add(AssignmentRecipient(assignment_id=assignment.id, user_id=recipient.id))
        db.session.commit()
        assignment_id = assignment.id
        owner_id = owner.id

    target_path = tmp_path / "AYT" / "zimmet" / "Belge_Alan" / "zimmet_owner.pdf"
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(pdf_bytes)

    with client.session_transaction() as sess:
        sess["_user_id"] = str(owner_id)
        sess["_fresh"] = True

    response = client.get(f"/zimmetler/{assignment_id}/signed-document/download")

    assert response.status_code == 200
    assert response.data == pdf_bytes
    assert response.mimetype == "application/pdf"
    assert "attachment;" in response.headers.get("Content-Disposition", "")
    assert "zimmet_owner.pdf" in response.headers.get("Content-Disposition", "")
    assert response.headers.get("Location") is None
    assert response.headers.get("Cache-Control") == "no-store, no-cache, must-revalidate, max-age=0, private"


def test_assignment_signed_document_download_uses_recipient_scope_and_local_public_url_fallback(client, app, tmp_path):
    pdf_bytes = b"%PDF-1.4 assignment-recipient"
    app.config["STORAGE_BACKEND"] = "local"
    app.config["LOCAL_UPLOAD_ROOT"] = str(tmp_path)

    with app.app_context():
        airport = HavalimaniFactory(ad="Adana Havalimanı", kodu="ADA")
        owner = KullaniciFactory(rol="sistem_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="zimmet-manager@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Belge Alıcısı")
        db.session.add_all([airport, owner, recipient])
        db.session.flush()
        assignment = AssignmentRecord(
            assignment_no="ZMT-RECIPIENT-001",
            airport_id=airport.id,
            created_by_id=owner.id,
            signed_document_key=None,
            signed_document_url="/static/uploads/ADA/zimmet/Belge_Alicisi/zimmet_recipient.pdf",
            signed_document_name="zimmet_recipient.pdf",
        )
        db.session.add(assignment)
        db.session.flush()
        db.session.add(AssignmentRecipient(assignment_id=assignment.id, user_id=recipient.id))
        db.session.commit()
        assignment_id = assignment.id
        recipient_id = recipient.id

    target_path = tmp_path / "ADA" / "zimmet" / "Belge_Alicisi" / "zimmet_recipient.pdf"
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(pdf_bytes)

    with client.session_transaction() as sess:
        sess["_user_id"] = str(recipient_id)
        sess["_fresh"] = True

    response = client.get(f"/zimmetler/{assignment_id}/signed-document/download")

    assert response.status_code == 200
    assert response.data == pdf_bytes
    assert response.mimetype == "application/pdf"
    assert "zimmet_recipient.pdf" in response.headers.get("Content-Disposition", "")


def test_assignment_signed_document_download_denies_unrelated_user(client, app, tmp_path):
    app.config["STORAGE_BACKEND"] = "local"
    app.config["LOCAL_UPLOAD_ROOT"] = str(tmp_path)

    with app.app_context():
        airport = HavalimaniFactory(ad="İzmir Havalimanı", kodu="ADB")
        owner = KullaniciFactory(rol="sistem_sorumlusu", havalimani=airport, is_deleted=False, kullanici_adi="zimmet-owner-2@sarx.com")
        recipient = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Yetkili Personel")
        unrelated = KullaniciFactory(rol="ekip_uyesi", havalimani=airport, is_deleted=False, tam_ad="Yetkisiz Personel")
        db.session.add_all([airport, owner, recipient, unrelated])
        db.session.flush()
        assignment = AssignmentRecord(
            assignment_no="ZMT-DENY-001",
            airport_id=airport.id,
            created_by_id=owner.id,
            signed_document_key="ADB/zimmet/Yetkili_Personel/zimmet_deny.pdf",
            signed_document_url="/static/uploads/ADB/zimmet/Yetkili_Personel/zimmet_deny.pdf",
            signed_document_name="zimmet_deny.pdf",
        )
        db.session.add(assignment)
        db.session.flush()
        db.session.add(AssignmentRecipient(assignment_id=assignment.id, user_id=recipient.id))
        db.session.commit()
        assignment_id = assignment.id
        unrelated_id = unrelated.id

    target_path = tmp_path / "ADB" / "zimmet" / "Yetkili_Personel" / "zimmet_deny.pdf"
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(b"%PDF-1.4 assignment-deny")

    with client.session_transaction() as sess:
        sess["_user_id"] = str(unrelated_id)
        sess["_fresh"] = True

    response = client.get(f"/zimmetler/{assignment_id}/signed-document/download")

    assert response.status_code == 404


def test_asset_qr_image_payload_uses_detail_url_even_when_legacy_qr_code_exists(client, app, monkeypatch):
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="ESB")
    template = EquipmentTemplateFactory(name="QR Legacy Template")
    asset = InventoryAssetFactory(
        equipment_template=template,
        airport=airport,
        status="aktif",
        qr_code="LEGACY-QR-CODE-001",
    )
    db.session.add_all([user, airport, template, asset])
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(user.id)
        sess["_fresh"] = True

    captured = {}

    def _fake_generate_qr_data(payload):
        captured["payload"] = payload
        return io.BytesIO(b"png")

    monkeypatch.setattr("routes.inventory.generate_qr_data", _fake_generate_qr_data)

    response = client.get(f"/api/qr-img/asset/{asset.id}")
    assert response.status_code == 200
    assert response.mimetype == "image/png"
    assert captured["payload"].startswith("http")
    assert f"/asset/{asset.id}/quick" in captured["payload"]

# 5. YETKİ KONTROLÜ: Düz personel malzeme ekleyemez
def test_personel_cannot_add_material(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    
    user = KullaniciFactory(rol="ekip_uyesi")
    db.session.add(user)
    db.session.commit() 
    
    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True
        
    response = client.post('/malzeme-ekle', data={
        'ad': 'Yeni Hortum',
        'kutu_kodu': 'K-01'
    })
    
    assert response.status_code == 403

# 6. BAKIM KAYDI: Bakım kaydı başarıyla girilebilir (Kapsam Artırıcı)
def test_bakim_kaydet_success(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    user = KullaniciFactory(rol="sistem_sorumlusu")
    m = MalzemeFactory(ad="Bakım Cihazı")
    db.session.add_all([user, m])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    response = client.post(f'/bakim-kaydet/{m.id}', data={
        'not': 'Yıllık genel bakım yapıldı',
        'maliyet': '500.50'
    }, follow_redirects=True)

    assert response.status_code == 200
    assert "Bakım kaydı başarıyla işlendi" in response.data.decode('utf-8')


def test_bakim_kaydet_offline_sync_idempotent_by_request_id(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    user = KullaniciFactory(rol="sistem_sorumlusu")
    m = MalzemeFactory(ad="Offline Bakım Cihazı")
    db.session.add_all([user, m])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    request_id = "offline-test-req-001"
    headers = {
        "X-SARX-Offline-Sync": "1",
        "X-SARX-Offline-Request-Id": request_id,
    }

    first = client.post(
        f'/bakim-kaydet/{m.id}',
        data={'not': 'Offline bakım denemesi', 'maliyet': '10'},
        headers=headers,
        follow_redirects=False,
    )
    second = client.post(
        f'/bakim-kaydet/{m.id}',
        data={'not': 'Offline bakım denemesi', 'maliyet': '10'},
        headers=headers,
        follow_redirects=False,
    )

    assert first.status_code == 200
    assert first.get_json()["status"] == "success"
    assert second.status_code == 200
    assert second.get_json()["status"] == "success"
    assert second.get_json().get("duplicate") is True
    assert BakimKaydi.query.filter_by(malzeme_id=m.id).count() == 1

# 7. RAPORLAMA: Excel ve PDF çıktıları (Kapsam Artırıcı)
def test_export_routes(client, app):
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="ERZ", ad="Erzurum Havalimanı")
    box = KutuFactory(kodu="ERZ-SAR-01", havalimani=airport)
    m = MalzemeFactory(
        ad="Export Test Malzemesi",
        havalimani=airport,
        kutu=box,
        seri_no="EXP-001",
    )
    db.session.add_all([user, airport, box, m])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    # Excel Testi
    excel_res = client.get('/envanter/excel')
    assert excel_res.status_code == 200
    assert excel_res.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    workbook = load_workbook(io.BytesIO(excel_res.data))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == ("Birim", "Kutu", "Malzeme Adı", "Seri No", "Durum", "Son Bakım", "Gelecek Bakım")
    assert ("ERZ", "ERZ-SAR-01", "Export Test Malzemesi", "EXP-001", m.durum, "-", "-") in rows[1:]

    # PDF Testi
    pdf_res = client.get('/envanter/pdf')
    assert pdf_res.status_code == 200
    assert pdf_res.mimetype == 'application/pdf'

# 8. B PLANI (MANUEL BULMA) KONTROLÜ
def test_kutu_bul_manual(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    
    h1 = HavalimaniFactory(kodu="ESB")
    kutu = KutuFactory(kodu="K-99", havalimani=h1)
    user = KullaniciFactory(rol="ekip_uyesi", havalimani=h1)
    
    db.session.add_all([h1, kutu, user])
    db.session.commit() 
    
    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True
    
    response = client.post('/kutu-bul', data={'kutu_kodu': 'K-99'}, follow_redirects=True)
    data_str = response.data.decode('utf-8')
    
    assert response.status_code == 200
    assert "K-99" in data_str

# 9. QR API: QR resim üretme rotası (Kapsam Artırıcı)
def test_qr_api_route(client, app):
    user = KullaniciFactory(rol="sistem_sorumlusu")
    k = KutuFactory(kodu="QR-TEST")
    db.session.add_all([user, k])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    response = client.get('/api/qr-img/QR-TEST')
    assert response.status_code == 200
    assert response.mimetype == 'image/png'


def test_malzeme_ekle_accepts_canonical_keys(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="AYT")
    box = KutuFactory(kodu="AYT-BOX-1", havalimani=airport)
    template = EquipmentTemplateFactory(name="Gaz Dedektörü", category="Elektronik")
    db.session.add_all([user, airport, box, template])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    response = client.post(
        '/malzeme-ekle',
        data={
            'airport_id': airport.id,
            'box_id': box.id,
            'template_id': template.id,
            'category': 'Elektronik',
            'asset_name': 'Gaz Dedektörü',
            'serial_no': 'CANON-001',
            'unit_count': 3,
            'status': 'aktif',
            'maintenance_period_months': 4,
            'notes': 'canonical-contract',
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    asset = InventoryAsset.query.filter_by(serial_no="CANON-001").first()
    assert asset is not None
    assert asset.status == "aktif"
    assert asset.unit_count == 3
    assert asset.maintenance_period_months == 4


def test_asset_duzenle_accepts_canonical_status_and_notes(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    airport = HavalimaniFactory(kodu="RZE")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport)
    template = EquipmentTemplateFactory(name="Kamera", category="Elektronik")
    asset = InventoryAssetFactory(equipment_template=template, airport=airport, status="aktif", serial_no="UPD-001", unit_count=1)
    db.session.add_all([airport, manager, template, asset])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(manager.id)
        sess['_fresh'] = True

    response = client.post(
        f"/asset-duzenle/{asset.id}",
        data={
            "status": "pasif",
            "unit_count": 2,
            "notes": "canonical-edit-note",
            "manual_url": "https://example.com/guide",
            "maintenance_period_months": 5,
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    refreshed = db.session.get(InventoryAsset, asset.id)
    assert refreshed.status == "pasif"
    assert refreshed.unit_count == 2
    assert refreshed.notes == "canonical-edit-note"


def test_asset_duzenle_accepts_legacy_keys(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    airport = HavalimaniFactory(kodu="ERZ")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport)
    template = EquipmentTemplateFactory(name="Pompa", category="Mekanik")
    asset = InventoryAssetFactory(
        equipment_template=template,
        airport=airport,
        status="aktif",
        serial_no="LEG-UPD-001",
        unit_count=1,
    )
    db.session.add_all([airport, manager, template, asset])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(manager.id)
        sess['_fresh'] = True

    response = client.post(
        f"/asset-duzenle/{asset.id}",
        data={
            "durum": "arizali",
            "stok": 4,
            "notlar": "legacy-edit-note",
            "bakim": "2026-02-14",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    refreshed = db.session.get(InventoryAsset, asset.id)
    assert refreshed.status == "pasif"
    assert refreshed.unit_count == 4
    assert refreshed.notes == "legacy-edit-note"
    assert refreshed.last_maintenance_date == date(2026, 2, 14)


def test_quick_detail_accepts_legacy_and_canonical_note_keys(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    airport = HavalimaniFactory(kodu="ADA")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport)
    template = EquipmentTemplateFactory(name="Projektör")
    asset = InventoryAssetFactory(equipment_template=template, airport=airport, status="aktif", notes="ilk")
    db.session.add_all([airport, manager, template, asset])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(manager.id)
        sess['_fresh'] = True

    first = client.post(
        f"/asset/{asset.id}/quick",
        data={"status": "aktif", "note": "legacy-note"},
        follow_redirects=True,
    )
    assert first.status_code == 200
    second = client.post(
        f"/asset/{asset.id}/quick",
        data={"status": "aktif", "notes": "canonical-note"},
        follow_redirects=True,
    )
    assert second.status_code == 200
    refreshed = db.session.get(InventoryAsset, asset.id)
    assert "legacy-note" in (refreshed.notes or "")
    assert "canonical-note" in (refreshed.notes or "")


def test_quick_detail_accepts_legacy_durum_key(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    airport = HavalimaniFactory(kodu="ASR")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport)
    template = EquipmentTemplateFactory(name="Termal Kamera")
    asset = InventoryAssetFactory(equipment_template=template, airport=airport, status="aktif")
    db.session.add_all([airport, manager, template, asset])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(manager.id)
        sess['_fresh'] = True

    response = client.post(
        f"/asset/{asset.id}/quick",
        data={"durum": "pasif", "note": "legacy-status-key"},
        follow_redirects=True,
    )
    assert response.status_code == 200
    refreshed = db.session.get(InventoryAsset, asset.id)
    assert refreshed.status == "pasif"


def test_quick_detail_merges_maintenance_date_aliases(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    airport = HavalimaniFactory(kodu="GZT")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport)
    template = EquipmentTemplateFactory(name="Kompresor")
    asset = InventoryAssetFactory(equipment_template=template, airport=airport, status="aktif")
    db.session.add_all([airport, manager, template, asset])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(manager.id)
        sess['_fresh'] = True

    r1 = client.post(
        f"/asset/{asset.id}/quick",
        data={"status": "aktif", "bakim": "2026-01-01"},
        follow_redirects=True,
    )
    assert r1.status_code == 200
    assert db.session.get(InventoryAsset, asset.id).last_maintenance_date == date(2026, 1, 1)

    r2 = client.post(
        f"/asset/{asset.id}/quick",
        data={"status": "aktif", "son_bakim_tarihi": "2026-02-02"},
        follow_redirects=True,
    )
    assert r2.status_code == 200
    assert db.session.get(InventoryAsset, asset.id).last_maintenance_date == date(2026, 2, 2)

    r3 = client.post(
        f"/asset/{asset.id}/quick",
        data={"status": "aktif", "last_maintenance_date": "2026-03-03"},
        follow_redirects=True,
    )
    assert r3.status_code == 200
    assert db.session.get(InventoryAsset, asset.id).last_maintenance_date == date(2026, 3, 3)


def test_malzeme_ekle_accepts_legacy_kutu_id_resolution(client, app):
    app.config['WTF_CSRF_ENABLED'] = False
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="BJV")
    box = KutuFactory(kodu="BJV-BOX-7", havalimani=airport)
    template = EquipmentTemplateFactory(name="Hortum", category="Kurtarma")
    db.session.add_all([user, airport, box, template])
    db.session.commit()

    with client.session_transaction() as sess:
        sess['_user_id'] = str(user.id)
        sess['_fresh'] = True

    response = client.post(
        "/malzeme-ekle",
        data={
            "havalimani_id": airport.id,
            "kutu_id": box.id,
            "template_id": template.id,
            "kategori": "Kurtarma",
            "ad": "Legacy Kutu ID Asset",
            "seri_no": "LEG-KUTU-001",
            "stok": 2,
            "durum": "Aktif",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    created = InventoryAsset.query.filter_by(serial_no="LEG-KUTU-001").first()
    assert created is not None
    assert created.legacy_material is not None
    assert created.legacy_material.kutu_id == box.id


def test_envanter_renders_accordion_and_no_work_order_filter(client, app):
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="KCO", ad="Kocaeli Cengiz Topel")
    box = KutuFactory(kodu="KCO-SAR-01", havalimani=airport)
    material = MalzemeFactory(ad="Akordiyon Test", kutu=box, havalimani=airport, is_deleted=False)
    template = EquipmentTemplateFactory(name="Akordiyon Test")
    asset = InventoryAssetFactory(equipment_template=template, airport=airport, status="aktif")
    db.session.add_all([user, airport, box, material, template, asset])
    db.session.flush()
    asset.legacy_material_id = material.id
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(user.id)
        sess["_fresh"] = True

    response = client.get("/envanter")
    html = response.data.decode("utf-8")
    assert response.status_code == 200
    assert 'name="is_emri_durumu"' not in html
    assert "Sıra No" in html
    assert "MİKTAR" not in html
    assert 'data-accordion-target="inventory-row-' in html
    assert "openBakimModal(" not in html
    assert "🔍 Envanter Detayı" not in html
    assert "📱 Hızlı" not in html
    assert "Hızlı Zimmet" in html
    assert "QR Etiketi" in html
    assert "/qr-uret/asset/" in html
    assert 'class="inventory-page-actions"' in html
    assert 'id="inventoryFilterPanel"' in html
    assert 'id="inventoryFilterToggle"' in html
    assert "🛠️ Bakım" in html
    assert "🗑️ Sil" in html
    assert "kritik ekipman" not in html.lower()
    assert "is_critical" not in html
    assert "kritik_mi" not in html
    assert "function closeAll(exceptId)" in html
    assert "panel.setAttribute('aria-hidden'" in html
    assert "toggle.setAttribute('aria-expanded'" in html
    assert "const OFFLINE_DB_NAME = 'SAR_Offline_DB';" in html
    assert "const OFFLINE_STORE_NAME = 'bekleyen_bakimlar';" in html
    assert "const OFFLINE_SYNC_TAG = 'bakim-senkronize-et';" in html
    assert "X-SARX-Offline-Sync" in html
    assert "X-SARX-Offline-Request-Id" in html
    assert "Bağlantı yok" in html
    assert "tbody tr.inventory-summary-row:hover" in html
    assert "window.location.assign(qrHref);" in html
    assert "QR kullanılamıyor" in html
    assert "inventory-table-compact" in html
    assert 'col-no' in html and 'col-material' in html and 'col-airport' in html and 'col-code' in html and 'col-status' in html and 'col-maintenance' in html and 'col-action' in html

    response_with_legacy_work_order = client.get("/envanter?is_emri_durumu=acik")
    html_with_legacy_work_order = response_with_legacy_work_order.data.decode("utf-8")
    assert response_with_legacy_work_order.status_code == 200
    assert "Akordiyon Test" in html_with_legacy_work_order


def test_envanter_airport_abbreviation_has_title_tooltip(client, app):
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="KCO", ad="Kocaeli Cengiz Topel")
    box = KutuFactory(kodu="KCO-SAR-22", havalimani=airport)
    material = MalzemeFactory(ad="Tooltip Test", kutu=box, havalimani=airport, is_deleted=False)
    db.session.add_all([user, airport, box, material])
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(user.id)
        sess["_fresh"] = True

    response = client.get("/envanter")
    html = response.data.decode("utf-8")
    assert response.status_code == 200
    assert 'title="Kocaeli Cengiz Topel"' in html


def test_status_options_hide_hurda_in_create_and_detail(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    airport = HavalimaniFactory(kodu="ERZ")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport)
    box = KutuFactory(kodu="ERZ-SAR-41", havalimani=airport)
    template = EquipmentTemplateFactory(name="Durum Test")
    asset = InventoryAssetFactory(equipment_template=template, airport=airport, status="aktif")
    db.session.add_all([airport, manager, box, template, asset])
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(manager.id)
        sess["_fresh"] = True

    create_page = client.get("/malzeme-ekle")
    create_html = create_page.data.decode("utf-8")
    assert create_page.status_code == 200
    assert 'value="hurda"' not in create_html

    detail_page = client.get(f"/asset/{asset.id}/detay")
    detail_html = detail_page.data.decode("utf-8")
    assert detail_page.status_code == 200
    assert 'name="status"' in detail_html
    assert 'value="hurda"' not in detail_html


def test_envanter_status_labels_render_only_aktif_pasif(client, app):
    user = KullaniciFactory(rol="sistem_sorumlusu")
    airport = HavalimaniFactory(kodu="AYT", ad="Antalya Havalimanı")
    box = KutuFactory(kodu="AYT-SAR-11", havalimani=airport)
    material = MalzemeFactory(ad="Legacy Durum", kutu=box, havalimani=airport, durum="Arızalı", is_deleted=False)
    db.session.add_all([user, airport, box, material])
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(user.id)
        sess["_fresh"] = True

    response = client.get("/envanter")
    html = response.data.decode("utf-8")
    assert response.status_code == 200
    assert "Arızalı" not in html
    assert "Bakımda" not in html
    assert "Pasif" in html


def test_malzeme_ekle_page_renders_bulk_excel_panel_and_ordered_labels(client, app):
    airport = HavalimaniFactory(kodu="ESB")
    manager = KullaniciFactory(rol="ekip_sorumlusu", havalimani=airport)
    box = KutuFactory(kodu="ESB-KUTU-9", havalimani=airport)
    template = EquipmentTemplateFactory(name="Label Test", category="Elektronik")
    db.session.add_all([airport, manager, box, template])
    db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(manager.id)
        sess["_fresh"] = True

    response = client.get("/malzeme-ekle")
    html = response.data.decode("utf-8")
    assert response.status_code == 200
    assert "Toplu Malzeme Ekle (Excel)" in html
    assert "1) Kayıt Tipi" in html
    assert "2) Merkezi Şablon" in html
    assert "3) Kategori" in html
