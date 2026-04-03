import io

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from decorators import update_user_permission_overrides
from extensions import db
from models import Kullanici
from tests.factories import HavalimaniFactory, KullaniciFactory


def _login(client, user_id):
    with client.session_transaction() as session:
        session.clear()
        session["_user_id"] = str(user_id)
        session["_fresh"] = True


def _extract_select_markup(html, element_id):
    start = html.index(f'<select id="{element_id}"')
    end = html.index("</select>", start)
    return html[start:end]


def _extract_update_form_markup(html, user_id):
    marker = f'<form method="POST" action="/kullanici-guncelle/{user_id}"'
    start = html.index(marker)
    end = html.index("</form>", start)
    return html[start:end]


def test_selected_user_loads_current_role_and_permission_overrides(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-ux@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Bakım Uzmanı",
            kullanici_adi="maint-ux@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.flush()
        update_user_permission_overrides(staff.id, ["inventory.export"], ["logs.view"])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert f'data-selected-user-id="{staff_id}"' in html
    assert 'value="ekip_uyesi" selected' in html
    assert 'name="h_id"' in html
    assert 'name="allow_permissions" value="inventory.export" checked' in html
    assert 'name="deny_permissions" value="logs.view" checked' in html


def test_user_management_renders_success_toast_after_create(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-toast@sarx.com")
        db.session.add(owner)
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.post(
        "/kullanici-ekle",
        data={
            "tam_ad": "Toast Kullanıcısı",
            "k_adi": "toast-user@sarx.com",
            "sifre": "GucluTest@123",
            "rol": "admin",
            "h_id": "",
        },
        follow_redirects=True,
    )
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'class="toast-stack"' in html
    assert 'flash-msg flash-success' in html
    assert "TOAST KULLANICISI personeli sisteme eklendi." in html
    assert "Beklenmedik Bir Hata" not in html
    assert "Sistem bağlantı hatası oluştu." not in html


def test_user_management_renders_error_toast_for_invalid_selection(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-danger@sarx.com")
        db.session.add(owner)
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar?user_id=999999")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'flash-msg flash-danger' in html
    assert "Yetkiniz olmayan kayıt görüntülenemedi." in html


def test_site_owner_can_see_all_users_across_airports(client, app):
    with app.app_context():
        erzurum = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        trabzon = HavalimaniFactory(ad="Trabzon Havalimanı", kodu="TZX")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-all@sarx.com")
        local_user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Erzurum Personeli",
            kullanici_adi="erzurum@sarx.com",
            havalimani=erzurum,
        )
        remote_user = KullaniciFactory(
            rol="depo_sorumlusu",
            is_deleted=False,
            tam_ad="Trabzon Personeli",
            kullanici_adi="trabzon@sarx.com",
            havalimani=trabzon,
        )
        db.session.add_all([erzurum, trabzon, owner, local_user, remote_user])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Erzurum Personeli" in html
    assert "Trabzon Personeli" in html


def test_non_owner_user_management_scope_is_limited_to_same_airport(client, app):
    with app.app_context():
        erzurum = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        trabzon = HavalimaniFactory(ad="Trabzon Havalimanı", kodu="TZX")
        admin_user = KullaniciFactory(
            rol="ekip_sorumlusu",
            is_deleted=False,
            tam_ad="Erzurum Yonetici",
            kullanici_adi="admin-erzurum@sarx.com",
            havalimani=erzurum,
        )
        local_user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Ayni Havalimani Personeli",
            kullanici_adi="local@sarx.com",
            havalimani=erzurum,
        )
        remote_user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Diger Havalimani Personeli",
            kullanici_adi="remote@sarx.com",
            havalimani=trabzon,
        )
        db.session.add_all([erzurum, trabzon, admin_user, local_user, remote_user])
        db.session.commit()
        admin_user_id = admin_user.id
        remote_user_id = remote_user.id

    _login(client, admin_user_id)
    response = client.get(f"/kullanicilar?user_id={remote_user_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Ayni Havalimani Personeli" in html
    assert "Diger Havalimani Personeli" not in html
    assert "Yetkiniz olmayan kayıt görüntülenemedi." in html


def test_role_filter_limits_users_and_displays_selected_role(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-filter@sarx.com")
        maintenance_user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Bakim Ekibi",
            kullanici_adi="bakim@sarx.com",
            havalimani=airport,
        )
        warehouse_user = KullaniciFactory(
            rol="yetkili",
            is_deleted=False,
            tam_ad="Depo Ekibi",
            kullanici_adi="depo@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, maintenance_user, warehouse_user])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar?role=ekip_uyesi")
    html = response.data.decode("utf-8")
    quick_select_html = _extract_select_markup(html, "userQuickSelect")

    assert response.status_code == 200
    assert "Bakim Ekibi" in html
    assert "Depo Ekibi" not in html
    assert "Rol: Ekip Üyesi" in html
    assert "Bakim Ekibi" in quick_select_html
    assert "Depo Ekibi" not in quick_select_html


def test_airport_filter_limits_users_and_displays_selected_airport(client, app):
    with app.app_context():
        erzurum = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        trabzon = HavalimaniFactory(ad="Trabzon Havalimanı", kodu="TZX")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-airport@sarx.com")
        erzurum_user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Erzurum Teknisyeni",
            kullanici_adi="erzurum-filter@sarx.com",
            havalimani=erzurum,
        )
        trabzon_user = KullaniciFactory(
            rol="depo_sorumlusu",
            is_deleted=False,
            tam_ad="Trabzon Depo",
            kullanici_adi="trabzon-filter@sarx.com",
            havalimani=trabzon,
        )
        db.session.add_all([erzurum, trabzon, owner, erzurum_user, trabzon_user])
        db.session.commit()
        owner_id = owner.id
        erzurum_id = erzurum.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?airport_id={erzurum_id}")
    html = response.data.decode("utf-8")
    quick_select_html = _extract_select_markup(html, "userQuickSelect")

    assert response.status_code == 200
    assert "Erzurum Teknisyeni" in html
    assert "Trabzon Depo" not in html
    assert "Havalimanı: Erzurum Havalimanı" in html
    assert "Erzurum Teknisyeni" in quick_select_html
    assert "Trabzon Depo" not in quick_select_html


def test_status_filter_returns_archived_users(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-status@sarx.com")
        active_user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Aktif Kullanıcı",
            kullanici_adi="aktif@sarx.com",
            havalimani=airport,
        )
        archived_user = KullaniciFactory(
            rol="depo_sorumlusu",
            is_deleted=True,
            tam_ad="Arşiv Kullanıcısı",
            kullanici_adi="arsiv@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, active_user, archived_user])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar?status=archived")
    html = response.data.decode("utf-8")
    quick_select_html = _extract_select_markup(html, "userQuickSelect")

    assert response.status_code == 200
    assert "Arşiv Kullanıcısı" in html
    assert "Aktif Kullanıcı" not in html
    assert 'value="archived" selected' in html
    assert "Arşiv Kullanıcısı" in quick_select_html
    assert "Aktif Kullanıcı" not in quick_select_html


def test_combined_filters_reduce_user_list_and_update_result_summary(client, app):
    with app.app_context():
        erzurum = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        trabzon = HavalimaniFactory(ad="Trabzon Havalimanı", kodu="TZX")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-combined@sarx.com")
        matching_user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Usta Teknisyen",
            kullanici_adi="usta@sarx.com",
            havalimani=erzurum,
        )
        wrong_airport = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Usta Trabzon",
            kullanici_adi="usta-trabzon@sarx.com",
            havalimani=trabzon,
        )
        wrong_role = KullaniciFactory(
            rol="yetkili",
            is_deleted=False,
            tam_ad="Usta Depo",
            kullanici_adi="usta-depo@sarx.com",
            havalimani=erzurum,
        )
        archived_match = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=True,
            tam_ad="Usta Arşiv",
            kullanici_adi="usta-arsiv@sarx.com",
            havalimani=erzurum,
        )
        db.session.add_all([erzurum, trabzon, owner, matching_user, wrong_airport, wrong_role, archived_match])
        db.session.commit()
        owner_id = owner.id
        erzurum_id = erzurum.id

    _login(client, owner_id)
    response = client.get(
        f"/kullanicilar?q=usta&role=ekip_uyesi&airport_id={erzurum_id}&status=active"
    )
    html = response.data.decode("utf-8")
    quick_select_html = _extract_select_markup(html, "userQuickSelect")

    assert response.status_code == 200
    assert "Usta Teknisyen" in html
    assert "Usta Trabzon" not in html
    assert "Usta Depo" not in html
    assert "Usta Arşiv" not in html
    assert 'data-result-count="1"' in html
    assert "1 kayıt filtreyle listeleniyor" in html
    assert "Arama: usta" in html
    assert "Rol: Ekip Üyesi" in html
    assert "Havalimanı: Erzurum Havalimanı" in html
    assert "Usta Teknisyen" in quick_select_html
    assert "Usta Trabzon" not in quick_select_html


def test_empty_state_renders_when_filters_return_no_users(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-empty@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Mevcut Personel",
            kullanici_adi="mevcut@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar?q=olmayan-sonuc")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'id="userServerEmpty"' in html
    assert "Seçili filtrelerle kullanıcı bulunamadı" in html
    assert 'id="userQuickSelect" class="form-control" aria-label="Kullanıcı seç" disabled' in html
    assert 'data-result-count="0"' in html


def test_user_search_matches_turkish_character_variants(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="İzmir Çiğli Havalimanı", kodu="IGL")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-tr-search@sarx.com")
        user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Şule Çağrı Öztürk",
            kullanici_adi="sule.cagri@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, user])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar?q=sule cagri ozturk izmir cigli")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Şule Çağrı Öztürk" in html
    assert "İzmir Çiğli Havalimanı" in html


def test_detail_panel_is_rendered_after_filter_and_selection_blocks(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-layout@sarx.com")
        db.session.add(owner)
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Filtrele ve kullanıcıyı bul" in html
    assert "Kullanıcı seç" in html
    assert "Detay ve yetki düzenleme" in html
    assert "Toplu yetki düzenleme" in html
    assert html.index('class="panel filter-panel stage-accordion"') < html.index('class="panel user-directory-panel stage-accordion"')
    assert html.index('class="panel user-directory-panel stage-accordion"') < html.index('id="userDetailPanel"')
    assert html.index('id="userDetailPanel"') < html.index('id="newUserPanel"')
    assert 'data-filter-summary' in html
    assert 'data-detail-empty-state' in html
    assert 'data-bulk-role-panel' in html
    assert html.count('data-stage-accordion') >= 3
    stage_select_start = html.index('data-admin-stage="select"')
    stage_select_markup = html[stage_select_start:html.index('>', stage_select_start)]
    assert " open" not in stage_select_markup
    assert 'id="userDirectoryShell"' in html


def test_bulk_role_panel_lists_only_selected_airport_staff(client, app):
    with app.app_context():
        erzurum = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        trabzon = HavalimaniFactory(ad="Trabzon Havalimanı", kodu="TZX")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-bulk-panel@sarx.com")
        local_staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Erzurum Toplu Personel",
            kullanici_adi="bulk-erzurum@sarx.com",
            havalimani=erzurum,
        )
        remote_staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Trabzon Toplu Personel",
            kullanici_adi="bulk-trabzon@sarx.com",
            havalimani=trabzon,
        )
        db.session.add_all([erzurum, trabzon, owner, local_staff, remote_staff])
        db.session.commit()
        owner_id = owner.id
        erzurum_id = erzurum.id
        local_staff_id = local_staff.id
        remote_staff_id = remote_staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?bulk_airport_id={erzurum_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Toplu yetki düzenleme" in html
    assert "Seçilen havalimanındaki personeller için toplu rol/kapsam düzenleme alanı." in html
    assert 'id="userDetailPanel" data-stage-accordion open' in html
    assert f'data-bulk-staff-id="{local_staff_id}"' in html
    assert f'data-bulk-staff-id="{remote_staff_id}"' not in html
    assert 'name="bulk_roles"' in html
    assert 'id="bulkRoleSelect"' in html


def test_bulk_role_apply_updates_selected_airport_users(client, app):
    with app.app_context():
        erzurum = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        trabzon = HavalimaniFactory(ad="Trabzon Havalimanı", kodu="TZX")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-bulk-apply@sarx.com")
        staff_one = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Toplu Bir",
            kullanici_adi="bulk-one@sarx.com",
            havalimani=erzurum,
        )
        staff_two = KullaniciFactory(
            rol="depo_sorumlusu",
            is_deleted=False,
            tam_ad="Toplu Iki",
            kullanici_adi="bulk-two@sarx.com",
            havalimani=erzurum,
        )
        remote_staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Uzak Personel",
            kullanici_adi="bulk-remote@sarx.com",
            havalimani=trabzon,
        )
        db.session.add_all([erzurum, trabzon, owner, staff_one, staff_two, remote_staff])
        db.session.commit()
        owner_id = owner.id
        erzurum_id = erzurum.id
        trabzon_id = trabzon.id
        staff_one_id = staff_one.id
        staff_two_id = staff_two.id
        remote_staff_id = remote_staff.id

    _login(client, owner_id)
    response = client.post(
        "/kullanicilar/toplu-yetki-guncelle",
        data={
            "bulk_airport_id": str(erzurum_id),
            "bulk_user_ids": [str(staff_one_id), str(staff_two_id), str(remote_staff_id)],
            "bulk_roles": ["ekip_sorumlusu", "ekip_uyesi"],
        },
        follow_redirects=True,
    )
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Birden fazla rol seçildiği için ilk rol uygulandı" in html
    assert "2 kullanıcı için toplu rol/kapsam güncellendi." in html
    assert "1 kullanıcı kapsam/yetki kontrolü nedeniyle atlandı." in html

    with app.app_context():
        refreshed_one = db.session.get(Kullanici, staff_one_id)
        refreshed_two = db.session.get(Kullanici, staff_two_id)
        refreshed_remote = db.session.get(Kullanici, remote_staff_id)
        assert refreshed_one.rol == "ekip_sorumlusu"
        assert refreshed_two.rol == "ekip_sorumlusu"
        assert refreshed_one.havalimani_id == erzurum_id
        assert refreshed_two.havalimani_id == erzurum_id
        assert refreshed_remote.rol == "bakim_sorumlusu"
        assert refreshed_remote.havalimani_id == trabzon_id


def test_non_owner_detail_panel_copy_hides_permission_wording(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        actor = KullaniciFactory(
            rol="ekip_sorumlusu",
            is_deleted=False,
            tam_ad="Erzurum Yonetici",
            kullanici_adi="lead-stage3@sarx.com",
            havalimani=airport,
        )
        selected_user = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Secilen Kullanici",
            kullanici_adi="selected-stage3@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, actor, selected_user])
        db.session.commit()
        actor_id = actor.id
        selected_user_id = selected_user.id

    _login(client, actor_id)
    response = client.get(f"/kullanicilar?user_id={selected_user_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Kullanıcı güncelle" in html
    assert "Seçilen kullanıcının temel bilgilerini, kapsamını bu alandan yönetin." in html
    assert "Detay ve yetki düzenleme" not in html
    assert "Seçilen kullanıcının temel bilgilerini, rolünü, kapsamını ve gerekiyorsa özel yetkilerini bu alandan yönetin." not in html


def test_user_selector_hides_email_and_shows_name_airport_and_role(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-selector@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Bakım Uzmanı",
            kullanici_adi="maint-selector@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar")
    html = response.data.decode("utf-8")
    quick_select_html = _extract_select_markup(html, "userQuickSelect")

    assert response.status_code == 200
    assert "Bakım Uzmanı • Erzurum Havalimanı • Ekip Üyesi" in quick_select_html
    assert "maint-selector@sarx.com" not in quick_select_html


def test_user_cards_render_core_identity_fields_cleanly(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-card@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Kart Kullanıcısı",
            kullanici_adi="card-user@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get("/kullanicilar")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert f'data-user-card="{staff_id}"' in html
    assert "Kart Kullanıcısı" in html
    assert "card-user@sarx.com" in html
    assert "Havalimanı" in html
    assert "Rol" in html
    assert "Ekip Üyesi" in html


def test_selected_user_keeps_selection_panel_open_and_detail_panel_ready(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Sivas Havalimanı", kodu="VAS")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-select@sarx.com")
        staff = KullaniciFactory(
            rol="ekip_uyesi",
            is_deleted=False,
            tam_ad="Seçili Personel",
            kullanici_adi="selected-user@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    stage_select_start = html.index('data-admin-stage="select"')
    stage_select_markup = html[stage_select_start:html.index('>', stage_select_start)]
    assert " open" in stage_select_markup
    assert "Detay paneli" in html
    assert "için hazır" in html
    assert 'id="userDetailPanel" data-stage-accordion open' in html


def test_user_management_renders_login_email_label_in_create_and_edit_forms(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-label@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Etiket Kullanıcısı",
            kullanici_adi="etiket@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert html.count("Giriş E-postası") >= 2
    assert "Bu alan kullanıcının sisteme girişte kullandığı e-posta adresidir." in html
    assert "Kullanıcı Adı" not in html
    assert html.count('type="email"') >= 2


def test_override_summary_is_split_into_allowed_and_withdrawn_sections(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-override@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Override Kullanıcısı",
            kullanici_adi="override@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.flush()
        update_user_permission_overrides(staff.id, ["inventory.export"], ["logs.view"])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'data-override-summary="allow"' in html
    assert 'data-override-summary="deny"' in html
    assert "Bu kullanıcıya ek olarak verilenler" in html
    assert "Bu kullanıcıdan özellikle kaldırılanlar" in html
    assert "İzin Verilenler" in html
    assert "Geri Çekilenler" in html


def test_override_summary_empty_states_are_rendered_cleanly(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-emptyoverride@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Boş Override Kullanıcısı",
            kullanici_adi="empty-override@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Ek özel izin yok." in html
    assert "Geri çekilen izin yok." in html


def test_detail_form_is_grouped_into_clear_sections(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-sections@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Bölüm Kullanıcısı",
            kullanici_adi="section-user@sarx.com",
            havalimani=airport,
            telefon_numarasi="+905551112233",
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Temel bilgiler" in html
    assert "Rol ve havalimanı" in html
    assert "Telefon / iletişim" in html
    assert "Override özeti" in html
    assert "Yetki blokları" in html


def test_site_owner_can_view_phone_number_field(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-phone@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Telefonlu Personel",
            kullanici_adi="phone-user@sarx.com",
            havalimani=airport,
            telefon_numarasi="+905551112233",
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    owner_response = client.get(f"/kullanicilar?user_id={staff_id}")
    owner_html = owner_response.data.decode("utf-8")

    assert owner_response.status_code == 200
    assert 'name="telefon_numarasi"' in owner_html
    assert "+905551112233" in owner_html


def test_non_owner_cannot_view_phone_number_field(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        admin_user = KullaniciFactory(
            rol="admin",
            is_deleted=False,
            kullanici_adi="admin-phone@sarx.com",
            havalimani=airport,
        )
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Telefonlu Personel",
            kullanici_adi="phone-user@sarx.com",
            havalimani=airport,
            telefon_numarasi="+905551112233",
        )
        db.session.add_all([airport, admin_user, staff])
        db.session.commit()
        admin_user_id = admin_user.id
        staff_id = staff.id

    _login(client, admin_user_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'name="telefon_numarasi"' not in html
    assert "+905551112233" not in html


def test_phone_number_is_saved_and_success_toasts_are_rendered(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-save@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Kayit Personeli",
            kullanici_adi="save-user@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        airport_id = airport.id
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.post(
        f"/kullanici-guncelle/{staff_id}",
        data={
            "tam_ad": "Kayit Personeli",
            "k_adi": "save-user@sarx.com",
            "rol": "ekip_uyesi",
            "h_id": str(airport_id),
            "telefon_numarasi": "+90 555 111 22 33",
        },
        follow_redirects=True,
    )
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Kullanıcı yetkileri güncellendi." in html
    assert "Telefon numarası kaydedildi." in html
    assert 'flash-msg flash-success' in html


def test_new_user_panel_renders_as_collapsed_details_panel(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-newpanel@sarx.com")
        db.session.add(owner)
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert '<details class="panel new-user-panel" id="newUserPanel">' in html
    assert 'id="newUserSummaryLabel">Formu aç<' in html
    assert 'data-password-guidance' in html
    assert 'data-password-rule="special"' in html
    assert 'data-close-new-user-panel' in html


def test_user_management_inputs_render_validation_hooks(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-hooks@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Doğrulama Kullanıcısı",
            kullanici_adi="hooks-user@sarx.com",
            havalimani=airport,
            telefon_numarasi="+905551112233",
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'data-uppercase-name' in html
    assert 'data-email-input' in html
    assert 'data-phone-input' in html
    assert 'inputmode="numeric"' in html
    assert 'autocomplete="tel"' in html
    assert '+90 5__ ___ __ __' in html


def test_user_management_renders_blood_and_measurement_fields(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-profile@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Profil Kullanıcısı",
            kullanici_adi="profile-user@sarx.com",
            havalimani=airport,
            kan_grubu_harf="A",
            kan_grubu_rh="+",
            boy_cm=181,
            kilo_kg=79,
            ayak_numarasi=42.5,
            ust_beden="L",
            alt_beden="M",
            beden="L",
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert 'name="kan_grubu_harf"' in html
    assert 'name="kan_grubu_rh"' in html
    assert 'name="boy_cm"' in html
    assert 'name="kilo_kg"' in html
    assert 'name="ayak_numarasi"' in html
    assert 'name="ust_beden"' in html
    assert 'name="alt_beden"' in html
    assert '<option value="+" selected>+</option>' in html
    assert "Kan Grubu:" in html
    assert "A Rh+" in html
    assert "Acil Durum" not in html
    assert "Ekip İçi Rol" not in html


def test_user_measurement_fields_are_saved_with_profile_update(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Kars Havalimanı", kodu="KSY")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-measure@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Olcu Personeli",
            kullanici_adi="olcu-user@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id
        airport_id = airport.id

    _login(client, owner_id)
    response = client.post(
        f"/kullanici-guncelle/{staff_id}",
        data={
            "tam_ad": "Olcu Personeli",
            "k_adi": "olcu-user@sarx.com",
            "rol": "ekip_uyesi",
            "h_id": str(airport_id),
            "kan_grubu_harf": "AB",
            "kan_grubu_rh": "-",
            "boy_cm": "182",
            "kilo_kg": "84",
            "ayak_numarasi": "42.5",
            "ust_beden": "XL",
            "alt_beden": "L",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Kullanıcı yetkileri güncellendi." in response.data.decode("utf-8")

    with app.app_context():
        updated = db.session.get(Kullanici, staff_id)
        assert updated.kan_grubu_harf == "AB"
        assert updated.kan_grubu_rh == "-"
        assert updated.kan_grubu == "AB Rh-"
        assert updated.boy_cm == 182
        assert updated.kilo_kg == 84
        assert updated.ayak_numarasi == 42.5
        assert updated.beden == "XL"
        assert updated.ust_beden == "XL"
        assert updated.alt_beden == "L"


def test_user_permission_override_helper_does_not_clear_pending_profile_changes(app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Kayseri Havalimanı", kodu="ASR")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Yetki Yardimcisi",
            kullanici_adi="helper-user@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, staff])
        db.session.commit()
        staff_id = staff.id

        target = db.session.get(Kullanici, staff_id)
        target.rol = "ekip_uyesi"
        target.kan_grubu_harf = "AB"
        target.kan_grubu_rh = "-"

        update_user_permission_overrides(target.id, [], [])
        db.session.commit()
        db.session.expire_all()

        updated = db.session.get(Kullanici, staff_id)
        assert updated.rol == "ekip_uyesi"
        assert updated.kan_grubu_harf == "AB"
        assert updated.kan_grubu_rh == "-"


def test_legacy_rh_values_render_with_new_plus_minus_selector(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Sinop Havalimanı", kodu="NOP")
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-rh-legacy@sarx.com")
        staff = KullaniciFactory(
            rol="bakim_sorumlusu",
            is_deleted=False,
            tam_ad="Legacy Rh Kullanıcısı",
            kullanici_adi="legacy-rh@sarx.com",
            havalimani=airport,
            kan_grubu_harf="0",
            kan_grubu_rh="Rh-",
        )
        db.session.add_all([airport, owner, staff])
        db.session.commit()
        owner_id = owner.id
        staff_id = staff.id

    _login(client, owner_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert '<option value="-" selected>-</option>' in html
    assert "0 Rh-" in html


def test_non_owner_cannot_change_user_role_scope_or_permission_matrix(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Van Havalimanı", kodu="VAN")
        lead = KullaniciFactory(
            rol="ekip_sorumlusu",
            is_deleted=False,
            kullanici_adi="lead-permission@sarx.com",
            havalimani=airport,
        )
        staff = KullaniciFactory(
            rol="ekip_uyesi",
            is_deleted=False,
            tam_ad="Yetki Korunan",
            kullanici_adi="guard-user@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, lead, staff])
        db.session.commit()
        lead_id = lead.id
        staff_id = staff.id
        airport_id = airport.id

    _login(client, lead_id)
    response = client.post(
        f"/kullanici-guncelle/{staff_id}",
        data={
            "tam_ad": "Yetki Korunan",
            "k_adi": "guard-user@sarx.com",
            "rol": "ekip_sorumlusu",
            "h_id": str(airport_id),
            "allow_permissions": ["inventory.export"],
        },
        follow_redirects=False,
    )

    assert response.status_code == 403

    with app.app_context():
        updated = db.session.get(Kullanici, staff_id)
        assert updated.rol == "ekip_uyesi"


def test_non_system_user_detail_panel_hides_role_scope_and_override_edit_controls(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Van Havalimanı", kodu="VAN")
        lead = KullaniciFactory(
            rol="ekip_sorumlusu",
            is_deleted=False,
            kullanici_adi="lead-readonly@sarx.com",
            havalimani=airport,
        )
        staff = KullaniciFactory(
            rol="ekip_uyesi",
            is_deleted=False,
            tam_ad="Readonly Personel",
            kullanici_adi="readonly-user@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, lead, staff])
        db.session.flush()
        update_user_permission_overrides(staff.id, ["inventory.export"], ["logs.view"])
        db.session.commit()
        lead_id = lead.id
        staff_id = staff.id

    _login(client, lead_id)
    response = client.get(f"/kullanicilar?user_id={staff_id}")
    html = response.data.decode("utf-8")
    form_html = _extract_update_form_markup(html, staff_id)

    assert response.status_code == 200
    assert 'data-selected-role=' not in form_html
    assert 'type="checkbox" name="allow_permissions"' not in form_html
    assert 'type="checkbox" name="deny_permissions"' not in form_html
    assert "yalnızca Sistem Sorumlusu tarafından düzenlenebilir" in form_html
    assert 'type="hidden" name="rol" value="ekip_uyesi"' in form_html


def test_team_lead_create_form_is_locked_to_team_member_and_own_airport(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Rize Havalimanı", kodu="RZV")
        lead = KullaniciFactory(
            rol="ekip_sorumlusu",
            is_deleted=False,
            kullanici_adi="lead-create@sarx.com",
            havalimani=airport,
        )
        db.session.add_all([airport, lead])
        db.session.commit()
        lead_id = lead.id

    _login(client, lead_id)
    response = client.get("/kullanicilar")
    html = response.data.decode("utf-8")
    create_role_select = html[html.index('data-create-role-select'):html.index("</select>", html.index('data-create-role-select'))]
    create_airport_select = html[html.index('data-create-airport-select'):html.index("</select>", html.index('data-create-airport-select'))]

    assert response.status_code == 200
    assert 'value="ekip_uyesi" selected' in create_role_select
    assert "Sistem Sorumlusu" not in create_role_select
    assert "Ekip Sorumlusu" not in create_role_select
    assert "Global" not in create_airport_select
    assert "Rize Havalimanı" in create_airport_select


def test_team_lead_can_create_only_team_member_in_own_airport_without_403(client, app):
    with app.app_context():
        own_airport = HavalimaniFactory(ad="Antalya Havalimanı", kodu="AYT")
        other_airport = HavalimaniFactory(ad="Isparta Havalimanı", kodu="ISE")
        lead = KullaniciFactory(
            rol="ekip_sorumlusu",
            is_deleted=False,
            kullanici_adi="lead-own-airport@sarx.com",
            havalimani=own_airport,
        )
        db.session.add_all([own_airport, other_airport, lead])
        db.session.commit()
        lead_id = lead.id
        own_airport_id = own_airport.id
        other_airport_id = other_airport.id

    _login(client, lead_id)
    response = client.post(
        "/kullanici-ekle",
        data={
            "tam_ad": "Yeni Takim Arkadasi",
            "k_adi": "new-team-member@sarx.com",
            "sifre": "GucluTest@123",
            "rol": "sistem_sorumlusu",
            "h_id": str(other_airport_id),
        },
        follow_redirects=True,
    )
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "personeli sisteme eklendi" in html

    with app.app_context():
        created = Kullanici.query.filter_by(kullanici_adi="new-team-member@sarx.com").first()
        assert created is not None
        assert created.rol == "ekip_uyesi"
        assert created.havalimani_id == own_airport_id


def test_role_switched_team_lead_can_open_user_management_and_create_scoped_user(client, app):
    with app.app_context():
        app.config["ROLE_SWITCH_ALLOWED_USERS"] = "mehmetcinocevi@gmail.com"
        airport = HavalimaniFactory(ad="Bodrum Havalimanı", kodu="BJV")
        owner = KullaniciFactory(
            rol="sahip",
            is_deleted=False,
            kullanici_adi="mehmetcinocevi@gmail.com",
            havalimani=airport,
        )
        db.session.add_all([airport, owner])
        db.session.commit()
        owner_id = owner.id
        airport_id = airport.id

    _login(client, owner_id)
    with client.session_transaction() as session:
        session["temporary_role_override"] = "ekip_sorumlusu"

    page = client.get("/kullanicilar")
    page_html = page.data.decode("utf-8")
    assert page.status_code == 200
    assert 'value="ekip_uyesi" selected' in page_html

    response = client.post(
        "/kullanici-ekle",
        data={
            "tam_ad": "Rol Gecis Kullanıcısı",
            "k_adi": "switched-team@sarx.com",
            "sifre": "GucluTest@123",
            "rol": "admin",
            "h_id": "",
            "telefon_numarasi": "+90 555 111 22 33",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    with app.app_context():
        created = Kullanici.query.filter_by(kullanici_adi="switched-team@sarx.com").first()
        assert created is not None
        assert created.rol == "ekip_uyesi"
        assert created.havalimani_id == airport_id
        assert created.telefon_numarasi is None


def test_invalid_email_is_rejected_in_user_create_form(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-invalidmail@sarx.com")
        db.session.add(owner)
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.post(
        "/kullanici-ekle",
        data={
            "tam_ad": "Hatali Mail",
            "k_adi": "hatali-mail",
            "sifre": "GucluTest@123",
            "rol": "admin",
            "h_id": "",
        },
        follow_redirects=True,
    )
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Geçerli bir e-posta adresi girin." in html


def test_common_email_is_trimmed_normalized_and_accepted_in_user_create_form(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-validmail@sarx.com")
        db.session.add(owner)
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.post(
        "/kullanici-ekle",
        data={
            "tam_ad": "Emre Baykan",
            "k_adi": "  Emre.Baykan54@Gmail.com  ",
            "sifre": "GucluTest@123",
            "rol": "admin",
            "h_id": "",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "personeli sisteme eklendi" in response.data.decode("utf-8")

    with app.app_context():
        created = Kullanici.query.filter_by(kullanici_adi="emre.baykan54@gmail.com").first()
        assert created is not None
        assert created.tam_ad == "EMRE BAYKAN"


def test_user_import_template_includes_profile_columns_and_dropdown_validations(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-template@sarx.com")
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        db.session.add_all([owner, airport])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/kullanicilar/template.xlsx")
    assert response.status_code == 200

    workbook = load_workbook(io.BytesIO(response.data))
    sheet = workbook["Kullanicilar"]
    headers = [cell.value for cell in sheet[1]]

    assert "kan_grubu_harf" in headers
    assert "kan_grubu_rh" in headers
    assert "boy_cm" in headers
    assert "kilo_kg" in headers
    assert "ayak_numarasi" in headers
    assert "ust_beden" in headers
    assert "alt_beden" in headers
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == f"A1:{get_column_letter(len(headers))}1"

    formulas = sorted({validation.formula1 for validation in sheet.data_validations.dataValidation if validation.formula1})
    assert "'Listeler'!$A$2:$A$5000" in formulas  # rol
    assert "'Listeler'!$B$2:$B$3" in formulas  # aktif/pasif
    assert "'Listeler'!$E$2:$E$5000" in formulas  # ayak_numarasi
    assert "'Listeler'!$F$2:$F$5000" in formulas  # ust_beden
    assert "'Listeler'!$G$2:$G$5000" in formulas  # alt_beden


def test_bulk_user_import_applies_optional_profile_fields(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-bulk-profile@sarx.com")
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        db.session.add_all([owner, airport])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    template_response = client.get("/kullanicilar/template.xlsx")
    workbook = load_workbook(io.BytesIO(template_response.data))
    sheet = workbook["Kullanicilar"]
    headers = [cell.value for cell in sheet[1]]
    row_payload = {
        "ad": "Selin",
        "soyad": "Yılmaz",
        "e-posta": "selin.bulk@sarx.com",
        "telefon": "+90 555 111 22 33",
        "rol": "ekip_uyesi",
        "havalimani": "ERZ",
        "aktif/pasif": "aktif",
        "not": "toplu içe aktarma",
        "gecici_sifre": "Gecici@123",
        "kan_grubu_harf": "A",
        "kan_grubu_rh": "+",
        "boy_cm": "171",
        "kilo_kg": "63",
        "ayak_numarasi": "39",
        "ust_beden": "M",
        "alt_beden": "S",
    }
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=column_index, value=row_payload.get(header, ""))
    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)

    preview_response = client.post(
        "/kullanicilar/import/preview",
        data={"import_file": (payload, "kullanici_import_sablonu.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert preview_response.status_code == 200
    assert "1 satır önizleme için hazırlandı." in preview_response.data.decode("utf-8")

    commit_response = client.post("/kullanicilar/import/commit", follow_redirects=True)
    assert commit_response.status_code == 200
    assert "1 kullanıcı içe aktarıldı." in commit_response.data.decode("utf-8")

    with app.app_context():
        created = Kullanici.query.filter_by(kullanici_adi="selin.bulk@sarx.com").first()
        assert created is not None
        assert created.kan_grubu_harf == "A"
        assert created.kan_grubu_rh == "+"
        assert created.boy_cm == 171
        assert created.kilo_kg == 63
        assert created.ayak_numarasi == 39.0
        assert created.ust_beden == "M"
        assert created.alt_beden == "S"


def test_roles_page_renders_row_action_alignment_fix(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", is_deleted=False, kullanici_adi="owner-rolecss@sarx.com")
        db.session.add(owner)
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/admin/roles")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert ".row-actions:not(td)" in html
    assert "td.row-actions, .data-table td.row-actions" in html
