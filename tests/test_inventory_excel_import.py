import io

from openpyxl import load_workbook

from extensions import db
from models import EquipmentTemplate, InventoryAsset, InventoryBulkImportJob, InventoryCategory
from services.inventory_bulk_import_service import parse_flexible_bool
from services.inventory_excel_service import build_inventory_template_workbook
from services.text_normalization_service import normalize_lookup_key, turkish_upper
from tests.factories import (
    EquipmentTemplateFactory,
    HavalimaniFactory,
    KutuFactory,
    KullaniciFactory,
    MaintenanceFormTemplateFactory,
)


def _login(client, user):
    with client.session_transaction() as sess:
        sess.clear()
        sess["_user_id"] = str(user.id)
        sess["_fresh"] = True


def _build_excel_file(row_dicts, *, lists_context):
    template = build_inventory_template_workbook(lists_context=lists_context)
    wb = load_workbook(template)
    ws = wb["VERI_GIRISI"]
    headers = [cell.value for cell in ws[1]]
    for row in row_dicts:
        ws.append([row.get(header) for header in headers])
    payload = io.BytesIO()
    wb.save(payload)
    payload.seek(0)
    return payload


def test_excel_template_download_has_expected_sheets(client, app):
    owner = KullaniciFactory(rol="sahip")
    airport = HavalimaniFactory(kodu="ESB")
    box = KutuFactory(kodu="ESB-KUTU-1", havalimani=airport)
    db.session.add_all([owner, airport, box])
    db.session.commit()
    _login(client, owner)

    response = client.get("/malzeme-ekle/excel-sablon")
    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.data))
    assert {"VERI_GIRISI", "LISTELER", "ACIKLAMA"} <= set(wb.sheetnames)


def test_excel_template_contains_dropdown_validations_for_core_lists(client, app):
    owner = KullaniciFactory(rol="sahip")
    airport = HavalimaniFactory(kodu="ESB")
    box = KutuFactory(kodu="ESB-KUTU-VAL", havalimani=airport)
    template = EquipmentTemplateFactory(name="Val Template", category="Elektronik")
    db.session.add_all([owner, airport, box, template])
    db.session.commit()
    _login(client, owner)

    response = client.get("/malzeme-ekle/excel-sablon")
    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.data))
    ws = wb["VERI_GIRISI"]
    formulas = sorted({dv.formula1 for dv in ws.data_validations.dataValidation if dv.formula1})
    assert "'LISTELER'!$A$2:$A$5000" in formulas  # merkezi_sablon
    assert "'LISTELER'!$B$2:$B$5000" in formulas  # havalimani
    assert "'LISTELER'!$C$2:$C$5000" in formulas  # kategori
    assert "'LISTELER'!$D$2:$D$5000" in formulas  # kullanim_durumu
    assert "'LISTELER'!$E$2:$E$5000" in formulas  # bakim_formu
    assert "'LISTELER'!$F$2:$F$5000" in formulas  # bakim_periyodu
    assert "'LISTELER'!$G$2:$G$5000" in formulas  # kutu_kodu


def test_excel_import_success_with_turkish_chars_and_bool_normalization(client, app):
    owner = KullaniciFactory(rol="sahip")
    airport = HavalimaniFactory(kodu="ESB", ad="Esenboğa")
    box = KutuFactory(kodu="ESB-KUTU-2", havalimani=airport)
    template = EquipmentTemplateFactory(name="Gaz Ölçer", category="Elektronik")
    form = MaintenanceFormTemplateFactory(name="Standart Form")
    db.session.add_all([owner, airport, box, template, form])
    db.session.commit()
    _login(client, owner)

    payload = _build_excel_file(
        [
            {
                "kayıt_tipi": "tekil",
                "merkezi_sablon": "Gaz Ölçer",
                "merkezi_sablondan_olustur": "Evet",
                "havalimani": "ESB - Esenboğa",
                "kategori": "Elektronik",
                "malzeme_adi": "İstasyon Ölçer",
                "marka": "Ölçüm AŞ",
                "model": "X-1",
                "demirbas_mi": "Evet",
                "demirbas_no": "DMR-1",
                "seri_no": "TR-EXCEL-001",
                "stok_birim_sayisi": "2",
                "kullanim_durumu": "aktif",
                "kalibrasyon_gerekli_mi": "1",
                "kalibrasyon_periyodu_ay": "6",
                "bakim_formu": "Standart Form",
                "bakim_periyodu": "4",
                "kutu_kodu": "ESB-KUTU-2",
                "aciklama_notlar": "ilk yükleme",
                "ad_soyad": "ibrahim çelik",
            }
        ],
        lists_context={
            "templates": ["Gaz Ölçer"],
            "airports": ["ESB - Esenboğa"],
            "categories": ["Elektronik"],
            "statuses": ["aktif", "pasif"],
            "maintenance_forms": ["Standart Form"],
            "month_values": list(range(1, 13)),
            "boxes": ["ESB-KUTU-2"],
        },
    )

    response = client.post(
        "/malzeme-ekle/excel-yukle",
        data={"excel_file": (payload, "import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200
    asset = InventoryAsset.query.filter_by(serial_no="TR-EXCEL-001").first()
    assert asset is not None
    assert asset.is_demirbas is True
    assert asset.calibration_required is True
    assert asset.asset_code is not None
    assert (asset.qr_code or "").startswith("http")
    assert "İBRAHİM ÇELİK" in (asset.notes or "")


def test_excel_import_bool_variations_in_real_flow(client, app):
    owner = KullaniciFactory(rol="sahip")
    airport = HavalimaniFactory(kodu="DLM", ad="Dalaman")
    box = KutuFactory(kodu="DLM-KUTU-1", havalimani=airport)
    template = EquipmentTemplateFactory(name="Bool Variant Template", category="Elektronik")
    db.session.add_all([owner, airport, box, template])
    db.session.commit()
    _login(client, owner)

    payload = _build_excel_file(
        [
            {
                "merkezi_sablon": "Bool Variant Template",
                "merkezi_sablondan_olustur": "true",
                "havalimani": "DLM - Dalaman",
                "kategori": "Elektronik",
                "malzeme_adi": "Bool V1",
                "seri_no": "BOOL-V1",
                "stok_birim_sayisi": "1",
                "kullanim_durumu": "aktif",
                "demirbas_mi": "true",
                "demirbas_no": "DMR-BOOL-1",
                "kalibrasyon_gerekli_mi": "false",
                "kutu_kodu": "DLM-KUTU-1",
            },
            {
                "merkezi_sablon": "Bool Variant Template",
                "merkezi_sablondan_olustur": "1",
                "havalimani": "DLM - Dalaman",
                "kategori": "Elektronik",
                "malzeme_adi": "Bool V2",
                "seri_no": "BOOL-V2",
                "stok_birim_sayisi": "1",
                "kullanim_durumu": "aktif",
                "demirbas_mi": "Hayır",
                "kalibrasyon_gerekli_mi": "x",
                "kutu_kodu": "DLM-KUTU-1",
            },
        ],
        lists_context={
            "templates": ["Bool Variant Template"],
            "airports": ["DLM - Dalaman"],
            "categories": ["Elektronik"],
            "statuses": ["aktif", "pasif"],
            "maintenance_forms": [],
            "month_values": list(range(1, 13)),
            "boxes": ["DLM-KUTU-1"],
        },
    )
    response = client.post(
        "/malzeme-ekle/excel-yukle",
        data={"excel_file": (payload, "import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200
    v1 = InventoryAsset.query.filter_by(serial_no="BOOL-V1").first()
    v2 = InventoryAsset.query.filter_by(serial_no="BOOL-V2").first()
    assert v1 is not None and v2 is not None
    assert v1.is_demirbas is True and v1.calibration_required is False
    assert v2.is_demirbas is False and v2.calibration_required is True


def test_excel_import_uppercases_ad_soyad_in_real_flow(client, app):
    owner = KullaniciFactory(rol="sahip")
    airport = HavalimaniFactory(kodu="ADB", ad="Adnan Menderes")
    box = KutuFactory(kodu="ADB-KUTU-1", havalimani=airport)
    template = EquipmentTemplateFactory(name="Uppercase Template", category="Elektronik")
    db.session.add_all([owner, airport, box, template])
    db.session.commit()
    _login(client, owner)

    payload = _build_excel_file(
        [
            {
                "merkezi_sablon": "Uppercase Template",
                "merkezi_sablondan_olustur": "Evet",
                "havalimani": "ADB - Adnan Menderes",
                "kategori": "Elektronik",
                "malzeme_adi": "Uppercase Test",
                "seri_no": "UPPER-001",
                "stok_birim_sayisi": "1",
                "kullanim_durumu": "aktif",
                "kutu_kodu": "ADB-KUTU-1",
                "ad_soyad": "ışık şahin",
            }
        ],
        lists_context={
            "templates": ["Uppercase Template"],
            "airports": ["ADB - Adnan Menderes"],
            "categories": ["Elektronik"],
            "statuses": ["aktif", "pasif"],
            "maintenance_forms": [],
            "month_values": list(range(1, 13)),
            "boxes": ["ADB-KUTU-1"],
        },
    )
    response = client.post(
        "/malzeme-ekle/excel-yukle",
        data={"excel_file": (payload, "import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200
    asset = InventoryAsset.query.filter_by(serial_no="UPPER-001").first()
    assert asset is not None
    assert "IŞIK ŞAHİN" in (asset.notes or "")


def test_excel_import_partial_failure_records_job_rows(client, app):
    owner = KullaniciFactory(rol="sahip")
    airport = HavalimaniFactory(kodu="AYT", ad="Antalya")
    box = KutuFactory(kodu="AYT-KUTU-1", havalimani=airport)
    template = EquipmentTemplateFactory(name="Termal Kamera", category="Elektronik")
    db.session.add_all([owner, airport, box, template])
    db.session.commit()
    _login(client, owner)

    payload = _build_excel_file(
        [
            {
                "merkezi_sablon": "Termal Kamera",
                "merkezi_sablondan_olustur": "Evet",
                "havalimani": "AYT - Antalya",
                "kategori": "Elektronik",
                "malzeme_adi": "Termal Kamera",
                "seri_no": "TR-EXCEL-OK",
                "stok_birim_sayisi": "1",
                "kullanim_durumu": "aktif",
                "kutu_kodu": "AYT-KUTU-1",
            },
            {
                "merkezi_sablon": "Termal Kamera",
                "merkezi_sablondan_olustur": "Evet",
                "havalimani": "AYT - Antalya",
                "kategori": "Bilinmeyen",
                "malzeme_adi": "Hatalı Satır",
                "seri_no": "TR-EXCEL-FAIL",
                "stok_birim_sayisi": "1",
                "kullanim_durumu": "aktif",
                "kutu_kodu": "AYT-KUTU-1",
            },
        ],
        lists_context={
            "templates": ["Termal Kamera"],
            "airports": ["AYT - Antalya"],
            "categories": ["Elektronik"],
            "statuses": ["aktif", "pasif"],
            "maintenance_forms": [],
            "month_values": list(range(1, 13)),
            "boxes": ["AYT-KUTU-1"],
        },
    )

    response = client.post(
        "/malzeme-ekle/excel-yukle",
        data={"excel_file": (payload, "import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200
    job = InventoryBulkImportJob.query.order_by(InventoryBulkImportJob.id.desc()).first()
    assert job is not None
    assert job.total_rows == 2
    assert job.success_rows == 1
    assert job.failed_rows == 1


def test_bulk_import_assigns_code_and_qr_for_each_success_row(client, app):
    owner = KullaniciFactory(rol="sahip")
    airport = HavalimaniFactory(kodu="BJV", ad="Bodrum")
    box = KutuFactory(kodu="BJV-KUTU-1", havalimani=airport)
    template = EquipmentTemplateFactory(name="Bulk QR", category="Elektronik")
    db.session.add_all([owner, airport, box, template])
    db.session.commit()
    _login(client, owner)

    payload = _build_excel_file(
        [
            {
                "merkezi_sablon": "Bulk QR",
                "merkezi_sablondan_olustur": "Evet",
                "havalimani": "BJV - Bodrum",
                "kategori": "Elektronik",
                "malzeme_adi": "Bulk A",
                "seri_no": "BULK-QR-1",
                "stok_birim_sayisi": "1",
                "kullanim_durumu": "aktif",
                "kutu_kodu": "BJV-KUTU-1",
            },
            {
                "merkezi_sablon": "Bulk QR",
                "merkezi_sablondan_olustur": "Evet",
                "havalimani": "BJV - Bodrum",
                "kategori": "Elektronik",
                "malzeme_adi": "Bulk B",
                "seri_no": "BULK-QR-2",
                "stok_birim_sayisi": "1",
                "kullanim_durumu": "aktif",
                "kutu_kodu": "BJV-KUTU-1",
            },
        ],
        lists_context={
            "templates": ["Bulk QR"],
            "airports": ["BJV - Bodrum"],
            "categories": ["Elektronik"],
            "statuses": ["aktif", "pasif"],
            "maintenance_forms": [],
            "month_values": list(range(1, 13)),
            "boxes": ["BJV-KUTU-1"],
        },
    )
    response = client.post(
        "/malzeme-ekle/excel-yukle",
        data={"excel_file": (payload, "import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200
    assets = InventoryAsset.query.filter(InventoryAsset.serial_no.in_(["BULK-QR-1", "BULK-QR-2"])).all()
    assert len(assets) == 2
    for asset in assets:
        assert (asset.asset_code or "").startswith("ARFF-SAR-")
        assert (asset.qr_code or "").startswith("http")


def test_inventory_code_uniqueness_under_same_input(client, app):
    owner = KullaniciFactory(rol="sahip")
    airport = HavalimaniFactory(kodu="ERC", ad="Erzincan")
    box = KutuFactory(kodu="ERC-KUTU-1", havalimani=airport)
    template = EquipmentTemplateFactory(name="Race Safe", category="Elektronik")
    db.session.add_all([owner, airport, box, template])
    db.session.commit()
    _login(client, owner)

    for serial in ("RACE-001", "RACE-002"):
        response = client.post(
            "/malzeme-ekle",
            data={
                "havalimani_id": airport.id,
                "template_id": template.id,
                "kategori": "Elektronik",
                "ad": "Race Safe Item",
                "seri_no": serial,
                "kutu_id": box.id,
                "stok": 1,
                "durum": "aktif",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200

    assets = InventoryAsset.query.filter(InventoryAsset.serial_no.in_(["RACE-001", "RACE-002"])).order_by(InventoryAsset.id.asc()).all()
    assert len(assets) == 2
    assert assets[0].asset_code != assets[1].asset_code
    assert assets[0].qr_code != assets[1].qr_code


def test_excel_import_rejects_unauthorized_airport_row(client, app):
    manager_airport = HavalimaniFactory(kodu="ESB", ad="Esenboğa")
    other_airport = HavalimaniFactory(kodu="SAW", ad="Sabiha Gökçen")
    manager = KullaniciFactory(rol="yetkili", havalimani=manager_airport)
    box = KutuFactory(kodu="ESB-KUTU-4", havalimani=manager_airport)
    template = EquipmentTemplateFactory(name="Pompa", category="Mekanik")
    db.session.add_all([manager_airport, other_airport, manager, box, template])
    db.session.commit()
    _login(client, manager)

    payload = _build_excel_file(
        [
            {
                "merkezi_sablon": "Pompa",
                "merkezi_sablondan_olustur": "Evet",
                "havalimani": "SAW - Sabiha Gökçen",
                "kategori": "Mekanik",
                "malzeme_adi": "Pompa",
                "seri_no": "TR-EXCEL-UNAUTH",
                "stok_birim_sayisi": "1",
                "kullanim_durumu": "aktif",
                "kutu_kodu": "ESB-KUTU-4",
            }
        ],
        lists_context={
            "templates": ["Pompa"],
            "airports": ["SAW - Sabiha Gökçen", "ESB - Esenboğa"],
            "categories": ["Mekanik"],
            "statuses": ["aktif", "pasif"],
            "maintenance_forms": [],
            "month_values": list(range(1, 13)),
            "boxes": ["ESB-KUTU-4"],
        },
    )

    response = client.post(
        "/malzeme-ekle/excel-yukle",
        data={"excel_file": (payload, "import.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert InventoryAsset.query.filter_by(serial_no="TR-EXCEL-UNAUTH").first() is None
    job = InventoryBulkImportJob.query.order_by(InventoryBulkImportJob.id.desc()).first()
    assert job.failed_rows == 1


def test_category_create_requires_system_owner(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    airport = HavalimaniFactory(kodu="TST")
    manager = KullaniciFactory(rol="yetkili", havalimani=airport)
    db.session.add_all([airport, manager])
    db.session.commit()

    manager_client = app.test_client()
    _login(manager_client, manager)
    forbidden = manager_client.post("/envanter/kategori-ekle", data={"name": "Özel Kategori"})
    assert forbidden.status_code == 403

    assert InventoryCategory.query.filter_by(name="Özel Kategori").first() is None


def test_central_template_create_permission(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    airport = HavalimaniFactory(kodu="ESB")
    manager = KullaniciFactory(rol="yetkili", havalimani=airport)
    db.session.add_all([airport, manager])
    db.session.commit()

    manager_client = app.test_client()
    _login(manager_client, manager)
    forbidden = manager_client.post(
        "/envanter/merkezi-sablon-ekle",
        data={"name": "Yetkisiz Şablon", "category": "Elektronik"},
    )
    assert forbidden.status_code == 403

    assert EquipmentTemplate.query.filter_by(name="Yetkisiz Şablon").first() is None


def test_category_create_blocks_turkish_variant_duplicates(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    owner = KullaniciFactory(rol="sahip", is_deleted=False)
    category = InventoryCategory(name="Özel Kategori", is_active=True, is_deleted=False, created_by_user_id=owner.id)
    db.session.add_all([owner, category])
    db.session.commit()

    _login(client, owner)
    response = client.post(
        "/envanter/kategori-ekle",
        data={"name": "Ozel Kategori", "description": "varyant"},
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Bu kategori zaten mevcut." in response.data.decode("utf-8")
    with app.app_context():
        assert InventoryCategory.query.filter_by(is_deleted=False).count() == 1


def test_turkish_upper_helper_and_bool_normalizer():
    assert turkish_upper("ibrahim çelik") == "İBRAHİM ÇELİK"
    assert turkish_upper("ışık şahin") == "IŞIK ŞAHİN"
    assert parse_flexible_bool("x") is True
    assert parse_flexible_bool("0") is False


def test_lookup_normalizer_matches_turkish_character_variants():
    assert normalize_lookup_key("İzmir") == normalize_lookup_key("izmir") == normalize_lookup_key("İZMİR")
    assert normalize_lookup_key("İzmir") == normalize_lookup_key("Izmir") == normalize_lookup_key("ızmir")
    assert normalize_lookup_key("Işık") == normalize_lookup_key("ışık") == normalize_lookup_key("ISIK")
    assert normalize_lookup_key("Şule") == normalize_lookup_key("sule")
    assert normalize_lookup_key("Çağrı") == normalize_lookup_key("cagri")
    assert normalize_lookup_key("Öztürk") == normalize_lookup_key("ozturk")
    assert normalize_lookup_key("Ünal") == normalize_lookup_key("unal")
    assert normalize_lookup_key("Çelik") == normalize_lookup_key("celik")
