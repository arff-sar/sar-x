from datetime import datetime, timezone
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy.exc import OperationalError

from extensions import db
from models import IslemLog, LoginVisualChallenge
from tests.factories import HavalimaniFactory, KullaniciFactory


def _login(client, user_id):
    with client.session_transaction() as session:
        session.clear()
        session["_user_id"] = str(user_id)
        session["_fresh"] = True


def _captcha_answer(client, app):
    client.get("/login")
    with client.session_transaction() as session:
        token = session.get("login_visual_captcha_token")
    assert token
    with app.app_context():
        challenge = LoginVisualChallenge.query.filter_by(token=token, invalidated_at=None).first()
        if challenge:
            return challenge.code
        fallback_store = app.extensions.get("login_visual_challenge_store", {})
        return fallback_store[token]["code"]


def test_captcha_failure_only_shows_safe_message_and_code(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    response = client.post(
        "/login",
        data={
            "kullanici_adi": "unknown@sarx.com",
            "sifre": "wrong",
            "security_verification": "hata",
        },
        follow_redirects=True,
    )
    html = response.data.decode("utf-8")

    assert response.status_code == 400
    assert "SAR-X-AUTH-1202" in html
    assert "Güvenlik doğrulaması başarısız oldu." in html
    assert "Traceback" not in html
    assert "Login captcha verification failed" not in html


def test_password_reset_mail_failure_returns_safe_error_code(client, app):
    app.config["WTF_CSRF_ENABLED"] = False
    user = KullaniciFactory(kullanici_adi="mailer@sarx.com", is_deleted=False)
    db.session.add(user)
    db.session.commit()

    with patch("routes.auth.mail_gonder", return_value=False):
        response = client.post(
            "/sifre-sifirla-talep",
            data={"kullanici_adi": "mailer@sarx.com"},
            follow_redirects=True,
        )

    html = response.data.decode("utf-8")
    assert response.status_code == 200
    assert "SAR-X-MAIL-4101" in html
    assert "Şifre sıfırlama isteği şu an gönderilemedi." in html


def test_owner_can_view_error_log_detail(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzurum Havalimanı", kodu="ERZ")
        owner = KullaniciFactory(rol="sahip", kullanici_adi="owner-errors@sarx.com", is_deleted=False, havalimani=airport)
        db.session.add_all([airport, owner])
        db.session.flush()
        log = IslemLog(
            kullanici_id=owner.id,
            islem_tipi="Sistem",
            detay="İşlem tamamlanamadı. Hata kodu: SAR-X-SYSTEM-5101",
            outcome="failed",
            error_code="SAR-X-SYSTEM-5101",
            title="Beklenmeyen Sunucu Hatası",
            user_message="İşlem tamamlanamadı.",
            owner_message="Beklenmeyen exception global fallback tarafından yakalandı.",
            module="SYSTEM",
            severity="critical",
            exception_type="RuntimeError",
            exception_message="boom",
            traceback_summary="Traceback summary",
            route="/test",
            method="GET",
            request_id="sarx-owner-detail",
            user_email=owner.kullanici_adi,
        )
        db.session.add(log)
        db.session.commit()
        owner_id = owner.id
        log_id = log.id

    _login(client, owner_id)
    listing = client.get("/hata-kayitlari")
    detail = client.get(f"/hata-kayitlari/{log_id}")
    detail_html = detail.data.decode("utf-8")

    assert listing.status_code == 200
    assert "Hata Kayıtları" in listing.data.decode("utf-8")
    assert "SAR-X-SYSTEM-5101" in listing.data.decode("utf-8")
    assert detail.status_code == 200
    assert "Traceback summary" in detail_html
    assert "RuntimeError" in detail_html


def test_non_owner_cannot_view_error_log_detail(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Samsun Havalimanı", kodu="SZF")
        readonly = KullaniciFactory(rol="readonly", kullanici_adi="readonly-errors@sarx.com", is_deleted=False, havalimani=airport)
        db.session.add_all([airport, readonly])
        db.session.flush()
        log = IslemLog(
            kullanici_id=readonly.id,
            islem_tipi="Sistem",
            detay="İşlem tamamlanamadı. Hata kodu: SAR-X-SYSTEM-5101",
            outcome="failed",
            error_code="SAR-X-SYSTEM-5101",
            title="Beklenmeyen Sunucu Hatası",
            user_message="İşlem tamamlanamadı.",
            owner_message="Beklenmeyen exception global fallback tarafından yakalandı.",
            module="SYSTEM",
            severity="critical",
            traceback_summary="Sensitive traceback summary",
            route="/admin/test",
            method="GET",
            request_id="sarx-no-detail",
        )
        db.session.add(log)
        db.session.commit()
        readonly_id = readonly.id
        log_id = log.id

    _login(client, readonly_id)
    listing = client.get("/hata-kayitlari")
    detail = client.get(f"/hata-kayitlari/{log_id}")
    listing_html = listing.data.decode("utf-8")
    detail_html = detail.data.decode("utf-8")

    assert listing.status_code == 403
    assert "SAR-X-ADMIN-6101" in listing_html
    assert "SAR-X-SYSTEM-5101" not in listing_html
    assert detail.status_code == 403
    assert "SAR-X-ADMIN-6101" in detail_html
    assert "Sensitive traceback summary" not in detail_html


def test_owner_error_log_listing_gracefully_handles_query_failure(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Erzincan Havalimanı", kodu="ERC")
        owner = KullaniciFactory(rol="sahip", kullanici_adi="owner-errors-safe@sarx.com", is_deleted=False, havalimani=airport)
        db.session.add_all([airport, owner])
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    with patch("routes.admin.logs._load_error_log_listing_data", side_effect=RuntimeError("listing exploded")):
        response = client.get("/hata-kayitlari")

    html = response.data.decode("utf-8")
    assert response.status_code == 200
    assert "Hata Kayıtları" in html
    assert "Hata kaydı bulunamadı" in html


def test_public_unexpected_error_fallback_hides_technical_details(client, app):
    app.config["PROPAGATE_EXCEPTIONS"] = False

    def boom():
        raise RuntimeError("postgres://user:secret@db.example.com/app failed")

    app.add_url_rule("/__test-boom", "test_boom", boom)
    response = client.get("/__test-boom")
    html = response.data.decode("utf-8")

    assert response.status_code == 500
    assert "SAR-X-SYSTEM-5101" in html
    assert "İşlem tamamlanamadı." in html
    assert "secret" not in html
    assert "RuntimeError" not in html
    assert "postgres://" not in html


def test_db_connection_error_returns_safe_message(client, app):
    app.config["PROPAGATE_EXCEPTIONS"] = False

    def boom_db():
        raise OperationalError(
            "SELECT 1",
            {},
            Exception("could not connect to server with postgres://user:secret@db.example.com/app"),
        )

    app.add_url_rule("/__test-db-boom", "test_db_boom", boom_db)
    response = client.get("/__test-db-boom")
    html = response.data.decode("utf-8")

    assert response.status_code == 503
    assert "SAR-X-DB-2101" in html
    assert "Sistem bağlantı hatası oluştu." in html
    assert "secret" not in html
    assert "postgres://" not in html


def test_error_listing_uses_tr_module_labels_summary_and_timezone(client, app):
    with app.app_context():
        airport = HavalimaniFactory(ad="Adana Havalimanı", kodu="ADA")
        owner = KullaniciFactory(rol="sahip", kullanici_adi="owner-errors-tr@sarx.com", is_deleted=False, havalimani=airport)
        team_member = KullaniciFactory(
            rol="ekip_uyesi",
            kullanici_adi="member-errors-tr@sarx.com",
            is_deleted=False,
            havalimani=airport,
        )
        db.session.add_all([airport, owner, team_member])
        db.session.flush()
        db.session.add(
            IslemLog(
                kullanici_id=team_member.id,
                islem_tipi="Sistem",
                detay="Login akışı tamamlanamadı.",
                outcome="failed",
                error_code="SAR-X-AUTH-1202",
                title="Güvenlik Doğrulaması Başarısız",
                user_message="Güvenlik doğrulaması başarısız oldu.",
                owner_message="Captcha doğrulaması geçmedi.",
                module="AUTH",
                severity="warning",
                route="/login/passkey/finish",
                method="POST",
                request_id="sarx-passkey-error",
                zaman=datetime(2026, 3, 31, 9, 15, 0, tzinfo=timezone.utc),
            )
        )
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/hata-kayitlari")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "Kimlik Doğrulama" in html
    assert "31.03.2026" in html
    assert "Tarih bilgisi yok" not in html
    assert "Ekip üyesi hesabında passkey giriş adımı tamamlanamadı." in html


def test_error_logs_pagination_and_excel_export_preserve_filters(client, app):
    with app.app_context():
        owner = KullaniciFactory(rol="sahip", kullanici_adi="owner-errors-pages@sarx.com", is_deleted=False)
        db.session.add(owner)
        db.session.flush()
        db.session.add(
            IslemLog(
                kullanici_id=owner.id,
                islem_tipi="Sistem",
                detay="Filtre dışı hata",
                outcome="failed",
                error_code="SAR-X-DB-2101",
                title="Veritabanı Bağlantı Hatası",
                user_message="Sistem bağlantı hatası oluştu.",
                owner_message="Veritabanı bağlantısı kurulamadı.",
                module="DB",
                severity="critical",
                route="/admin/test",
                method="GET",
                request_id="db-outside",
            )
        )
        for index in range(25):
            db.session.add(
                IslemLog(
                    kullanici_id=owner.id,
                    islem_tipi="Sistem",
                    detay=f"AUTH hata {index}",
                    outcome="failed",
                    error_code="SAR-X-AUTH-1202",
                    title=f"AUTH Hata {index}",
                    user_message="Güvenlik doğrulaması başarısız oldu.",
                    owner_message="Captcha doğrulaması geçmedi.",
                    module="AUTH",
                    severity="warning",
                    route="/login",
                    method="POST",
                    request_id=f"auth-{index}",
                )
            )
        db.session.commit()
        owner_id = owner.id

    _login(client, owner_id)
    response = client.get("/hata-kayitlari?module=AUTH&page=2")
    html = response.data.decode("utf-8")

    assert response.status_code == 200
    assert "AUTH Hata 24" not in html
    assert "AUTH Hata 4" in html
    assert "Filtre dışı hata" not in html
    assert "module=AUTH" in html
    assert "Sayfa 2 / 2" in html

    export = client.get("/hata-kayitlari/excel?module=AUTH")
    workbook = load_workbook(filename=BytesIO(export.data))
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    assert export.status_code == 200
    assert headers == ["Durum", "Tarih", "Modül", "Hata Kodu", "Başlık", "Kısa Açıklama", "Kullanıcı", "Sayfa", "Request ID"]
    assert len(rows) == 25
    assert all(row[2] == "Kimlik Doğrulama" for row in rows)
